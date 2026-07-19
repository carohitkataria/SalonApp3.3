from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Body, Request, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response, HTMLResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import socketio
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
from enum import Enum
import uuid
from datetime import datetime, timezone, time, timedelta, date
import calendar
from apscheduler.schedulers.asyncio import AsyncIOScheduler
import io
import base64
import qrcode
import jwt
from passlib.context import CryptContext
import random
import secrets  # Secure random number generation
import math
import json
import httpx

# Import Twilio service
from twilio_service import (
    send_whatsapp_otp, 
    send_whatsapp_notification,
    send_booking_confirmation_template,
    send_booking_completed_template,
    send_your_turn_now_template,
    send_token_approaching_template,
    verify_whatsapp_otp,
    is_verify_configured,
    format_booking_confirmation,
    format_queue_status,
    format_token_near,
    format_token_called,
    format_token_cancelled,
    format_token_rescheduled,
    format_salon_calling,
)

# Import invoice service
from invoice_service import generate_invoice_pdf, save_invoice_pdf

# Import platform admin module (Phase 1 — Part A)
import platform_admin as platform_admin_mod

# Import platform admin management module (Phase 5 — Part A)
import platform_admin_management as platform_admin_mgmt_mod

# Import discount codes module (Phase 4 — Part D)
import discount_codes as discount_codes_mod

# Phase 8 — Supplier auth/signup
import supplier_auth as supplier_auth_mod

# Phase 9 — Supplier products/dashboard
import supplier_products as supplier_products_mod

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Socket.IO setup
sio = socketio.AsyncServer(async_mode='asgi', cors_allowed_origins='*')

# Create the main FastAPI app
fastapi_app = FastAPI()
api_router = APIRouter(prefix="/api")

# Socket.IO app wraps FastAPI app
socket_app = socketio.ASGIApp(sio, fastapi_app)

# Export socket_app as the main app for ASGI server
# This prevents recursion issues during deployment
app = socket_app

# ============ MODELS ============

# Salon Models
class SalonCreate(BaseModel):
    salon_name: str
    owner_name: str
    phone: str
    email: Optional[str] = None
    address: str
    city: Optional[str] = None  # City for filtering
    latitude: float
    longitude: float
    upi_id: Optional[str] = None
    payment_timing: str = "after"  # before/after
    password: Optional[str] = None  # Optional password for login
    gender_tag: str = "Unisex"  # Unisex/Men/Women

class SalonUpdate(BaseModel):
    model_config = ConfigDict(extra="allow")
    salon_name: Optional[str] = None
    owner_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None  # City for filtering
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    upi_id: Optional[str] = None
    payment_timing: Optional[str] = None
    is_gst_registered: Optional[bool] = None
    gstin: Optional[str] = None
    logo_url: Optional[str] = None
    photo_gallery: Optional[List[str]] = None
    tax_rate: Optional[float] = None
    invoice_prefix: Optional[str] = None
    invoice_start_number: Optional[int] = None
    password: Optional[str] = None  # To update password
    attendance_rules: Optional[Dict[str, Any]] = None  # Module 3 — Staff Settings attendance config
    attendance_mode: Optional[str] = None  # Module 4 — direct mode set (rare; usually via /attendance-mode)
    geo_settings: Optional[Dict[str, Any]] = None  # Module 4 — geo check-in config

class SalonPasswordLogin(BaseModel):
    phone: str
    password: str

class DayOperationalHours(BaseModel):
    """Operational hours for a specific day"""
    is_holiday: bool = False
    opening_time: Optional[str] = "09:00"  # Format: "HH:MM"
    closing_time: Optional[str] = "20:00"  # Format: "HH:MM"
    lunch_start: Optional[str] = None  # Format: "HH:MM"
    lunch_end: Optional[str] = None  # Format: "HH:MM"

class OperationalHours(BaseModel):
    """Weekly operational hours"""
    monday: DayOperationalHours = Field(default_factory=DayOperationalHours)
    tuesday: DayOperationalHours = Field(default_factory=DayOperationalHours)
    wednesday: DayOperationalHours = Field(default_factory=DayOperationalHours)
    thursday: DayOperationalHours = Field(default_factory=DayOperationalHours)
    friday: DayOperationalHours = Field(default_factory=DayOperationalHours)
    saturday: DayOperationalHours = Field(default_factory=DayOperationalHours)
    sunday: DayOperationalHours = Field(default_factory=DayOperationalHours)

class ManualToggle(BaseModel):
    """Manual open/close override.

    closed_mode values:
      - None / not set: not closed (is_open=True)
      - 'full': salon is fully closed (online + offline). No bookings via any channel.
      - 'online_only': salon is closed for online bookings; walk-in / QR / salon-side
        manual bookings still work.
    """
    is_overridden: bool = False
    is_open: bool = True
    closed_mode: Optional[str] = None  # 'full' | 'online_only' | None
    overridden_at: Optional[str] = None  # ISO timestamp

class Salon(BaseModel):
    model_config = ConfigDict(extra="allow")
    id: str
    salon_name: str
    owner_name: str
    phone: str
    email: Optional[str] = None
    address: str
    city: Optional[str] = None  # City for filtering
    latitude: float
    longitude: float
    upi_id: Optional[str] = None
    payment_timing: str
    is_active: bool = True
    is_gst_registered: bool = False
    gstin: Optional[str] = None
    logo_url: Optional[str] = None
    photo_gallery: List[str] = []  # Array of image URLs
    rating: float = 0  # Average rating out of 5
    total_reviews: int = 0  # Total number of reviews
    gender_tag: Optional[str] = "Unisex"  # Unisex/Men/Women
    tax_rate: float = 2.5  # Default GST rate (CGST + SGST = 5%)
    invoice_prefix: str = "INV"  # Invoice number prefix
    invoice_start_number: int = 1  # Starting invoice number
    current_invoice_number: int = 1  # Current invoice counter
    operational_hours: Optional[OperationalHours] = None
    manual_toggle: Optional[ManualToggle] = None
    attendance_rules: Optional[Dict[str, Any]] = None  # Module 3 — Staff Settings (legacy placement config)
    attendance_mode: Optional[str] = "service_completion"  # Module 4 — "service_completion" | "geo_checkin"
    attendance_mode_history: Optional[List[Dict[str, Any]]] = None  # Module 4 — change audit
    geo_settings: Optional[Dict[str, Any]] = None  # Module 4 — Mode B configuration
    created_at: str

# Service Models
class ServiceCreate(BaseModel):
    service_name: str
    description: Optional[str] = None
    category: str = "General"  # Category for grouping
    sub_category: Optional[str] = None  # Fine-grained bucket under the category
    gender_tag: str = "Unisex"  # Men/Women/Unisex
    default_duration: int = 30  # minutes
    base_price: float = 0
    price_type: str = "fixed"  # fixed/onwards
    images: List[str] = []  # List of image URLs
    thumbnail_url: Optional[str] = None  # Circular thumbnail for category display
    is_favorite: bool = False
    is_enabled: bool = True  # Salon can enable/disable
    available_at_home: bool = False  # Can be delivered at home
    # Item 4 — At-home settings (effective only when available_at_home=True)
    home_price: Optional[float] = None         # ₹ when delivered at home (else falls back to base_price)
    home_min_order_value: Optional[float] = None  # ₹ MOQ on the cart before at-home checkout is allowed
    home_min_items: Optional[int] = None       # Minimum number of services for at-home
    home_travel_fee: Optional[float] = None    # ₹ flat travel fee added at checkout
    home_service_radius_km: Optional[float] = None  # Service area radius

class ServiceUpdate(BaseModel):
    service_name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    sub_category: Optional[str] = None
    gender_tag: Optional[str] = None
    default_duration: Optional[int] = None
    base_price: Optional[float] = None
    price_type: Optional[str] = None
    images: Optional[List[str]] = None
    thumbnail_url: Optional[str] = None
    is_favorite: Optional[bool] = None
    favorite_order: Optional[int] = None
    is_enabled: Optional[bool] = None
    available_at_home: Optional[bool] = None
    home_price: Optional[float] = None
    home_min_order_value: Optional[float] = None
    home_min_items: Optional[int] = None
    home_travel_fee: Optional[float] = None
    home_service_radius_km: Optional[float] = None

class Service(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    service_name: str
    description: Optional[str] = None
    category: str = "Services"
    sub_category: Optional[str] = None
    gender_tag: str = "Unisex"
    default_duration: int
    base_price: float
    price_type: str = "fixed"
    images: List[str] = []
    thumbnail_url: Optional[str] = None
    is_favorite: bool = False
    favorite_order: Optional[int] = None
    is_active: bool = True
    is_enabled: bool = True
    available_at_home: bool = False
    home_price: Optional[float] = None
    home_min_order_value: Optional[float] = None
    home_min_items: Optional[int] = None
    home_travel_fee: Optional[float] = None
    home_service_radius_km: Optional[float] = None

# Package Models
class PackageService(BaseModel):
    service_id: str
    service_name: str

class PackageCreate(BaseModel):
    package_name: str
    description: Optional[str] = None
    service_ids: List[str]
    total_price: float
    image_url: Optional[str] = None
    gender_tag: str = "Unisex"

class Package(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    package_name: str
    description: Optional[str] = None
    service_ids: List[str]
    total_price: float
    image_url: Optional[str] = None
    gender_tag: str = "Unisex"
    is_active: bool = True
    created_at: str

# Salon-Specific Package Models (for customized packages per salon)
class SalonPackageCreate(BaseModel):
    salon_id: str
    package_name: str
    description: Optional[str] = None
    service_ids: List[str]
    total_price: float
    image_url: Optional[str] = None
    gender_tag: str = "Unisex"

class SalonPackageUpdate(BaseModel):
    package_name: Optional[str] = None
    description: Optional[str] = None
    service_ids: Optional[List[str]] = None
    total_price: Optional[float] = None
    image_url: Optional[str] = None
    gender_tag: Optional[str] = None
    is_active: Optional[bool] = None

class SalonPackage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    package_name: str
    description: Optional[str] = None
    service_ids: List[str]
    total_price: float
    image_url: Optional[str] = None
    gender_tag: str = "Unisex"
    is_active: bool = True
    created_at: str

# Staff Models (formerly Barber - keeping backward compatibility)
class BarberCreate(BaseModel):
    name: str
    salon_id: str
    branch_id: Optional[str] = None  # Branch this staff is assigned to
    experience: int
    category: str
    specialization: Optional[str] = None
    gender_specialization: Optional[str] = None  # Men/Women/Unisex/Kids
    mobile: str
    profile_image: Optional[str] = None
    is_barber: bool = True  # True if staff is a barber (visible to customers)
    # New employee fields
    department: Optional[str] = None  # e.g., "Hairstyling", "Spa", "Reception"
    designation: Optional[str] = None  # e.g., "Senior Stylist", "Receptionist"
    emergency_contact: Optional[str] = None
    aadhar_number: Optional[str] = None
    doj: Optional[str] = None  # Date of Joining (YYYY-MM-DD)
    dob: Optional[str] = None  # Date of Birth
    last_working_date: Optional[str] = None  # YYYY-MM-DD; barber stops being visible to customers after this date
    leave_dates: Optional[List[str]] = None  # YYYY-MM-DD list of leave days (incl. future)
    compensation: Optional[float] = None
    documents: Optional[List[str]] = None  # URLs to uploaded documents

class BarberUpdate(BaseModel):
    name: Optional[str] = None
    branch_id: Optional[str] = None  # For transfers / assignment changes
    experience: Optional[int] = None
    category: Optional[str] = None
    specialization: Optional[str] = None
    gender_specialization: Optional[str] = None  # Men/Women/Unisex/Kids
    mobile: Optional[str] = None
    queue_status: Optional[str] = None
    profile_image: Optional[str] = None
    photo_url: Optional[str] = None
    on_leave: Optional[bool] = None
    intro: Optional[str] = None
    gallery: Optional[List[str]] = None
    is_barber: Optional[bool] = None  # Can update barber visibility
    # New employee fields
    department: Optional[str] = None
    designation: Optional[str] = None
    emergency_contact: Optional[str] = None
    aadhar_number: Optional[str] = None
    doj: Optional[str] = None
    dob: Optional[str] = None
    last_working_date: Optional[str] = None
    leave_dates: Optional[List[str]] = None
    compensation: Optional[float] = None
    documents: Optional[List[str]] = None

class Barber(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    salon_id: str
    branch_id: Optional[str] = None
    experience: int
    category: str
    specialization: Optional[str] = None
    gender_specialization: Optional[str] = None  # Men/Women/Unisex/Kids
    mobile: str
    profile_image: Optional[str] = None
    photo_url: Optional[str] = None  # Alias for profile_image
    queue_status: str = "available"  # available/busy/offline
    on_leave: bool = False  # Legacy global "on leave today" flag (kept for backward compat)
    is_active: bool = True
    is_barber: bool = True  # True if staff is a barber (visible to customers)
    intro: Optional[str] = None  # Staff's story/about
    gallery: List[str] = []  # Portfolio images
    rating: float = 4.5  # Average rating
    total_reviews: int = 0  # Total number of reviews
    # New employee fields
    department: Optional[str] = None
    designation: Optional[str] = None
    emergency_contact: Optional[str] = None
    aadhar_number: Optional[str] = None
    doj: Optional[str] = None  # Date of Joining
    dob: Optional[str] = None  # Date of Birth
    last_working_date: Optional[str] = None  # If set + today > this date, barber is hidden from customers
    leave_dates: List[str] = []  # YYYY-MM-DD list; barber on leave on those days
    compensation: Optional[float] = None
    documents: List[str] = []  # URLs to uploaded documents (legacy)
    staff_documents: List[Dict[str, Any]] = []  # Structured docs: {id, doc_type, label, file_data, mime_type, file_name, size_kb, uploaded_at}
    # Customer-view-only flag: True when this barber is on leave for the requested date.
    # Set by GET /salons/{id}/barbers?customer_view=true&date=YYYY-MM-DD so the UI
    # can render the staff card greyed-out with an "On Leave" tag.
    is_on_leave: Optional[bool] = None

class BarberServicePrice(BaseModel):
    barber_id: str
    service_id: str
    price: float
    is_available: bool = True

class BarberServiceAssignment(BaseModel):
    service_id: str
    price: float
    is_available: bool = True

# Auth Models
class SalonOTPRequest(BaseModel):
    phone: str

class SalonOTPVerify(BaseModel):
    phone: str
    otp: str

class SalonLogin(BaseModel):
    phone: str
    password: Optional[str] = None

class SalonToken(BaseModel):
    access_token: str
    token_type: str = "bearer"
    salon_id: str

# User Models
class UserLogin(BaseModel):
    name: str
    phone: str
    gender: Optional[str] = None  # Men/Women

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    gender: Optional[str] = None  # Men/Women
    dob: Optional[str] = None  # YYYY-MM-DD
    email: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    pincode: Optional[str] = None
    # Marketing / customer master fields (M1)
    wedding_anniversary: Optional[str] = None  # YYYY-MM-DD
    spouse_name: Optional[str] = None
    spouse_date_of_birth: Optional[str] = None  # YYYY-MM-DD
    important_dates: Optional[List[Dict[str, Any]]] = None  # [{label, date}, ...]
    is_otp_verified: Optional[bool] = False  # OTP verification status
    otp_verified_at: Optional[str] = None  # When OTP was verified
    # Whether the user has set a password for password-based login. We never
    # return the hash; just a boolean flag so the frontend can show the
    # right CTA ("Set password" vs "Reset password").
    has_password: Optional[bool] = False
    created_at: str


class CustomerOTPRequest(BaseModel):
    phone: str

class CustomerOTPVerify(BaseModel):
    phone: str
    otp: str


# ============ Module 8 — Customer password-login & long-lived sessions ============

class CustomerCheckAccountIn(BaseModel):
    phone: str


class CustomerSendOtpV2In(BaseModel):
    phone: str
    # Used purely for telemetry + UI hints; backend behaviour is the same for
    # all three purposes. Accepted: "login", "set_password", "reset_password".
    purpose: Optional[str] = "login"


class CustomerVerifyOtpV2In(BaseModel):
    phone: str
    otp: str
    purpose: Optional[str] = "login"


class CustomerLoginPasswordIn(BaseModel):
    phone: str
    password: str


class CustomerSetPasswordIn(BaseModel):
    phone: str
    password: str
    # JWT issued by /verify-otp when purpose ∈ {set_password, reset_password}.
    # Short-lived (10 min) — proves the caller just verified an OTP.
    password_reset_token: str


class UserProfileUpdate(BaseModel):
    name: Optional[str] = None
    gender: Optional[str] = None
    dob: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    pincode: Optional[str] = None
    # Marketing / customer master fields (M1)
    wedding_anniversary: Optional[str] = None
    spouse_name: Optional[str] = None
    spouse_date_of_birth: Optional[str] = None
    important_dates: Optional[List[Dict[str, Any]]] = None

# Token/Booking Models
class BookingCreate(BaseModel):
    salon_id: str
    branch_id: Optional[str] = None  # Defaults to main branch if absent
    user_id: str
    customer_name: str
    phone: str
    date: str
    shift: str  # Morning/Noon/Evening
    time_slot: Optional[str] = None  # Keep for backward compatibility
    barber_id: str  # can be "any"
    selected_services: List[str]
    source: str = "online"
    booking_type: str = "instant"  # instant/future
    booking_for_self: bool = True  # True = self, False = others
    customer_gender: Optional[str] = None  # For guest bookings
    is_guest: bool = False  # True if booking as guest (no auth)
    payment_mode: Optional[str] = None  # cash/upi/wallet/card
    customer_gender: Optional[str] = None

class TokenModel(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    branch_id: Optional[str] = None  # Branch where this booking belongs
    token_number: str  # Global/salon-wide token: M1, M2, N1, E1...
    customer_name: str
    phone: str
    user_id: Optional[str] = None
    date: str
    shift: str  # Morning/Noon/Evening
    time_slot: Optional[str] = None  # Keep for backward compatibility
    barber_id: str
    barber_name: str
    selected_services: List[str]
    total_amount: float
    # Module 7 — per-service barber assignment (additive, backward-compatible)
    # When present, `total_amount` = sum(line_total). When absent/empty, behave
    # as legacy: all services credited to `barber_id`.
    service_assignments: Optional[List[Dict[str, Any]]] = None
    order_discount_percent: Optional[float] = 0.0
    order_discount_amount: Optional[float] = 0.0
    subtotal: Optional[float] = None  # Pre-discount sum of line prices
    # 75%-rule capacity tracking
    total_service_minutes: Optional[int] = None
    blocked_minutes: Optional[int] = None
    status: str = "waiting"  # waiting | called | completed | skipped
    payment_status: str = "pending"
    payment_mode: Optional[str] = None
    payment_confirmed: bool = False  # Salon must confirm payment before completing
    upi_transaction_id: Optional[str] = None
    source: str
    booking_type: str
    booking_for_self: bool = True
    allocated_at: Optional[str] = None
    called_at: Optional[str] = None
    completed_at: Optional[str] = None
    recall_count: int = 0
    invoice_id: Optional[str] = None  # Link to invoice
    # Frictionless booking — snapshot of the booking-time OTP-verification state
    # of the customer. False for "non-OTP-verified" express bookings made by
    # guests/unverified users; true once they verify via OTP. Used by the
    # history UI to surface a small badge.
    is_otp_verified_at_booking: Optional[bool] = False
    created_at: str

class AddServicesRequest(BaseModel):
    service_ids: List[str]  # List of new service IDs to add
    final_amount: Optional[float] = None  # Override final amount (salon can adjust)

# Invoice Model
class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    token_id: str
    salon_id: str
    invoice_number: str
    customer_name: str
    customer_phone: str
    invoice_data: dict  # Complete invoice data including services, amounts, etc.
    pdf_base64: Optional[str] = None  # Base64 encoded PDF
    created_at: str

# Rating/Review Models
class RatingCreate(BaseModel):
    token_id: str  # The completed booking token ID
    barber_id: str
    salon_id: str
    rating: int = Field(..., ge=1, le=5)  # 1-5 stars
    review: str = Field(..., min_length=1)  # Review comment is required

class RatingResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    token_id: str
    user_id: str
    user_name: str
    barber_id: str
    barber_name: str
    salon_id: str
    rating: int
    review: str
    created_at: str

class BarberRatingSummary(BaseModel):
    barber_id: str
    barber_name: str
    average_rating: float
    total_reviews: int
    reviews: List[RatingResponse]

# ============ SALON USER MODELS (Multi-User Access) ============

class SalonUserPermissions(BaseModel):
    # ---- LEGACY flat keys — kept for backward compatibility ----
    can_edit_salon: bool = False
    can_access_analytics: bool = False
    can_access_financials: bool = False
    can_delete_salon: bool = False
    can_access_services: bool = False   # See Services & Offerings section
    can_access_gallery: bool = False    # See Gallery section
    can_access_staff: bool = False      # See Staff Management section (own profile by default)
    can_view_all_staff: bool = False    # When can_access_staff, see ALL staff (not just own)
    can_access_marketing: bool = False  # See Marketing section (campaigns/coupons/rewards)
    # ---- NEW: Granular per-module action-level permissions ----
    # Structure: { "staff": {"view": true, "edit": false, ...}, "financials": {...}, ... }
    # If empty, we fall back to the legacy flat keys above (so old salon_users
    # keep working unchanged). See has_module_permission() for canonical mapping.
    modules: Dict[str, Dict[str, bool]] = {}

class SalonUserCreate(BaseModel):
    salon_id: str
    name: str
    mobile: str
    login_id: str  # Free text login ID
    password: str
    role: str = "staff"  # "admin" | "staff" | "branch_manager"
    staff_id: Optional[str] = None  # Link to staff member in barbers collection
    assigned_branch_ids: Optional[List[str]] = None  # For branch_manager
    permissions: Optional[SalonUserPermissions] = None

class SalonUserUpdate(BaseModel):
    name: Optional[str] = None
    mobile: Optional[str] = None
    login_id: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    staff_id: Optional[str] = None
    assigned_branch_ids: Optional[List[str]] = None
    permissions: Optional[SalonUserPermissions] = None
    status: Optional[str] = None  # active/inactive

class SalonUser(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    name: str
    mobile: str
    login_id: str
    password_hash: str
    role: str  # "admin" | "staff" | "branch_manager"
    staff_id: Optional[str] = None  # Link to staff member
    assigned_branch_ids: List[str] = []  # Branches this user can access (branch_manager)
    permissions: SalonUserPermissions
    status: str = "active"  # active/inactive
    created_at: str

class SalonUserLogin(BaseModel):
    identifier: str  # Can be mobile number or login_id
    password: str

class SalonUserToken(BaseModel):
    access_token: str
    token_type: str = "bearer"
    salon_id: str
    user_id: str
    role: str
    permissions: SalonUserPermissions
    assigned_branch_ids: List[str] = []
    staff_id: Optional[str] = None  # Linked barber id (for self check-in)

# ============ BRANCH MODELS (Multi-Branch / Multi-Location) ============

class BranchCreate(BaseModel):
    branch_name: str
    branch_code: Optional[str] = None  # short code e.g. "BLR-01"
    address: Optional[str] = None
    city: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    branch_manager_staff_id: Optional[str] = None  # link to barbers.id
    is_main_branch: bool = False
    status: str = "active"  # active/inactive
    operational_hours: Optional[OperationalHours] = None

class BranchUpdate(BaseModel):
    branch_name: Optional[str] = None
    branch_code: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    branch_manager_staff_id: Optional[str] = None
    is_main_branch: Optional[bool] = None
    status: Optional[str] = None
    operational_hours: Optional[OperationalHours] = None

class Branch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    branch_name: str
    branch_code: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    branch_manager_staff_id: Optional[str] = None
    is_main_branch: bool = False
    status: str = "active"
    operational_hours: Optional[OperationalHours] = None
    created_at: str

class StaffBranchTransferCreate(BaseModel):
    staff_id: str
    from_branch_id: Optional[str] = None  # null if first assignment
    to_branch_id: str
    transfer_date: str  # YYYY-MM-DD
    remarks: Optional[str] = None

class StaffBranchTransfer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    staff_id: str
    from_branch_id: Optional[str] = None
    to_branch_id: str
    transfer_date: str
    remarks: Optional[str] = None
    created_by: Optional[str] = None
    created_at: str

# ============ SUBSCRIPTION MODELS (SalonHub Pro) ============

class SubscriptionPlanCreate(BaseModel):
    plan_name: str
    price: float
    price_per_branch: Optional[float] = None  # Phase 2 — Part C
    billing_cycle: str = "monthly"  # monthly | yearly
    max_staff: Optional[int] = None
    max_branches: Optional[int] = None
    features: List[str] = []
    status: str = "active"  # active | inactive

class SubscriptionPlanUpdate(BaseModel):
    plan_name: Optional[str] = None
    price: Optional[float] = None
    price_per_branch: Optional[float] = None  # Phase 2 — Part C
    billing_cycle: Optional[str] = None
    max_staff: Optional[int] = None
    max_branches: Optional[int] = None
    features: Optional[List[str]] = None
    status: Optional[str] = None

class SubscriptionPlan(BaseModel):
    id: str
    plan_name: str
    price: float
    price_per_branch: Optional[float] = None  # Phase 2 (Part C) — per-branch pricing
    billing_cycle: str
    max_staff: Optional[int] = None
    max_branches: Optional[int] = None
    features: List[str] = []
    status: str
    created_at: str

class SalonSubscription(BaseModel):
    id: str
    salon_id: str
    plan_id: str
    plan_name: Optional[str] = None
    price: Optional[float] = None
    subscription_status: str  # active | expired | cancelled | payment_failed | pending
    start_date: Optional[str] = None
    expiry_date: Optional[str] = None
    payment_status: str  # paid | pending | failed
    cashfree_order_id: Optional[str] = None
    cashfree_payment_id: Optional[str] = None
    auto_renew: bool = False
    # Phase 2 (Part C) — per-branch pricing snapshot fields
    billable_branch_count: Optional[int] = None
    price_per_branch_snapshot: Optional[float] = None
    branch_ids_snapshot: Optional[List[str]] = None
    base_amount: Optional[float] = None
    discount_code_applied: Optional[str] = None
    discount_amount: Optional[float] = None
    total_amount: Optional[float] = None
    created_at: str
    updated_at: Optional[str] = None

class PaymentTransaction(BaseModel):
    id: str
    salon_id: str
    subscription_id: Optional[str] = None
    plan_id: Optional[str] = None
    amount: float
    currency: str = "INR"
    payment_gateway: str = "cashfree"
    gateway_order_id: Optional[str] = None
    gateway_payment_id: Optional[str] = None
    payment_status: str  # pending | success | failed | cancelled
    payment_response: Optional[Dict[str, Any]] = None
    payment_method: Optional[str] = None
    created_at: str
    updated_at: Optional[str] = None

class CreateOrderRequest(BaseModel):
    plan_id: Optional[str] = None  # if None, use default active plan
    discount_code: Optional[str] = None  # Phase 7 — applies discount to base_amount



# ============ AUTH HELPERS ============

def create_access_token(data: dict):
    to_encode = data.copy()
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(token: str):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.InvalidTokenError:
        return None

async def get_current_salon(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Legacy salon auth — accepts both the original `role=salon` JWT (single-user)
    AND the multi-user `salon_admin` / `salon_branch_manager` JWTs issued by
    `/api/salon/users/login`. Staff role is intentionally excluded — they are not
    allowed to mutate catalog data."""
    token = credentials.credentials
    payload = verify_token(token)
    if not payload or payload.get("role") not in ("salon", "salon_admin", "salon_branch_manager"):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication credentials"
        )
    # Normalise `sub` so downstream handlers can keep treating it as the salon id.
    if payload.get("role") != "salon" and payload.get("salon_id"):
        payload["sub"] = payload["salon_id"]
    return payload

async def get_current_salon_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get current authenticated salon user (admin / staff / branch_manager)"""
    token = credentials.credentials
    payload = verify_token(token)
    # Accept legacy "salon" role as well as the multi-user salon roles
    if not payload or payload.get("role") not in [
        "salon_admin", "salon_staff", "salon_branch_manager", "salon"
    ]:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication credentials"
        )

    # Phase-5 enforcement — block suspended salons on EVERY salon-side request.
    # (Previously the suspend check was only at login time, so already-logged-in
    #  admins kept access to mutations like adding branches/barbers.)
    salon_id = payload.get("salon_id") or payload.get("sub")
    if salon_id:
        salon_doc = await db.salons.find_one({"id": salon_id}, {"_id": 0, "status": 1})
        if salon_doc and salon_doc.get("status") == "suspended":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "code": "salon_suspended",
                    "message": "This salon has been suspended by the platform administrator. Please contact support.",
                },
            )
    return payload

async def get_current_salon_admin(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Get current authenticated salon user - admin only"""
    token = credentials.credentials
    payload = verify_token(token)
    # Accept legacy "salon" role as admin (backward compatibility)
    if not payload or payload.get("role") not in ["salon_admin", "salon"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    
    # For legacy salon tokens, salon_id is in 'sub' field
    if payload.get("role") == "salon" and "salon_id" not in payload:
        payload["salon_id"] = payload.get("sub")
    
    return payload

def check_permission(user_payload: dict, permission: str) -> bool:
    """Check if user has specific permission"""
    if user_payload.get("role") == "salon_admin":
        return True  # Admin has all permissions
    
    permissions = user_payload.get("permissions", {})
    return permissions.get(permission, False)


# ---------------------------------------------------------------------------
# Granular module-level permissions (view / create / edit / delete / …)
#
# The salon-user's `permissions.modules` dict is the authoritative source.
# Example shape:
#   {
#     "staff":          {"view": True, "view_all": True, "edit": False, "delete": False,
#                        "attendance": True, "salary_view": True, "salary_pay": False,
#                        "documents": False, "access_control": False},
#     "financials":     {"view_dashboard": True, "view_transactions": True,
#                        "create_transaction": False, "edit_transaction": False,
#                        "delete_transaction": False},
#     "analytics":      {"view": True},
#     "services":       {"view": True, "create": False, "edit": False, "delete": False,
#                        "toggle": False, "upload_csv": False, "manage_categories": False,
#                        "manage_packages": False, "manage_memberships": False},
#     "gallery":        {"view": True, "upload": False, "delete": False},
#     "marketing":      {"view": True, "create_campaign": False, "edit_campaign": False,
#                        "delete_campaign": False, "manage_coupons": False,
#                        "manage_loyalty": False},
#     "salon_settings": {"view": True, "edit_profile": False, "edit_hours": False,
#                        "edit_notifications": False, "edit_branches": False,
#                        "manage_users": False, "manage_subscription": False},
#     "delete_salon":   {"allowed": False},
#   }
#
# has_module_permission() falls back to the legacy flat "can_access_*" keys
# whenever the modules dict is empty for that action, so pre-existing
# salon_users continue to work without any data migration.
# ---------------------------------------------------------------------------

# Map (module, action) -> legacy flat key that grants the same capability.
_MODULE_LEGACY_MAP: Dict[str, Dict[str, str]] = {
    "staff": {
        "view": "can_access_staff",
        "view_all": "can_view_all_staff",
        "create": "can_access_staff",
        "edit": "can_access_staff",
        "delete": "can_access_staff",
        "attendance": "can_access_staff",
        "salary_view": "can_access_staff",
        "salary_pay": "can_access_staff",
        "documents": "can_access_staff",
        "access_control": "can_access_staff",
    },
    "financials": {
        "view_dashboard": "can_access_financials",
        "view_transactions": "can_access_financials",
        "create_transaction": "can_access_financials",
        "edit_transaction": "can_access_financials",
        "delete_transaction": "can_access_financials",
    },
    "analytics": {
        "view": "can_access_analytics",
    },
    "reports": {
        # Reports = merged Financials + Analytics. Grants view via either legacy flag.
        "view": "can_access_analytics",
        "view_dashboard": "can_access_financials",
        "view_transactions": "can_access_financials",
        "create_transaction": "can_access_financials",
        "edit_transaction": "can_access_financials",
        "delete_transaction": "can_access_financials",
        "edit_targets": "can_access_financials",
        "edit_prefs": "can_access_analytics",
    },
    "services": {
        "view": "can_access_services",
        "create": "can_access_services",
        "edit": "can_access_services",
        "delete": "can_access_services",
        "toggle": "can_access_services",
        "upload_csv": "can_access_services",
        "manage_categories": "can_access_services",
        "manage_packages": "can_access_services",
        "manage_memberships": "can_access_services",
    },
    "gallery": {
        "view": "can_access_gallery",
        "upload": "can_access_gallery",
        "delete": "can_access_gallery",
    },
    "marketing": {
        "view": "can_access_marketing",
        "create_campaign": "can_access_marketing",
        "edit_campaign": "can_access_marketing",
        "delete_campaign": "can_access_marketing",
        "manage_coupons": "can_access_marketing",
        "manage_loyalty": "can_access_marketing",
    },
    "salon_settings": {
        "view": "can_edit_salon",
        "edit_profile": "can_edit_salon",
        "edit_hours": "can_edit_salon",
        "edit_notifications": "can_edit_salon",
        "edit_branches": "can_edit_salon",
        "manage_users": "can_edit_salon",
        "manage_subscription": "can_edit_salon",
    },
    "delete_salon": {
        "allowed": "can_delete_salon",
    },
}


def has_module_permission(user_payload: dict, module: str, action: str) -> bool:
    """
    Check whether the current salon user has permission for a given
    module action (e.g. `has_module_permission(u, "staff", "attendance")`).

    Admins always pass. Branch managers get everything within their assigned
    branches (branch scoping is enforced separately by enforce_branch_for_manager).
    For staff, the granular `permissions.modules[module][action]` flag is
    consulted first; if unset, we fall back to the legacy flat `can_access_*`
    key so existing users don't lose access.
    """
    role = user_payload.get("role")
    if role in ("salon_admin", "admin", "salon"):
        return True
    if role == "salon_branch_manager":
        # Branch managers get admin-equivalent access inside their assigned branches.
        return True

    perms = user_payload.get("permissions") or {}
    modules = perms.get("modules") or {}
    module_perms = modules.get(module) or {}
    if module_perms.get(action):
        return True

    # Legacy fallback — grants if the old flat "can_access_*" flag is true.
    legacy_key = _MODULE_LEGACY_MAP.get(module, {}).get(action)
    if legacy_key and perms.get(legacy_key):
        return True

    # Reports module = union of legacy Financials + Analytics access.
    if module == "reports":
        if perms.get("can_access_financials") or perms.get("can_access_analytics"):
            return True
    return False


def require_module_permission(module: str, action: str = "view"):
    """
    FastAPI dependency factory. Returns a dependency that verifies the JWT
    AND enforces the given module+action permission. 403 on failure.
    """
    async def _dep(current_user=Depends(get_current_salon_user)):
        if not has_module_permission(current_user, module, action):
            raise HTTPException(
                status_code=403,
                detail=f"Permission denied: {module}.{action}"
            )
        return current_user
    return _dep


def is_branch_manager(user_payload: dict) -> bool:
    return user_payload.get("role") == "salon_branch_manager"


def assigned_branch_ids_for(user_payload: dict) -> List[str]:
    return list(user_payload.get("assigned_branch_ids") or [])


def check_branch_access(user_payload: dict, branch_id: Optional[str]) -> bool:
    """RBAC for branch managers: every branch_id passed in queries must be in
    their assigned list. Admins and legacy salon roles are unrestricted."""
    if not is_branch_manager(user_payload):
        return True
    assigned = assigned_branch_ids_for(user_payload)
    # If the manager has no branches assigned at all, deny everything (safer default).
    if not assigned:
        return False
    if branch_id is None:
        # Cross-branch access is not allowed for managers. The caller is expected
        # to substitute the manager's first assigned branch.
        return False
    return branch_id in assigned


def enforce_branch_for_manager(user_payload: dict, branch_id: Optional[str]) -> str:
    """Resolve the effective branch_id for a request while enforcing manager scope.
    - For admins/staff/legacy: returns branch_id as-is (may be None).
    - For branch managers: validates branch_id is in assigned_branch_ids, else 403.
      If branch_id is None, returns the first assigned branch (so cross-branch
      requests transparently scope to one branch).
    """
    if not is_branch_manager(user_payload):
        return branch_id
    assigned = assigned_branch_ids_for(user_payload)
    if not assigned:
        raise HTTPException(status_code=403, detail="Branch manager has no assigned branches")
    if branch_id is None:
        return assigned[0]
    if branch_id not in assigned:
        raise HTTPException(status_code=403, detail="Access denied for this branch")
    return branch_id


# ============ BRANCH HELPERS ============

# Collections that are part of the branch-aware data model. Each gets a `branch_id`
# field. When migration runs, all existing documents in these collections are
# stamped with the salon's main branch id so old single-location salons keep working.
BRANCH_AWARE_COLLECTIONS = [
    "tokens",
    "barbers",
    "attendance",
    "financial_transactions",
    "salon_customers",
    "invoices",
    "salon_users",
    "customer_memberships",
    "wallet_transactions",
    "incentive_payouts",
    "salary_records",
]


async def get_main_branch(salon_id: str) -> Optional[dict]:
    """Return the main branch for a salon (creating one if missing).
    Only returns ACTIVE branches; an inactive branch should never become the
    silent default for new bookings/staff/etc."""
    branch = await db.salon_branches.find_one(
        {"salon_id": salon_id, "is_main_branch": True, "status": "active"}, {"_id": 0}
    )
    if branch:
        return branch
    # Fallback: any active branch (covers edge cases where main was wrongly
    # deactivated outside the API).
    branch = await db.salon_branches.find_one(
        {"salon_id": salon_id, "status": "active"}, {"_id": 0}
    )
    return branch


async def resolve_branch_id(salon_id: str, branch_id: Optional[str]) -> Optional[str]:
    """Resolve `branch_id` for a request.

    If a `branch_id` is provided, return it as-is (caller should validate it
    belongs to the salon if security-sensitive). If not provided, fall back to
    the salon's main branch id so legacy clients keep working.
    """
    if branch_id:
        return branch_id
    main = await get_main_branch(salon_id)
    return main["id"] if main else None


async def ensure_main_branch_for_salon(salon: dict) -> dict:
    """Idempotently create a "Main Branch" for a salon if none exists. Returns the branch dict."""
    salon_id = salon["id"]
    existing_main = await db.salon_branches.find_one(
        {"salon_id": salon_id, "is_main_branch": True}, {"_id": 0}
    )
    if existing_main:
        return existing_main

    # If any branches exist for this salon (without a main flag yet), promote the
    # first one to main; otherwise create a fresh "Main Branch" inheriting salon
    # address details so existing single-location data remains coherent.
    any_branch = await db.salon_branches.find_one({"salon_id": salon_id}, {"_id": 0})
    if any_branch:
        await db.salon_branches.update_one(
            {"id": any_branch["id"]},
            {"$set": {"is_main_branch": True}},
        )
        any_branch["is_main_branch"] = True
        return any_branch

    new_branch = {
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "branch_name": "Main Branch",
        "branch_code": "MAIN",
        "address": salon.get("address"),
        "city": salon.get("city"),
        "latitude": salon.get("latitude"),
        "longitude": salon.get("longitude"),
        "phone": salon.get("phone"),
        "email": salon.get("email"),
        "branch_manager_staff_id": None,
        "is_main_branch": True,
        "status": "active",
        "operational_hours": salon.get("operational_hours"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.salon_branches.insert_one(new_branch.copy())
    new_branch.pop("_id", None)
    return new_branch


async def migrate_branches():
    """Idempotent migration:
    1. Ensure every salon has a "Main Branch" record in `salon_branches`.
    2. Back-fill `branch_id` on every existing document in branch-aware collections
       so legacy single-location data stays accessible after the multi-branch upgrade.
    """
    salons = await db.salons.find({}, {"_id": 0}).to_list(length=None)
    if not salons:
        return

    total_backfilled = 0
    for salon in salons:
        salon_id = salon["id"]
        main_branch = await ensure_main_branch_for_salon(salon)
        main_branch_id = main_branch["id"]

        for coll in BRANCH_AWARE_COLLECTIONS:
            res = await db[coll].update_many(
                {"salon_id": salon_id, "branch_id": {"$in": [None, ""]}},
                {"$set": {"branch_id": main_branch_id}},
            )
            # Also handle docs missing the field entirely
            res2 = await db[coll].update_many(
                {"salon_id": salon_id, "branch_id": {"$exists": False}},
                {"$set": {"branch_id": main_branch_id}},
            )
            total_backfilled += (res.modified_count + res2.modified_count)
    if total_backfilled:
        logger.info(f"[BRANCH MIGRATION] Back-filled branch_id on {total_backfilled} legacy docs")


async def migrate_subscription_pricing_v2():
    """Phase 2 (Part C) — idempotent migration.

    1. Backfill `price_per_branch` on every subscription_plans doc that doesn't
       have it (copy from `price`).
    2. Backfill `billable_branch_count`, `price_per_branch_snapshot`,
       `branch_ids_snapshot`, `base_amount`, `total_amount`,
       `discount_code_applied`, `discount_amount` on existing paid salon
       subscriptions.

    Guarded by a `system_flags.subscription_v2_migrated = True` doc so it only
    runs once.
    """
    flag = await db.system_flags.find_one({"key": "subscription_v2_migrated"}, {"_id": 0})
    if flag and flag.get("value") is True:
        return

    # 1) Plans
    plans_updated = 0
    async for plan in db.subscription_plans.find({}, {"_id": 0}):
        if plan.get("price_per_branch") in (None, "", 0, 0.0):
            await db.subscription_plans.update_one(
                {"id": plan["id"]},
                {"$set": {"price_per_branch": float(plan.get("price") or 0)}},
            )
            plans_updated += 1

    # 2) Active / paid subscriptions
    subs_updated = 0
    async for sub in db.salon_subscriptions.find(
        {"payment_status": "paid"}, {"_id": 0}
    ):
        if sub.get("billable_branch_count") is not None:
            continue

        salon_id = sub.get("salon_id")
        plan_id = sub.get("plan_id")
        plan = await db.subscription_plans.find_one({"id": plan_id}, {"_id": 0}) if plan_id else None
        price_per_branch = float(
            (plan or {}).get("price_per_branch")
            or (plan or {}).get("price")
            or sub.get("price")
            or 0
        )

        # Try to use the salon's current active branches as the snapshot;
        # fall back to 1 if none.
        branch_ids = await get_active_branch_ids(salon_id) if salon_id else []
        if not branch_ids:
            billable = 1
            branch_ids = []
        else:
            billable = len(branch_ids)

        base_amount = round(price_per_branch * billable, 2)
        await db.salon_subscriptions.update_one(
            {"id": sub["id"]},
            {
                "$set": {
                    "billable_branch_count": billable,
                    "price_per_branch_snapshot": price_per_branch,
                    "branch_ids_snapshot": branch_ids,
                    "base_amount": base_amount,
                    "discount_code_applied": None,
                    "discount_amount": 0.0,
                    "total_amount": base_amount,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
            },
        )
        subs_updated += 1

    await db.system_flags.update_one(
        {"key": "subscription_v2_migrated"},
        {"$set": {"value": True, "ran_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    logger.info(
        f"[SUBSCRIPTION V2 MIGRATION] plans updated: {plans_updated}, "
        f"subscriptions backfilled: {subs_updated}"
    )


# Optional authentication dependencies (don't raise if no auth)
async def get_current_salon_user_optional(credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False))):
    """Get current authenticated salon user if present, otherwise return None"""
    if not credentials:
        return None
    token = credentials.credentials
    payload = verify_token(token)
    if not payload or payload.get("role") not in [
        "salon_admin", "salon_staff", "salon_branch_manager", "salon"
    ]:
        return None
    return payload


async def get_current_salon_optional(credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False))):
    """Get current authenticated salon if present, otherwise return None"""
    if not credentials:
        return None
    token = credentials.credentials
    payload = verify_token(token)
    if not payload or payload.get("role") != "salon":
        return None
    return payload

# ============ HELPER FUNCTIONS ============

def generate_otp():
    """Generate a cryptographically secure random 6-digit OTP"""
    return str(secrets.randbelow(900000) + 100000)


async def _otp_is_valid(phone: str, code: str, db_collection) -> bool:
    """Unified OTP validation.

    Production path:  ask Twilio Verify (`verify_whatsapp_otp`).
    Dev/mock path:    fall back to the locally-stored OTP doc in
                      `db_collection` (which we keep populating for audit and
                      for the legacy mock flow).

    The fallback **only** kicks in when Twilio Verify is not configured
    (`verify_not_configured`).  Any other Verify failure (wrong code, expired,
    max attempts) returns False immediately — we never bypass Verify with the
    stale DB OTP, since the user would have received Twilio's code, not ours.
    """
    if not code or not phone:
        return False

    result = await verify_whatsapp_otp(phone, code)
    if result.get("valid"):
        return True
    if result.get("error") != "verify_not_configured":
        # Twilio answered → its verdict is final.
        return False

    # Mock / dev path — check the local OTP store.
    stored = await db_collection.find_one({"phone": phone})
    if not stored or stored.get("otp") != code:
        return False
    try:
        expires_at = datetime.fromisoformat(
            str(stored.get("expires_at", "")).replace("Z", "+00:00")
        )
    except Exception:
        return False
    return datetime.now(timezone.utc) <= expires_at

def calculate_distance(lat1, lon1, lat2, lon2):
    """Calculate distance between two coordinates in km"""
    R = 6371  # Earth's radius in km
    
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    
    a = (math.sin(dlat / 2) * math.sin(dlat / 2) +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) * math.sin(dlon / 2))
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    return R * c

def get_shifts():
    """Get available shifts with timings"""
    return [
        {"id": "Morning", "name": "Morning", "time": "7 AM - 11 AM"},
        {"id": "Noon", "name": "Noon", "time": "11 AM - 4 PM"},
        {"id": "Evening", "name": "Evening", "time": "4 PM - 9 PM"}
    ]

def get_shift_prefix(shift: str) -> str:
    """Get token prefix for shift"""
    prefixes = {
        "Morning": "M",
        "Noon": "N",
        "Evening": "E"
    }
    return prefixes.get(shift, "M")

def extract_token_sequence(token_number: str) -> int:
    """Extract numeric sequence from token (e.g., M001 -> 1)"""
    try:
        return int(token_number[1:])  # Skip first character (prefix)
    except (ValueError, IndexError):
        return 0

async def get_next_token_number(salon_id: str, date: str, shift: str) -> str:
    """Get next SALON-WIDE token number for date/shift.

    Token format: {shift_prefix}{seq} where seq is a salon-wide sequence per shift.
    Example: M1, M2, N1, E1, ...
    This is the GLOBAL (total) token number visible to customer & salon.
    Barber-wise position is computed separately (see compute_barber_queue_position).
    """
    prefix = get_shift_prefix(shift)

    tokens = await db.tokens.find(
        {"salon_id": salon_id, "date": date, "shift": shift},
        {"_id": 0, "token_number": 1}
    ).to_list(5000)

    # Only consider tokens starting with this prefix
    shift_tokens = [t for t in tokens if t.get("token_number", "").startswith(prefix)]
    if shift_tokens:
        max_seq = max([extract_token_sequence(t["token_number"]) for t in shift_tokens])
        next_seq = max_seq + 1
    else:
        next_seq = 1

    # No zero-padding; customers see M12, N3 etc. (matches PRD examples like M12)
    return f"{prefix}{next_seq}"

def generate_2hour_slots():
    """DEPRECATED: Generate 2-hour time slots (kept for backward compatibility)"""
    slots = []
    start_times = list(range(8, 22, 2))  # 8AM to 10PM, every 2 hours
    
    for start in start_times:
        end = start + 2
        slot = f"{start:02d}:00-{end:02d}:00"
        slots.append(slot)
    
    return slots


# ============================================================================
# ADVANCED TOKEN MANAGEMENT — Shift capacity, 75% rule, barber-wise queue
# ============================================================================

# Token system tuning constants
BLOCKING_FACTOR = 0.75  # 75% rule: blocked time = service duration × 0.75
DEFAULT_SERVICE_DURATION = 30  # fallback minutes if service has no duration

# Default shift window when salon has no operational_hours configured
DEFAULT_OPEN_HOUR = 9
DEFAULT_CLOSE_HOUR = 21


def _parse_hour(hhmm: Optional[str], default_h: int) -> int:
    """Return integer hour portion of 'HH:MM'. Defaults if empty/malformed."""
    try:
        if not hhmm:
            return default_h
        return int(hhmm.split(":")[0])
    except Exception:
        return default_h


def _weekday_key(date_str: str) -> str:
    """Return lowercase day name (monday..sunday) for a date string YYYY-MM-DD."""
    try:
        return datetime.fromisoformat(date_str).strftime("%A").lower()
    except Exception:
        return datetime.now(timezone.utc).strftime("%A").lower()


async def get_salon_shift_windows(salon_id: str, date: str) -> Dict[str, Dict[str, int]]:
    """
    Compute shift windows for a salon on a given date based on its operational hours.

    Rules (per user spec):
      - Base: 4 hours per shift (Morning, Noon, Evening) = 12h baseline.
      - Shift durations derived from salon's opening/closing time.
      - Extra hours go to Morning first, then Evening (lunch counts as part of Evening).
      - All hours kept as whole numbers.

    Returns:
        {
          "Morning": {"start": 7, "end": 11, "duration_hours": 4, "duration_minutes": 240},
          "Noon":    {"start": 11, "end": 15, ...},
          "Evening": {"start": 15, "end": 19, ...},
        }
    """
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0}) or {}
    op_hours = salon.get("operational_hours") or {}

    day_key = _weekday_key(date)
    day_hours = op_hours.get(day_key) if isinstance(op_hours, dict) else None

    if isinstance(day_hours, dict) and not day_hours.get("is_holiday"):
        open_h = _parse_hour(day_hours.get("opening_time"), DEFAULT_OPEN_HOUR)
        close_h = _parse_hour(day_hours.get("closing_time"), DEFAULT_CLOSE_HOUR)
    else:
        open_h, close_h = DEFAULT_OPEN_HOUR, DEFAULT_CLOSE_HOUR

    if close_h <= open_h:
        # fall back to defaults if malformed
        open_h, close_h = DEFAULT_OPEN_HOUR, DEFAULT_CLOSE_HOUR

    total_hours = close_h - open_h

    # Distribute hours: base even split, then extras -> Morning, then Evening
    base = total_hours // 3
    extras = total_hours % 3
    durations = [base, base, base]  # [M, N, E]
    if extras >= 1:
        durations[0] += 1  # morning
    if extras >= 2:
        durations[2] += 1  # evening (lunch stays in evening window)
    # (extras==3 not possible since %3)

    m_start = open_h
    m_end = m_start + durations[0]
    n_start = m_end
    n_end = n_start + durations[1]
    e_start = n_end
    e_end = e_start + durations[2]

    return {
        "Morning": {
            "start": m_start, "end": m_end,
            "duration_hours": durations[0],
            "duration_minutes": durations[0] * 60,
        },
        "Noon": {
            "start": n_start, "end": n_end,
            "duration_hours": durations[1],
            "duration_minutes": durations[1] * 60,
        },
        "Evening": {
            "start": e_start, "end": e_end,
            "duration_hours": durations[2],
            "duration_minutes": durations[2] * 60,
        },
    }


async def calc_service_total_minutes(service_ids: List[str]) -> int:
    """Sum default_duration of provided services in minutes. Unknown services get DEFAULT_SERVICE_DURATION."""
    if not service_ids:
        return 0
    total = 0
    for sid in service_ids:
        svc = await db.services.find_one({"id": sid}, {"_id": 0, "default_duration": 1})
        if svc and svc.get("default_duration"):
            total += int(svc["default_duration"])
        else:
            total += DEFAULT_SERVICE_DURATION
    return total


def calc_blocked_minutes_from_total(total_minutes: int) -> int:
    """Apply 75% rule. Ceil so partial minutes always count toward capacity."""
    if total_minutes <= 0:
        return 0
    return math.ceil(total_minutes * BLOCKING_FACTOR)


async def get_barber_blocked_minutes_used(salon_id: str, barber_id: str, date: str, shift: str) -> int:
    """Sum blocked minutes of *pending* tokens assigned to this barber for this shift.

    Only bookings that still occupy the chair count toward capacity:
      • waiting / future / in_progress / called → COUNT (pending)
      • completed / cancelled / skipped         → DO NOT count (slot is free again)

    This means a barber who finishes a booking inside a shift can immediately
    accept a new booking in the same shift if the shift duration permits.
    """
    if barber_id == "any":
        return 0
    tokens = await db.tokens.find(
        {
            "salon_id": salon_id,
            "barber_id": barber_id,
            "date": date,
            "shift": shift,
            "status": {"$nin": ["cancelled", "skipped", "completed"]},
        },
        {"_id": 0, "selected_services": 1, "blocked_minutes": 1, "total_service_minutes": 1},
    ).to_list(1000)

    used = 0
    for t in tokens:
        if t.get("blocked_minutes") is not None:
            used += int(t["blocked_minutes"])
        else:
            mins = await calc_service_total_minutes(t.get("selected_services", []))
            used += calc_blocked_minutes_from_total(mins)
    return used


async def can_fit_in_barber_shift(
    salon_id: str, barber_id: str, date: str, shift: str, new_blocked_minutes: int
) -> Dict[str, Any]:
    """Shift-availability check per the user's rounding rule:

      Shift is declared FULL once the cumulative blocked time crosses the shift duration.
      The booking that causes the crossover is still allowed (last one may cross).
      The next booking is blocked.

    Returns dict with: allowed, used_before, used_after, capacity_minutes, is_full_after
    """
    windows = await get_salon_shift_windows(salon_id, date)
    shift_info = windows.get(shift) or {}
    capacity = int(shift_info.get("duration_minutes", 0))

    used_before = await get_barber_blocked_minutes_used(salon_id, barber_id, date, shift)

    # Not allowed if the shift was ALREADY full before this booking
    already_full = used_before >= capacity if capacity > 0 else False
    allowed = (not already_full) and capacity > 0

    used_after = used_before + (new_blocked_minutes if allowed else 0)
    is_full_after = used_after >= capacity if capacity > 0 else True

    return {
        "allowed": allowed,
        "used_before": used_before,
        "used_after": used_after,
        "capacity_minutes": capacity,
        "is_full_before": already_full,
        "is_full_after": is_full_after,
    }


def is_barber_available_on(barber: Dict[str, Any], date_str: str) -> bool:
    """Return True if this barber is allowed to take bookings / be visible to customers on date_str.
    Date format: YYYY-MM-DD.
    Conditions for unavailability:
    - Barber's joining date (doj) is after `date_str`
    - last_working_date is set and `date_str` > last_working_date
    - `date_str` is in barber's leave_dates list
    - Legacy global on_leave flag is True (only for "today" — kept for back-compat).
    """
    if not date_str:
        return True
    try:
        # Joining date
        doj = (barber.get("doj") or "").strip()
        if doj and date_str < doj:
            return False
        # Last working date
        lwd = (barber.get("last_working_date") or "").strip()
        if lwd and date_str > lwd:
            return False
        # Per-date leave list
        leave_dates = barber.get("leave_dates") or []
        if isinstance(leave_dates, list) and date_str in leave_dates:
            return False
        # Legacy global on_leave flag — only treat as "leave today"
        ist = timezone(timedelta(hours=5, minutes=30))
        today_ist = datetime.now(ist).strftime("%Y-%m-%d")
        if barber.get("on_leave") is True and date_str == today_ist:
            return False
    except Exception:
        # On any parsing error, do not block the barber
        return True
    return True


async def pick_fastest_barber(
    salon_id: str, date: str, shift: str, required_blocked_minutes: int,
    customer_gender: Optional[str] = None
) -> Optional[Dict[str, Any]]:
    """Auto-assign the 'fastest' barber using the priority rules.

    Priority 1: Barber with shortest ACTIVE queue today (waiting + called),
                AND enough capacity in this shift.
    Priority 2: Barber with fewest bookings YESTERDAY, with enough capacity.
    Priority 3: Any random active barber with enough capacity.

    Returns the chosen barber doc or None if no barber has capacity.
    Skips: barbers on leave on this date, barbers past last_working_date,
           barbers whose joining date is later than this date.
    """
    barbers = await db.barbers.find(
        {
            "salon_id": salon_id,
            "is_active": True,
            "$or": [{"on_leave": False}, {"on_leave": None}, {"on_leave": {"$exists": False}}],
        },
        {"_id": 0},
    ).to_list(500)

    if not barbers:
        return None

    # Filter out barbers unavailable on this date (joining date / last working day / leave_dates)
    barbers = [b for b in barbers if is_barber_available_on(b, date)]
    if not barbers:
        return None

    # Filter barbers with capacity for this shift
    eligible = []
    for b in barbers:
        fit = await can_fit_in_barber_shift(salon_id, b["id"], date, shift, required_blocked_minutes)
        if fit["allowed"]:
            eligible.append(b)

    if not eligible:
        return None

    # Priority 1: shortest active queue today
    today_counts = []
    for b in eligible:
        cnt = await db.tokens.count_documents({
            "salon_id": salon_id,
            "barber_id": b["id"],
            "date": date,
            "status": {"$in": ["waiting", "called", "in_progress", "in_service"]},
        })
        today_counts.append((cnt, b))

    if today_counts:
        min_today = min(c for c, _ in today_counts)
        # If there are barbers with zero/low counts, pick any among those.
        # If ALL are zero (no data today), fall to priority 2.
        if min_today > 0 or any(c > 0 for c, _ in today_counts):
            tied = [b for c, b in today_counts if c == min_today]
            return secrets.choice(tied)

    # Priority 2: fewest bookings YESTERDAY
    try:
        yest = (datetime.fromisoformat(date) - timedelta(days=1)).strftime("%Y-%m-%d")
    except Exception:
        yest = (datetime.now(timezone.utc) - timedelta(days=1)).strftime("%Y-%m-%d")

    yest_counts = []
    for b in eligible:
        cnt = await db.tokens.count_documents({
            "salon_id": salon_id,
            "barber_id": b["id"],
            "date": yest,
            "status": {"$nin": ["cancelled", "skipped"]},
        })
        yest_counts.append((cnt, b))

    if yest_counts and any(c > 0 for c, _ in yest_counts):
        min_yest = min(c for c, _ in yest_counts)
        tied = [b for c, b in yest_counts if c == min_yest]
        return secrets.choice(tied)

    # Priority 3: random active barber
    return secrets.choice(eligible)


async def compute_barber_queue_snapshot(salon_id: str, barber_id: str, date: str) -> Dict[str, Any]:
    """Return barber queue context used for position & wait calculations.

    Shape:
      {
        "currently_serving": {token_number, customer_name, id, blocked_minutes, ...} | None,
        "waiting_queue": [token, ...] ordered by created_at ascending,
        "remaining_current_minutes": int (75%-rule residual for the token being served)
      }
    """
    if not barber_id or barber_id == "any":
        return {"currently_serving": None, "waiting_queue": [], "remaining_current_minutes": 0}

    # Currently serving: status 'called' or 'in_progress'/'in_service'
    serving = await db.tokens.find_one(
        {
            "salon_id": salon_id,
            "barber_id": barber_id,
            "date": date,
            "status": {"$in": ["called", "in_progress", "in_service"]},
        },
        {"_id": 0},
        sort=[("called_at", -1)],
    )

    remaining = 0
    if serving:
        if serving.get("blocked_minutes") is not None:
            remaining = int(serving["blocked_minutes"])
        else:
            mins = await calc_service_total_minutes(serving.get("selected_services", []))
            remaining = calc_blocked_minutes_from_total(mins)
        # If called_at present, subtract elapsed; floor at 0
        try:
            if serving.get("called_at"):
                called_dt = datetime.fromisoformat(serving["called_at"].replace("Z", "+00:00"))
                elapsed = (datetime.now(timezone.utc) - called_dt).total_seconds() / 60
                remaining = max(0, int(remaining - elapsed))
        except Exception:
            pass

    waiting = await db.tokens.find(
        {
            "salon_id": salon_id,
            "barber_id": barber_id,
            "date": date,
            "status": "waiting",
        },
        {"_id": 0},
    ).sort("created_at", 1).to_list(1000)

    return {
        "currently_serving": serving,
        "waiting_queue": waiting,
        "remaining_current_minutes": remaining,
    }


async def compute_queue_status_for_token(token: Dict[str, Any]) -> Dict[str, Any]:
    """Build customer-facing queue view for a single token.

    Returns:
      {
        total_token, barber_name, barber_id,
        position, people_before, estimated_wait_minutes,
        currently_serving_token, currently_serving_services, approx_finish_time_minutes,
        status_message, is_current (True when THIS customer is being served)
      }
    """
    salon_id = token.get("salon_id")
    barber_id = token.get("barber_id")
    date = token.get("date")
    token_id = token.get("id")

    snapshot = await compute_barber_queue_snapshot(salon_id, barber_id, date)
    serving = snapshot["currently_serving"]
    waiting = snapshot["waiting_queue"]
    remaining_current = snapshot["remaining_current_minutes"]

    status = token.get("status")
    is_current_customer = bool(serving and serving.get("id") == token_id)

    position = 0
    people_before = 0
    wait_minutes = 0
    status_message = ""

    if is_current_customer:
        position = 1
        people_before = 0
        wait_minutes = 0
        status_message = "It's your turn — you're being served now."
    elif status in ("called", "in_progress", "in_service"):
        # Called but maybe not the 'serving' lookup returned them — treat as current
        position = 1
        people_before = 0
        wait_minutes = 0
        status_message = "Please proceed to the salon chair."
    elif status == "waiting":
        # Find index in waiting queue (FIFO by created_at)
        idx = next((i for i, t in enumerate(waiting) if t.get("id") == token_id), None)
        if idx is None:
            position = 1
            people_before = 0
        else:
            people_before = idx + (1 if serving else 0)
            position = people_before + 1
        # wait = remaining_current + blocked minutes of all waiting tokens ahead
        wait_minutes = remaining_current
        if idx is not None:
            for t in waiting[:idx]:
                if t.get("blocked_minutes") is not None:
                    wait_minutes += int(t["blocked_minutes"])
                else:
                    mins = await calc_service_total_minutes(t.get("selected_services", []))
                    wait_minutes += calc_blocked_minutes_from_total(mins)

        if people_before >= 3:
            status_message = f"{people_before} customers before you."
        elif people_before == 2:
            status_message = "Your turn is coming soon. Please be ready."
        elif people_before == 1:
            status_message = "Please proceed to the salon chair."
        else:
            status_message = "You're next!"
    elif status == "completed":
        status_message = "Service completed."
    elif status == "cancelled":
        status_message = "Booking cancelled."
    elif status == "skipped":
        status_message = "Booking skipped."
    else:
        status_message = "Waiting for the salon to start the shift."

    # Currently-serving info (for Live Chair Status section)
    serving_payload = None
    approx_finish = None
    if serving:
        # Fetch service names briefly
        svc_names = []
        for sid in serving.get("selected_services", []) or []:
            svc = await db.services.find_one({"id": sid}, {"_id": 0, "service_name": 1})
            if svc:
                svc_names.append(svc.get("service_name"))
        serving_payload = {
            "token_number": serving.get("token_number"),
            "customer_name_masked": _mask_name(serving.get("customer_name")),
            "services": svc_names,
            "started_at": serving.get("called_at"),
        }
        approx_finish = remaining_current

    return {
        "total_token": token.get("token_number"),
        "barber_id": barber_id,
        "barber_name": token.get("barber_name"),
        "position": position,
        "people_before": people_before,
        "estimated_wait_minutes": wait_minutes,
        "currently_serving": serving_payload,
        "approx_finish_minutes": approx_finish,
        "status": status,
        "is_current_customer": is_current_customer,
        "status_message": status_message,
    }


def _mask_name(name: Optional[str]) -> str:
    """Privacy: show first name only or initials to avoid leaking identities."""
    if not name:
        return ""
    first = name.strip().split(" ")[0]
    if len(first) <= 2:
        return first
    return first[0] + "*" * (len(first) - 2) + first[-1]

async def check_and_apply_loyalty_reward(salon_id: str, customer_phone: str, booking_amount: float, payment_mode: str = None):
    """Check if customer qualifies for loyalty reward and auto top-up wallet (Multi-Tier with individual periods)
    
    IMPORTANT: Only counts completed bookings paid by non-wallet methods (cash, upi, pay_later, card).
    Wallet balance is independent of membership - stored in customer_wallets collection.
    """
    # Get loyalty program settings
    loyalty_program = await db.loyalty_programs.find_one({"salon_id": salon_id, "enabled": True}, {"_id": 0})
    if not loyalty_program or not loyalty_program.get("tiers"):
        return None
    
    # Normalize phone number
    if not customer_phone.startswith("+91"):
        customer_phone = f"+91{customer_phone}"
    
    # Find highest qualifying tier (each tier has its own period)
    qualifying_tier = None
    total_spend_for_tier = 0
    
    for tier in sorted(loyalty_program["tiers"], key=lambda x: x["spend_amount"], reverse=True):
        # Calculate date range for THIS tier's period
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=tier["period_months"] * 30)
        
        # Get customer's completed bookings in THIS tier's period
        # ONLY count bookings paid by non-wallet methods (cash, upi, pay_later, card)
        completed_bookings = await db.tokens.find({
            "salon_id": salon_id,
            "phone": customer_phone,
            "status": "completed",
            "completed_at": {"$gte": cutoff_date.isoformat()},
            "payment_mode": {"$nin": ["wallet", None]}  # Exclude wallet payments
        }, {"_id": 0, "total_amount": 1, "payment_mode": 1}).to_list(1000)
        
        # Calculate total spend in this period (only non-wallet payments)
        total_spend = sum([b.get("total_amount", 0) for b in completed_bookings])
        
        # Check if threshold is met for this tier
        if total_spend >= tier["spend_amount"]:
            qualifying_tier = tier
            total_spend_for_tier = total_spend
            break  # Found highest qualifying tier
    
    if not qualifying_tier:
        return None
    
    # Calculate top-up amount based on tier
    topup_amount = (qualifying_tier["spend_amount"] * qualifying_tier["topup_percentage"]) / 100
    
    # Check if customer already received this tier's reward in current period
    # to avoid duplicate rewards
    cutoff_for_check = datetime.now(timezone.utc) - timedelta(days=qualifying_tier["period_months"] * 30)
    existing_reward = await db.loyalty_rewards.find_one({
        "salon_id": salon_id,
        "customer_phone": customer_phone,
        "tier_name": qualifying_tier["name"],
        "created_at": {"$gte": cutoff_for_check.isoformat()}
    })
    
    if existing_reward:
        # Already rewarded for this tier in current period
        return None
    
    # Get or create customer wallet (independent of membership)
    customer_wallet = await db.customer_wallets.find_one({
        "salon_id": salon_id,
        "customer_phone": customer_phone
    }, {"_id": 0})
    
    if customer_wallet:
        # Top up existing wallet
        new_balance = customer_wallet["wallet_balance"] + topup_amount
        await db.customer_wallets.update_one(
            {"id": customer_wallet["id"]},
            {"$set": {"wallet_balance": new_balance, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        # Create new customer wallet
        new_balance = topup_amount
        customer_wallet = {
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            "customer_phone": customer_phone,
            "wallet_balance": new_balance,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.customer_wallets.insert_one(customer_wallet)
    
    # Record the loyalty reward to prevent duplicates
    await db.loyalty_rewards.insert_one({
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "customer_phone": customer_phone,
        "tier_name": qualifying_tier["name"],
        "topup_amount": topup_amount,
        "total_spend": total_spend_for_tier,
        "period_months": qualifying_tier["period_months"],
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Record wallet transaction
    await db.wallet_transactions.insert_one({
        "id": str(uuid.uuid4()),
        "customer_phone": customer_phone,
        "salon_id": salon_id,
        "transaction_type": "credit",
        "amount": topup_amount,
        "balance_after": new_balance,
        "description": f"🎉 Loyalty reward ({qualifying_tier['name']} tier): Spent ₹{total_spend_for_tier:.2f} in {qualifying_tier['period_months']} months",
        "source": "loyalty_reward",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Create in-app notification for customer
    await create_in_app_notification(
        user_type="customer",
        user_id=customer_phone,
        title="🎉 Loyalty Bonus Credited!",
        message=f"Congratulations! You've earned ₹{topup_amount:.2f} as loyalty reward ({qualifying_tier['name']} tier). Your new wallet balance is ₹{new_balance:.2f}.",
        notification_type="loyalty_reward",
        setting_key="membership_added",
        salon_id=salon_id,
        related_id=customer_wallet.get("id", "")
    )
    
    # Send WhatsApp notification for loyalty bonus
    try:
        # Get salon name for the message
        salon = await db.salons.find_one({"id": salon_id}, {"_id": 0, "salon_name": 1})
        salon_name = salon.get("salon_name", "The Salon") if salon else "The Salon"
        
        whatsapp_message = f"""🎉 *Loyalty Bonus Credited!*

Hi! Great news from *{salon_name}*!

You've earned a loyalty reward:
• *Tier*: {qualifying_tier['name']}
• *Bonus*: ₹{topup_amount:.2f}
• *Total Spend*: ₹{total_spend_for_tier:.2f} in {qualifying_tier['period_months']} months

💰 *New Wallet Balance*: ₹{new_balance:.2f}

Thank you for being a valued customer! Use your wallet balance on your next visit.

Book your next appointment on SalonHub! 💇"""
        
        await send_whatsapp_notification(customer_phone, whatsapp_message, "loyalty_bonus")
        logger.info(f"WhatsApp loyalty notification sent to {customer_phone}")
    except Exception as e:
        logger.error(f"Failed to send WhatsApp loyalty notification to {customer_phone}: {e}")
    
    return {
        "rewarded": True,
        "tier": qualifying_tier["name"],
        "topup_amount": topup_amount,
        "new_balance": new_balance,
        "total_spend": total_spend_for_tier
    }

async def calculate_booking_total(service_ids: List[str], barber_id: str) -> float:
    """Calculate total amount for selected services"""
    total = 0.0
    for service_id in service_ids:
        pricing = await db.barber_services.find_one({
            "barber_id": barber_id,
            "service_id": service_id
        })
        if pricing:
            total += pricing.get("price", 0)
    return total


# ============ Module 7 — Per-service barber revenue attribution ============

def attribute_token_revenue_to_barbers(token: dict) -> Dict[str, float]:
    """Return {barber_id: revenue_credited} for one token after pro-rata
    order-discount allocation across service lines.

    Legacy tokens (no `service_assignments` or empty list) → full credit to the
    main `barber_id`. Split tokens credit each line to its assigned barber.
    Invariant: sum(returned.values()) ≈ token.total_amount.
    """
    assignments = token.get("service_assignments") or []
    if not assignments:
        bid = token.get("barber_id")
        if not bid or bid == "any":
            return {}
        return {bid: float(token.get("total_amount") or 0)}

    subtotal = sum(float(a.get("service_price") or 0) for a in assignments)
    discount_amt = float(token.get("order_discount_amount") or 0)
    out: Dict[str, float] = {}
    for a in assignments:
        sp = float(a.get("service_price") or 0)
        share = (discount_amt * (sp / subtotal)) if subtotal > 0 else 0.0
        line_attr = sp - share
        bid = a.get("barber_id")
        if not bid:
            continue
        out[bid] = out.get(bid, 0.0) + line_attr
    return out


def attribute_token_revenue_to_services(token: dict) -> List[Dict[str, Any]]:
    """Return per-service revenue (after pro-rata discount) for one token.

    Each row: {service_id, barber_id, service_price, line_total}.
    Legacy tokens distribute total_amount evenly across selected_services.
    """
    assignments = token.get("service_assignments") or []
    if assignments:
        subtotal = sum(float(a.get("service_price") or 0) for a in assignments)
        discount_amt = float(token.get("order_discount_amount") or 0)
        rows: List[Dict[str, Any]] = []
        for a in assignments:
            sp = float(a.get("service_price") or 0)
            share = (discount_amt * (sp / subtotal)) if subtotal > 0 else 0.0
            rows.append({
                "service_id": a.get("service_id"),
                "barber_id": a.get("barber_id"),
                "service_price": sp,
                "line_total": round(sp - share, 2),
            })
        return rows

    services = token.get("selected_services") or []
    if not services:
        return []
    per = float(token.get("total_amount") or 0) / len(services)
    return [
        {
            "service_id": sid,
            "barber_id": token.get("barber_id"),
            "service_price": per,
            "line_total": round(per, 2),
        }
        for sid in services
    ]


async def broadcast_update(event_type: str, data: dict):
    """Broadcast updates via WebSocket"""
    await sio.emit(event_type, data)

# ============ NOTIFICATION SETTINGS ============

# Default notification preferences
DEFAULT_SALON_NOTIFICATION_SETTINGS = {
    "new_booking": True,
    "booking_change": True,
    "membership_purchase": True,
    "review_added": True,
    # ---- WhatsApp event toggles (controlled by the salon, applied to OUTBOUND
    # ---- customer WhatsApp messages from this salon) ----
    "whatsapp_enabled": True,  # Master switch
    "whatsapp_booking_confirmation": True,
    "whatsapp_booking_completed": True,
    "whatsapp_booking_cancelled": True,
    "whatsapp_booking_rescheduled": True,
    "whatsapp_your_turn_now": True,
    "whatsapp_token_approaching": True,
    "whatsapp_salon_calling": True,
}

DEFAULT_CUSTOMER_NOTIFICATION_SETTINGS = {
    # In-app toggles
    "payment_confirmation": True,
    "turn_approaching": True,
    "manual_notify": True,
    "membership_added": True,
    "membership_expiry": True,  # 30-day and 7-day reminders
    "booking_status_change": True,
    "custom_package": True,
    # WhatsApp toggles (independent on/off for whatsapp messages)
    "whatsapp_payment_confirmation": True,
    "whatsapp_turn_approaching": True,
    "whatsapp_manual_notify": True,
    "whatsapp_membership_added": True,
    "whatsapp_membership_expiry": True,
    "whatsapp_booking_status_change": True,
    "whatsapp_booking_confirmation": True,
    "whatsapp_booking_cancelled": True,
    "whatsapp_booking_rescheduled": True,
}


async def get_salon_notification_settings(salon_id: str) -> dict:
    """Get salon notification settings, creating defaults if missing."""
    settings = await db.salon_notification_settings.find_one({"salon_id": salon_id}, {"_id": 0})
    if not settings:
        return {**DEFAULT_SALON_NOTIFICATION_SETTINGS, "salon_id": salon_id}
    # Merge with defaults to ensure new keys default to True
    merged = {**DEFAULT_SALON_NOTIFICATION_SETTINGS, **settings}
    return merged


async def get_customer_notification_settings(phone: str) -> dict:
    """Get customer notification settings, creating defaults if missing."""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    settings = await db.customer_notification_settings.find_one({"phone": phone}, {"_id": 0})
    if not settings:
        return {**DEFAULT_CUSTOMER_NOTIFICATION_SETTINGS, "phone": phone}
    merged = {**DEFAULT_CUSTOMER_NOTIFICATION_SETTINGS, **settings}
    return merged


async def create_in_app_notification(
    user_type: str,
    user_id: str,
    title: str,
    message: str,
    notification_type: str,
    setting_key: str,
    salon_id: str = "",
    related_id: str = "",
):
    """Create an in-app notification, respecting user's notification preferences."""
    try:
        # Normalize phone for customer
        if user_type == "customer" and not user_id.startswith("+91"):
            user_id = f"+91{user_id}"

        # Check user preferences
        if user_type == "salon":
            settings = await get_salon_notification_settings(user_id)
        else:
            settings = await get_customer_notification_settings(user_id)

        if not settings.get(setting_key, True):
            logger.info(f"In-app notification suppressed for {user_type}/{user_id} (setting {setting_key} is OFF)")
            return None

        notification = {
            "id": str(uuid.uuid4()),
            "user_type": user_type,
            "user_id": user_id,
            "salon_id": salon_id,
            "title": title,
            "message": message,
            "type": notification_type,
            "related_id": related_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.notifications.insert_one(notification)
        # Best-effort live broadcast (non-blocking)
        try:
            await sio.emit("notification_created", {
                "user_type": user_type,
                "user_id": user_id,
                "notification": {**notification},
            })
        except Exception:
            pass
        return notification
    except Exception as e:
        logger.error(f"Failed to create in-app notification: {e}")
        return None


async def should_send_customer_whatsapp(phone: str, setting_key: str) -> bool:
    """Check if customer has WhatsApp notifications enabled for given key."""
    settings = await get_customer_notification_settings(phone)
    return bool(settings.get(setting_key, True))


async def should_send_salon_whatsapp(salon_id: str, event_key: str) -> bool:
    """Check if the salon has the given WhatsApp event enabled.

    Respects both the master `whatsapp_enabled` switch and the per-event
    `whatsapp_<event>` flag. Defaults to True when the setting is missing.
    """
    if not salon_id:
        return True
    settings = await get_salon_notification_settings(salon_id)
    if not settings.get("whatsapp_enabled", True):
        return False
    return bool(settings.get(event_key, True))


def build_action_links(token_id: str, salon_id: str) -> str:
    """Build clickable Reschedule and Cancel action links for WhatsApp messages."""
    base = os.getenv("FRONTEND_BASE_URL") or os.getenv("REACT_APP_BACKEND_URL", "")
    if not base:
        return ""
    base = base.rstrip("/")
    api_base = base
    # Use relative API path for cancel that returns confirmation HTML
    reschedule_url = f"{base}/book/{salon_id}?modify={token_id}"
    cancel_url = f"{api_base}/api/tokens/{token_id}/cancel-link"
    return f"\n\n🔁 *Reschedule:* {reschedule_url}\n❌ *Cancel:* {cancel_url}"


async def send_booking_notification(token_data: dict, notification_type: str):
    """Send WhatsApp notification for booking events"""
    try:
        phone = token_data.get('phone')
        customer_name = token_data.get('customer_name')
        
        if not phone:
            logger.warning(f"No phone number for notification type: {notification_type}")
            return
        
        # Get salon details
        salon = await db.salons.find_one({"id": token_data.get('salon_id')}, {"_id": 0})
        salon_name = salon.get('salon_name', 'The Looks Salon') if salon else 'The Looks Salon'
        
        message = None
        whatsapp_setting_key = None
        action_links = ""

        token_id = token_data.get('id')
        salon_id_for_links = token_data.get('salon_id', '')
        
        if notification_type == 'booking_confirmation':
            # Route through the approved WhatsApp Content Template
            # (HX4ec6d831674ce97cc1dc209327445b81).  This is required for any
            # business-initiated message outside the 24h reply window.
            whatsapp_setting_key = 'whatsapp_booking_confirmation'
            if not await should_send_salon_whatsapp(token_data.get('salon_id', ''), whatsapp_setting_key):
                logger.info(
                    f"WhatsApp suppressed by salon settings for {phone} (salon {token_data.get('salon_id')} - {whatsapp_setting_key})"
                )
                return
            if whatsapp_setting_key and not await should_send_customer_whatsapp(phone, whatsapp_setting_key):
                logger.info(
                    f"WhatsApp notification suppressed for {phone} (setting {whatsapp_setting_key} is OFF)"
                )
                return
            result = await send_booking_confirmation_template(
                phone_number=phone,
                customer_name=customer_name or 'Customer',
                salon_name=salon_name,
                token_number=token_data.get('token_number', 0),
                date=token_data.get('date') or '',
                # `time_slot` may be None on bookings that only picked a shift.
                # Twilio Content Templates reject empty variables → fall back to
                # the shift name (e.g. "Morning") so the template always renders.
                time_slot=(token_data.get('time_slot') or token_data.get('shift') or 'TBD'),
                barber_name=(token_data.get('barber_name') or 'Any available'),
            )
            logger.info(
                f"Notification sent: booking_confirmation to {phone}, status: {result.get('status')}"
            )
            return
        elif notification_type == 'booking_completed':
            # Approved 'booking_completed' Content template
            # (default SID: HXa417403d8b7ff32ce17fcadc6fe1c19a)
            whatsapp_setting_key = 'whatsapp_booking_completed'
            if not await should_send_salon_whatsapp(token_data.get('salon_id', ''), whatsapp_setting_key):
                logger.info(
                    f"WhatsApp suppressed by salon settings for {phone} (salon {token_data.get('salon_id')} - {whatsapp_setting_key})"
                )
                return
            if not await should_send_customer_whatsapp(phone, 'whatsapp_booking_status_change'):
                logger.info(
                    f"WhatsApp suppressed by customer settings for {phone} (whatsapp_booking_status_change OFF)"
                )
                return
            result = await send_booking_completed_template(
                phone_number=phone,
                customer_name=customer_name or 'Customer',
                salon_name=salon_name,
                token_number=token_data.get('token_number', 0),
                barber_name=token_data.get('barber_name') or '',
                amount=token_data.get('total_amount') or token_data.get('amount') or 0,
            )
            logger.info(
                f"Notification sent: booking_completed to {phone}, status: {result.get('status')}"
            )
            return
        elif notification_type == 'token_called':
            # 'your_turn_now' — customer must come to the chair immediately.
            # Use the approved Content template
            # (default SID: HXce2a0648ccfc5d259615714b7f49457b).
            whatsapp_setting_key = 'whatsapp_your_turn_now'
            if not await should_send_salon_whatsapp(token_data.get('salon_id', ''), whatsapp_setting_key):
                logger.info(
                    f"WhatsApp suppressed by salon settings for {phone} (salon {token_data.get('salon_id')} - {whatsapp_setting_key})"
                )
                return
            if not await should_send_customer_whatsapp(phone, 'whatsapp_turn_approaching'):
                logger.info(
                    f"WhatsApp suppressed by customer settings for {phone} (whatsapp_turn_approaching OFF)"
                )
                return
            result = await send_your_turn_now_template(
                phone_number=phone,
                customer_name=customer_name or 'Customer',
                salon_name=salon_name,
                barber_name=token_data.get('barber_name') or 'your stylist',
                token_number=token_data.get('token_number', 0),
            )
            logger.info(
                f"Notification sent: your_turn_now to {phone}, status: {result.get('status')}"
            )
            return
        elif notification_type == 'salon_calling':
            # Triggered by the salon's "Send Notification to Customer" button on the token row.
            message = format_salon_calling(
                customer_name=customer_name,
                salon_name=salon_name,
                barber_name=token_data.get('barber_name')
            )
            whatsapp_setting_key = 'whatsapp_salon_calling'
            if token_id and salon_id_for_links:
                action_links = build_action_links(token_id, salon_id_for_links)
        elif notification_type == 'token_cancelled':
            message = format_token_cancelled(
                customer_name=customer_name,
                token_number=token_data.get('token_number'),
                reason=token_data.get('cancellation_reason')
            )
            whatsapp_setting_key = 'whatsapp_booking_cancelled'
            # No reschedule/cancel link on cancelled — already cancelled.
        elif notification_type == 'token_rescheduled':
            message = format_token_rescheduled(
                customer_name=customer_name,
                old_date=token_data.get('old_date', ''),
                new_date=token_data.get('date', ''),
                new_slot=token_data.get('time_slot', ''),
                token_number=token_data.get('token_number', 0)
            )
            whatsapp_setting_key = 'whatsapp_booking_rescheduled'
            if token_id and salon_id_for_links:
                action_links = build_action_links(token_id, salon_id_for_links)
        elif notification_type == 'token_skipped':
            message = format_token_cancelled(
                customer_name=customer_name,
                token_number=token_data.get('token_number'),
                reason="Skipped"
            )
            whatsapp_setting_key = 'whatsapp_booking_cancelled'
            if token_id and salon_id_for_links:
                action_links = build_action_links(token_id, salon_id_for_links)
        
        if message:
            # Check salon-side master + per-event toggle first (Item 8).
            if whatsapp_setting_key and not await should_send_salon_whatsapp(token_data.get('salon_id', ''), whatsapp_setting_key):
                logger.info(
                    f"WhatsApp suppressed by salon settings for {phone} (salon {token_data.get('salon_id')} - {whatsapp_setting_key})"
                )
                return
            # Then check customer WhatsApp preference (default ON).
            # Customer settings use the legacy 'whatsapp_booking_status_change' bucket
            # for events not covered by their per-event toggle.
            customer_key_map = {
                'whatsapp_salon_calling': 'whatsapp_turn_approaching',
                'whatsapp_booking_cancelled': 'whatsapp_booking_cancelled',
                'whatsapp_booking_rescheduled': 'whatsapp_booking_rescheduled',
            }
            customer_key = customer_key_map.get(whatsapp_setting_key, whatsapp_setting_key)
            if customer_key and not await should_send_customer_whatsapp(phone, customer_key):
                logger.info(f"WhatsApp notification suppressed for {phone} (customer setting {customer_key} is OFF)")
                return
            full_message = message + action_links if action_links else message
            result = await send_whatsapp_notification(phone, full_message, notification_type)
            logger.info(f"Notification sent: {notification_type} to {phone}, status: {result.get('status')}")
            
    except Exception as e:
        logger.error(f"Failed to send notification {notification_type}: {str(e)}")

async def check_and_notify_nearby_tokens(salon_id: str, barber_id: str, date: str, current_token_number: str):
    """Check and notify customers who are 3, 2 or 1 tokens away in THIS barber's queue (in-app + WhatsApp).

    Note: With salon-wide token numbering, 'tokens_away' is computed by barber-queue FIFO position
    (waiting + called), NOT by raw numeric token sequence.
    """
    try:
        waiting_tokens = await db.tokens.find(
            {"salon_id": salon_id, "barber_id": barber_id, "date": date, "status": "waiting"},
            {"_id": 0}
        ).sort("created_at", 1).to_list(100)

        for idx, token in enumerate(waiting_tokens):
            # Position in barber's queue (1-indexed), where 1 is "next up"
            # idx==0 means next in line → 1 away, idx==1 → 2 away, idx==2 → 3 away
            tokens_away = idx + 1
            phone = token.get('phone')
            token_number = token.get('token_number')

            # User spec (Item 8): only send WhatsApp 'token_approaching' for
            # tokens_away of 1 or 2. In-app notifications still go for 1/2/3.
            if tokens_away not in (3, 2, 1) or not phone:
                continue

            notified_field = f"notified_{tokens_away}_away"
            if token.get(notified_field):
                continue

            # Message text matches PRD
            if tokens_away == 1:
                title = "Please proceed to the salon chair"
                msg_body = "Please proceed to the salon chair. You're up next."
            elif tokens_away == 2:
                title = "Your turn is coming soon"
                msg_body = "Your turn is coming soon. Please be ready."
            else:
                title = "You're 3 customers away"
                msg_body = f"You are 3 customers away. Your token {token_number} will be called shortly."

            await create_in_app_notification(
                user_type="customer",
                user_id=phone,
                title=title,
                message=msg_body,
                notification_type=f"turn_{tokens_away}_away",
                setting_key="turn_approaching",
                salon_id=salon_id,
                related_id=token.get('id', ''),
            )

            # WhatsApp 'token_approaching' template — only for 1 or 2 away.
            if tokens_away in (1, 2):
                salon_ok = await should_send_salon_whatsapp(salon_id, 'whatsapp_token_approaching')
                customer_ok = await should_send_customer_whatsapp(phone, 'whatsapp_turn_approaching')
                if salon_ok and customer_ok:
                    # Figure out the currently-being-served token at this barber.
                    serving_doc = await db.tokens.find_one(
                        {
                            "salon_id": salon_id,
                            "barber_id": barber_id,
                            "date": date,
                            "status": "in_progress",
                        },
                        {"_id": 0, "token_number": 1},
                    )
                    current_serving = str(serving_doc.get("token_number")) if serving_doc else "—"

                    salon_doc = await db.salons.find_one({"id": salon_id}, {"_id": 0, "name": 1})
                    salon_name_for_msg = (salon_doc or {}).get("name") or "the salon"

                    await send_token_approaching_template(
                        phone_number=phone,
                        customer_name=token.get('customer_name') or 'Customer',
                        token_number=token_number,
                        tokens_away=tokens_away,
                        salon_name=salon_name_for_msg,
                        barber_name=token.get('barber_name') or 'your stylist',
                        current_serving=current_serving,
                    )

            await db.tokens.update_one(
                {"id": token.get('id')},
                {"$set": {notified_field: True}}
            )
            logger.info(f"Notified token #{token_number} ({tokens_away} away) via barber queue position")

    except Exception as e:
        logger.error(f"Failed to check nearby tokens: {str(e)}")

async def generate_and_send_invoice(token_id: str):
    """Generate invoice PDF and send via WhatsApp"""
    try:
        # Get token details
        token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
        if not token:
            raise Exception("Token not found")
        
        # Get salon details
        salon = await db.salons.find_one({"id": token['salon_id']}, {"_id": 0})
        if not salon:
            raise Exception("Salon not found")
        
        # Get services details
        services_data = []
        total_amount = 0
        
        for service_id in token.get('selected_services', []):
            service = await db.services.find_one({"id": service_id}, {"_id": 0})
            if service:
                # For simplicity, no discount here - can be added later
                price = service.get('base_price', 0)
                services_data.append({
                    "name": service.get('service_name'),
                    "price": price,
                    "discount": 0,
                    "amount": price
                })
                total_amount += price
        
        # Calculate tax if GST registered
        is_tax_invoice = salon.get('is_gst_registered', False)
        tax_rate = salon.get('tax_rate', 9.0)
        
        if is_tax_invoice:
            subtotal = total_amount
            cgst = subtotal * (tax_rate / 100)
            sgst = subtotal * (tax_rate / 100)
            total = subtotal + cgst + sgst
        else:
            subtotal = total_amount
            cgst = 0
            sgst = 0
            total = subtotal
        
        # Generate invoice number using salon's prefix and counter
        invoice_prefix = salon.get('invoice_prefix', 'INV')
        current_number = salon.get('current_invoice_number')
        
        # If current_invoice_number not set, initialize from invoice_start_number
        if current_number is None:
            current_number = salon.get('invoice_start_number', 1)
            # Set it in database
            await db.salons.update_one(
                {"id": salon['id']},
                {"$set": {"current_invoice_number": current_number}}
            )
        
        # Calculate padding based on number size (minimum 4 digits)
        padding = max(4, len(str(current_number)))
        invoice_no = f"{invoice_prefix}{str(current_number).zfill(padding)}"
        
        # Increment invoice counter for next time
        await db.salons.update_one(
            {"id": salon['id']},
            {"$inc": {"current_invoice_number": 1}}
        )        
        # Prepare invoice data
        invoice_data = {
            "salon": {
                "salon_name": salon.get('salon_name', 'Salon'),
                "address": salon.get('address', ''),
                "gstin": salon.get('gstin'),
                "logo_url": salon.get('logo_url')
            },
            "customer": {
                "name": token.get('customer_name', 'Customer'),
                "phone": token.get('phone', '')
            },
            "invoice_no": invoice_no,
            "date": datetime.now().strftime('%d/%m/%Y'),
            "services": services_data,
            "subtotal": subtotal,
            "cgst": cgst,
            "sgst": sgst,
            "tax_rate": tax_rate,
            "total": total,
            "payment_method": "UPI",
            "is_tax_invoice": is_tax_invoice
        }
        
        # Generate PDF
        pdf_data = generate_invoice_pdf(invoice_data)
        
        # Convert PDF to base64 for storage
        import base64
        pdf_base64 = base64.b64encode(pdf_data).decode('utf-8')
        
        # Create invoice ID
        invoice_id = str(uuid.uuid4())
        
        # Generate invoice link
        invoice_link = f"{os.getenv('REACT_APP_BACKEND_URL', 'http://localhost:8001')}/api/invoices/{invoice_id}/view"
        
        # Send link via WhatsApp
        message = f"""
📄 *Invoice Generated*

Hello {token.get('customer_name')}!

Your service at {salon.get('salon_name')} is complete.

*Invoice #{invoice_no}*
Total Amount: ₹{total:.2f}

🔗 View/Download Invoice:
{invoice_link}

Thank you for visiting us! 💈
        """.strip()
        
        # Send message with link
        result = await send_whatsapp_notification(
            token.get('phone'),
            message,
            'invoice_sent'
        )
        
        # Store invoice record with base64 PDF
        invoice_record = {
            "id": invoice_id,
            "token_id": token_id,
            "invoice_no": invoice_no,
            "salon_id": token['salon_id'],
            "branch_id": token.get('branch_id'),
            "customer_name": token.get('customer_name'),
            "customer_phone": token.get('phone'),
            "invoice_data": invoice_data,
            "pdf_base64": pdf_base64,
            "amount": total,
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "sent_status": result.get('status'),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.invoices.insert_one(invoice_record)
        
        # Update token with invoice_id
        await db.tokens.update_one(
            {"id": token_id},
            {"$set": {"invoice_id": invoice_id}}
        )
        
        logger.info(f"Invoice {invoice_no} generated and link sent to {token.get('phone')}")
        
        return {
            "success": True,
            "invoice_id": invoice_id,
            "invoice_number": invoice_no,
            "invoice_link": invoice_link
        }
        
    except Exception as e:
        logger.error(f"Failed to generate and send invoice: {str(e)}")
        return {"success": False, "error": str(e)}


# ============ SUBSCRIPTION HELPERS ============

import cashfree_service

GRACE_PERIOD_DAYS = 0  # No grace period; subscription becomes premium immediately on payment

async def get_active_plan() -> dict:
    """Return the default active plan, or fall back to any active plan.

    Always normalises so that `price_per_branch` is populated — legacy plans
    that only have `price` will have `price_per_branch = price`.
    """
    plan = await db.subscription_plans.find_one(
        {"status": "active", "is_default": True}, {"_id": 0}
    )
    if not plan:
        plan = await db.subscription_plans.find_one({"status": "active"}, {"_id": 0})
    if plan:
        # Phase 2 (Part C): ensure price_per_branch is always available
        if plan.get("price_per_branch") in (None, 0, 0.0):
            plan["price_per_branch"] = float(plan.get("price") or 0)
    return plan


async def count_billable_branches(salon_id: str) -> int:
    """Phase 2 (Part C): how many branches the salon should be charged for.

    Rule: every branch with status == 'active' counts. The main branch counts.
    The minimum is 1 (so a brand-new salon with no branches yet is billed for 1).
    """
    n = await db.salon_branches.count_documents(
        {"salon_id": salon_id, "status": "active"}
    )
    return max(int(n or 0), 1)


async def get_active_branch_ids(salon_id: str) -> List[str]:
    docs = await db.salon_branches.find(
        {"salon_id": salon_id, "status": "active"}, {"_id": 0, "id": 1}
    ).to_list(length=None)
    return [d["id"] for d in docs if d.get("id")]

async def get_current_subscription(salon_id: str) -> Optional[dict]:
    """
    Return the most recent (latest expiry) subscription record for the salon.
    Returns None if salon has never paid.

    Phase 5: 'granted' payments (platform-comp / grant-pro) count as valid too.
    """
    sub = await db.salon_subscriptions.find_one(
        {"salon_id": salon_id, "payment_status": {"$in": ["paid", "granted"]}},
        {"_id": 0},
        sort=[("expiry_date", -1)],
    )
    return sub

async def get_subscription_status(salon_id: str) -> dict:
    """
    Compute subscription state for a salon.

    Phase 2 (Part C): also returns per-branch pricing fields so the frontend
    can render "₹X/month/branch × N branches = ₹Y/month" and the
    "Next renewal will be ₹Z" banner.

    Returns dict with: is_premium, status, expiry_date, days_remaining,
    subscription, plan, plus per-branch fields:
      - price_per_branch
      - billable_branch_count
      - active_branch_count (always current, used for next_renewal_amount)
      - next_renewal_amount
      - base_amount / total_amount / discount_code_applied / discount_amount
        (from the active subscription if any)
      - branches_added_mid_cycle (bool — to show the deferred-billing banner)
      - trial_used (bool — salon already redeemed its 30-day free trial)
      - is_trial (bool — current active subscription is the free trial)
    """
    sub = await get_current_subscription(salon_id)
    plan = await get_active_plan()
    salon_doc = await db.salons.find_one({"id": salon_id}, {"_id": 0, "trial_used": 1})
    now = datetime.now(timezone.utc)
    is_premium = False
    status = "free"
    expiry_date = None
    days_remaining = None

    if sub and sub.get("expiry_date"):
        try:
            expiry_dt = datetime.fromisoformat(sub["expiry_date"].replace("Z", "+00:00"))
            if expiry_dt.tzinfo is None:
                expiry_dt = expiry_dt.replace(tzinfo=timezone.utc)
            expiry_date = sub["expiry_date"]
            delta = expiry_dt - now
            days_remaining = max(0, delta.days)
            if delta.total_seconds() > 0:
                is_premium = True
                status = "active"
            else:
                status = "expired"
        except Exception as e:
            logger.warning(f"[Subscription] Bad expiry_date for salon {salon_id}: {e}")

    # Phase 2 (Part C) — per-branch pricing context
    price_per_branch = float((plan or {}).get("price_per_branch") or (plan or {}).get("price") or 0)
    active_branch_count = await count_billable_branches(salon_id)
    next_renewal_amount = round(price_per_branch * active_branch_count, 2)

    # Pull snapshot fields from the active sub (if any)
    sub_billable = (sub or {}).get("billable_branch_count")
    sub_base_amount = (sub or {}).get("base_amount")
    sub_total_amount = (sub or {}).get("total_amount")
    sub_discount_code = (sub or {}).get("discount_code_applied")
    sub_discount_amount = (sub or {}).get("discount_amount")

    # Mid-cycle branch addition detection: salon is currently premium AND has
    # more active branches now than were paid for on the current subscription.
    branches_added_mid_cycle = False
    if is_premium and sub_billable is not None:
        try:
            branches_added_mid_cycle = active_branch_count > int(sub_billable)
        except Exception:
            branches_added_mid_cycle = False

    return {
        "is_premium": is_premium,
        "status": status,
        "expiry_date": expiry_date,
        "days_remaining": days_remaining,
        "subscription": sub,
        "plan": plan,
        # Phase 2 (Part C) additions
        "price_per_branch": price_per_branch,
        "billable_branch_count": int(sub_billable) if sub_billable is not None else active_branch_count,
        "active_branch_count": active_branch_count,
        "next_renewal_amount": next_renewal_amount,
        "base_amount": float(sub_base_amount) if sub_base_amount is not None else round(price_per_branch * active_branch_count, 2),
        "total_amount": float(sub_total_amount) if sub_total_amount is not None else round(price_per_branch * active_branch_count, 2),
        "discount_code_applied": sub_discount_code,
        "discount_amount": float(sub_discount_amount) if sub_discount_amount is not None else 0.0,
        "branches_added_mid_cycle": branches_added_mid_cycle,
        # Phase 5 (Part A) — platform admin override summary
        "grant_type": (sub or {}).get("grant_type"),  # None | "comp" | "grant_pro"
        "max_branches_override": (sub or {}).get("max_branches_override"),
        "max_branches_effective": (
            (sub or {}).get("max_branches_override")
            or (sub or {}).get("max_branches")
            or (plan or {}).get("max_branches")
        ),
        "trial_ends_at": (sub or {}).get("trial_ends_at"),
        "is_platform_granted": bool((sub or {}).get("payment_status") == "granted"),
        # 30-day free trial fields (Feb 2026)
        "trial_used": bool((salon_doc or {}).get("trial_used") or (sub or {}).get("is_trial")),
        "is_trial": bool((sub or {}).get("is_trial")),
    }

async def enforce_premium_or_within_limit(salon_id: str, *, resource: str) -> None:
    """
    Block creation of premium-only resources when salon is on free plan.
    resource in {"staff", "branch"}
    Free-plan caps: max 1 staff, 0 branches.
    Raises HTTPException(402) on block.
    """
    sub_status = await get_subscription_status(salon_id)
    if sub_status["is_premium"]:
        return  # premium = no limit

    plan = sub_status.get("plan") or {}
    if resource == "staff":
        active_count = await db.barbers.count_documents(
            {"salon_id": salon_id, "is_active": True}
        )
        max_allowed = 1
        if active_count >= max_allowed:
            raise HTTPException(
                status_code=402,
                detail={
                    "error": "subscription_required",
                    "message": "Free plan allows only 1 staff member. Upgrade to SalonHub Pro for unlimited staff.",
                    "limit_type": "max_staff",
                    "current_count": active_count,
                    "max_allowed": max_allowed,
                    "plan_price": float(plan.get("price", 499.0)),
                    "plan_name": plan.get("plan_name", "SalonHub Pro"),
                },
            )
    elif resource == "branch":
        existing_count = await db.salon_branches.count_documents({"salon_id": salon_id})
        max_allowed = 1  # only the auto-created main branch is allowed on free plan
        if existing_count >= max_allowed:
            raise HTTPException(
                status_code=402,
                detail={
                    "error": "subscription_required",
                    "message": "Free plan does not allow multiple branches. Upgrade to SalonHub Pro to add branches.",
                    "limit_type": "max_branches",
                    "current_count": existing_count,
                    "max_allowed": max_allowed,
                    "plan_price": float(plan.get("price", 499.0)),
                    "plan_name": plan.get("plan_name", "SalonHub Pro"),
                },
            )

def _add_billing_cycle(start_dt: datetime, billing_cycle: str) -> datetime:
    """Add billing cycle to a datetime."""
    if billing_cycle == "yearly":
        try:
            return start_dt.replace(year=start_dt.year + 1)
        except ValueError:
            return start_dt.replace(year=start_dt.year + 1, day=28)
    return start_dt + timedelta(days=30)


# ============ INITIALIZATION ============

async def initialize_data():
    """Initialize default data"""
    
    # Initialize default salon
    salon_count = await db.salons.count_documents({})
    if salon_count == 0:
        default_salon = {
            "id": str(uuid.uuid4()),
            "salon_name": "The Looks Unisex Salon",
            "owner_name": "Owner Name",
            "phone": "+917503070727",
            "email": "salon@example.com",
            "address": "123 Main Street, Bangalore, Karnataka",
            "latitude": 12.9716,
            "longitude": 77.5946,
            "upi_id": "salon@upi",
            "payment_timing": "after",
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.salons.insert_one(default_salon)
        salon_id = default_salon["id"]
    else:
        salon = await db.salons.find_one({}, {"_id": 0})
        salon_id = salon["id"]
    
    # Initialize services (Jul 2026 — only starter "General" services)
    service_count = await db.services.count_documents({})
    if service_count == 0:
        from predefined_services import PREDEFINED_SERVICES as _STARTER_SERVICES
        services = []
        for sd in _STARTER_SERVICES:
            services.append({
                "id": str(uuid.uuid4()),
                "service_name": sd["service_name"],
                "description": sd.get("description") or "",
                "category": sd.get("category") or "General",
                "gender_tag": sd.get("gender_tag") or "Unisex",
                "default_duration": sd.get("default_duration") or 30,
                "base_price": sd.get("base_price") or 0,
                "price_type": sd.get("price_type") or "fixed",
                "is_active": True,
            })
        if services:
            await db.services.insert_many(services)
        service_ids = [s["id"] for s in services]
    else:
        services = await db.services.find({}, {"_id": 0}).to_list(100)
        service_ids = [s["id"] for s in services]
    
    # Initialize barbers
    barber_count = await db.barbers.count_documents({})
    if barber_count == 0:
        barbers = [
            {
                "id": str(uuid.uuid4()),
                "name": "Imran",
                "salon_id": salon_id,
                "experience": 8,
                "category": "master",
                "mobile": "+919876543211",
                "queue_status": "available",
                "is_active": True,
                "is_barber": True  # Visible to customers
            },
            {
                "id": str(uuid.uuid4()),
                "name": "Abdul",
                "salon_id": salon_id,
                "experience": 5,
                "category": "star",
                "mobile": "+919876543212",
                "queue_status": "available",
                "is_active": True,
                "is_barber": True  # Visible to customers
            }
        ]
        await db.barbers.insert_many(barbers)
        
        # Set pricing for each barber
        for barber in barbers:
            for i, service_id in enumerate(service_ids):
                # Imran (master) has higher prices
                if barber["name"] == "Imran":
                    price = services[i]["base_price"] * 1.2
                else:
                    price = services[i]["base_price"]
                
                pricing = {
                    "id": str(uuid.uuid4()),
                    "barber_id": barber["id"],
                    "service_id": service_id,
                    "price": price,
                    "is_available": True
                }
                await db.barber_services.insert_one(pricing)
        
        # Create salon admin user
        admin_user = {
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            "name": "Admin",
            "login_id": "admin",
            "mobile": "+917503070727",
            "password_hash": pwd_context.hash("salon123"),
            "role": "admin",
            "status": "active",
            "permissions": {
                "can_edit_salon": True,
                "can_access_analytics": True,
                "can_access_financials": True,
                "can_delete_salon": True,
                "can_access_services": True,
                "can_access_gallery": True,
                "can_access_staff": True,
                "can_view_all_staff": True,
                "can_access_marketing": True
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.salon_users.insert_one(admin_user)
        logger.info("Created admin user: login_id='admin' (password from seed config)")

    # Seed default subscription plan if none exists
    plan_count = await db.subscription_plans.count_documents({})
    if plan_count == 0:
        default_plan = {
            "id": str(uuid.uuid4()),
            "plan_name": "SalonHub Pro",
            "price": 999.0,
            "price_per_branch": 999.0,  # Phase 2 (Part C)
            "billing_cycle": "monthly",
            "max_staff": None,
            "max_branches": None,
            "features": [
                "Unlimited Staff",
                "Multiple Branches",
                "Branch Management",
                "Staff Transfers",
                "Attendance System",
            ],
            "status": "active",
            "is_default": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.subscription_plans.insert_one(default_plan)
        logger.info("Seeded default SalonHub Pro plan @ ₹999/month/branch")

    # Seed yearly plan if none exists (₹9999/year/branch — save 17%)
    yearly_existing = await db.subscription_plans.find_one(
        {"billing_cycle": "yearly"}, {"_id": 0, "id": 1}
    )
    if not yearly_existing:
        yearly_plan = {
            "id": str(uuid.uuid4()),
            "plan_name": "SalonHub Pro (Yearly)",
            "price": 9999.0,
            "price_per_branch": 9999.0,
            "billing_cycle": "yearly",
            "max_staff": None,
            "max_branches": None,
            "features": [
                "Unlimited Staff",
                "Multiple Branches",
                "Branch Management",
                "Staff Transfers",
                "Attendance System",
                "Save ~17% vs monthly",
            ],
            "status": "active",
            "is_default": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.subscription_plans.insert_one(yearly_plan)
        logger.info("Seeded SalonHub Pro Yearly plan @ ₹9999/year/branch")

    # One-shot price migration: bump legacy ₹499 default plan to ₹999.
    # Guarded by a system_flags doc so it only runs once.
    price_flag = await db.system_flags.find_one(
        {"key": "subscription_default_price_v3_999"}, {"_id": 0}
    )
    if not price_flag or not price_flag.get("value"):
        res = await db.subscription_plans.update_many(
            {"price": 499.0, "billing_cycle": "monthly"},
            {"$set": {"price": 999.0, "price_per_branch": 999.0}},
        )
        await db.system_flags.update_one(
            {"key": "subscription_default_price_v3_999"},
            {"$set": {"value": True, "ran_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True,
        )
        if res.modified_count:
            logger.info(f"[Subscription V3] Bumped {res.modified_count} legacy monthly plans 499→999")

    # Seed/refresh a 1-year premium subscription for the test salon (+917503070727)
    # so the admin doesn't see "Upgrade to SalonHub Pro" prompts when using premium features.
    try:
        test_salon = await db.salons.find_one({"phone": "+917503070727"}, {"_id": 0, "id": 1})
        if test_salon:
            test_salon_id = test_salon["id"]
            # Look for an existing paid subscription that is still valid for at least ~11 months.
            existing = await db.salon_subscriptions.find_one(
                {"salon_id": test_salon_id, "payment_status": "paid"},
                {"_id": 0},
                sort=[("expiry_date", -1)],
            )
            needs_seed = True
            if existing and existing.get("expiry_date"):
                try:
                    exp_dt = datetime.fromisoformat(existing["expiry_date"].replace("Z", "+00:00"))
                    if exp_dt.tzinfo is None:
                        exp_dt = exp_dt.replace(tzinfo=timezone.utc)
                    if (exp_dt - datetime.now(timezone.utc)).days >= 330:
                        needs_seed = False
                except Exception:
                    needs_seed = True

            if needs_seed:
                # Pick the default active plan (or any active plan)
                plan = await db.subscription_plans.find_one(
                    {"status": "active", "is_default": True}, {"_id": 0}
                )
                if not plan:
                    plan = await db.subscription_plans.find_one({"status": "active"}, {"_id": 0})

                now_dt = datetime.now(timezone.utc)
                expiry_dt = now_dt + timedelta(days=365)
                # Phase 2 (Part C): include per-branch pricing snapshot
                price_per_branch = float(
                    (plan or {}).get("price_per_branch")
                    or (plan or {}).get("price")
                    or 0
                )
                # Use real branches if they exist; otherwise default to 1.
                snapshot_branch_ids = await get_active_branch_ids(test_salon_id)
                billable = max(len(snapshot_branch_ids), 1)
                base_amount = round(price_per_branch * billable, 2)

                sub_doc = {
                    "id": str(uuid.uuid4()),
                    "salon_id": test_salon_id,
                    "plan_id": (plan or {}).get("id"),
                    "plan_name": (plan or {}).get("plan_name", "SalonHub Pro"),
                    "price": float((plan or {}).get("price", 0.0)),
                    "subscription_status": "active",
                    "start_date": now_dt.isoformat(),
                    "expiry_date": expiry_dt.isoformat(),
                    "payment_status": "paid",
                    "cashfree_order_id": None,
                    "cashfree_payment_id": "TEST_SEED_PREMIUM_1Y",
                    "auto_renew": False,
                    "is_test_seed": True,
                    # Phase 2 (Part C) snapshot fields
                    "billable_branch_count": billable,
                    "price_per_branch_snapshot": price_per_branch,
                    "branch_ids_snapshot": snapshot_branch_ids,
                    "base_amount": base_amount,
                    "discount_code_applied": None,
                    "discount_amount": 0.0,
                    "total_amount": base_amount,
                    "created_at": now_dt.isoformat(),
                    "updated_at": now_dt.isoformat(),
                }
                await db.salon_subscriptions.insert_one(sub_doc.copy())
                logger.info(
                    f"[Subscription Seed] Granted 1-year premium subscription to test salon "
                    f"{test_salon_id} (phone +917503070727), expires {expiry_dt.isoformat()}"
                )
    except Exception as e:
        logger.warning(f"[Subscription Seed] Skipped seeding test premium subscription: {e}")

async def allocate_future_tokens():
    """Run at 5-6 AM to allocate tokens for future bookings"""
    logger.info("Running future token allocation...")
    today = datetime.now(timezone.utc).date().isoformat()
    
    # Get all unallocated future bookings for today
    future_bookings = await db.tokens.find({
        "booking_type": "future",
        "date": today,
        "allocated_at": None
    }, {"_id": 0}).to_list(1000)
    
    # Group by salon and barber
    grouped = {}
    for booking in future_bookings:
        key = f"{booking['salon_id']}_{booking['barber_id']}"
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(booking)
    
    # Allocate tokens
    for key, bookings in grouped.items():
        # Sort by time slot
        bookings.sort(key=lambda x: x['time_slot'])
        
        for i, booking in enumerate(bookings, start=1):
            await db.tokens.update_one(
                {"id": booking["id"]},
                {"$set": {
                    "token_number": i,
                    "allocated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
    
    logger.info(f"Allocated {len(future_bookings)} future tokens")

# ============ WEBSOCKET EVENTS ============

@sio.event
async def connect(sid, environ):
    logger.info(f"Client connected: {sid}")

@sio.event
async def disconnect(sid):
    logger.info(f"Client disconnected: {sid}")



# ============ PREDEFINED SERVICES INITIALIZATION ============

from predefined_services import PREDEFINED_SERVICES, PREDEFINED_PACKAGES

async def initialize_predefined_services_for_salon(salon_id: str):
    """Initialize predefined services and packages for a salon on first access"""
    
    # Check if already initialized
    initialized = await db.salon_initialized.find_one({"salon_id": salon_id})
    if initialized:
        return {"already_initialized": True}
    
    # Always ensure predefined services exist (add missing ones)
    existing_services = await db.services.find({}, {"_id": 0, "service_name": 1, "category": 1}).to_list(1000)
    existing_set = {(s["service_name"], s.get("category", "General")) for s in existing_services}
    
    services_to_insert = []
    for service_data in PREDEFINED_SERVICES:
        service_key = (service_data["service_name"], service_data.get("category", "General"))
        if service_key not in existing_set:
            service = {
                "id": str(uuid.uuid4()),
                **service_data,
                "is_active": True,
                "images": [],
                "is_favorite": False
            }
            services_to_insert.append(service)
    
    if services_to_insert:
        await db.services.insert_many(services_to_insert)
        logger.info(f"Inserted {len(services_to_insert)} new predefined services")
    
    # Get all service IDs
    all_services = await db.services.find({}, {"_id": 0, "id": 1}).to_list(1000)
    service_ids = [s["id"] for s in all_services]
    
    # Initialize salon_services (all disabled by default - salon can enable them)
    # Only add services that don't already exist for this salon
    existing_salon_services = await db.salon_services.find(
        {"salon_id": salon_id},
        {"_id": 0, "service_id": 1}
    ).to_list(1000)
    existing_salon_service_ids = {ss["service_id"] for ss in existing_salon_services}
    
    salon_services_to_insert = []
    for service_id in service_ids:
        if service_id not in existing_salon_service_ids:
            salon_service = {
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "service_id": service_id,
                "is_enabled": False,  # Salon must enable services they want to offer
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            salon_services_to_insert.append(salon_service)
    
    if salon_services_to_insert:
        await db.salon_services.insert_many(salon_services_to_insert)
        logger.info(f"Inserted {len(salon_services_to_insert)} salon_service mappings")
    
    # Initialize predefined packages for this salon
    packages_to_insert = []
    for package_data in PREDEFINED_PACKAGES:
        package = {
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            **package_data,
            "service_ids": [],  # Salon will add services to packages
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        packages_to_insert.append(package)
    
    if packages_to_insert:
        await db.salon_packages.insert_many(packages_to_insert)
    
    # Mark as initialized
    await db.salon_initialized.insert_one({
        "salon_id": salon_id,
        "initialized_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "initialized": True,
        "new_services_added": len(services_to_insert),
        "services_count": len(service_ids),
        "packages_count": len(packages_to_insert)
    }

# ============ API ROUTES ============

@api_router.get("/")
async def root():
    return {"message": "The Looks Salon API v3.0 - Multi-Salon Edition"}

# ============ SALON ROUTES ============

@api_router.get("/salons", response_model=List[Salon])
async def get_salons(lat: Optional[float] = None, lng: Optional[float] = None, radius: float = 2):
    """Get all salons, optionally filtered by location"""
    salons = await db.salons.find({"is_active": True}, {"_id": 0}).to_list(100)
    
    if lat and lng:
        # Filter by distance
        nearby_salons = []
        for salon in salons:
            distance = calculate_distance(lat, lng, salon["latitude"], salon["longitude"])
            if distance <= radius:
                salon["distance"] = round(distance, 2)
                nearby_salons.append(salon)
        return sorted(nearby_salons, key=lambda x: x["distance"])
    
    return salons

@api_router.get("/salons/search")
async def search_salons(name: Optional[str] = None, city: Optional[str] = None):
    """Search salons by name and/or city (case-insensitive partial match)"""
    if not name and not city:
        raise HTTPException(status_code=400, detail="Please provide a search query (name or city)")
    
    query = {"is_active": True}
    
    # Build combined search query
    conditions = []
    if name and len(name) >= 1:
        conditions.append({"salon_name": {"$regex": name, "$options": "i"}})
    if city and len(city) >= 1:
        conditions.append({"city": {"$regex": city, "$options": "i"}})
    
    if len(conditions) == 1:
        query.update(conditions[0])
    elif len(conditions) > 1:
        query["$and"] = conditions
    
    salons = await db.salons.find(query, {"_id": 0}).limit(50).to_list(50)
    
    return {"salons": [Salon(**s) for s in salons]}

@api_router.get("/salons/by-city")
async def get_salons_by_city(city: str):
    """Get salons by city"""
    if not city or len(city) < 2:
        raise HTTPException(status_code=400, detail="City name must be at least 2 characters")
    
    salons = await db.salons.find(
        {
            "is_active": True,
            "city": {"$regex": city, "$options": "i"}
        },
        {"_id": 0}
    ).limit(50).to_list(50)
    
    return {"salons": [Salon(**s) for s in salons]}

@api_router.get("/cities")
async def get_cities():
    """Get list of unique cities with salons"""
    cities = await db.salons.distinct("city", {"is_active": True})
    return {"cities": [c for c in cities if c]}


def _build_location_row(salon: dict, branch: dict, lat: Optional[float], lng: Optional[float]) -> dict:
    """Compose one customer-facing location row per branch.
    Uses the BRANCH's address/coords when present (so each branch is its own
    bookable destination), with fallbacks to the salon for fields the branch
    doesn't override (logo, gallery, gender_tag, ratings)."""
    # Each branch is shown as a separate "salon" on the Find a Salon screen.
    branch_addr = branch.get("address") or salon.get("address")
    branch_city = branch.get("city") or salon.get("city")
    branch_lat = branch.get("latitude") if branch.get("latitude") is not None else salon.get("latitude")
    branch_lng = branch.get("longitude") if branch.get("longitude") is not None else salon.get("longitude")
    branch_phone = branch.get("phone") or salon.get("phone")

    salon_name_base = salon.get("salon_name") or "Salon"
    branch_name = branch.get("branch_name") or "Main Branch"
    # If only a Main Branch, just show the salon name.
    display_name = salon_name_base if branch.get("is_main_branch") else f"{salon_name_base} – {branch_name}"

    row = {
        "id": branch["id"],  # the BRANCH id is what the customer card represents
        "salon_id": salon["id"],
        "branch_id": branch["id"],
        "branch_name": branch_name,
        "is_main_branch": bool(branch.get("is_main_branch")),
        "salon_name": display_name,
        "address": branch_addr,
        "city": branch_city,
        "latitude": branch_lat,
        "longitude": branch_lng,
        "phone": branch_phone,
        "logo_url": salon.get("logo_url"),
        "photo_gallery": salon.get("photo_gallery") or [],
        "gender_tag": salon.get("gender_tag"),
        "rating": salon.get("rating") or 0,
        "total_reviews": salon.get("total_reviews") or 0,
        "manual_toggle": salon.get("manual_toggle"),
    }
    if lat is not None and lng is not None and branch_lat is not None and branch_lng is not None:
        row["distance"] = round(calculate_distance(lat, lng, branch_lat, branch_lng), 2)
    return row


async def _list_salon_locations(
    name: Optional[str] = None,
    city: Optional[str] = None,
    lat: Optional[float] = None,
    lng: Optional[float] = None,
    radius: Optional[float] = None,
) -> List[dict]:
    """Fan out one row per active branch joined with salon details. Used by the
    customer-side 'Find a salon' screen so chains appear as one entry per branch."""
    salon_query = {"is_active": True}
    if name:
        salon_query["salon_name"] = {"$regex": name, "$options": "i"}
    salons = await db.salons.find(salon_query, {"_id": 0}).to_list(500)
    if not salons:
        return []

    salon_ids = [s["id"] for s in salons]
    branch_query = {"salon_id": {"$in": salon_ids}, "status": "active"}
    if city:
        # Match either the branch city or the salon city below
        branch_query_or = [{"city": {"$regex": city, "$options": "i"}}]
        branch_query["$or"] = branch_query_or
    branches = await db.salon_branches.find(branch_query, {"_id": 0}).to_list(2000)

    salon_by_id = {s["id"]: s for s in salons}
    rows: List[dict] = []
    for b in branches:
        s = salon_by_id.get(b["salon_id"])
        if not s:
            continue
        # When city was passed but the BRANCH didn't match, also accept salon-level city match
        if city and not b.get("city") and s.get("city"):
            if city.lower() not in s["city"].lower():
                continue
        rows.append(_build_location_row(s, b, lat, lng))

    if lat is not None and lng is not None and radius is not None:
        rows = [r for r in rows if r.get("distance") is not None and r["distance"] <= radius]
        rows.sort(key=lambda r: r.get("distance", float("inf")))
    elif lat is not None and lng is not None:
        rows.sort(key=lambda r: r.get("distance", float("inf")))
    else:
        rows.sort(key=lambda r: ((r.get("salon_name") or "").lower(), not r.get("is_main_branch")))

    return rows


@api_router.get("/public/salon-locations")
async def list_salon_locations(
    lat: Optional[float] = None,
    lng: Optional[float] = None,
    radius: Optional[float] = None,
    name: Optional[str] = None,
    city: Optional[str] = None,
):
    """Public 'Find a salon' listing — one row per active branch.

    Used by the customer-facing salon picker. A salon with N branches appears N
    times here (each row carries `salon_id` + `branch_id`). The card title
    follows the rule: main branch → just the salon name; non-main → "Salon – Branch".
    """
    rows = await _list_salon_locations(name=name, city=city, lat=lat, lng=lng, radius=radius)
    return rows


@api_router.get("/public/salons/{salon_id}/branches", response_model=List[Branch])
async def list_public_branches(salon_id: str):
    """Public listing of a salon's active branches — used by the customer
    booking page to populate per-branch selectors / verification."""
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0, "id": 1})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    branches = await db.salon_branches.find(
        {"salon_id": salon_id, "status": "active"}, {"_id": 0}
    ).to_list(length=None)
    branches.sort(key=lambda b: (not b.get("is_main_branch"), b.get("created_at", "")))
    return [Branch(**b) for b in branches]


@api_router.get("/salons/{salon_id}", response_model=Salon)
async def get_salon(salon_id: str):
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    return Salon(**salon)

@api_router.post("/salons", response_model=Salon)
async def create_salon(salon: SalonCreate, current_salon=Depends(get_current_salon)):
    salon_dict = salon.model_dump()
    salon_dict["id"] = str(uuid.uuid4())
    salon_dict["is_active"] = True
    salon_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.salons.insert_one(salon_dict)
    # Auto-create Main Branch so customer search can find the salon immediately.
    try:
        await ensure_main_branch_for_salon(salon_dict)
    except Exception as e:
        logger.warning(f"[create_salon] Main Branch auto-create failed for salon {salon_dict['id']}: {e}")
    return Salon(**salon_dict)

@api_router.put("/salons/{salon_id}", response_model=Salon)
@api_router.patch("/salons/{salon_id}", response_model=Salon)
async def update_salon(salon_id: str, salon: SalonUpdate, current_user=Depends(get_current_salon_admin)):
    """Update salon profile (now supports partial updates)"""
    # Authorise: admin token must belong to the same salon (legacy 'salon' tokens
    # carry salon_id in 'sub'; new 'salon_admin' tokens carry it as 'salon_id').
    token_salon_id = current_user.get("salon_id") or current_user.get("sub")
    if token_salon_id != salon_id:
        raise HTTPException(status_code=403, detail="Not allowed for this salon")
    # Enforce granular RBAC: even admin-tokens flow through this — legacy admin
    # is fine, but staff-role admins with restricted salon_settings.edit_profile
    # will be blocked here.
    if not has_module_permission(current_user, "salon_settings", "edit_profile"):
        raise HTTPException(status_code=403, detail="Permission denied: salon_settings.edit_profile")

    existing = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    # Only update fields that are provided
    update_data = {k: v for k, v in salon.model_dump().items() if v is not None}
    
    # If invoice_start_number is being changed, reset current_invoice_number
    if 'invoice_start_number' in update_data:
        update_data['current_invoice_number'] = update_data['invoice_start_number']
    
    if update_data:
        await db.salons.update_one({"id": salon_id}, {"$set": update_data})
    
    updated = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    return Salon(**updated)

@api_router.delete("/salons/{salon_id}")
async def delete_salon(salon_id: str, current_salon=Depends(get_current_salon)):
    """Delete a salon and all associated data"""
    # Verify the salon exists
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    # Verify the requesting salon is deleting its own profile
    if current_salon.get("id") != salon_id:
        raise HTTPException(status_code=403, detail="You can only delete your own salon profile")
    
    # Delete all associated data
    await db.barbers.delete_many({"salon_id": salon_id})
    await db.barber_services.delete_many({"salon_id": salon_id})
    await db.salon_services.delete_many({"salon_id": salon_id})
    await db.tokens.delete_many({"salon_id": salon_id})
    await db.ratings.delete_many({"salon_id": salon_id})
    await db.salon_initialized.delete_many({"salon_id": salon_id})
    await db.salon_packages.delete_many({"salon_id": salon_id})
    
    # Delete the salon itself
    await db.salons.delete_one({"id": salon_id})
    
    return {"message": "Salon and all associated data deleted successfully"}

# ============================================================================
# BRANCH MANAGEMENT (Multi-Location)
# ============================================================================

def _check_salon_admin_for_salon(current_user: dict, salon_id: str):
    """Authorise a salon-admin (or legacy `salon` token) for a given salon_id.
    Raises 403 if not authorised."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    role = current_user.get("role")
    user_salon_id = current_user.get("salon_id") or current_user.get("sub")
    if user_salon_id != salon_id:
        raise HTTPException(status_code=403, detail="Access denied for this salon")
    if role not in ("salon_admin", "salon", "admin"):
        # Allow staff with edit permission for read-ish operations? No — branch admin
        # CRUD is admin-only by design; staff can list via the read endpoint which
        # uses get_current_salon_user.
        raise HTTPException(status_code=403, detail="Admin access required")


async def _generate_branch_qr(booking_url: str) -> str:
    """Generate a base64 PNG QR code for a given booking URL."""
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(booking_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return f"data:image/png;base64,{base64.b64encode(buffer.getvalue()).decode()}"


@api_router.get("/salons/{salon_id}/branches", response_model=List[Branch])
async def list_branches(
    salon_id: str,
    include_inactive: bool = False,
    current_user: dict = Depends(get_current_salon_user),
):
    """List all branches for a salon. Any authenticated salon user can list.
    Branch managers only see branches in their assigned_branch_ids."""
    user_salon_id = current_user.get("salon_id") or current_user.get("sub")
    if user_salon_id != salon_id:
        raise HTTPException(status_code=403, detail="Access denied for this salon")

    # Make sure a Main Branch exists (lazy migration safeguard)
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if salon:
        await ensure_main_branch_for_salon(salon)

    query = {"salon_id": salon_id}
    if not include_inactive:
        query["status"] = "active"

    # Branch-manager scoping — only see assigned branches.
    if is_branch_manager(current_user):
        assigned = assigned_branch_ids_for(current_user)
        if not assigned:
            return []
        query["id"] = {"$in": assigned}

    branches = await db.salon_branches.find(query, {"_id": 0}).to_list(length=None)
    # Sort: main branch first, then by created_at
    branches.sort(key=lambda b: (not b.get("is_main_branch"), b.get("created_at", "")))
    return [Branch(**b) for b in branches]


@api_router.post("/salons/{salon_id}/branches", response_model=Branch)
async def create_branch(
    salon_id: str,
    payload: BranchCreate,
    current_user: dict = Depends(get_current_salon_user),
):
    """Create a new branch under a salon."""
    _check_salon_admin_for_salon(current_user, salon_id)

    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")

    # Ensure main branch exists first (so the very first create might not collide)
    await ensure_main_branch_for_salon(salon)

    # Subscription paywall: free plan disallows adding branches beyond the auto-created main
    await enforce_premium_or_within_limit(salon_id, resource="branch")

    branch_dict = payload.model_dump()

    # Enforce per-salon uniqueness of branch_code (case-insensitive). Empty/None codes are skipped.
    code = (branch_dict.get("branch_code") or "").strip()
    if code:
        existing = await db.salon_branches.find_one(
            {"salon_id": salon_id, "branch_code": {"$regex": f"^{code}$", "$options": "i"}},
            {"_id": 0, "id": 1},
        )
        if existing:
            raise HTTPException(status_code=400, detail=f"Branch code '{code}' already exists for this salon")

    # If user requests to create a new "main", demote any existing main first
    if branch_dict.get("is_main_branch"):
        await db.salon_branches.update_many(
            {"salon_id": salon_id, "is_main_branch": True},
            {"$set": {"is_main_branch": False}},
        )

    branch_dict["id"] = str(uuid.uuid4())
    branch_dict["salon_id"] = salon_id
    branch_dict.setdefault("status", "active")
    branch_dict["created_at"] = datetime.now(timezone.utc).isoformat()

    insert_doc = branch_dict.copy()
    await db.salon_branches.insert_one(insert_doc)
    branch_dict.pop("_id", None)
    return Branch(**branch_dict)


@api_router.get("/salons/{salon_id}/branches/{branch_id}", response_model=Branch)
async def get_branch(
    salon_id: str,
    branch_id: str,
    current_user: dict = Depends(get_current_salon_user),
):
    user_salon_id = current_user.get("salon_id") or current_user.get("sub")
    if user_salon_id != salon_id:
        raise HTTPException(status_code=403, detail="Access denied for this salon")

    branch = await db.salon_branches.find_one(
        {"id": branch_id, "salon_id": salon_id}, {"_id": 0}
    )
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    return Branch(**branch)


@api_router.put("/salons/{salon_id}/branches/{branch_id}", response_model=Branch)
async def update_branch(
    salon_id: str,
    branch_id: str,
    payload: BranchUpdate,
    current_user: dict = Depends(get_current_salon_user),
):
    _check_salon_admin_for_salon(current_user, salon_id)

    branch = await db.salon_branches.find_one(
        {"id": branch_id, "salon_id": salon_id}, {"_id": 0}
    )
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")

    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}

    # If promoting to main branch, demote others
    if update_data.get("is_main_branch") is True:
        await db.salon_branches.update_many(
            {"salon_id": salon_id, "is_main_branch": True, "id": {"$ne": branch_id}},
            {"$set": {"is_main_branch": False}},
        )

    if update_data:
        await db.salon_branches.update_one(
            {"id": branch_id, "salon_id": salon_id},
            {"$set": update_data},
        )

    updated = await db.salon_branches.find_one(
        {"id": branch_id, "salon_id": salon_id}, {"_id": 0}
    )
    return Branch(**updated)


@api_router.delete("/salons/{salon_id}/branches/{branch_id}")
async def delete_branch(
    salon_id: str,
    branch_id: str,
    current_user: dict = Depends(get_current_salon_user),
):
    """Soft-delete a branch (status=inactive). Cannot delete the main branch.
    The branch must have no active tokens for today/future to be deactivated.
    """
    _check_salon_admin_for_salon(current_user, salon_id)

    branch = await db.salon_branches.find_one(
        {"id": branch_id, "salon_id": salon_id}, {"_id": 0}
    )
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    if branch.get("is_main_branch"):
        raise HTTPException(status_code=400, detail="Cannot delete the main branch. Set another branch as main first.")

    today = datetime.now(timezone.utc).date().isoformat()
    active_tokens = await db.tokens.count_documents({
        "salon_id": salon_id,
        "branch_id": branch_id,
        "date": {"$gte": today},
        "status": {"$nin": ["completed", "cancelled", "skipped"]},
    })
    if active_tokens > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot deactivate branch with {active_tokens} active/upcoming bookings. Cancel or reassign them first.",
        )

    await db.salon_branches.update_one(
        {"id": branch_id, "salon_id": salon_id},
        {"$set": {"status": "inactive"}},
    )
    return {"message": "Branch deactivated", "branch_id": branch_id}


@api_router.post("/salons/{salon_id}/branches/{branch_id}/set-main")
async def set_main_branch(
    salon_id: str,
    branch_id: str,
    current_user: dict = Depends(get_current_salon_user),
):
    """Promote a branch to be the salon's main branch. Demotes the previous main."""
    _check_salon_admin_for_salon(current_user, salon_id)

    branch = await db.salon_branches.find_one(
        {"id": branch_id, "salon_id": salon_id}, {"_id": 0}
    )
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    if branch.get("status") != "active":
        raise HTTPException(status_code=400, detail="Cannot set an inactive branch as main")

    await db.salon_branches.update_many(
        {"salon_id": salon_id, "is_main_branch": True},
        {"$set": {"is_main_branch": False}},
    )
    await db.salon_branches.update_one(
        {"id": branch_id, "salon_id": salon_id},
        {"$set": {"is_main_branch": True}},
    )
    return {"message": "Main branch updated", "branch_id": branch_id}


@api_router.get("/salons/{salon_id}/branches/{branch_id}/qr-code")
async def get_branch_qr(
    salon_id: str,
    branch_id: str,
    base_url: Optional[str] = None,
):
    """Public endpoint: return per-branch booking QR code (base64 PNG).

    Pass `base_url` query param (e.g. `https://app.example.com`) to embed an
    absolute URL in the QR. Falls back to a relative path if missing.
    """
    branch = await db.salon_branches.find_one(
        {"id": branch_id, "salon_id": salon_id}, {"_id": 0}
    )
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")

    effective_base = (base_url or os.environ.get("PUBLIC_BASE_URL") or "").rstrip("/")
    booking_url = (
        f"{effective_base}/salon/{salon_id}?branch={branch_id}"
        if effective_base
        else f"/salon/{salon_id}?branch={branch_id}"
    )
    qr_code = await _generate_branch_qr(booking_url)

    return {
        "qr_code": qr_code,
        "booking_url": booking_url,
        "branch_name": branch.get("branch_name"),
    }


@api_router.get("/salons/{salon_id}/branches/{branch_id}/services-menu-qr")
async def get_branch_services_menu_qr(
    salon_id: str,
    branch_id: str,
    base_url: Optional[str] = None,
):
    """Public endpoint: return a Services-Menu QR code (base64 PNG) for the
    branch. Salons can print this like a menu — scanning it lands the customer
    on the salon's services page where they can pick services and tap "Book".
    """
    branch = await db.salon_branches.find_one(
        {"id": branch_id, "salon_id": salon_id}, {"_id": 0}
    )
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")

    effective_base = (base_url or os.environ.get("PUBLIC_BASE_URL") or "").rstrip("/")
    menu_url = (
        f"{effective_base}/salon/{salon_id}/menu?branch={branch_id}"
        if effective_base
        else f"/salon/{salon_id}/menu?branch={branch_id}"
    )
    qr_code = await _generate_branch_qr(menu_url)

    return {
        "qr_code": qr_code,
        "menu_url": menu_url,
        "branch_name": branch.get("branch_name"),
    }


# ----- Staff Branch Transfers (Phase 2) -----

@api_router.post("/salons/{salon_id}/staff-branch-transfers", response_model=StaffBranchTransfer)
async def create_staff_branch_transfer(
    salon_id: str,
    payload: StaffBranchTransferCreate,
    current_user: dict = Depends(get_current_salon_user),
):
    """Transfer a staff member to a different branch.
    - Admin: can transfer any staff to any branch within the salon.
    - Branch Manager: can only transfer staff INTO one of their assigned branches
      AND the staff being transferred must currently be in one of their assigned branches.
    - Staff: 403.
    """
    user_salon_id = current_user.get("salon_id") or current_user.get("sub")
    if user_salon_id != salon_id:
        raise HTTPException(status_code=403, detail="Access denied for this salon")

    role = current_user.get("role")
    if role == "salon_staff":
        raise HTTPException(status_code=403, detail="Staff cannot transfer barbers")

    # Validate staff
    staff = await db.barbers.find_one(
        {"id": payload.staff_id, "salon_id": salon_id, "is_active": True}, {"_id": 0}
    )
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")

    # Validate target branch
    target = await db.salon_branches.find_one(
        {"id": payload.to_branch_id, "salon_id": salon_id, "status": "active"}, {"_id": 0}
    )
    if not target:
        raise HTTPException(status_code=400, detail="Target branch not found or inactive")

    # Resolve from_branch_id (default to staff's current branch)
    from_branch_id = payload.from_branch_id or staff.get("branch_id")

    # Branch manager scope
    if is_branch_manager(current_user):
        assigned = assigned_branch_ids_for(current_user)
        if payload.to_branch_id not in assigned:
            raise HTTPException(status_code=403, detail="Cannot transfer to a branch you do not manage")
        if from_branch_id and from_branch_id not in assigned:
            raise HTTPException(status_code=403, detail="Cannot transfer staff from a branch you do not manage")

    # No-op guard
    if from_branch_id == payload.to_branch_id:
        raise HTTPException(status_code=400, detail="Staff is already at this branch")

    transfer = {
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "staff_id": payload.staff_id,
        "from_branch_id": from_branch_id,
        "to_branch_id": payload.to_branch_id,
        "transfer_date": payload.transfer_date,
        "remarks": payload.remarks,
        "created_by": current_user.get("sub") or current_user.get("user_id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.staff_branch_transfers.insert_one(transfer.copy())
    transfer.pop("_id", None)

    # Apply the transfer immediately on the staff record
    await db.barbers.update_one(
        {"id": payload.staff_id, "salon_id": salon_id},
        {"$set": {"branch_id": payload.to_branch_id}},
    )

    return StaffBranchTransfer(**transfer)


@api_router.get("/salons/{salon_id}/staff-branch-transfers", response_model=List[StaffBranchTransfer])
async def list_staff_branch_transfers(
    salon_id: str,
    staff_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_salon_user),
):
    """List staff transfers. Staff can read their salon's transfers; branch managers
    are auto-scoped to transfers involving their assigned branches."""
    user_salon_id = current_user.get("salon_id") or current_user.get("sub")
    if user_salon_id != salon_id:
        raise HTTPException(status_code=403, detail="Access denied for this salon")

    query = {"salon_id": salon_id}
    if staff_id:
        query["staff_id"] = staff_id
    if branch_id:
        query["$or"] = [{"from_branch_id": branch_id}, {"to_branch_id": branch_id}]

    if is_branch_manager(current_user):
        assigned = assigned_branch_ids_for(current_user)
        if not assigned:
            return []
        # Restrict to transfers touching any assigned branch.
        scoped = {"$or": [{"from_branch_id": {"$in": assigned}}, {"to_branch_id": {"$in": assigned}}]}
        # Combine with caller filters using $and so both clauses apply.
        if "$or" in query:
            query = {"$and": [query, scoped]}
        else:
            query.update(scoped)

    rows = await db.staff_branch_transfers.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [StaffBranchTransfer(**r) for r in rows]


# ============================================================================
# END BRANCH MANAGEMENT
# ============================================================================

@api_router.get("/salons/{salon_id}/operational-hours")
async def get_operational_hours(salon_id: str):
    """Get operational hours for a salon"""
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    # Return default operational_hours if not set, but always use real manual_toggle from DB
    actual_manual_toggle = salon.get("manual_toggle") or {"is_overridden": False, "is_open": True, "closed_mode": None, "overridden_at": None}
    if not salon.get("operational_hours"):
        default_hours = OperationalHours().model_dump()
        return {"operational_hours": default_hours, "manual_toggle": actual_manual_toggle}
    
    return {
        "operational_hours": salon.get("operational_hours", OperationalHours().model_dump()),
        "manual_toggle": actual_manual_toggle
    }

@api_router.put("/salons/{salon_id}/operational-hours")
async def update_operational_hours(
    salon_id: str, 
    hours: OperationalHours, 
    current_user: Optional[dict] = Depends(get_current_salon_user_optional),
    current_salon: Optional[dict] = Depends(get_current_salon_optional)
):
    """Update operational hours for a salon"""
    existing = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    # Allow if admin/staff user with edit permission OR salon owner
    is_authorized = False
    
    if current_user:
        # Multi-user salon roles: 'salon_admin' has full access; 'salon_staff' needs 'can_edit_salon' permission.
        # Also accept legacy 'admin' / 'salon' role values for backward compatibility.
        user_role = current_user.get("role")
        if user_role in ("salon_admin", "admin", "salon"):
            # Make sure the user belongs to this salon
            if current_user.get("salon_id") in (None, salon_id):
                is_authorized = True
            else:
                # If salon_id is set on user, must match
                if current_user.get("salon_id") == salon_id:
                    is_authorized = True
        elif has_module_permission(current_user, "salon_settings", "edit_hours"):
            if current_user.get("salon_id") == salon_id:
                is_authorized = True
    
    if current_salon:
        salon_id_from_token = current_salon.get("sub") or current_salon.get("id")
        if salon_id_from_token == salon_id:
            is_authorized = True
    
    if not is_authorized:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    await db.salons.update_one(
        {"id": salon_id},
        {"$set": {"operational_hours": hours.model_dump()}}
    )
    
    return {"message": "Operational hours updated successfully", "operational_hours": hours.model_dump()}

@api_router.put("/salons/{salon_id}/manual-toggle")
async def update_manual_toggle(salon_id: str, toggle_data: dict, current_user=Depends(get_current_salon_user)):
    """Toggle salon open/close manually (admin / branch_manager)"""
    existing = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    # Verify the authenticated user belongs to this salon
    user_salon_id = current_user.get("salon_id") or current_user.get("sub") or current_user.get("id")
    if user_salon_id != salon_id:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    is_overridden = toggle_data.get("is_overridden", True)
    is_open = toggle_data.get("is_open", True)
    closed_mode = toggle_data.get("closed_mode")  # 'full' | 'online_only' | None

    # Validate closed_mode
    if closed_mode is not None and closed_mode not in ("full", "online_only"):
        raise HTTPException(status_code=400, detail="closed_mode must be 'full' or 'online_only' or null")

    # When salon is open, clear closed_mode
    if is_open:
        closed_mode = None
    else:
        # When closing, require a closed_mode (default to 'full' for backward compat)
        if not closed_mode:
            closed_mode = "full"

    manual_toggle = {
        "is_overridden": is_overridden,
        "is_open": is_open,
        "closed_mode": closed_mode,
        "overridden_at": datetime.now(timezone.utc).isoformat() if is_overridden else None
    }

    await db.salons.update_one(
        {"id": salon_id},
        {"$set": {"manual_toggle": manual_toggle}}
    )

    if is_overridden:
        if is_open:
            msg = "Salon manually opened"
        elif closed_mode == "online_only":
            msg = "Salon closed for online bookings only (walk-in & QR still allowed)"
        else:
            msg = "Salon manually closed (online + offline)"
    else:
        msg = "Manual override cleared"

    return {"message": msg, "manual_toggle": manual_toggle}

@api_router.get("/salons/{salon_id}/is-accepting-bookings")
async def check_booking_availability(salon_id: str):
    """Check if salon is accepting bookings based on operational hours and holidays"""
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    # Check manual toggle first and auto-clear if day has changed
    manual_toggle = salon.get("manual_toggle", {})
    if manual_toggle.get("is_overridden"):
        # Auto-clear override if day has changed
        overridden_at = manual_toggle.get("overridden_at")
        if overridden_at:
            overridden_date = datetime.fromisoformat(overridden_at).date()
            current_date = datetime.now(timezone.utc).date()
            if overridden_date < current_date:
                # Day has changed, clear the override
                await db.salons.update_one(
                    {"id": salon_id},
                    {"$set": {"manual_toggle": {"is_overridden": False, "is_open": True, "closed_mode": None, "overridden_at": None}}}
                )
                manual_toggle = {"is_overridden": False, "is_open": True, "closed_mode": None, "overridden_at": None}
        
        # If still overridden, return manual status
        if manual_toggle.get("is_overridden"):
            closed_mode = manual_toggle.get("closed_mode")
            is_open = manual_toggle.get("is_open", True)
            if is_open:
                return {
                    "is_accepting_bookings": True,
                    "reason": "manual_override",
                    "closed_mode": None,
                    "message": "Salon is manually open"
                }
            if closed_mode == "online_only":
                return {
                    "is_accepting_bookings": False,
                    "reason": "closed_online_only",
                    "closed_mode": "online_only",
                    "message": "Closed Online — Visit Salon"
                }
            # full close
            return {
                "is_accepting_bookings": False,
                "reason": "manual_override",
                "closed_mode": "full",
                "message": "Salon is manually closed"
            }
    
    # Check operational hours
    operational_hours = salon.get("operational_hours")
    if not operational_hours:
        # Default: always open
        return {"is_accepting_bookings": True, "reason": "default", "message": "Salon is open"}
    
    # Get current day
    current_day = datetime.now(timezone.utc).strftime('%A').lower()
    day_hours = operational_hours.get(current_day)
    
    if not day_hours:
        return {"is_accepting_bookings": True, "reason": "no_config", "message": "Salon is open"}
    
    # Check if today is a holiday
    if day_hours.get("is_holiday"):
        return {"is_accepting_bookings": False, "reason": "holiday", "message": "Salon is closed (Holiday)"}
    
    # Salon is open (we don't block during lunch as per requirements)
    return {"is_accepting_bookings": True, "reason": "operational_hours", "message": "Salon is open"}


@api_router.get("/salons/{salon_id}/ratings")
async def get_salon_ratings(salon_id: str, limit: int = 50, skip: int = 0):
    """Get all ratings for a salon (across all barbers), sorted by latest"""
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    # Get all ratings for this salon sorted by latest
    reviews = await db.ratings.find(
        {"salon_id": salon_id},
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Calculate salon average rating from all reviews
    pipeline = [
        {"$match": {"salon_id": salon_id}},
        {"$group": {
            "_id": "$salon_id",
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1}
        }}
    ]
    stats = await db.ratings.aggregate(pipeline).to_list(1)
    
    avg_rating = round(stats[0]["average_rating"], 1) if stats else 0
    total_reviews = stats[0]["total_reviews"] if stats else 0
    
    # Also update the salon's rating field
    if stats:
        await db.salons.update_one(
            {"id": salon_id},
            {"$set": {"rating": avg_rating, "total_reviews": total_reviews}}
        )
    
    return {
        "salon_id": salon_id,
        "salon_name": salon.get("salon_name", ""),
        "average_rating": avg_rating,
        "total_reviews": total_reviews,
        "reviews": [RatingResponse(**r) for r in reviews]
    }

@api_router.post("/salons/{salon_id}/initialize")
async def initialize_salon_services(salon_id: str, current_salon=Depends(get_current_salon)):
    """Initialize predefined services and packages for a salon"""
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    
    result = await initialize_predefined_services_for_salon(salon_id)
    return result

@api_router.post("/salons/{salon_id}/reset-initialization")
async def reset_salon_initialization(salon_id: str, current_salon=Depends(get_current_salon)):
    """Reset initialization flag - allows re-initialization of services"""
    await db.salon_initialized.delete_many({"salon_id": salon_id})
    return {"message": "Initialization reset. Services can be re-initialized now."}

@api_router.get("/salons/{salon_id}/services/enabled")
async def get_salon_enabled_services(salon_id: str):
    """Get all services enabled for a specific salon"""
    # Get salon's enabled services
    salon_services = await db.salon_services.find(
        {"salon_id": salon_id, "is_enabled": True},
        {"_id": 0}
    ).to_list(1000)
    
    service_ids = [ss["service_id"] for ss in salon_services]
    
    # Get service details
    services = await db.services.find(
        {"id": {"$in": service_ids}, "is_active": True},
        {"_id": 0}
    ).to_list(1000)
    
    return services


@api_router.get("/salons/{salon_id}/menu")
async def get_salon_menu(salon_id: str, branch: Optional[str] = None):
    """Public endpoint that powers the printable Services-Menu QR landing page.

    Returns salon + branch + enabled services + active barbers in one call so
    customers landing from the QR can see everything they need to pick services
    and start a booking — no login required.
    """
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")

    # Resolve effective branch (explicit > main). Branch is optional but if a
    # branch_id is supplied we surface its metadata.
    branch_doc = None
    if branch:
        branch_doc = await db.salon_branches.find_one(
            {"id": branch, "salon_id": salon_id}, {"_id": 0}
        )
    if not branch_doc:
        branch_doc = await db.salon_branches.find_one(
            {"salon_id": salon_id, "is_main_branch": True}, {"_id": 0}
        )

    # Enabled services with category info
    salon_services = await db.salon_services.find(
        {"salon_id": salon_id, "is_enabled": True}, {"_id": 0}
    ).to_list(1000)
    service_ids = [ss["service_id"] for ss in salon_services]
    services = await db.services.find(
        {"id": {"$in": service_ids}, "is_active": True}, {"_id": 0}
    ).to_list(1000)

    # Apply per-salon price overrides (if any)
    override_map = {
        ss["service_id"]: ss.get("salon_price")
        for ss in salon_services
        if ss.get("salon_price") is not None
    }
    for svc in services:
        if svc["id"] in override_map:
            svc["base_price"] = override_map[svc["id"]]

    # Group by category, sorted
    services.sort(key=lambda s: (s.get("category") or "General", s.get("service_name") or ""))

    return {
        "salon": {
            "id": salon.get("id"),
            "salon_name": salon.get("salon_name"),
            "logo_url": salon.get("logo_url"),
            "address": salon.get("address"),
            "city": salon.get("city"),
            "phone": salon.get("phone"),
            "operational_hours": salon.get("operational_hours"),
        },
        "branch": (
            {
                "id": branch_doc.get("id"),
                "branch_name": branch_doc.get("branch_name"),
                "address": branch_doc.get("address"),
                "city": branch_doc.get("city"),
                "phone": branch_doc.get("phone"),
                "is_main_branch": branch_doc.get("is_main_branch", False),
            }
            if branch_doc
            else None
        ),
        "services": services,
    }

@api_router.get("/salons/{salon_id}/services/all")
async def get_all_services_with_salon_status(salon_id: str):
    """Get services for the given salon.

    A service is visible to a salon if either:
      • the salon created it (services.salon_id == salon_id), OR
      • a salon_services row exists linking it to the salon (legacy).
    Predefined / master services that were never explicitly loaded by the
    salon are NOT auto-shown here (that was the "pre-filled services"
    problem for fresh onboards).
    """
    # Get salon's linked services from salon_services (legacy / imported / explicitly loaded)
    salon_services = await db.salon_services.find(
        {"salon_id": salon_id},
        {"_id": 0, "service_id": 1, "is_enabled": 1}
    ).to_list(1000)
    linked_ids = [ss["service_id"] for ss in salon_services]
    linked_enabled_map = {ss["service_id"]: ss.get("is_enabled", False) for ss in salon_services}

    # Fetch services either owned by this salon OR linked via salon_services.
    query = {
        "is_active": True,
        "$or": [
            {"salon_id": salon_id},
            {"id": {"$in": linked_ids}} if linked_ids else {"id": "__no_match__"},
        ],
    }
    all_services = await db.services.find(query, {"_id": 0}).to_list(1000)

    # Add is_enabled_for_salon:
    #   - Own service (salon_id == salon_id) is considered enabled by default
    #   - Linked service uses salon_services flag
    for service in all_services:
        if service.get("salon_id") == salon_id:
            service["is_enabled_for_salon"] = True
            service["is_owned"] = True
        else:
            service["is_enabled_for_salon"] = linked_enabled_map.get(service["id"], False)
            service["is_owned"] = False

    return all_services

@api_router.put("/salons/{salon_id}/services/{service_id}/toggle")
async def toggle_service_for_salon(
    salon_id: str, 
    service_id: str, 
    is_enabled: bool,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Enable or disable a service for a specific salon"""
    # Verify token (supports both legacy and multi-user auth)
    token = credentials.credentials
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid authentication")
    # RBAC: services.toggle
    if not has_module_permission(payload, "services", "toggle"):
        raise HTTPException(status_code=403, detail="Permission denied: services.toggle")
    
    # Upsert: ensure a single entry per (salon_id, service_id).
    # Clean up any pre-existing duplicates first.
    await db.salon_services.delete_many({
        "salon_id": salon_id,
        "service_id": service_id
    })
    await db.salon_services.insert_one({
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "service_id": service_id,
        "is_enabled": is_enabled,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    
    return {"success": True, "is_enabled": is_enabled}


# ============ SERVICE ROUTES ============

@api_router.get("/services", response_model=List[Service])
async def get_services():
    services = await db.services.find({"is_active": True}, {"_id": 0}).to_list(1000)
    return services

@api_router.post("/services", response_model=Service)
async def create_service(service: ServiceCreate, current_salon=Depends(get_current_salon)):
    service_dict = service.model_dump()
    service_dict["id"] = str(uuid.uuid4())
    service_dict["is_active"] = True
    # Scope to the creating salon so it doesn't leak into other salons' lists.
    salon_id = current_salon.get("salon_id") or current_salon.get("sub")
    if salon_id:
        service_dict["salon_id"] = salon_id

    await db.services.insert_one(service_dict)
    return Service(**service_dict)

@api_router.put("/services/{service_id}", response_model=Service)
async def update_service(service_id: str, service: ServiceUpdate, current_salon=Depends(get_current_salon)):
    existing = await db.services.find_one({"id": service_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Service not found")
    
    update_data = {k: v for k, v in service.model_dump().items() if v is not None}
    if update_data:
        await db.services.update_one({"id": service_id}, {"$set": update_data})
    
    updated = await db.services.find_one({"id": service_id}, {"_id": 0})
    return Service(**updated)

@api_router.delete("/services/{service_id}")
async def delete_service(service_id: str, current_salon=Depends(get_current_salon)):
    """Soft delete a service"""
    existing = await db.services.find_one({"id": service_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Service not found")
    
    await db.services.update_one({"id": service_id}, {"$set": {"is_active": False}})
    # Also remove from all barber_services
    await db.barber_services.delete_many({"service_id": service_id})
    return {"message": "Service deleted"}


@api_router.post("/salons/{salon_id}/services/bulk-delete")
async def bulk_delete_salon_services(
    salon_id: str,
    body: dict,
    current_user=Depends(get_current_salon_user),
):
    """Bulk delete services for a salon.

    Behaviour (Jul 2026 requirement — salon bulk-delete):
    * Salon-owned services (`salon_id == salon_id`) are HARD-deleted.
    * Global services (no salon_id) are disabled for THIS salon only (via
      `salon_services.is_enabled=False`) — they remain in the global catalog.
    * All barber_services links for these services within this salon are cleared.

    Body: {"service_ids": ["...", "..."]}
    Returns: {ok, hard_deleted, disabled_for_salon, barber_links_removed}.
    """
    if not has_module_permission(current_user, "services", "delete"):
        raise HTTPException(status_code=403, detail="Permission denied: services.delete")
    ids = list((body or {}).get("service_ids") or [])
    if not ids:
        raise HTTPException(status_code=400, detail="service_ids is required")
    now_iso = datetime.now(timezone.utc).isoformat()

    services_docs = await db.services.find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "salon_id": 1}).to_list(length=None)
    salon_owned = [s["id"] for s in services_docs if s.get("salon_id") == salon_id]
    global_ids = [s["id"] for s in services_docs if not s.get("salon_id")]

    hard_res = None
    if salon_owned:
        hard_res = await db.services.delete_many({"id": {"$in": salon_owned}, "salon_id": salon_id})
        await db.salon_services.delete_many({"salon_id": salon_id, "service_id": {"$in": salon_owned}})

    disabled_res = None
    if global_ids:
        disabled_res = await db.salon_services.update_many(
            {"salon_id": salon_id, "service_id": {"$in": global_ids}},
            {"$set": {"is_enabled": False, "updated_at": now_iso}},
        )

    # Clean up barber_services links (only for this salon's barbers)
    salon_barber_ids = await db.barbers.distinct("id", {"salon_id": salon_id})
    bs_res = await db.barber_services.delete_many({
        "barber_id": {"$in": salon_barber_ids},
        "service_id": {"$in": ids},
    })

    return {
        "ok": True,
        "hard_deleted": (hard_res.deleted_count if hard_res else 0),
        "disabled_for_salon": (disabled_res.modified_count if disabled_res else 0),
        "barber_links_removed": bs_res.deleted_count,
    }

@api_router.get("/services/categories")
async def get_service_categories():
    """Get all unique service categories with thumbnails"""
    services = await db.services.find({"is_active": True}, {"_id": 0, "category": 1, "thumbnail_url": 1}).to_list(1000)
    
    # Group by category and get first thumbnail for each
    category_thumbnails = {}
    for s in services:
        cat = s.get("category", "General")
        if cat not in category_thumbnails:
            category_thumbnails[cat] = s.get("thumbnail_url")
    
    # Updated default thumbnails with highly relevant images for each category
    default_thumbnails = {
        "General": "https://images.pexels.com/photos/7781850/pexels-photo-7781850.jpeg?w=200&h=200&fit=crop",
        "Hair Treatments": "https://images.pexels.com/photos/3993146/pexels-photo-3993146.jpeg?w=200&h=200&fit=crop",
        "Hair Color": "https://images.pexels.com/photos/3993146/pexels-photo-3993146.jpeg?w=200&h=200&fit=crop",
        "Massage & Spa": "https://images.pexels.com/photos/3757952/pexels-photo-3757952.jpeg?w=200&h=200&fit=crop",
        "Men's Grooming": "https://images.pexels.com/photos/9992819/pexels-photo-9992819.jpeg?w=200&h=200&fit=crop",
        "Manicure & Pedicure": "https://images.pexels.com/photos/3997379/pexels-photo-3997379.jpeg?w=200&h=200&fit=crop",
        "Waxing & Threading": "https://images.pexels.com/photos/6135615/pexels-photo-6135615.jpeg?w=200&h=200&fit=crop",
        "Packages": "https://images.pexels.com/photos/3065171/pexels-photo-3065171.jpeg?w=200&h=200&fit=crop",
        "Bridal": "https://images.pexels.com/photos/3065171/pexels-photo-3065171.jpeg?w=200&h=200&fit=crop",
        "Facial": "https://images.pexels.com/photos/3985325/pexels-photo-3985325.jpeg?w=200&h=200&fit=crop",
        "Advance/Hydra Facial": "https://images.pexels.com/photos/3985325/pexels-photo-3985325.jpeg?w=200&h=200&fit=crop",
        "Favorites": "https://images.pexels.com/photos/7755651/pexels-photo-7755651.jpeg?w=200&h=200&fit=crop",
        "Bleach": "https://images.pexels.com/photos/3993146/pexels-photo-3993146.jpeg?w=200&h=200&fit=crop",
        "Body Care": "https://images.pexels.com/photos/3757952/pexels-photo-3757952.jpeg?w=200&h=200&fit=crop",
        "Nail Art": "https://images.pexels.com/photos/3997379/pexels-photo-3997379.jpeg?w=200&h=200&fit=crop",
        "Hair Spa": "https://images.pexels.com/photos/3993146/pexels-photo-3993146.jpeg?w=200&h=200&fit=crop",
        "Keratin": "https://images.pexels.com/photos/3993146/pexels-photo-3993146.jpeg?w=200&h=200&fit=crop",
        "Smoothening": "https://images.pexels.com/photos/3993146/pexels-photo-3993146.jpeg?w=200&h=200&fit=crop",
        "Threading": "https://images.pexels.com/photos/6135615/pexels-photo-6135615.jpeg?w=200&h=200&fit=crop"
    }
    
    categories = []
    for cat in sorted(category_thumbnails.keys()):
        categories.append({
            "name": cat,
            "thumbnail_url": category_thumbnails[cat] or default_thumbnails.get(cat, default_thumbnails["General"])
        })
    
    return {"categories": categories}


# ---- Sub-categories master (per salon, free-form) ----
@api_router.get("/salons/{salon_id}/services/subcategories")
async def list_service_subcategories(salon_id: str):
    """Return list of sub-categories used by this salon under Services and Packages,
    plus any explicitly-added sub-categories saved in service_subcategories."""
    # collected from actual services
    pipeline = [
        {"$match": {"salon_id": salon_id, "is_active": True}},
        {"$group": {"_id": {"category": "$category", "sub_category": "$sub_category"}}}
    ]
    seen: Dict[str, set] = {"Services": set(), "Packages": set()}
    try:
        async for row in db.services.aggregate(pipeline):
            k = row["_id"]
            cat = k.get("category") or "Services"
            sub = k.get("sub_category")
            if cat not in ("Services", "Packages"):
                cat = "Services"
            if sub:
                seen[cat].add(sub)
    except Exception:
        pass
    # extras saved manually
    try:
        docs = await db.service_subcategories.find(
            {"salon_id": salon_id}, {"_id": 0}
        ).to_list(500)
        for d in docs:
            cat = d.get("category") or "Services"
            if cat not in ("Services", "Packages"):
                cat = "Services"
            if d.get("name"):
                seen[cat].add(d["name"])
    except Exception:
        pass
    return {
        "Services": sorted(list(seen["Services"])),
        "Packages": sorted(list(seen["Packages"])),
    }


@api_router.post("/salons/{salon_id}/services/subcategories")
async def add_service_subcategory(
    salon_id: str,
    payload: dict,
    current_user=Depends(get_current_salon_user),
):
    """Add a free-form sub-category to service_subcategories (idempotent)."""
    if not has_module_permission(current_user, "services", "manage_categories"):
        raise HTTPException(403, "Permission denied")
    cat = payload.get("category") or "Services"
    name = (payload.get("name") or "").strip()
    if cat not in ("Services", "Packages"):
        cat = "Services"
    if not name:
        raise HTTPException(400, "name required")
    await db.service_subcategories.update_one(
        {"salon_id": salon_id, "category": cat, "name": name},
        {"$set": {"salon_id": salon_id, "category": cat, "name": name,
                  "updated_at": datetime.utcnow().isoformat()}},
        upsert=True,
    )
    return {"success": True, "category": cat, "name": name}


@api_router.delete("/salons/{salon_id}/services/subcategories")
async def delete_service_subcategory(
    salon_id: str,
    category: str,
    name: str,
    current_user=Depends(get_current_salon_user),
):
    if not has_module_permission(current_user, "services", "manage_categories"):
        raise HTTPException(403, "Permission denied")
    await db.service_subcategories.delete_one(
        {"salon_id": salon_id, "category": category, "name": name}
    )
    return {"success": True}


@api_router.get("/services/by-category")
async def get_services_by_category(gender: Optional[str] = None):
    """Get services grouped by category"""
    query = {"is_active": True}
    if gender and gender != "all":
        query["$or"] = [{"gender_tag": gender}, {"gender_tag": "Unisex"}]
    
    services = await db.services.find(query, {"_id": 0}).to_list(1000)
    
    # Group by category
    categorized = {}
    for service in services:
        category = service.get("category", "General")
        if category not in categorized:
            categorized[category] = []
        categorized[category].append(service)
    
    return categorized

@api_router.get("/services/favorites", response_model=List[Service])
async def get_favorite_services():
    """Get all favorite services"""
    services = await db.services.find(
        {"is_active": True, "is_favorite": True}, 
        {"_id": 0}
    ).sort("favorite_order", 1).to_list(100)
    return services

@api_router.put("/services/{service_id}/favorite")
async def toggle_favorite(service_id: str, is_favorite: bool, current_salon=Depends(get_current_salon)):
    """Mark/unmark service as favorite"""
    existing = await db.services.find_one({"id": service_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Service not found")
    
    update_data = {"is_favorite": is_favorite}
    if is_favorite:
        # Get max favorite_order and increment
        favorites = await db.services.find({"is_favorite": True}, {"_id": 0, "favorite_order": 1}).to_list(100)
        max_order = max([f.get("favorite_order", 0) for f in favorites]) if favorites else 0
        update_data["favorite_order"] = max_order + 1
    else:
        update_data["favorite_order"] = None
    
    await db.services.update_one({"id": service_id}, {"$set": update_data})
    return {"message": "Favorite status updated", "is_favorite": is_favorite}

@api_router.put("/services/favorites/reorder")
async def reorder_favorites(service_ids: List[str], current_salon=Depends(get_current_salon)):
    """Reorder favorites based on array order"""
    for index, service_id in enumerate(service_ids):
        await db.services.update_one(
            {"id": service_id},
            {"$set": {"favorite_order": index}}
        )
    return {"message": "Favorites reordered successfully"}

# ============ PACKAGE ROUTES ============

@api_router.get("/packages", response_model=List[Package])
async def get_packages(gender: Optional[str] = None):
    """Get all packages"""
    query = {"is_active": True}
    if gender and gender != "all":
        query["$or"] = [{"gender_tag": gender}, {"gender_tag": "Unisex"}]
    
    packages = await db.packages.find(query, {"_id": 0}).to_list(100)
    return packages

@api_router.post("/packages", response_model=Package)
async def create_package(package: PackageCreate, current_salon=Depends(get_current_salon)):
    """Create a new package"""
    package_dict = package.model_dump()
    package_dict["id"] = str(uuid.uuid4())
    package_dict["is_active"] = True
    package_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.packages.insert_one(package_dict)
    return Package(**package_dict)

@api_router.put("/packages/{package_id}", response_model=Package)
async def update_package(package_id: str, package: PackageCreate, current_salon=Depends(get_current_salon)):
    """Update a package"""
    existing = await db.packages.find_one({"id": package_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Package not found")
    
    update_data = package.model_dump()
    await db.packages.update_one({"id": package_id}, {"$set": update_data})
    
    updated = await db.packages.find_one({"id": package_id}, {"_id": 0})
    return Package(**updated)

@api_router.delete("/packages/{package_id}")
async def delete_package(package_id: str, current_salon=Depends(get_current_salon)):
    """Delete a package"""
    existing = await db.packages.find_one({"id": package_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Package not found")
    
    await db.packages.update_one({"id": package_id}, {"$set": {"is_active": False}})
    return {"message": "Package deleted"}

# ============ SALON-SPECIFIC PACKAGE ROUTES ============

@api_router.get("/salons/{salon_id}/packages", response_model=List[SalonPackage])
async def get_salon_packages(salon_id: str, gender: Optional[str] = None):
    """Get all packages for a specific salon"""
    query = {"salon_id": salon_id, "is_active": True}
    if gender and gender != "all":
        query["$or"] = [{"gender_tag": gender}, {"gender_tag": "Unisex"}]
    
    packages = await db.salon_packages.find(query, {"_id": 0}).to_list(100)
    return packages

@api_router.get("/salons/{salon_id}/packages/with-services")
async def get_salon_packages_with_services(salon_id: str, gender: Optional[str] = None):
    """Get all active packages for a salon with service details"""
    query = {"salon_id": salon_id, "is_active": True}
    if gender and gender != "all":
        query["$or"] = [{"gender_tag": gender}, {"gender_tag": "Unisex"}]
    
    packages = await db.salon_packages.find(query, {"_id": 0}).to_list(100)
    
    # Populate service details for each package
    for package in packages:
        if package.get('service_ids'):
            services = []
            for service_id in package['service_ids']:
                service = await db.services.find_one({"id": service_id}, {"_id": 0})
                if service:
                    services.append(service)
            package['services'] = services
    
    return {"packages": packages}

@api_router.post("/salons/{salon_id}/packages", response_model=SalonPackage)
async def create_salon_package(
    salon_id: str,
    package: SalonPackageCreate,
    current_salon=Depends(get_current_salon)
):
    """Create a new package for a salon"""
    package_dict = package.model_dump()
    package_dict["id"] = str(uuid.uuid4())
    package_dict["is_active"] = True
    package_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.salon_packages.insert_one(package_dict)
    return SalonPackage(**package_dict)

@api_router.put("/salons/{salon_id}/packages/{package_id}", response_model=SalonPackage)
async def update_salon_package(
    salon_id: str,
    package_id: str,
    package: SalonPackageUpdate,
    current_salon=Depends(get_current_salon)
):
    """Update a salon's package"""
    existing = await db.salon_packages.find_one(
        {"id": package_id, "salon_id": salon_id},
        {"_id": 0}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Package not found")
    
    # Only update provided fields
    update_data = {k: v for k, v in package.model_dump().items() if v is not None}
    
    if update_data:
        await db.salon_packages.update_one(
            {"id": package_id, "salon_id": salon_id},
            {"$set": update_data}
        )
    
    updated = await db.salon_packages.find_one(
        {"id": package_id, "salon_id": salon_id},
        {"_id": 0}
    )
    return SalonPackage(**updated)

@api_router.delete("/salons/{salon_id}/packages/{package_id}")
async def delete_salon_package(
    salon_id: str,
    package_id: str,
    current_salon=Depends(get_current_salon)
):
    """Delete a salon's package"""
    existing = await db.salon_packages.find_one(
        {"id": package_id, "salon_id": salon_id},
        {"_id": 0}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Package not found")
    
    await db.salon_packages.update_one(
        {"id": package_id, "salon_id": salon_id},
        {"$set": {"is_active": False}}
    )
    return {"message": "Package deleted"}


# ============ BARBER ROUTES ============

def _is_barber_employed_on(barber: Dict[str, Any], date_str: str) -> bool:
    """Subset of is_barber_available_on: only checks employment window
    (doj / last_working_date). Used when we want to keep on-leave barbers
    visible (greyed-out) but hide barbers who haven't joined or have left.
    """
    if not date_str:
        return True
    try:
        doj = (barber.get("doj") or "").strip()
        if doj and date_str < doj:
            return False
        lwd = (barber.get("last_working_date") or "").strip()
        if lwd and date_str > lwd:
            return False
    except Exception:
        return True
    return True


def _is_barber_on_leave_on(barber: Dict[str, Any], date_str: str) -> bool:
    """True if barber is on leave for the given date (per-date list or legacy flag)."""
    if not date_str:
        return False
    try:
        leave_dates = barber.get("leave_dates") or []
        if isinstance(leave_dates, list) and date_str in leave_dates:
            return True
        ist = timezone(timedelta(hours=5, minutes=30))
        today_ist = datetime.now(ist).strftime("%Y-%m-%d")
        if barber.get("on_leave") is True and date_str == today_ist:
            return True
    except Exception:
        return False
    return False


def normalize_barber_data(barber: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normalize barber data to ensure all fields match Pydantic model expectations.
    Converts None values to appropriate defaults for list fields.
    """
    if barber.get("leave_dates") is None:
        barber["leave_dates"] = []
    if barber.get("gallery") is None:
        barber["gallery"] = []
    if barber.get("documents") is None:
        barber["documents"] = []
    return barber


@api_router.get("/salons/{salon_id}/barbers", response_model=List[Barber])
async def get_salon_barbers(
    salon_id: str,
    available_only: bool = False,
    customer_view: bool = False,
    date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: Optional[dict] = Depends(get_current_salon_user_optional),
):
    """
    Get barbers for a salon
    available_only=True: Strict filter — only return barbers fully available on the given date.
    customer_view=True : Return all employed barbers (is_barber=True OR is_barber not set - for backward compatibility).
                         Barbers on leave for the target date are KEPT in the response but flagged with `is_on_leave=True`,
                         so the customer UI can show them greyed-out instead of hiding them.
    available_only=False & customer_view=False: Return all active staff (admin view).
    branch_id (optional): When set, returns only staff assigned to this branch.
    """
    try:
        # Branch-manager scoping (skip for customer_view so customers still see everything)
        if not customer_view and current_user and is_branch_manager(current_user):
            branch_id = enforce_branch_for_manager(current_user, branch_id)
        query = {"salon_id": salon_id, "is_active": True}
        if customer_view:
            # For customer view, include barbers where is_barber=True OR is_barber field doesn't exist (backward compatibility)
            query["$or"] = [{"is_barber": True}, {"is_barber": {"$exists": False}}]
        if branch_id:
            query["branch_id"] = branch_id

        barbers = await db.barbers.find(query, {"_id": 0}).to_list(500)
        logger.info(f"Fetched {len(barbers)} barbers for salon {salon_id}, query: {query}")

        # Normalize barber data to ensure fields match Pydantic model expectations
        barbers = [normalize_barber_data(b) for b in barbers]

        # Resolve target date — defaults to today (IST)
        ist = timezone(timedelta(hours=5, minutes=30))
        target_date = (date or "").strip() or datetime.now(ist).strftime("%Y-%m-%d")

        if available_only:
            # Strict filter (used by booking-engine endpoints / admin "available" toggle).
            barbers = [b for b in barbers if is_barber_available_on(b, target_date)]
            logger.info(f"After available_only filter: {len(barbers)} barbers")
        elif customer_view:
            # Customer view: hide barbers who haven't joined yet or have left employment,
            # but KEEP on-leave barbers visible — flag them so the UI can grey them out.
            out = []
            for b in barbers:
                if not _is_barber_employed_on(b, target_date):
                    continue
                b["is_on_leave"] = _is_barber_on_leave_on(b, target_date)
                out.append(b)
            barbers = out
            logger.info(f"After customer_view filter: {len(barbers)} barbers")

        return barbers
    except HTTPException:
        # Don't swallow auth/RBAC errors raised intentionally above.
        raise
    except Exception as e:
        logger.error(f"Error fetching barbers for salon {salon_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch barbers: {str(e)}")

@api_router.post("/salons/{salon_id}/barbers", response_model=Barber)
async def create_barber(salon_id: str, barber: BarberCreate, current_user=Depends(get_current_salon_user)):
    # Only admin/branch_manager can add staff
    if current_user.get("role") not in ("salon", "salon_admin", "salon_branch_manager"):
        raise HTTPException(status_code=403, detail="Admin access required to add staff")
    # Subscription paywall: free plan allows max 1 active staff
    await enforce_premium_or_within_limit(salon_id, resource="staff")

    barber_dict = barber.model_dump()
    barber_dict["id"] = str(uuid.uuid4())
    barber_dict["salon_id"] = salon_id  # Override with URL param
    # Resolve branch (defaults to main branch if not specified) so every staff is branch-tagged.
    barber_dict["branch_id"] = await resolve_branch_id(salon_id, barber_dict.get("branch_id"))
    barber_dict["queue_status"] = "available"
    barber_dict["is_active"] = True
    
    await db.barbers.insert_one(barber_dict)
    # Normalize before returning to ensure Pydantic validation passes
    barber_dict = normalize_barber_data(barber_dict)
    return Barber(**barber_dict)

@api_router.put("/barbers/{barber_id}", response_model=Barber)
async def update_barber(barber_id: str, barber_update: BarberUpdate, current_user=Depends(get_current_salon_user)):
    """Update barber details. Accepts both legacy salon admin token and multi-user salon-admin tokens."""
    # RBAC: staff.edit
    if not has_module_permission(current_user, "staff", "edit"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.edit")
    existing = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    update_data = {k: v for k, v in barber_update.model_dump().items() if v is not None}
    if update_data:
        await db.barbers.update_one({"id": barber_id}, {"$set": update_data})
    
    updated = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    # Normalize before returning to ensure Pydantic validation passes
    updated = normalize_barber_data(updated)
    return Barber(**updated)

@api_router.delete("/barbers/{barber_id}")
async def delete_barber(
    barber_id: str,
    current_user=Depends(get_current_salon_user),
):
    """
    Hard delete a staff member.

    Removes the barber document and all operational data linked to them
    (services mapping, attendance overrides, branch transfers, ratings,
    notification settings, salon_user login access).

    PRESERVES financial history: financial_transactions, salary_records,
    incentive_payouts — for audit/accounting purposes.
    """
    if not has_module_permission(current_user, "staff", "delete"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.delete")

    existing = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Barber not found")

    salon_id = existing.get("salon_id")
    _check_salon_admin_for_salon(current_user, salon_id)

    # Aggregate counts of preserved items for the response (for transparency)
    preserved = {
        "financial_transactions": await db.financial_transactions.count_documents(
            {"barber_id": barber_id}
        ),
        "salary_records": await db.salary_records.count_documents(
            {"barber_id": barber_id}
        ),
        "incentive_payouts": await db.incentive_payouts.count_documents(
            {"barber_id": barber_id}
        ),
    }

    # Hard-delete the barber and operational data
    await db.barbers.delete_one({"id": barber_id})
    await db.barber_services.delete_many({"barber_id": barber_id})
    await db.staff_branch_transfers.delete_many({"barber_id": barber_id})
    await db.attendance.delete_many({"barber_id": barber_id})
    await db.notifications.delete_many({"barber_id": barber_id})
    await db.ratings.delete_many({"barber_id": barber_id})
    # Free up tokens that referenced this barber so customer queue stays clean
    await db.tokens.delete_many({"barber_id": barber_id, "status": {"$in": ["waiting", "in_progress"]}})

    # Remove staff login access (salon_user records linked to this barber)
    login_removed = await db.salon_users.delete_many({"barber_id": barber_id})

    # Anonymize barber_id references in preserved financial collections so future
    # joins don't blow up (we keep the row but null out the link).
    await db.financial_transactions.update_many(
        {"barber_id": barber_id},
        {"$set": {"barber_id": None, "barber_name_snapshot": existing.get("name")}},
    )
    await db.salary_records.update_many(
        {"barber_id": barber_id},
        {"$set": {"barber_id": None, "barber_name_snapshot": existing.get("name")}},
    )
    await db.incentive_payouts.update_many(
        {"barber_id": barber_id},
        {"$set": {"barber_id": None, "barber_name_snapshot": existing.get("name")}},
    )

    return {
        "message": "Staff deleted permanently",
        "barber_id": barber_id,
        "barber_name": existing.get("name"),
        "login_access_removed": login_removed.deleted_count > 0,
        "preserved_records": preserved,
    }


# ============ STAFF DOCUMENTS (Aadhar / PAN / Photo / etc) ============

class StaffDocumentUpload(BaseModel):
    doc_type: str  # 'aadhar_front' | 'aadhar_back' | 'pan' | 'photo' | 'other'
    label: Optional[str] = None  # custom label, e.g., for "other"
    file_data: str  # base64-encoded data URL or raw base64
    mime_type: Optional[str] = None
    file_name: Optional[str] = None


def _doc_size_kb(data: str) -> int:
    """Estimate base64 payload size in KB."""
    if not data:
        return 0
    payload = data.split(",", 1)[1] if "," in data else data
    return int(len(payload) * 3 / 4 / 1024)


@api_router.get("/barbers/{barber_id}/documents")
async def list_staff_documents(
    barber_id: str, current_user=Depends(get_current_salon_user)
):
    """List documents (without file_data — to keep payload small)."""
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    _check_salon_admin_for_salon(current_user, barber.get("salon_id"))
    docs = barber.get("staff_documents") or []
    summary = [
        {
            "id": d.get("id"),
            "doc_type": d.get("doc_type"),
            "label": d.get("label"),
            "file_name": d.get("file_name"),
            "mime_type": d.get("mime_type"),
            "size_kb": d.get("size_kb"),
            "uploaded_at": d.get("uploaded_at"),
        }
        for d in docs
    ]
    return {"barber_id": barber_id, "documents": summary}


@api_router.get("/barbers/{barber_id}/documents/{doc_id}")
async def get_staff_document(
    barber_id: str, doc_id: str, current_user=Depends(get_current_salon_user)
):
    """Return a single document with its file_data."""
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    _check_salon_admin_for_salon(current_user, barber.get("salon_id"))
    docs = barber.get("staff_documents") or []
    doc = next((d for d in docs if d.get("id") == doc_id), None)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return doc


@api_router.post("/barbers/{barber_id}/documents")
async def upload_staff_document(
    barber_id: str,
    payload: StaffDocumentUpload,
    current_user=Depends(get_current_salon_user),
):
    """Add a new document for a staff member (base64 inline)."""
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    _check_salon_admin_for_salon(current_user, barber.get("salon_id"))

    # Size guard (keep mongo doc reasonably small ~10MB total per barber)
    size_kb = _doc_size_kb(payload.file_data)
    if size_kb > 10 * 1024:  # 10MB per file
        raise HTTPException(status_code=400, detail="File too large (max 10 MB)")

    doc = {
        "id": str(uuid.uuid4()),
        "doc_type": payload.doc_type,
        "label": payload.label or payload.doc_type,
        "file_data": payload.file_data,
        "mime_type": payload.mime_type,
        "file_name": payload.file_name,
        "size_kb": size_kb,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.barbers.update_one(
        {"id": barber_id},
        {"$push": {"staff_documents": doc}},
    )
    summary = {k: v for k, v in doc.items() if k != "file_data"}
    return {"document": summary}


@api_router.delete("/barbers/{barber_id}/documents/{doc_id}")
async def delete_staff_document(
    barber_id: str, doc_id: str, current_user=Depends(get_current_salon_user)
):
    """Remove a document from a staff member."""
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    _check_salon_admin_for_salon(current_user, barber.get("salon_id"))

    res = await db.barbers.update_one(
        {"id": barber_id},
        {"$pull": {"staff_documents": {"id": doc_id}}},
    )
    if res.modified_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document removed", "doc_id": doc_id}

@api_router.get("/barbers/{barber_id}/services")
async def get_barber_services(barber_id: str):
    """Get salon-enabled services with barber-specific pricing"""
    # Get barber to find salon_id
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0, "salon_id": 1})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    salon_id = barber["salon_id"]
    
    # Get only services enabled for this salon
    salon_services = await db.salon_services.find(
        {"salon_id": salon_id, "is_enabled": True},
        {"_id": 0}
    ).to_list(1000)
    
    enabled_service_ids = [ss["service_id"] for ss in salon_services]
    
    # Get service details for enabled services only
    services = await db.services.find(
        {"id": {"$in": enabled_service_ids}, "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    result = []
    for service in services:
        pricing = await db.barber_services.find_one({
            "barber_id": barber_id,
            "service_id": service["id"]
        }, {"_id": 0})
        
        if pricing:
            result.append({
                **service,
                "barber_price": pricing["price"],
                "is_available": pricing["is_available"]
            })
        else:
            result.append({
                **service,
                "barber_price": service["base_price"],
                "is_available": False  # Not assigned yet
            })
    
    return result

@api_router.put("/barbers/{barber_id}/services")
async def update_barber_services(barber_id: str, services: List[BarberServiceAssignment], current_salon=Depends(get_current_salon)):
    """Bulk update barber services with pricing"""
    existing = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    # Clear existing service assignments for this barber
    await db.barber_services.delete_many({"barber_id": barber_id})
    
    # Insert new assignments
    for svc in services:
        if svc.is_available:  # Only add services that are enabled
            pricing = {
                "id": str(uuid.uuid4()),
                "barber_id": barber_id,
                "service_id": svc.service_id,
                "price": svc.price,
                "is_available": svc.is_available
            }
            await db.barber_services.insert_one(pricing)
    
    return {"message": f"Updated {len([s for s in services if s.is_available])} services for barber"}

@api_router.put("/barbers/{barber_id}/services/{service_id}/toggle")
async def toggle_barber_service(
    barber_id: str, 
    service_id: str, 
    is_available: bool,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Toggle service availability for a barber"""
    # Verify authentication
    token = credentials.credentials
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid authentication")
    
    # Check if barber exists
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    # Check if service exists
    service = await db.services.find_one({"id": service_id}, {"_id": 0})
    if not service:
        raise HTTPException(status_code=404, detail="Service not found")
    
    # Check if barber_service entry exists
    existing = await db.barber_services.find_one({
        "barber_id": barber_id,
        "service_id": service_id
    }, {"_id": 0})
    
    if existing:
        # Update existing
        await db.barber_services.update_one(
            {"barber_id": barber_id, "service_id": service_id},
            {"$set": {"is_available": is_available}}
        )
    else:
        # Create new entry with default price
        await db.barber_services.insert_one({
            "id": str(uuid.uuid4()),
            "barber_id": barber_id,
            "service_id": service_id,
            "price": service.get("base_price", 0),
            "is_available": is_available
        })
    
    return {"success": True, "is_available": is_available}


@api_router.put("/barbers/{barber_id}/services/{service_id}/price")
async def update_barber_service_price(barber_id: str, service_id: str, price: float, current_salon=Depends(get_current_salon)):
    existing = await db.barber_services.find_one({"barber_id": barber_id, "service_id": service_id})
    
    if existing:
        await db.barber_services.update_one(
            {"barber_id": barber_id, "service_id": service_id},
            {"$set": {"price": price}}
        )
    else:
        pricing = {
            "id": str(uuid.uuid4()),
            "barber_id": barber_id,
            "service_id": service_id,
            "price": price,
            "is_available": True
        }
        await db.barber_services.insert_one(pricing)
    
    return {"message": "Price updated"}

# ============ AUTH ROUTES ============

@api_router.post("/salon/send-otp")
async def send_otp(request: SalonOTPRequest):
    """Send OTP to salon phone number via WhatsApp"""
    phone = request.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    logger.info(f"OTP request received for phone: {phone}")
    
    # Check if salon exists with this phone
    salon = await db.salons.find_one({"phone": phone}, {"_id": 0})
    salon_exists = salon is not None
    
    # Generate OTP
    otp = generate_otp()
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
    
    logger.info(f"Generated OTP for {phone}: {otp}")
    
    # Store OTP in database
    await db.salon_otp.delete_many({"phone": phone})
    await db.salon_otp.insert_one({
        "phone": phone,
        "otp": otp,
        "expires_at": expires_at.isoformat(),
        "verified": False
    })
    
    logger.info(f"OTP stored in database for {phone}")
    
    # Send OTP via WhatsApp
    whatsapp_result = await send_whatsapp_otp(phone, otp)
    
    logger.info(f"WhatsApp send result for {phone}: {whatsapp_result}")
    
    # Build response — channel is whatever Verify actually used (e.g. "sms").
    channel = whatsapp_result.get('channel') or 'sms'
    channel_label = 'WhatsApp' if channel == 'whatsapp' else 'SMS'
    response = {
        "message": f"OTP sent successfully via {channel_label}",
        "salon_exists": salon_exists,
        "delivery_status": whatsapp_result.get('status'),
        "channel": channel,
    }
    
    # OTP is NEVER returned in the API response — delivery is handled by Twilio.
    # The local OTP is logged server-side for debugging/support.
    if whatsapp_result.get('status') in ['mock', 'failed']:
        if whatsapp_result.get('status') == 'mock':
            response['note'] = "⚠️ Messaging not configured. Please contact support."
            logger.warning(f"Mock OTP (Twilio not configured) for {phone}: {otp}")
        else:
            response['error'] = whatsapp_result.get('error')
            response['note'] = "OTP delivery failed. Please try again."
            logger.error(f"OTP delivery failed for {phone}: {whatsapp_result.get('error')}")
    else:
        logger.info(f"✅ OTP sent via {channel} to {phone}. (Local audit code: {otp})")
        response['note'] = f"OTP sent to your {channel_label}. Please check your messages."
    
    return response

@api_router.post("/salon/register", response_model=Salon)
async def register_salon(salon: SalonCreate):
    """Register a new salon"""
    # Normalize phone format first
    phone = salon.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    if len(phone) != 13:
        raise HTTPException(status_code=400, detail="Invalid phone number format")
    
    # Check if salon with this phone already exists (use normalized phone)
    existing = await db.salons.find_one({"phone": phone}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Salon already registered with this phone number")
    
    # Create salon
    salon_dict = salon.model_dump()
    salon_dict["phone"] = phone
    salon_dict["id"] = str(uuid.uuid4())
    salon_dict["is_active"] = True
    salon_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    # Hash password if provided
    if salon.password:
        salon_dict["password_hash"] = pwd_context.hash(salon.password)
        del salon_dict["password"]  # Don't store plain password
    
    await db.salons.insert_one(salon_dict)
    
    # Bug #3 fix: A new salon MUST have a Main Branch immediately so the
    # customer-side salon search (`/api/public/salon-locations`) — which
    # returns one row per active branch — can find it right after signup.
    try:
        await ensure_main_branch_for_salon(salon_dict)
    except Exception as e:
        logger.warning(f"[register_salon] Main Branch auto-create failed for salon {salon_dict['id']}: {e}")
    
    return Salon(**salon_dict)

@api_router.post("/salon/verify-otp", response_model=SalonToken)
async def verify_otp(request: SalonOTPVerify):
    """Verify OTP and return access token.

    Production: validates against Twilio Verify.  Mock/dev: validates against
    the locally-stored OTP in `db.salon_otp`.
    """
    phone = request.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"

    if not await _otp_is_valid(phone, request.otp, db.salon_otp):
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")

    # Audit: mark any matching local OTP doc as verified, ignore if absent.
    await db.salon_otp.update_one(
        {"phone": phone, "otp": request.otp},
        {"$set": {"verified": True}},
    )

    # Find salon by phone
    salon = await db.salons.find_one({"phone": phone}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found for this phone")

    # Generate token
    token = create_access_token({"sub": salon["id"], "role": "salon", "phone": phone})

    return SalonToken(access_token=token, salon_id=salon["id"])

@api_router.post("/salon/password-login", response_model=SalonToken)
async def salon_password_login(credentials: SalonPasswordLogin):
    """Login with phone and password"""
    phone = credentials.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Find salon by phone
    salon = await db.salons.find_one({"phone": phone}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found for this phone number")
    
    # Check if password exists
    if "password_hash" not in salon or not salon["password_hash"]:
        raise HTTPException(status_code=400, detail="Password not set for this salon. Please use OTP login.")
    
    # Verify password
    if not pwd_context.verify(credentials.password, salon["password_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect password")
    
    # Generate token
    token = create_access_token({"sub": salon["id"], "role": "salon", "phone": phone})
    
    return SalonToken(access_token=token, salon_id=salon["id"])

@api_router.put("/salon/{salon_id}/set-password")
async def set_salon_password(salon_id: str, new_password: str, current_salon=Depends(get_current_salon)):
    """Set or update salon password"""
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    # Hash password
    password_hash = pwd_context.hash(new_password)
    
    # Update salon
    await db.salons.update_one(
        {"id": salon_id},
        {"$set": {"password_hash": password_hash}}
    )
    
    return {"message": "Password updated successfully"}

# ============ SALON USER MULTI-USER AUTH ROUTES ============

@api_router.post("/salon/users/login", response_model=SalonUserToken)
async def salon_user_login(credentials: SalonUserLogin):
    """Multi-user salon login (staff/admin) - accepts mobile number or login ID"""
    identifier = credentials.identifier.strip()
    
    # Format phone if it looks like a phone number
    if identifier.isdigit():
        if not identifier.startswith("+91"):
            identifier = f"+91{identifier}"
    
    # Find user by login_id or mobile
    query = {
        "$or": [
            {"login_id": credentials.identifier.strip()},
            {"mobile": identifier}
        ],
        "status": "active"
    }
    
    salon_user = await db.salon_users.find_one(query, {"_id": 0})
    
    if not salon_user:
        raise HTTPException(status_code=404, detail="User not found or inactive")
    
    # Verify password
    if not pwd_context.verify(credentials.password, salon_user["password_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect password")

    # Phase 5 (Part A) — Block login if the salon is suspended by platform admin.
    salon_doc = await db.salons.find_one(
        {"id": salon_user["salon_id"]}, {"_id": 0, "status": 1, "suspension_reason": 1}
    )
    if salon_doc and salon_doc.get("status") == "suspended":
        raise HTTPException(
            status_code=403,
            detail={
                "code": "SALON_SUSPENDED",
                "message": "This salon has been suspended by the platform administrator.",
                "reason": salon_doc.get("suspension_reason"),
            },
        )
    
    # Generate token with role and permissions
    permissions = salon_user.get("permissions", {
        "can_edit_salon": False,
        "can_access_analytics": False,
        "can_access_financials": False,
        "can_delete_salon": False
    })
    # Ensure newly added keys default to False if missing on legacy records
    permissions.setdefault("can_access_financials", False)
    permissions.setdefault("can_access_analytics", False)
    permissions.setdefault("can_edit_salon", False)
    permissions.setdefault("can_delete_salon", False)
    permissions.setdefault("can_access_services", False)
    permissions.setdefault("can_access_gallery", False)
    permissions.setdefault("can_access_staff", False)
    permissions.setdefault("can_view_all_staff", False)
    permissions.setdefault("can_access_marketing", False)

    assigned_branch_ids = salon_user.get("assigned_branch_ids") or []
    staff_id = salon_user.get("staff_id")

    token = create_access_token({
        "sub": salon_user["id"],
        "role": f"salon_{salon_user['role']}",  # salon_admin / salon_staff / salon_branch_manager
        "salon_id": salon_user["salon_id"],
        "permissions": permissions,
        "assigned_branch_ids": assigned_branch_ids,
        "staff_id": staff_id,
    })
    
    return SalonUserToken(
        access_token=token,
        salon_id=salon_user["salon_id"],
        user_id=salon_user["id"],
        role=salon_user["role"],
        permissions=SalonUserPermissions(**permissions),
        assigned_branch_ids=assigned_branch_ids,
        staff_id=staff_id,
    )

@api_router.post("/salon/users", response_model=SalonUser)
async def create_salon_user(user_data: SalonUserCreate, current_user=Depends(get_current_salon_admin)):
    """Create new salon user (admin only)"""
    # Validate salon_id matches current user's salon
    if user_data.salon_id != current_user.get("salon_id"):
        raise HTTPException(status_code=403, detail="Cannot create user for different salon")
    
    # Format mobile
    mobile = user_data.mobile
    if not mobile.startswith("+91"):
        mobile = f"+91{mobile}"
    
    # Check if login_id already exists
    existing_login = await db.salon_users.find_one({"login_id": user_data.login_id}, {"_id": 0})
    if existing_login:
        raise HTTPException(status_code=400, detail="Login ID already exists")
    
    # Check if mobile already exists
    existing_mobile = await db.salon_users.find_one({"mobile": mobile}, {"_id": 0})
    if existing_mobile:
        raise HTTPException(status_code=400, detail="Mobile number already exists")
    
    # Check that staff mobile != salon phone
    salon = await db.salons.find_one({"id": user_data.salon_id}, {"_id": 0, "phone": 1})
    if salon and salon.get("phone") == mobile:
        raise HTTPException(status_code=400, detail="Staff mobile cannot be same as salon phone")
    
    # Validate staff_id if provided
    if user_data.staff_id:
        staff = await db.barbers.find_one({"id": user_data.staff_id, "salon_id": user_data.salon_id}, {"_id": 0})
        if not staff:
            raise HTTPException(status_code=404, detail="Staff member not found")
    
    # Hash password
    if len(user_data.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    password_hash = pwd_context.hash(user_data.password)
    
    # Set default permissions for staff
    if user_data.role == "staff" and not user_data.permissions:
        permissions = {
            "can_edit_salon": False,
            "can_access_analytics": False,
            "can_access_financials": False,
            "can_delete_salon": False,
            "can_access_services": False,
            "can_access_gallery": False,
            "can_access_staff": False,
            "can_view_all_staff": False,
            "can_access_marketing": False
        }
    else:
        permissions = user_data.permissions.dict() if user_data.permissions else {
            "can_edit_salon": False,
            "can_access_analytics": False,
            "can_access_financials": False,
            "can_delete_salon": False,
            "can_access_services": False,
            "can_access_gallery": False,
            "can_access_staff": False,
            "can_view_all_staff": False,
            "can_access_marketing": False
        }

    # Validate role
    if user_data.role not in ("admin", "staff", "branch_manager"):
        raise HTTPException(status_code=400, detail="Invalid role")

    # Validate assigned_branch_ids for branch_manager
    assigned_branch_ids = list(user_data.assigned_branch_ids or [])
    if user_data.role == "branch_manager":
        if not assigned_branch_ids:
            raise HTTPException(status_code=400, detail="Branch manager must have at least one assigned_branch_id")
        # Make sure each branch belongs to this salon
        valid = await db.salon_branches.find(
            {"salon_id": user_data.salon_id, "id": {"$in": assigned_branch_ids}},
            {"_id": 0, "id": 1},
        ).to_list(length=None)
        valid_ids = {b["id"] for b in valid}
        invalid = [bid for bid in assigned_branch_ids if bid not in valid_ids]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Invalid branch ids for this salon: {invalid}")

    # Create user
    new_user = {
        "id": str(uuid.uuid4()),
        "salon_id": user_data.salon_id,
        "branch_id": await resolve_branch_id(
            user_data.salon_id,
            assigned_branch_ids[0] if assigned_branch_ids else None,
        ),
        "name": user_data.name,
        "mobile": mobile,
        "login_id": user_data.login_id,
        "password_hash": password_hash,
        "role": user_data.role,
        "staff_id": user_data.staff_id,
        "assigned_branch_ids": assigned_branch_ids,
        "permissions": permissions,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.salon_users.insert_one(new_user)
    
    # Remove password_hash from response
    response_user = new_user.copy()
    response_user["permissions"] = SalonUserPermissions(**permissions)
    
    return SalonUser(**response_user)

@api_router.get("/salon/users")
async def get_salon_users(current_user=Depends(get_current_salon_admin)):
    """Get all users for a salon (admin only)"""
    salon_id = current_user.get("salon_id")
    
    users = await db.salon_users.find(
        {"salon_id": salon_id},
        {"_id": 0, "password_hash": 0}
    ).to_list(100)
    
    return {"users": users}

@api_router.put("/salon/users/{user_id}")
async def update_salon_user(user_id: str, update_data: SalonUserUpdate, current_user=Depends(get_current_salon_admin)):
    """Update salon user (admin only)"""
    salon_id = current_user.get("salon_id")
    
    # Check user exists and belongs to same salon
    user = await db.salon_users.find_one({"id": user_id, "salon_id": salon_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_fields = {}
    
    if update_data.name:
        update_fields["name"] = update_data.name
    
    if update_data.mobile:
        mobile = update_data.mobile
        if not mobile.startswith("+91"):
            mobile = f"+91{mobile}"
        
        # Check mobile doesn't exist for other users
        existing = await db.salon_users.find_one({"mobile": mobile, "id": {"$ne": user_id}}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Mobile number already exists")
        
        # Check against salon phone
        salon = await db.salons.find_one({"id": salon_id}, {"_id": 0, "phone": 1})
        if salon and salon.get("phone") == mobile:
            raise HTTPException(status_code=400, detail="Staff mobile cannot be same as salon phone")
        
        update_fields["mobile"] = mobile
    
    if update_data.login_id:
        # Check login_id doesn't exist for other users
        existing = await db.salon_users.find_one({"login_id": update_data.login_id, "id": {"$ne": user_id}}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Login ID already exists")
        update_fields["login_id"] = update_data.login_id
    
    if update_data.password:
        if len(update_data.password) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        update_fields["password_hash"] = pwd_context.hash(update_data.password)
    
    if update_data.staff_id:
        # Validate staff exists
        staff = await db.barbers.find_one({"id": update_data.staff_id, "salon_id": salon_id}, {"_id": 0})
        if not staff:
            raise HTTPException(status_code=404, detail="Staff member not found")
        update_fields["staff_id"] = update_data.staff_id
    
    if update_data.role is not None:
        if update_data.role not in ("admin", "staff", "branch_manager"):
            raise HTTPException(status_code=400, detail="Invalid role")
        update_fields["role"] = update_data.role

    if update_data.assigned_branch_ids is not None:
        bids = list(update_data.assigned_branch_ids)
        if bids:
            valid = await db.salon_branches.find(
                {"salon_id": salon_id, "id": {"$in": bids}},
                {"_id": 0, "id": 1},
            ).to_list(length=None)
            valid_ids = {b["id"] for b in valid}
            invalid = [bid for bid in bids if bid not in valid_ids]
            if invalid:
                raise HTTPException(status_code=400, detail=f"Invalid branch ids for this salon: {invalid}")
        update_fields["assigned_branch_ids"] = bids

    if update_data.permissions:
        update_fields["permissions"] = update_data.permissions.dict()
    
    if update_data.status:
        update_fields["status"] = update_data.status
    
    if update_fields:
        await db.salon_users.update_one(
            {"id": user_id},
            {"$set": update_fields}
        )
    
    # Return the fully updated user (without password_hash) so callers can render
    # the new role / branch list immediately.
    updated = await db.salon_users.find_one(
        {"id": user_id, "salon_id": salon_id},
        {"_id": 0, "password_hash": 0},
    )
    return updated or {"message": "User updated successfully"}

@api_router.delete("/salon/users/{user_id}")
async def delete_salon_user(user_id: str, current_user=Depends(get_current_salon_admin)):
    """Deactivate salon user (admin only)"""
    salon_id = current_user.get("salon_id")
    
    # Check user exists and belongs to same salon
    user = await db.salon_users.find_one({"id": user_id, "salon_id": salon_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Deactivate instead of delete
    await db.salon_users.update_one(
        {"id": user_id},
        {"$set": {"status": "inactive"}}
    )
    
    return {"message": "User deactivated successfully"}

@api_router.post("/user/login", response_model=User)
async def user_login(credentials: UserLogin):
    """User login with name and phone"""
    phone = credentials.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    if len(phone) != 13:
        raise HTTPException(status_code=400, detail="Invalid phone number format")
    
    # Check if user exists
    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    
    if user:
        # Update name and gender if changed
        update_fields = {}
        if user["name"] != credentials.name:
            update_fields["name"] = credentials.name
        if credentials.gender and user.get("gender") != credentials.gender:
            update_fields["gender"] = credentials.gender
        
        if update_fields:
            await db.users.update_one(
                {"phone": phone},
                {"$set": update_fields}
            )
            user.update(update_fields)
        
        return User(**user)
    else:
        new_user = {
            "id": str(uuid.uuid4()),
            "name": credentials.name,
            "phone": phone,
            "gender": credentials.gender,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(new_user)
        return User(**new_user)


@api_router.get("/users/by-phone/{phone}", response_model=User)
async def get_user_by_phone(phone: str):
    """Get a customer's profile by phone."""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return User(**user)


@api_router.put("/users/by-phone/{phone}", response_model=User)
async def update_user_profile(phone: str, payload: UserProfileUpdate):
    """Update a customer's profile fields (name, gender, dob, email, address, city, pincode)."""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    update_dict = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if update_dict:
        await db.users.update_one({"phone": phone}, {"$set": update_dict})
        user.update(update_dict)
    return User(**user)


# ============ CUSTOMER OTP VERIFICATION ============

@api_router.post("/customer/send-otp")
async def send_customer_otp(request: CustomerOTPRequest):
    """Send OTP to customer phone number via WhatsApp for verification"""
    phone = request.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    if len(phone) != 13:
        raise HTTPException(status_code=400, detail="Invalid phone number format")
    
    logger.info(f"Customer OTP request for phone: {phone}")
    
    # Check if user exists
    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found. Please login first.")
    
    # Generate OTP
    otp = generate_otp()
    logger.info(f"Generated OTP for customer {phone}: {otp}")
    
    # Store OTP in database
    await db.customer_otp.delete_many({"phone": phone})
    await db.customer_otp.insert_one({
        "phone": phone,
        "otp": otp,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
    })
    
    logger.info(f"OTP stored in database for customer {phone}")
    
    # Send OTP via Twilio Verify (SMS)
    whatsapp_result = await send_whatsapp_otp(phone, otp)

    channel = whatsapp_result.get('channel') or 'sms'
    channel_label = 'WhatsApp' if channel == 'whatsapp' else 'SMS'

    response = {
        "success": True,
        "message": f"OTP sent successfully via {channel_label}",
        "phone": phone,
        "delivery_status": whatsapp_result.get('status'),
        "channel": channel,
    }

    status = whatsapp_result.get('status')
    if status == 'mock':
        response['note'] = "⚠️ Messaging not configured. Please contact support."
        logger.warning(f"Mock OTP (Twilio not configured) for customer {phone}: {otp}")
    elif status == 'failed':
        response['note'] = "OTP delivery failed. Please try again."
        response['error'] = whatsapp_result.get('error')
        logger.error(f"OTP delivery failed for customer {phone}: {whatsapp_result.get('error')}")
    else:
        response['note'] = f"OTP sent to your {channel_label}. Please check your messages."
        logger.info(f"✅ OTP sent via {channel} to customer {phone}")

    return response


@api_router.post("/customer/verify-otp")
async def verify_customer_otp(request: CustomerOTPVerify):
    """Verify customer OTP and mark user as OTP verified.

    Production: validates against Twilio Verify.  Mock/dev: validates against
    the locally-stored OTP in `db.customer_otp`.
    """
    phone = request.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"

    if len(phone) != 13:
        raise HTTPException(status_code=400, detail="Invalid phone number format")

    if not await _otp_is_valid(phone, request.otp, db.customer_otp):
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")

    # OTP is valid — delete the local copy and mark user as verified
    await db.customer_otp.delete_many({"phone": phone})

    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    await db.users.update_one(
        {"phone": phone},
        {"$set": {
            "is_otp_verified": True,
            "otp_verified_at": datetime.now(timezone.utc).isoformat(),
        }},
    )

    user["is_otp_verified"] = True
    user["otp_verified_at"] = datetime.now(timezone.utc).isoformat()

    logger.info(f"Customer {phone} OTP verified successfully")

    return {
        "success": True,
        "message": "OTP verified successfully",
        "user": User(**user),
    }


@api_router.get("/customer/{phone}/otp-status")
async def get_customer_otp_status(phone: str):
    """Check if customer is OTP verified"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "phone": phone,
        "is_otp_verified": user.get("is_otp_verified", False),
        "otp_verified_at": user.get("otp_verified_at")
    }


# ============ Module 8 — Customer Auth (OTP + Password) ============
# Long-lived session tokens + password-based login + set/reset password via OTP.

# Customer JWT TTL — 365 days so the app feels "never logged out" unless the
# customer explicitly logs out. (Frontend stores the token in localStorage.)
CUSTOMER_JWT_TTL_DAYS = 365
# Short-lived JWT issued after OTP verification when purpose ∈ {set_password,
# reset_password}. The frontend MUST include it in the next set-password call.
PASSWORD_RESET_TOKEN_TTL_MIN = 10


def _normalize_phone_e164(raw: str) -> str:
    """Normalise to +91XXXXXXXXXX. Raises 400 on bad input."""
    if not raw:
        raise HTTPException(status_code=400, detail="Phone number is required")
    p = raw.strip().replace(" ", "").replace("-", "")
    if not p.startswith("+91"):
        digits = p.lstrip("+")
        if digits.startswith("91") and len(digits) == 12:
            p = "+" + digits
        elif digits.startswith("0") and len(digits) == 11:
            p = "+91" + digits[1:]
        elif len(digits) == 10:
            p = "+91" + digits
        else:
            raise HTTPException(status_code=400, detail="Invalid phone number format")
    if len(p) != 13 or not p[1:].isdigit():
        raise HTTPException(status_code=400, detail="Invalid phone number format")
    return p


def _issue_customer_session_token(user: dict) -> str:
    """Mint a long-lived customer JWT. The frontend treats this as the session."""
    now = datetime.now(timezone.utc)
    payload = {
        "sub": user.get("id"),
        "phone": user.get("phone"),
        "role": "customer",
        "iat": int(now.timestamp()),
        "exp": int((now + timedelta(days=CUSTOMER_JWT_TTL_DAYS)).timestamp()),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def _issue_password_reset_token(phone: str) -> str:
    """Short-lived JWT proving the caller just verified an OTP for password ops."""
    now = datetime.now(timezone.utc)
    payload = {
        "phone": phone,
        "scope": "password_reset",
        "iat": int(now.timestamp()),
        "exp": int((now + timedelta(minutes=PASSWORD_RESET_TOKEN_TTL_MIN)).timestamp()),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def _verify_password_reset_token(token: str, expected_phone: str) -> None:
    """Raise HTTPException 401 if the token is invalid / wrong phone / expired."""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Reset link expired. Please request a new OTP.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid reset token")
    if payload.get("scope") != "password_reset":
        raise HTTPException(status_code=401, detail="Invalid reset token scope")
    if payload.get("phone") != expected_phone:
        raise HTTPException(status_code=401, detail="Reset token does not match this phone")


def _user_to_public(user: dict) -> dict:
    """Strip sensitive fields and add the derived `has_password` flag."""
    out = {k: v for k, v in user.items() if k not in ("password_hash", "_id")}
    out["has_password"] = bool(user.get("password_hash"))
    return out


async def _ensure_user_for_otp(phone: str) -> dict:
    """Get-or-create a lightweight user document so first-contact OTP works."""
    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if user:
        return user
    new_user = {
        "id": str(uuid.uuid4()),
        "name": "",
        "phone": phone,
        "is_otp_verified": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(new_user)
    return new_user


@api_router.post("/auth/customer/check-account")
async def auth_customer_check_account(body: CustomerCheckAccountIn):
    """Tells the frontend which CTA to show on the password tab — set or reset."""
    phone = _normalize_phone_e164(body.phone)
    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    return {
        "phone": phone,
        "exists": bool(user),
        "has_password": bool(user and user.get("password_hash")),
        "is_otp_verified": bool(user and user.get("is_otp_verified")),
    }


@api_router.post("/auth/customer/send-otp")
async def auth_customer_send_otp(body: CustomerSendOtpV2In):
    """Send a WhatsApp OTP for customer login, set-password, or reset-password.

    Auto-creates the user on first-contact login OTP so first-time signup works
    on the same screen.
    """
    phone = _normalize_phone_e164(body.phone)
    purpose = (body.purpose or "login").lower()
    if purpose not in {"login", "set_password", "reset_password"}:
        raise HTTPException(status_code=400, detail="Invalid purpose")

    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if purpose in {"set_password", "reset_password"} and not user:
        raise HTTPException(status_code=404, detail="No account found for this phone")
    if purpose == "reset_password" and not (user and user.get("password_hash")):
        raise HTTPException(status_code=400, detail="No password set. Use 'set password' instead.")
    if purpose == "login" and not user:
        await _ensure_user_for_otp(phone)

    otp = generate_otp()
    logger.info(f"[auth/customer/send-otp] phone={phone} purpose={purpose} otp={otp}")
    await db.customer_otp.delete_many({"phone": phone})
    await db.customer_otp.insert_one({
        "phone": phone,
        "otp": otp,
        "purpose": purpose,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat(),
    })

    whatsapp_result = await send_whatsapp_otp(phone, otp)
    status_str = whatsapp_result.get("status")
    resp = {
        "success": True,
        "phone": phone,
        "purpose": purpose,
        "delivery_status": status_str,
    }
    if status_str == "mock":
        resp["note"] = "⚠️ Messaging not configured. Please contact support."
    elif status_str == "failed":
        resp["note"] = "OTP delivery failed. Please try again."
        resp["error"] = whatsapp_result.get("error")
    else:
        resp["note"] = "OTP sent to your mobile. Please check your messages."
    return resp


@api_router.post("/auth/customer/verify-otp")
async def auth_customer_verify_otp(body: CustomerVerifyOtpV2In):
    """Verify a WhatsApp OTP. Depending on `purpose`:
      • login            → issue a long-lived session token + return user
      • set_password     → issue a short-lived password_reset_token
      • reset_password   → issue a short-lived password_reset_token
    """
    phone = _normalize_phone_e164(body.phone)
    purpose = (body.purpose or "login").lower()
    if purpose not in {"login", "set_password", "reset_password"}:
        raise HTTPException(status_code=400, detail="Invalid purpose")

    # Production: Twilio Verify.  Mock/dev: db.customer_otp.
    if not await _otp_is_valid(phone, body.otp, db.customer_otp):
        raise HTTPException(status_code=400, detail="Invalid or expired OTP")

    # Clean up any local OTP doc (audit only).
    await db.customer_otp.delete_many({"phone": phone})

    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user:
        user = await _ensure_user_for_otp(phone)

    await db.users.update_one(
        {"phone": phone},
        {"$set": {
            "is_otp_verified": True,
            "otp_verified_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    user["is_otp_verified"] = True
    user["otp_verified_at"] = datetime.now(timezone.utc).isoformat()

    if purpose == "login":
        token = _issue_customer_session_token(user)
        logger.info(f"[auth/customer/verify-otp] login OK phone={phone} user={user.get('id')}")
        return {
            "success": True,
            "purpose": "login",
            "access_token": token,
            "token_type": "bearer",
            "user": _user_to_public(user),
            "needs_profile": not (user.get("name") or "").strip(),
        }

    reset_token = _issue_password_reset_token(phone)
    return {
        "success": True,
        "purpose": purpose,
        "password_reset_token": reset_token,
        "expires_in_minutes": PASSWORD_RESET_TOKEN_TTL_MIN,
    }


@api_router.post("/auth/customer/set-password")
async def auth_customer_set_password(body: CustomerSetPasswordIn):
    """Set (or reset) a customer's login password. Requires the short-lived
    `password_reset_token` returned by /auth/customer/verify-otp.
    """
    phone = _normalize_phone_e164(body.phone)
    pw = (body.password or "").strip()
    if len(pw) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    if len(pw) > 128:
        raise HTTPException(status_code=400, detail="Password is too long")

    _verify_password_reset_token(body.password_reset_token, phone)

    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="No account found for this phone")

    password_hash = pwd_context.hash(pw)
    await db.users.update_one(
        {"phone": phone},
        {"$set": {
            "password_hash": password_hash,
            "password_set_at": datetime.now(timezone.utc).isoformat(),
            "is_otp_verified": True,
        }},
    )
    user["password_hash"] = password_hash
    user["is_otp_verified"] = True

    session_token = _issue_customer_session_token(user)
    return {
        "success": True,
        "access_token": session_token,
        "token_type": "bearer",
        "user": _user_to_public(user),
    }


@api_router.post("/auth/customer/login-password")
async def auth_customer_login_password(body: CustomerLoginPasswordIn):
    """Customer login with phone + password. Issues a long-lived session JWT."""
    phone = _normalize_phone_e164(body.phone)
    pw = (body.password or "").strip()
    if not pw:
        raise HTTPException(status_code=400, detail="Password is required")

    user = await db.users.find_one({"phone": phone}, {"_id": 0})
    if not user or not user.get("password_hash"):
        raise HTTPException(status_code=401, detail="Invalid phone or password")
    if not pwd_context.verify(pw, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid phone or password")

    token = _issue_customer_session_token(user)
    logger.info(f"[auth/customer/login-password] OK phone={phone}")
    return {
        "success": True,
        "access_token": token,
        "token_type": "bearer",
        "user": _user_to_public(user),
    }


# ============================================================================
# GOOGLE LOGIN — Emergent-managed Google OAuth (Item 9).
# ----------------------------------------------------------------------------
# The frontend redirects to `https://auth.emergentagent.com/?redirect=<callback>`
# After Google sign-in, the user lands at <callback>#session_id=<id>. The
# frontend POSTs the session_id (plus an `audience`) to this endpoint. We
# exchange it with Emergent's identity service to get the user's email/name,
# find-or-create the corresponding user record in the per-audience collection,
# and return the audience's existing JWT so the rest of the app keeps working
# without any cookie / session changes.
# ============================================================================

EMERGENT_AUTH_BASE = os.environ.get(
    "EMERGENT_AUTH_BASE",
    "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
)


async def _exchange_emergent_session(session_id: str) -> dict:
    """Call Emergent's auth backend to exchange a session_id for user data."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id is required")
    try:
        async with httpx.AsyncClient(timeout=10.0) as http:
            res = await http.get(
                EMERGENT_AUTH_BASE,
                headers={"X-Session-ID": session_id},
            )
        if res.status_code != 200:
            logger.warning(
                f"[google] Emergent session exchange failed status={res.status_code} body={res.text[:200]}"
            )
            raise HTTPException(status_code=401, detail="Google sign-in could not be completed. Please try again.")
        data = res.json()
        if not data.get("email"):
            raise HTTPException(status_code=401, detail="Google sign-in returned no email")
        return data
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[google] Emergent exchange error: {e}")
        raise HTTPException(status_code=502, detail="Google sign-in service is temporarily unavailable")


class GoogleAuthIn(BaseModel):
    session_id: str
    audience: str  # one of: customer | salon | platform | supplier


@api_router.post("/auth/google")
async def auth_google(payload: GoogleAuthIn):
    """Unified Google sign-in for all four audiences (Item 9).

    Returns the SAME shape each audience's existing login endpoint returns, so
    the frontend stores the token in its existing auth context.
    """
    audience = (payload.audience or "").strip().lower()
    if audience not in ("customer", "salon", "platform", "supplier"):
        raise HTTPException(status_code=400, detail="audience must be customer | salon | platform | supplier")

    data = await _exchange_emergent_session(payload.session_id)
    email = (data.get("email") or "").strip().lower()
    name = (data.get("name") or "").strip() or email.split("@")[0]
    picture = data.get("picture")
    google_sub = data.get("id")
    now_iso = datetime.now(timezone.utc).isoformat()

    # --- CUSTOMER ----------------------------------------------------------
    if audience == "customer":
        user = await db.users.find_one({"email": email}, {"_id": 0})
        if not user:
            user = {
                "id": str(uuid.uuid4()),
                "name": name,
                "email": email,
                "phone": None,
                "picture": picture,
                "google_sub": google_sub,
                "is_otp_verified": True,
                "auth_provider": "google",
                "created_at": now_iso,
            }
            await db.users.insert_one(user)
        else:
            await db.users.update_one(
                {"id": user["id"]},
                {"$set": {
                    "google_sub": google_sub,
                    "picture": picture or user.get("picture"),
                    "is_otp_verified": True,
                    "last_login_at": now_iso,
                }},
            )
            user["is_otp_verified"] = True
        token = _issue_customer_session_token(user)
        return {
            "success": True,
            "access_token": token,
            "token_type": "bearer",
            "user": _user_to_public(user),
        }

    # --- SALON ADMIN (multi-user) -----------------------------------------
    if audience == "salon":
        # Salon users authenticate via `salon_users`. Match on email; deny if
        # this email isn't linked to any salon user.
        salon_user = await db.salon_users.find_one({"email": email}, {"_id": 0})
        if not salon_user:
            raise HTTPException(
                status_code=403,
                detail="No salon account is linked to this Google email. Ask your salon admin to add you.",
            )
        await db.salon_users.update_one(
            {"id": salon_user["id"]},
            {"$set": {"last_login_at": now_iso, "google_sub": google_sub}},
        )
        # Mirror the existing salon-user login response.
        token = create_access_token({
            "sub": salon_user.get("salon_id"),
            "salon_user_id": salon_user.get("id"),
            "role": salon_user.get("role"),
            "login_id": salon_user.get("login_id"),
        })
        return {
            "access_token": token,
            "token_type": "bearer",
            "salon_id": salon_user.get("salon_id"),
            "user": {
                "id": salon_user.get("id"),
                "name": salon_user.get("name"),
                "email": salon_user.get("email"),
                "role": salon_user.get("role"),
                "assigned_branch_ids": salon_user.get("assigned_branch_ids", []),
            },
        }

    # --- PLATFORM ADMIN ----------------------------------------------------
    if audience == "platform":
        admin = await db.platform_admins.find_one({"email": email, "status": "active"}, {"_id": 0})
        if not admin:
            raise HTTPException(
                status_code=403,
                detail="This Google account is not authorised as a platform admin.",
            )
        await db.platform_admins.update_one(
            {"id": admin["id"]},
            {"$set": {"last_login_at": now_iso, "google_sub": google_sub, "updated_at": now_iso}},
        )
        token = platform_admin_mod._make_access_token(admin)
        return {
            "access_token": token,
            "token_type": "bearer",
            "admin": {
                "id": admin["id"],
                "mobile": admin.get("mobile"),
                "name": admin.get("name"),
                "email": admin.get("email"),
                "is_owner": bool(admin.get("is_owner")),
            },
        }

    # --- SUPPLIER ----------------------------------------------------------
    # audience == "supplier"
    supplier = await db.suppliers.find_one({"email": email}, {"_id": 0})
    if not supplier:
        raise HTTPException(
            status_code=403,
            detail="No supplier account is linked to this Google email. Please sign up first.",
        )
    if supplier.get("status") != "active":
        raise HTTPException(
            status_code=403,
            detail=f"Supplier account is {supplier.get('status', 'pending')}. Please contact support.",
        )
    await db.suppliers.update_one(
        {"id": supplier["id"]},
        {"$set": {"last_login_at": now_iso, "google_sub": google_sub}},
    )
    token = supplier_auth_mod._make_access_token(supplier)
    return {
        "access_token": token,
        "token_type": "bearer",
        "supplier": {
            "id": supplier["id"],
            "name": supplier.get("name"),
            "email": supplier.get("email"),
            "business_name": supplier.get("business_name"),
            "status": supplier.get("status"),
        },
    }


async def get_current_customer(
    credentials: HTTPAuthorizationCredentials = Depends(security),
) -> dict:
    """Bearer-token dependency that resolves the customer User document."""
    payload = verify_token(credentials.credentials)
    if not payload or payload.get("role") != "customer":
        raise HTTPException(status_code=401, detail="Invalid customer session")
    user = await db.users.find_one({"id": payload.get("sub")}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Customer not found")
    return user


@api_router.get("/auth/customer/me")
async def auth_customer_me(user=Depends(get_current_customer)):
    """Returns the currently-logged-in customer. Frontend calls this on boot
    (with the localStorage token) to rehydrate the session."""
    return {"user": _user_to_public(user)}


@api_router.get("/customer/{phone}/last-salon")
async def get_customer_last_salon(phone: str):
    """Get the last visited salon for smart routing"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    customer = await db.customers.find_one({"phone": phone}, {"_id": 0, "last_booking_salon_id": 1})
    if customer and customer.get("last_booking_salon_id"):
        return {"salon_id": customer["last_booking_salon_id"]}
    return {"salon_id": None}


@api_router.get("/salons/{salon_id}/customers")
async def get_salon_customers(salon_id: str, branch_id: Optional[str] = None, current_user=Depends(get_current_salon_user)):
    """Get all customers who have booked at this salon + manually added.
    Each customer also includes the active wallet balance (if any).
    If branch_id is provided, restrict to customers who have booked at (or were
    manually added at) that specific branch.
    """
    if is_branch_manager(current_user):
        branch_id = enforce_branch_for_manager(current_user, branch_id)
    # Get unique customers from tokens
    tokens_query = {"salon_id": salon_id, "customer_status": {"$ne": "deleted"}}
    if branch_id:
        tokens_query["branch_id"] = branch_id
    tokens = await db.tokens.find(
        tokens_query,
        {"_id": 0, "user_id": 1, "customer_name": 1, "phone": 1, "created_at": 1, "date": 1}
    ).to_list(10000)

    # Group by phone to get unique customers + track last visit date
    customers_map = {}
    last_visit_by_phone: Dict[str, str] = {}
    for token in tokens:
        phone = token.get('phone')
        if not phone:
            continue
        visit_dt = token.get('date') or token.get('created_at') or ''
        if isinstance(visit_dt, str) and visit_dt:
            prev = last_visit_by_phone.get(phone, '')
            if visit_dt > prev:
                last_visit_by_phone[phone] = visit_dt
        if phone not in customers_map:
            user_data = None
            if token.get('user_id'):
                user_data = await db.users.find_one({"id": token['user_id']}, {"_id": 0})
            customers_map[phone] = {
                "phone": phone,
                "name": token.get('customer_name'),
                "user_id": token.get('user_id'),
                "gender": user_data.get('gender') if user_data else None,
                "date_of_birth": user_data.get('date_of_birth') if user_data else None,
                # Pull profile photo from user account if guest uploaded via customer app
                "photo_url": (user_data or {}).get('profile_photo') or (user_data or {}).get('photo_url'),
            }

    # Also include manually added customers
    manual_query = {"salon_id": salon_id, "status": {"$ne": "deleted"}}
    if branch_id:
        manual_query["branch_id"] = branch_id
    manual_customers = await db.salon_customers.find(
        manual_query, {"_id": 0}
    ).to_list(10000)

    for mc in manual_customers:
        phone = mc.get('phone')
        if not phone:
            continue
        if phone not in customers_map:
            customers_map[phone] = {
                "phone": phone,
                "name": mc.get('name'),
                "user_id": None,
                "gender": mc.get('gender', 'Men'),
                "date_of_birth": mc.get('date_of_birth') or mc.get('dob'),
                "anniversary": mc.get('anniversary'),
                "source": mc.get('source') or "manual",
                "photo_url": mc.get('photo_url'),
                "instagram_id": mc.get('instagram_id'),
                "facebook_id": mc.get('facebook_id'),
                "preferred_barber_id": mc.get('preferred_barber_id'),
                "tags": mc.get('tags') or [],
                "id": mc.get('id'),
                # Seed-friendly aggregate fields (used by Guests V2 page)
                "visit_count": mc.get('visit_count'),
                "total_spend": mc.get('total_spend'),
                "notes": mc.get('notes'),
            }
        else:
            # Merge — salon-master fields take priority when the token record was thin.
            cust = customers_map[phone]
            if not cust.get('gender'):
                cust['gender'] = mc.get('gender', 'Men')
            if not cust.get('date_of_birth'):
                cust['date_of_birth'] = mc.get('date_of_birth') or mc.get('dob')
            if not cust.get('anniversary'):
                cust['anniversary'] = mc.get('anniversary')
            if not cust.get('photo_url') and mc.get('photo_url'):
                cust['photo_url'] = mc.get('photo_url')
            cust.setdefault('instagram_id', mc.get('instagram_id'))
            cust.setdefault('facebook_id', mc.get('facebook_id'))
            cust.setdefault('preferred_barber_id', mc.get('preferred_barber_id'))
            cust.setdefault('tags', mc.get('tags') or [])
            cust.setdefault('id', mc.get('id'))
            # Prefer master-doc aggregates when present (seeded data / manual overrides)
            if mc.get('visit_count') is not None:
                cust['visit_count'] = mc.get('visit_count')
            if mc.get('total_spend') is not None:
                cust['total_spend'] = mc.get('total_spend')
            if mc.get('notes'):
                cust['notes'] = mc.get('notes')

    # Compute visit_count + total_spend from tokens when master-doc value is missing
    if customers_map:
        phones_needing_agg = [p for p, c in customers_map.items() if c.get('visit_count') is None or c.get('total_spend') is None]
        if phones_needing_agg:
            token_agg_q = {"salon_id": salon_id, "phone": {"$in": phones_needing_agg}, "customer_status": {"$ne": "deleted"}}
            if branch_id:
                token_agg_q["branch_id"] = branch_id
            agg_tokens = await db.tokens.find(token_agg_q, {"_id": 0, "phone": 1, "final_amount": 1, "total_amount": 1, "status": 1}).to_list(20000)
            counts: Dict[str, int] = {}
            spends: Dict[str, float] = {}
            for t in agg_tokens:
                p = t.get('phone')
                if not p:
                    continue
                counts[p] = counts.get(p, 0) + 1
                if t.get('status') in ('completed', 'complete'):
                    amt = t.get('final_amount') or t.get('total_amount') or 0
                    spends[p] = spends.get(p, 0.0) + float(amt)
            for p, c in customers_map.items():
                if c.get('visit_count') is None:
                    c['visit_count'] = counts.get(p, 0)
                if c.get('total_spend') is None:
                    c['total_spend'] = spends.get(p, 0.0)

    # Attach last visit date computed from tokens
    for phone, cust in customers_map.items():
        cust['last_visit'] = last_visit_by_phone.get(phone)
    
    # Attach active wallet balance for each customer (if any active membership)
    if customers_map:
        phones = list(customers_map.keys())
        memberships = await db.customer_memberships.find(
            {"salon_id": salon_id, "customer_phone": {"$in": phones}, "is_active": True},
            {"_id": 0, "customer_phone": 1, "wallet_balance": 1, "membership_name": 1}
        ).to_list(20000)
        wallet_by_phone = {}
        for m in memberships:
            p = m.get("customer_phone")
            if p:
                wallet_by_phone[p] = {
                    "wallet_balance": float(m.get("wallet_balance") or 0),
                    "membership_name": m.get("membership_name") or ""
                }
        for phone, cust in customers_map.items():
            info = wallet_by_phone.get(phone, {"wallet_balance": 0, "membership_name": ""})
            # Membership overrides only when present
            cust["wallet_balance"] = info["wallet_balance"]
            cust["membership_name"] = info["membership_name"]

        # Fill from master doc if endpoint's membership lookup returned empty (for seeded / manual data)
        for mc in manual_customers:
            p = mc.get('phone')
            if p and p in customers_map:
                if not customers_map[p].get("wallet_balance"):
                    customers_map[p]["wallet_balance"] = float(mc.get("wallet_balance") or 0)
                if not customers_map[p].get("membership_name"):
                    customers_map[p]["membership_name"] = mc.get("membership_name") or ""
    
    return {"customers": list(customers_map.values())}

@api_router.post("/salons/{salon_id}/customers")
async def add_salon_customer(salon_id: str, body: dict, current_user=Depends(get_current_salon_user)):
    """Manually add a customer to the salon"""
    name = body.get("name", "").strip()
    phone = body.get("phone", "").strip()
    gender = body.get("gender", "Men")
    
    if not name:
        raise HTTPException(status_code=400, detail="Customer name is required")
    
    if phone:
        # Normalize phone
        phone = phone.replace(" ", "").replace("-", "")
        if not phone.startswith("+91"):
            phone = f"+91{phone}"
        
        # Check if already exists
        existing = await db.salon_customers.find_one({
            "salon_id": salon_id,
            "phone": phone
        })
        if existing:
            # Update name/gender
            await db.salon_customers.update_one(
                {"id": existing["id"]},
                {"$set": {"name": name, "gender": gender}}
            )
            return {"message": "Customer updated", "customer": {
                "id": existing["id"], "name": name, "phone": phone, "gender": gender
            }}
    
    customer = {
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "branch_id": await resolve_branch_id(salon_id, body.get("branch_id")),
        "name": name,
        "phone": phone or None,
        "gender": gender,
        "email": body.get("email") or None,
        "tags": body.get("tags") or [],
        "notes": body.get("notes") or "",
        # Extended master fields (added for Home v2 New Guest drawer)
        "photo_url": body.get("photo_url") or None,   # data-URL or CDN URL
        "dob": body.get("dob") or None,               # ISO date "YYYY-MM-DD"
        "anniversary": body.get("anniversary") or None,  # ISO date "YYYY-MM-DD"
        "preferred_barber_id": body.get("preferred_barber_id") or None,
        "instagram_id": (body.get("instagram_id") or "").strip() or None,
        "facebook_id":  (body.get("facebook_id")  or "").strip() or None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        # source enum: online|qr|owner|direct  (legacy value "manual" also allowed)
        "source": (body.get("source") or "owner"),
    }
    
    await db.salon_customers.insert_one(customer)
    customer.pop("_id", None)
    
    return {"message": "Customer added", "customer": customer}


@api_router.put("/salons/{salon_id}/customers/{phone}")
async def update_salon_customer(
    salon_id: str,
    phone: str,
    body: dict,
    current_user=Depends(get_current_salon_user),
):
    """Update a customer's master details (name, phone, gender, date_of_birth).

    Item 5b — Customer Master CRUD. Identified by their current phone in URL.
    Soft-blocks phone collisions inside the same salon. When phone changes,
    salon_customers and downstream tokens.phone are updated together.
    """
    # Normalise the lookup phone
    lookup_phone = (phone or "").strip()
    if not lookup_phone.startswith("+91"):
        lookup_phone = f"+91{lookup_phone}"

    new_name = (body.get("name") or "").strip()
    new_phone_raw = (body.get("phone") or "").replace(" ", "").replace("-", "").strip()
    new_phone = new_phone_raw
    if new_phone and not new_phone.startswith("+91"):
        new_phone = f"+91{new_phone}"
    new_gender = body.get("gender")
    new_dob = body.get("date_of_birth")

    updates = {}
    if new_name:
        updates["name"] = new_name
    if new_gender:
        updates["gender"] = new_gender
    if new_dob is not None:
        updates["date_of_birth"] = new_dob

    # If phone is being changed, validate no collision
    phone_changed = bool(new_phone) and new_phone != lookup_phone
    if phone_changed:
        collision = await db.salon_customers.find_one({
            "salon_id": salon_id,
            "phone": new_phone,
            "status": {"$ne": "deleted"},
        })
        if collision:
            raise HTTPException(status_code=409, detail="Another customer in this salon already uses that phone.")
        updates["phone"] = new_phone

    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Update or upsert the salon_customers master row
    existing = await db.salon_customers.find_one({"salon_id": salon_id, "phone": lookup_phone})
    if existing:
        await db.salon_customers.update_one(
            {"id": existing["id"]},
            {"$set": updates},
        )
    else:
        # The phone might only exist on tokens — create a master row so future edits work.
        await db.salon_customers.insert_one({
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            "branch_id": await resolve_branch_id(salon_id, None),
            "name": new_name or "Customer",
            "phone": new_phone or lookup_phone,
            "gender": new_gender or "Men",
            "date_of_birth": new_dob,
            "created_at": updates["updated_at"],
            "updated_at": updates["updated_at"],
            "source": "edited",
        })

    # Mirror phone changes onto existing tokens of this salon
    if phone_changed:
        await db.tokens.update_many(
            {"salon_id": salon_id, "phone": lookup_phone},
            {"$set": {"phone": new_phone}},
        )

    # Mirror name changes onto recent tokens (snapshot) when the customer renames
    if new_name:
        await db.tokens.update_many(
            {"salon_id": salon_id, "phone": new_phone or lookup_phone},
            {"$set": {"customer_name": new_name}},
        )

    return {
        "message": "Customer updated",
        "customer": {
            "phone": new_phone or lookup_phone,
            "name": new_name or (existing.get("name") if existing else None),
            "gender": new_gender or (existing.get("gender") if existing else None),
            "date_of_birth": new_dob if new_dob is not None else (existing.get("date_of_birth") if existing else None),
        },
    }


@api_router.delete("/salons/{salon_id}/customers/{phone}")
async def delete_salon_customer(
    salon_id: str,
    phone: str,
    current_user=Depends(get_current_salon_user),
):
    """Soft-delete a customer from the salon master (Item 5a).

    Marks status='deleted' on salon_customers AND tags tokens with
    customer_status='deleted' so the customer list query can exclude them.
    The booking history rows themselves are preserved.
    """
    role = (current_user.get("role") or "").lower()
    if role not in ("admin", "salon_admin", "salon"):
        raise HTTPException(status_code=403, detail="Only salon admins can delete customers")

    lookup_phone = (phone or "").strip()
    if not lookup_phone.startswith("+91"):
        lookup_phone = f"+91{lookup_phone}"

    now_iso = datetime.now(timezone.utc).isoformat()
    res = await db.salon_customers.update_many(
        {"salon_id": salon_id, "phone": lookup_phone},
        {"$set": {"status": "deleted", "deleted_at": now_iso}},
    )

    # Even if no master row exists, mark all booking tokens so listing hides them.
    await db.tokens.update_many(
        {"salon_id": salon_id, "phone": lookup_phone},
        {"$set": {"customer_status": "deleted"}},
    )

    if res.matched_count == 0:
        # Customer only existed via tokens; the tokens-update above is enough.
        return {"message": "Customer removed from master (booking history preserved)"}

    return {"message": "Customer deleted"}


# ============================================================================
# BULK CUSTOMER UPLOAD (Excel)
# ============================================================================

@api_router.get("/salons/{salon_id}/customers/template")
async def download_customer_template(salon_id: str, current_user=Depends(get_current_salon_user)):
    """Download an Excel template for bulk customer upload.

    Columns: Name, Mobile No., Gender, Date of Birth (YYYY-MM-DD)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl unavailable: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = ["Name", "Mobile No.", "Gender", "Date of Birth"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Sample rows (will be visible to user as a guide)
    samples = [
        ["Rahul Sharma", "9876543210", "Men", "1990-05-15"],
        ["Priya Singh", "9123456789", "Women", "1995-11-23"],
        ["Aman Verma", "+919999988888", "Men", ""],
    ]
    for r_idx, row in enumerate(samples, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Set column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18

    # Add an instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "Bulk Customer Upload — Instructions",
        "",
        "1. Fill rows starting from row 2 in the 'Customers' sheet.",
        "2. Required columns:",
        "   • Name           — full name of the customer",
        "   • Mobile No.     — 10 digits or +91 prefixed (used as the unique key)",
        "   • Gender         — Men / Women / Kids",
        "   • Date of Birth  — YYYY-MM-DD (optional, leave blank if unknown)",
        "3. Duplicate phone numbers are skipped (existing customers are NOT overwritten).",
        "4. Save as .xlsx and upload via the 'Bulk Upload' button.",
    ]
    for i, line in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_upload_template.xlsx"},
    )


def _normalize_phone_for_bulk(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Excel may give floats like 9876543210.0 — strip trailing .0
    if s.endswith(".0"):
        s = s[:-2]
    # Remove non-digits except leading +
    digits = "".join(ch for ch in s if ch.isdigit() or ch == "+")
    if digits.startswith("+91"):
        digits = digits[3:]
    elif digits.startswith("+"):
        digits = digits.lstrip("+")
    # Strip Indian country code if duplicated
    if len(digits) > 10 and digits.startswith("91"):
        digits = digits[-10:]
    if len(digits) != 10 or not digits.isdigit():
        return None
    return f"+91{digits}"


def _normalize_gender_for_bulk(raw) -> str:
    if raw is None:
        return "Men"
    s = str(raw).strip().lower()
    if s in ("m", "male", "men", "man", "boy"):
        return "Men"
    if s in ("f", "female", "women", "woman", "girl", "lady"):
        return "Women"
    if s in ("k", "kid", "kids", "child", "children"):
        return "Kids"
    if s in ("o", "other", "others"):
        return "Other"
    return "Men"


def _normalize_dob_for_bulk(raw) -> Optional[str]:
    if raw is None:
        return None
    # If Excel passes a datetime
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return None
    # Try ISO YYYY-MM-DD first
    fmts = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    return None


@api_router.post("/salons/{salon_id}/customers/bulk-upload")
async def bulk_upload_customers(
    salon_id: str,
    file: UploadFile = File(...),
    branch_id: Optional[str] = Form(None),
    current_user=Depends(get_current_salon_user),
):
    """Bulk upload customers from an Excel file.

    Accepts .xlsx and .xls. Required columns (case-insensitive):
    Name, Mobile No., Gender, Date of Birth.
    Duplicate phone numbers are skipped (existing customers preserved).
    """
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    fname = file.filename.lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls") or fname.endswith(".csv")):
        raise HTTPException(status_code=400, detail="File must be .xlsx, .xls or .csv")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB)")

    rows: List[dict] = []
    try:
        if fname.endswith(".csv"):
            import csv as _csv
            text = content.decode("utf-8", errors="ignore")
            reader = _csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append({k.strip().lower(): v for k, v in r.items() if k})
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)
            # Try the first sheet that looks like 'Customers' or just the active one
            ws = None
            for name in ("Customers", "customers"):
                if name in wb.sheetnames:
                    ws = wb[name]
                    break
            if ws is None:
                ws = wb.active

            # Header detection (first non-empty row)
            header_row = None
            for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if any((c is not None and str(c).strip() != "") for c in r):
                    header_row = [str(c).strip().lower() if c is not None else "" for c in r]
                    break
            if not header_row:
                raise HTTPException(status_code=400, detail="Excel file is empty")

            for r in ws.iter_rows(min_row=2, values_only=True):
                if all(c is None or str(c).strip() == "" for c in r):
                    continue
                row = {}
                for idx, val in enumerate(r):
                    if idx < len(header_row):
                        key = header_row[idx]
                        row[key] = val
                rows.append(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found")

    # Map flexible header names to canonical fields
    def _find(d: dict, *keys: str):
        for k in keys:
            if k in d:
                return d[k]
            for actual in d.keys():
                if actual.replace(" ", "").replace(".", "").replace("_", "").lower() == k.replace(" ", "").replace(".", "").replace("_", "").lower():
                    return d[actual]
        return None

    resolved_branch = await resolve_branch_id(salon_id, branch_id)
    now_iso = datetime.now(timezone.utc).isoformat()

    inserted = 0
    skipped_duplicate = 0
    skipped_invalid = 0
    errors: List[dict] = []

    # Pre-fetch existing phones for dedup
    existing_docs = await db.salon_customers.find(
        {"salon_id": salon_id},
        {"_id": 0, "phone": 1}
    ).to_list(20000)
    existing_phones = {d.get("phone") for d in existing_docs if d.get("phone")}

    # Also dedup against existing tokens phones
    token_phones = await db.tokens.distinct("phone", {"salon_id": salon_id})
    existing_phones.update(p for p in token_phones if p)

    new_docs: List[dict] = []
    for idx, row in enumerate(rows, start=2):
        name = _find(row, "name", "customername", "customer name")
        mobile = _find(row, "mobile no", "mobile", "mobileno", "phone", "mobile no.", "mobilenumber", "mobile number")
        gender = _find(row, "gender")
        dob = _find(row, "date of birth", "dob", "dateofbirth", "birthday", "birth date")

        name_str = (str(name).strip() if name is not None else "")
        phone_norm = _normalize_phone_for_bulk(mobile)
        gender_norm = _normalize_gender_for_bulk(gender)
        dob_norm = _normalize_dob_for_bulk(dob)

        if not name_str or not phone_norm:
            skipped_invalid += 1
            errors.append({"row": idx, "reason": "Missing/invalid Name or Mobile No.", "raw": {"name": name, "mobile": mobile}})
            continue

        if phone_norm in existing_phones:
            skipped_duplicate += 1
            continue

        new_docs.append({
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            "branch_id": resolved_branch,
            "name": name_str,
            "phone": phone_norm,
            "gender": gender_norm,
            "dob": dob_norm,
            "created_at": now_iso,
            "source": "bulk_upload",
        })
        existing_phones.add(phone_norm)

    if new_docs:
        await db.salon_customers.insert_many(new_docs)
        inserted = len(new_docs)

        # Optionally, upsert into users collection (for DOB tracking, no auth)
        for d in new_docs:
            try:
                await db.users.update_one(
                    {"phone": d["phone"]},
                    {"$setOnInsert": {
                        "id": str(uuid.uuid4()),
                        "name": d["name"],
                        "phone": d["phone"],
                        "gender": d["gender"],
                        "dob": d.get("dob"),
                        "is_otp_verified": False,
                        "created_at": now_iso,
                    },
                     "$set": {
                         "name": d["name"],
                         "gender": d["gender"],
                         **({"dob": d["dob"]} if d.get("dob") else {}),
                     }},
                    upsert=True,
                )
            except Exception:
                pass

    return {
        "inserted": inserted,
        "skipped_duplicate": skipped_duplicate,
        "skipped_invalid": skipped_invalid,
        "total_rows": len(rows),
        "errors": errors[:50],  # cap
        "message": f"Imported {inserted} customers · {skipped_duplicate} duplicates skipped · {skipped_invalid} invalid rows.",
    }


@api_router.get("/salons/{salon_id}/customers/csv-template")
async def get_customers_csv_template(salon_id: str):
    """Return a downloadable CSV template for the guest bulk-upload.
    Public endpoint (safe – it's just a template with dummy rows)."""
    import csv as _csv
    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(["Name", "Mobile No.", "Gender", "Date of Birth"])
    writer.writerow(["Priya Sharma", "9876543210", "Female", "1994-03-14"])
    writer.writerow(["Amit Kumar", "9123456789", "Male", "1988-11-02"])
    csv_bytes = buf.getvalue().encode("utf-8")
    from fastapi.responses import Response as _Response
    return _Response(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=guests_template.csv"},
    )




# ============================================================================
# MENU PARSING via OpenAI GPT-5 (Vision)
# ============================================================================

@api_router.post("/salons/{salon_id}/services/parse-menu")
async def parse_salon_menu(
    salon_id: str,
    file: UploadFile = File(...),
    current_user=Depends(get_current_salon_user),
):
    """Parse a salon menu (PDF/PNG/JPG) using GPT-5 vision and return a list of
    services and packages extracted from the menu. The salon can then choose to
    'add' (merge with existing) or 'replace' the predefined services.
    """
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    fname = file.filename.lower()
    allowed_exts = (".pdf", ".png", ".jpg", ".jpeg", ".webp")
    if not fname.endswith(allowed_exts):
        raise HTTPException(status_code=400, detail="File must be PDF, PNG, JPG, JPEG or WEBP")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10 MB)")

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="EMERGENT_LLM_KEY not configured")

    # Gemini supports both PDFs and images natively via FileContentWithMimeType,
    # and has a more generous budget on the Emergent Universal LLM Key than GPT-5.
    import tempfile
    suffix = os.path.splitext(fname)[1]
    mime_map = {
        ".pdf": "application/pdf",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }
    mime = mime_map.get(suffix, "application/octet-stream")

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

        system_message = (
            "You are an expert salon-services extraction assistant. The user will upload a "
            "salon menu (PDF, photo, or image). Extract every service and package on the menu. "
            "Return ONLY a valid JSON object with the following shape and NOTHING ELSE — no "
            "prose, no markdown fences, no explanations:\n"
            "{\n"
            '  "services": [\n'
            "    {\n"
            '      "service_name": "string",\n'
            '      "description": "string (optional, may be empty)",\n'
            '      "category": "string — best-fit category like Hair, Facial, Spa, Bleach, '
            'Pedicure, Manicure, Wax, Threading, Massage, etc.",\n'
            '      "gender": "Men | Women | Unisex",\n'
            '      "default_duration": 30,\n'
            '      "base_price": 0\n'
            "    }\n"
            "  ],\n"
            '  "packages": [\n'
            "    {\n"
            '      "package_name": "string",\n'
            '      "service_names": ["service 1", "service 2"],\n'
            '      "description": "string (optional)",\n'
            '      "gender": "Men | Women | Unisex",\n'
            '      "package_price": 0\n'
            "    }\n"
            "  ]\n"
            "}\n"
            "Rules:\n"
            "1. Strip currency symbols (Rs, INR, ₹) — return integer prices only.\n"
            "2. If gender is unclear, default to 'Unisex'.\n"
            "3. If duration is missing, estimate by service type (haircut=30, facial=60, "
            "spa=75, threading=15, massage=60).\n"
            "4. Skip non-service items (taxes, addresses, phone numbers, contact info).\n"
            "5. Ensure JSON is parseable — no trailing commas, no comments, no markdown."
        )

        chat = LlmChat(
            api_key=api_key,
            session_id=f"menu-parse-{salon_id}-{uuid.uuid4()}",
            system_message=system_message,
        ).with_model("gemini", "gemini-2.5-pro")

        file_content = FileContentWithMimeType(
            file_path=tmp_path,
            mime_type=mime,
        )
        msg = UserMessage(
            text=(
                "Extract every salon service and package from this menu and return strict JSON "
                "in the exact schema described in the system prompt. Do not include any markdown "
                "or prose outside the JSON object."
            ),
            file_contents=[file_content],
        )

        try:
            raw = await chat.send_message(msg)
        except Exception as e:
            logger.exception("Menu parse LLM error: %s", e)
            raise HTTPException(status_code=502, detail=f"AI parsing failed: {e}")

        # Best-effort JSON extraction (model may include code fences)
        raw_text = (raw or "").strip()
        if raw_text.startswith("```"):
            raw_text = raw_text.strip("`")
            if raw_text.lower().startswith("json"):
                raw_text = raw_text[4:].strip()
        first = raw_text.find("{")
        last = raw_text.rfind("}")
        if first == -1 or last == -1:
            raise HTTPException(status_code=502, detail="AI response was not valid JSON")
        json_str = raw_text[first:last + 1]
        try:
            data = json.loads(json_str)
        except Exception:
            import re as _re
            cleaned = _re.sub(r",\s*([}\]])", r"\1", json_str)
            try:
                data = json.loads(cleaned)
            except Exception as e:
                raise HTTPException(status_code=502, detail=f"AI returned invalid JSON: {e}")

        services_out = []
        for s in (data.get("services") or []):
            if not s.get("service_name"):
                continue
            services_out.append({
                "service_name": str(s.get("service_name")).strip()[:120],
                "description": str(s.get("description") or "").strip()[:500],
                "category": str(s.get("category") or "General").strip()[:60] or "General",
                "gender": (s.get("gender") if s.get("gender") in ("Men", "Women", "Unisex") else "Unisex"),
                "default_duration": int(s.get("default_duration") or 30) or 30,
                "base_price": int(round(float(s.get("base_price") or 0))),
            })

        packages_out = []
        for p in (data.get("packages") or []):
            if not p.get("package_name"):
                continue
            packages_out.append({
                "package_name": str(p.get("package_name")).strip()[:120],
                "service_names": [str(x).strip()[:120] for x in (p.get("service_names") or []) if x],
                "description": str(p.get("description") or "").strip()[:500],
                "gender": (p.get("gender") if p.get("gender") in ("Men", "Women", "Unisex") else "Unisex"),
                "package_price": int(round(float(p.get("package_price") or 0))),
            })

        return {
            "services": services_out,
            "packages": packages_out,
            "service_count": len(services_out),
            "package_count": len(packages_out),
            "message": f"Parsed {len(services_out)} services and {len(packages_out)} packages from the menu.",
        }
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@api_router.post("/salons/{salon_id}/services/apply-parsed")
async def apply_parsed_menu_services(
    salon_id: str,
    body: dict = Body(...),
    current_user=Depends(get_current_salon_user),
):
    """Apply previously parsed services/packages to the salon.

    Body:
      {
        "services": [...],            # list returned from parse-menu (or edited by salon)
        "packages": [...],            # optional list of packages
        "mode": "add" | "replace"     # 'add' merges with existing, 'replace' wipes salon services & packages first
      }
    Returns counts of created services, skipped duplicates, packages created.
    """
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")

    services_in = body.get("services") or []
    packages_in = body.get("packages") or []
    mode = (body.get("mode") or "add").lower()
    if mode not in ("add", "replace"):
        raise HTTPException(status_code=400, detail="mode must be 'add' or 'replace'")

    now_iso = datetime.now(timezone.utc).isoformat()

    if mode == "replace":
        # Disable all currently-enabled services for this salon and remove its custom services.
        # We do NOT delete global predefined services — just disable them for this salon.
        await db.salon_services.update_many(
            {"salon_id": salon_id},
            {"$set": {"is_enabled": False, "updated_at": now_iso}}
        )
        # Delete salon-only services that were imported previously
        await db.services.delete_many({"salon_id": salon_id})
        await db.salon_packages.delete_many({"salon_id": salon_id})

    # Insert new services as salon-scoped services (services collection stores both)
    created_services = 0
    skipped_services = 0
    name_to_service_id: Dict[str, str] = {}

    # Load existing services for dedup by name
    existing = await db.services.find(
        {"$or": [{"salon_id": salon_id}, {"salon_id": {"$exists": False}}]},
        {"_id": 0, "id": 1, "service_name": 1}
    ).to_list(5000)
    name_to_existing = {(s.get("service_name") or "").strip().lower(): s.get("id") for s in existing}

    docs_to_insert: List[dict] = []
    for s in services_in:
        sname = str(s.get("service_name") or "").strip()
        if not sname:
            continue
        key = sname.lower()
        if key in name_to_existing and mode == "add":
            skipped_services += 1
            name_to_service_id[key] = name_to_existing[key]
            continue
        sid = str(uuid.uuid4())
        doc = {
            "id": sid,
            "salon_id": salon_id,
            "service_name": sname[:120],
            "description": str(s.get("description") or "").strip()[:500],
            "category": str(s.get("category") or "General").strip()[:60] or "General",
            "gender": (s.get("gender") if s.get("gender") in ("Men", "Women", "Unisex") else "Unisex"),
            "default_duration": int(s.get("default_duration") or 30) or 30,
            "base_price": float(s.get("base_price") or 0),
            "is_active": True,
            "thumbnail_url": s.get("thumbnail_url") or None,
            "source": "menu_parse",
            "created_at": now_iso,
        }
        docs_to_insert.append(doc)
        name_to_service_id[key] = sid

    if docs_to_insert:
        await db.services.insert_many(docs_to_insert)
        created_services = len(docs_to_insert)
        # Enable each new service for this salon
        salon_service_docs = []
        for d in docs_to_insert:
            salon_service_docs.append({
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "service_id": d["id"],
                "is_enabled": True,
                "created_at": now_iso,
                "updated_at": now_iso,
            })
        if salon_service_docs:
            await db.salon_services.insert_many(salon_service_docs)

    # Packages
    created_packages = 0
    if packages_in:
        for p in packages_in:
            pname = str(p.get("package_name") or "").strip()
            if not pname:
                continue
            svc_ids: List[str] = []
            for sn in (p.get("service_names") or []):
                key = str(sn).strip().lower()
                if key in name_to_service_id:
                    svc_ids.append(name_to_service_id[key])
                elif key in name_to_existing:
                    svc_ids.append(name_to_existing[key])
            pkg_doc = {
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "package_name": pname[:120],
                "description": str(p.get("description") or "").strip()[:500],
                "service_ids": svc_ids,
                "gender": (p.get("gender") if p.get("gender") in ("Men", "Women", "Unisex") else "Unisex"),
                "package_price": float(p.get("package_price") or 0),
                "is_active": True,
                "source": "menu_parse",
                "created_at": now_iso,
            }
            await db.salon_packages.insert_one(pkg_doc)
            created_packages += 1

    return {
        "mode": mode,
        "created_services": created_services,
        "skipped_services": skipped_services,
        "created_packages": created_packages,
        "message": (
            f"Imported {created_services} new services"
            + (f" (skipped {skipped_services} duplicates)" if mode == "add" and skipped_services else "")
            + (f" and {created_packages} packages." if created_packages else ".")
        ),
    }


# CSV column headers for the service uploader (also used for the template).
SERVICE_CSV_HEADERS = [
    "service_name",
    "description",
    "category",
    "gender_tag",
    "default_duration",
    "base_price",
    "price_type",
    "is_favorite",
    "available_at_home",
    "thumbnail_url",
    "images",
]


@api_router.get("/salons/{salon_id}/services/csv-template")
async def download_services_csv_template(salon_id: str):
    """Return a ready-to-fill CSV template (headers + two example rows).

    Public so the salon can grab the template before authenticating in the UI.
    """
    import csv as _csv
    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(SERVICE_CSV_HEADERS)
    writer.writerow([
        "Haircut - Men", "Classic scissor cut & style", "Hair",
        "Men", "30", "250", "fixed", "false", "false", "", "",
    ])
    writer.writerow([
        "Hair Spa", "Relaxing nourishing hair spa", "Spa",
        "Unisex", "45", "600", "onwards", "true", "true", "", "",
    ])
    csv_text = buf.getvalue()
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="services_template.csv"'},
    )


@api_router.post("/salons/{salon_id}/services/upload-csv")
async def upload_services_csv(
    salon_id: str,
    file: UploadFile = File(...),
    current_user=Depends(get_current_salon_user),
):
    """Bulk-ADD services to a salon from a CSV (or Excel) file.

    This ALWAYS adds — it never replaces, overwrites or disables existing
    services. A row whose service_name already exists for this salon (or is
    repeated within the file) is skipped and reported.

    Columns (case-insensitive headers; only `service_name` is required):
      service_name*  — name of the service (required)
      description    — free text
      category       — grouping, defaults to "General"
      gender_tag     — Men / Women / Unisex (defaults Unisex)
      default_duration — minutes (integer, defaults 30)
      base_price     — number (defaults 0)
      price_type     — fixed / onwards (defaults fixed)
      is_favorite    — true/false (defaults false)
      available_at_home — true/false (defaults false)
      thumbnail_url  — image URL for the circular category thumbnail
      images         — one or more image URLs separated by | or ,
    """
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")

    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    fname = file.filename.lower()
    if not (fname.endswith(".csv") or fname.endswith(".xlsx") or fname.endswith(".xls")):
        raise HTTPException(status_code=400, detail="File must be .csv, .xlsx or .xls")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB)")

    # ---- Parse rows into list[dict] with lowercased header keys ----
    rows: List[dict] = []
    try:
        if fname.endswith(".csv"):
            import csv as _csv
            text = content.decode("utf-8-sig", errors="ignore")
            reader = _csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append({(k or "").strip().lower(): v for k, v in r.items() if k})
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            header_row = None
            for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if any((c is not None and str(c).strip() != "") for c in r):
                    header_row = [str(c).strip().lower() if c is not None else "" for c in r]
                    break
            if not header_row:
                raise HTTPException(status_code=400, detail="File is empty")
            for r in ws.iter_rows(min_row=2, values_only=True):
                if all(c is None or str(c).strip() == "" for c in r):
                    continue
                row = {}
                for idx, val in enumerate(r):
                    if idx < len(header_row) and header_row[idx]:
                        row[header_row[idx]] = val
                rows.append(row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in the file")

    # ---- Flexible header resolution + value coercion helpers ----
    def _norm(s: str) -> str:
        return str(s).replace(" ", "").replace("_", "").replace(".", "").lower()

    def _find(d: dict, *keys: str):
        for k in keys:
            if k in d:
                return d[k]
        normed = {_norm(k): v for k, v in d.items()}
        for k in keys:
            if _norm(k) in normed:
                return normed[_norm(k)]
        return None

    def _truthy(v) -> bool:
        return str(v).strip().lower() in ("1", "true", "yes", "y", "t", "on")

    def _to_int(v, default: int) -> int:
        try:
            s = str(v).strip()
            if s == "" or s is None:
                return default
            return int(float(s))
        except Exception:
            return default

    def _to_float(v, default: float) -> float:
        try:
            s = str(v).strip().replace(",", "")
            if s == "":
                return default
            return float(s)
        except Exception:
            return default

    def _norm_gender(v) -> str:
        s = str(v or "").strip().lower()
        if s in ("men", "man", "male", "m"):
            return "Men"
        if s in ("women", "woman", "female", "f", "w"):
            return "Women"
        return "Unisex"

    def _norm_price_type(v) -> str:
        s = str(v or "").strip().lower()
        return "onwards" if s in ("onwards", "onward", "starting", "from") else "fixed"

    def _split_images(v) -> List[str]:
        if not v:
            return []
        raw = str(v)
        parts = raw.split("|") if "|" in raw else raw.split(",")
        return [p.strip() for p in parts if p.strip()]

    # ---- Dedup against services already present for THIS salon ----
    existing = await db.services.find(
        {"salon_id": salon_id, "is_active": True},
        {"_id": 0, "service_name": 1},
    ).to_list(10000)
    existing_names = {(s.get("service_name") or "").strip().lower() for s in existing}

    now_iso = datetime.now(timezone.utc).isoformat()
    docs_to_insert: List[dict] = []
    seen_in_file: set = set()
    skipped_duplicates = 0
    errors: List[dict] = []

    # Row numbers reported to the user are 1-based and account for the header
    # line (so the first data row is row 2 — matches what they see in Excel).
    for idx, raw in enumerate(rows):
        row_no = idx + 2
        name = str(_find(raw, "service_name", "name", "service") or "").strip()
        if not name:
            errors.append({"row": row_no, "reason": "Missing service_name"})
            continue
        key = name.lower()
        if key in existing_names:
            skipped_duplicates += 1
            continue
        if key in seen_in_file:
            skipped_duplicates += 1
            continue

        seen_in_file.add(key)
        doc = {
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            "service_name": name[:120],
            "description": str(_find(raw, "description", "desc") or "").strip()[:500] or None,
            "category": (str(_find(raw, "category", "cat") or "").strip()[:60] or "General"),
            "sub_category": (str(_find(raw, "sub_category", "subcategory", "sub-cat", "subcat") or "").strip()[:60] or None),
            "gender_tag": _norm_gender(_find(raw, "gender_tag", "gender")),
            "default_duration": max(1, _to_int(_find(raw, "default_duration", "duration", "minutes"), 30)),
            "base_price": max(0.0, _to_float(_find(raw, "base_price", "price", "amount"), 0.0)),
            "price_type": _norm_price_type(_find(raw, "price_type", "pricetype")),
            "is_favorite": _truthy(_find(raw, "is_favorite", "favorite", "favourite")),
            "available_at_home": _truthy(_find(raw, "available_at_home", "athome", "home")),
            "thumbnail_url": (str(_find(raw, "thumbnail_url", "thumbnail", "thumb") or "").strip() or None),
            "images": _split_images(_find(raw, "images", "image_urls", "image")),
            "is_active": True,
            "is_enabled": True,
            "source": "csv_upload",
            "created_at": now_iso,
        }
        docs_to_insert.append(doc)

    created = 0
    batch_id: Optional[str] = None
    if docs_to_insert:
        await db.services.insert_many(docs_to_insert)
        created = len(docs_to_insert)
        # Enable each newly-created service for this salon (idempotent mapping).
        salon_service_docs = [
            {
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "service_id": d["id"],
                "is_enabled": True,
                "created_at": now_iso,
                "updated_at": now_iso,
            }
            for d in docs_to_insert
        ]
        await db.salon_services.insert_many(salon_service_docs)

        # Record an upload batch so the salon can see history & roll it back.
        batch_id = str(uuid.uuid4())
        uploader = (
            current_user.get("email")
            or current_user.get("phone")
            or current_user.get("sub")
            or "salon_admin"
        )
        await db.service_upload_batches.insert_one({
            "id": batch_id,
            "salon_id": salon_id,
            "filename": file.filename,
            "uploaded_by": uploader,
            "uploaded_at": now_iso,
            "created_count": created,
            "skipped_count": skipped_duplicates,
            "error_count": len(errors),
            "service_ids": [d["id"] for d in docs_to_insert],
            "status": "active",  # → "rolled_back" after undo
        })

    return {
        "success": True,
        "batch_id": batch_id,
        "total_rows": len(rows),
        "created": created,
        "skipped_duplicates": skipped_duplicates,
        "errors": errors,
        "message": (
            f"Added {created} new service(s)"
            + (f", skipped {skipped_duplicates} duplicate(s)" if skipped_duplicates else "")
            + (f", {len(errors)} row(s) had errors" if errors else "")
            + "."
        ),
    }


# ---- Upload template + history + rollback ------------------------------------
SERVICES_CSV_TEMPLATE = (
    "service_name,description,category,sub_category,gender_tag,default_duration,base_price,price_type,is_favorite,available_at_home,thumbnail_url,images\n"
    "Men's Haircut,Classic scissor cut with styling,Services,Hair,Men,30,300,fixed,true,false,,\n"
    "Beard Trim,Shape-up and hot towel,Services,Beard,Men,20,150,fixed,false,false,,\n"
    "Women's Haircut,Wash + cut + blow-dry,Services,Hair,Women,45,600,fixed,true,false,,\n"
    "Classic Manicure,Nail shaping + cuticle care,Services,Nails,Unisex,30,400,fixed,false,true,,\n"
    "Bridal Glow Package,Facial + hair spa + mani-pedi,Packages,,Women,180,4999,onwards,true,false,,\n"
)


@api_router.get("/services/upload-template.csv")
async def download_services_csv_template():
    """Return a small illustrative CSV so owners know the exact column headers.
    Available to any authenticated salon; no salon_id required."""
    from fastapi.responses import Response
    return Response(
        content=SERVICES_CSV_TEMPLATE,
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="services-upload-template.csv"'
        },
    )


@api_router.get("/salons/{salon_id}/services/upload-batches")
async def list_service_upload_batches(
    salon_id: str,
    current_user=Depends(get_current_salon_user),
):
    """List past service-upload batches for this salon, newest first."""
    batches = await db.service_upload_batches.find(
        {"salon_id": salon_id},
        {"_id": 0},
    ).sort("uploaded_at", -1).to_list(200)
    return {"batches": batches}


@api_router.delete("/salons/{salon_id}/services/upload-batches/{batch_id}")
async def rollback_service_upload_batch(
    salon_id: str,
    batch_id: str,
    current_user=Depends(get_current_salon_user),
):
    """Undo a service-upload batch — hard-delete the created services and their
    salon_services mappings. Idempotent: rolling back an already-rolled-back
    batch is a no-op."""
    batch = await db.service_upload_batches.find_one(
        {"id": batch_id, "salon_id": salon_id}, {"_id": 0}
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Upload batch not found")

    if batch.get("status") == "rolled_back":
        return {"success": True, "removed": 0, "already_rolled_back": True}

    svc_ids: List[str] = batch.get("service_ids") or []
    removed = 0
    if svc_ids:
        # Only delete services still owned by this salon (safety).
        result = await db.services.delete_many({
            "id": {"$in": svc_ids},
            "salon_id": salon_id,
        })
        removed = int(getattr(result, "deleted_count", 0) or 0)
        # Clean mappings + any barber links.
        await db.salon_services.delete_many({
            "salon_id": salon_id,
            "service_id": {"$in": svc_ids},
        })
        try:
            await db.barber_services.delete_many({"service_id": {"$in": svc_ids}})
        except Exception:
            pass

    await db.service_upload_batches.update_one(
        {"id": batch_id, "salon_id": salon_id},
        {"$set": {
            "status": "rolled_back",
            "rolled_back_at": datetime.now(timezone.utc).isoformat(),
            "removed_count": removed,
        }}
    )
    return {"success": True, "removed": removed, "batch_id": batch_id}



@api_router.post("/salons/{salon_id}/salon-booking")
async def create_salon_booking(salon_id: str, body: dict, current_user=Depends(get_current_salon_user)):
    """Create a booking from the salon side (walk-in, phone call, etc.)"""
    # === Manual close (full only) enforcement for salon-side manual bookings ===
    salon_for_close = await db.salons.find_one({"id": salon_id}, {"_id": 0, "manual_toggle": 1})
    mt = (salon_for_close or {}).get("manual_toggle") or {}
    if mt.get("is_overridden") and not mt.get("is_open", True):
        if (mt.get("closed_mode") or "full") == "full":
            raise HTTPException(
                status_code=400,
                detail="Salon is fully closed (online + offline). Please open the salon to take bookings."
            )
        # online_only: salon-side manual booking is allowed
    
    customer_name = body.get("customer_name", "Walk-in").strip()
    phone = body.get("phone", "").strip()
    gender = body.get("gender", "Men")
    barber_id = body.get("barber_id", "any")
    selected_services = body.get("selected_services", [])
    selected_products = body.get("selected_products") or []
    shift = body.get("shift", "")
    date = body.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    payment_mode = body.get("payment_mode")
    
    if not customer_name:
        customer_name = "Walk-in"
    
    # Normalize phone
    if phone:
        phone = phone.replace(" ", "").replace("-", "")
        if not phone.startswith("+91"):
            phone = f"+91{phone}"
    
    # Auto-detect current shift if not provided
    if not shift:
        from datetime import datetime as dt
        now = dt.now()
        hour = now.hour
        if hour < 12:
            shift = "Morning"
        elif hour < 16:
            shift = "Noon"
        else:
            shift = "Evening"
    
    # Calculate total
    total_amount = 0
    for service_id in selected_services:
        if barber_id and barber_id != "any":
            # Check barber-specific pricing
            barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
            if barber:
                barber_service = next((s for s in (barber.get("services") or []) if s.get("service_id") == service_id), None)
                if barber_service:
                    total_amount += barber_service.get("price", 0)
                    continue
        # Fallback to base price
        service = await db.services.find_one({"id": service_id}, {"_id": 0})
        if service:
            total_amount += service.get("base_price", 0)

    # Product lines (also decrement inventory)
    product_details = []
    for p in selected_products:
        pid = p.get("product_id") or p.get("id")
        qty = int(p.get("qty") or 0)
        unit_price = float(p.get("unit_price") or 0)
        if not pid or qty <= 0:
            continue
        if unit_price <= 0:
            item = await db.salon_inventory.find_one({"id": pid, "salon_id": salon_id}, {"_id": 0})
            if item:
                unit_price = float(item.get("retail_price") or item.get("selling_price") or 0)
        line_total = qty * unit_price
        total_amount += line_total
        product_details.append({
            "product_id": pid,
            "name": p.get("name"),
            "qty": qty,
            "unit_price": unit_price,
            "line_total": line_total,
        })
        try:
            await db.salon_inventory.update_one(
                {"id": pid, "salon_id": salon_id},
                {"$inc": {"stock_quantity": -qty}}
            )
        except Exception:
            pass

    # Auto-assign barber when "any" is selected, using the same fastest-barber
    # logic as the customer booking flow (priority: shortest active queue today
    # → fewest yesterday → random eligible).
    if not barber_id or barber_id == "any":
        service_total_minutes = await calc_service_total_minutes(selected_services)
        required_blocked = calc_blocked_minutes_from_total(service_total_minutes)
        chosen = await pick_fastest_barber(
            salon_id=salon_id,
            date=date,
            shift=shift,
            required_blocked_minutes=required_blocked,
            customer_gender=gender,
        )
        if not chosen:
            raise HTTPException(
                status_code=400,
                detail=f"All barbers are fully booked for {shift} shift. Please choose another shift or date."
            )
        barber_id = chosen["id"]
    
    # Get token number (signature: salon_id, date, shift)
    token_number = await get_next_token_number(salon_id, date, shift)
    
    # Get barber name (always set since barber_id is now resolved)
    barber_name = "Unknown"
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if barber:
        barber_name = barber.get("name", "Unknown")

    # Resolve branch: explicit body override > barber's branch > salon main
    requested_branch_id = body.get("branch_id")
    booking_branch_id = await resolve_branch_id(salon_id, requested_branch_id)
    if not requested_branch_id and barber and barber.get("branch_id"):
        booking_branch_id = barber.get("branch_id")

    # Handle wallet payment — must debit from customer's membership wallet first
    payment_status = "pending"
    payment_confirmed = False
    if payment_mode == "wallet":
        if not phone:
            raise HTTPException(status_code=400, detail="Customer mobile number is required for wallet payment.")
        membership = await db.customer_memberships.find_one({
            "salon_id": salon_id,
            "customer_phone": phone,
            "is_active": True,
        }, {"_id": 0})
        if not membership:
            raise HTTPException(status_code=400, detail="No active wallet/membership found for this customer.")
        current_balance = float(membership.get("wallet_balance") or 0)
        if current_balance < total_amount:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient wallet balance. Available: ₹{current_balance:.2f}, Required: ₹{total_amount:.2f}",
            )
        new_balance = current_balance - total_amount
        await db.customer_memberships.update_one(
            {"id": membership["id"]},
            {"$set": {"wallet_balance": new_balance}}
        )
        await db.wallet_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "customer_phone": phone,
            "salon_id": salon_id,
            "transaction_type": "debit",
            "amount": total_amount,
            "balance_after": new_balance,
            "description": f"Salon booking payment - {shift} shift",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        payment_status = "paid"
        payment_confirmed = True
    
    token_dict = {
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "branch_id": booking_branch_id,
        "token_number": token_number,
        "customer_name": customer_name,
        "phone": phone or "",
        "user_id": "",
        "barber_id": barber_id,
        "barber_name": barber_name,
        "selected_services": selected_services,
        "selected_products": product_details,
        "date": date,
        "shift": shift,
        "time_slot": shift,
        "total_amount": total_amount,
        "status": "waiting",
        "payment_status": payment_status,
        "payment_mode": payment_mode,
        "payment_confirmed": payment_confirmed,
        "source": "salon",
        "booking_type": "instant",
        "booking_for_self": True,
        "customer_gender": gender,
        "recall_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.tokens.insert_one(token_dict)
    token_dict.pop("_id", None)
    
    # Also save customer if phone provided
    if phone:
        existing_customer = await db.salon_customers.find_one({
            "salon_id": salon_id, "phone": phone
        })
        if not existing_customer:
            await db.salon_customers.insert_one({
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "branch_id": booking_branch_id,
                "name": customer_name,
                "phone": phone,
                "gender": gender,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "source": "salon_booking"
            })
    
    await broadcast_update("new_token", token_dict)
    # Also emit the standardized "token_created" so all dashboards / queues refresh.
    await broadcast_update("token_created", token_dict)
    
    return token_dict

@api_router.get("/salons/{salon_id}/customers/{phone}/bookings")
async def get_customer_bookings(salon_id: str, phone: str, current_user=Depends(get_current_salon_user)):
    """Get all bookings for a specific customer"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    bookings = await db.tokens.find(
        {"salon_id": salon_id, "phone": phone},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"bookings": bookings}

# Custom Package Models
class CustomerPackageService(BaseModel):
    service_id: str
    service_name: str
    original_price: float
    discounted_price: float

class CustomerPackageCreate(BaseModel):
    salon_id: str
    customer_phone: str
    customer_name: str
    package_name: str
    services: List[CustomerPackageService]
    total_original: float
    discount_percentage: float
    total_discounted: float
    notes: Optional[str] = None

class CustomerPackage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    customer_phone: str
    customer_name: str
    package_name: str
    services: List[CustomerPackageService]
    total_original: float
    discount_percentage: float
    total_discounted: float
    notes: Optional[str] = None
    is_active: bool = True
    created_at: str

# ============ MEMBERSHIP & WALLET MODELS ============

class MembershipPlanCreate(BaseModel):
    salon_id: str
    name: str  # e.g., "Gold", "Diamond", "Platinum"
    amount: float  # Amount customer pays
    credit: float  # Credit added to wallet
    validity_months: int  # Validity in months
    terms_conditions: str  # T&C text
    # Color-tier badge (shown to customer as a colored badge)
    tier: Optional[str] = "Custom"  # one of: Diamond, Gold, Silver, Custom
    color: Optional[str] = None  # hex color override, optional

class MembershipPlan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    name: str
    amount: float
    credit: float
    validity_months: int
    terms_conditions: str
    tier: Optional[str] = "Custom"
    color: Optional[str] = None
    is_active: bool = True
    created_at: str

class CustomerMembershipCreate(BaseModel):
    customer_phone: str
    customer_name: str
    membership_plan_id: str
    payment_mode: str  # "cash", "card", "upi", "wallet"
    paid_amount: float

class CustomerMembership(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    customer_phone: str
    customer_name: str
    membership_plan_id: str
    membership_name: str
    tier: Optional[str] = "Custom"
    color: Optional[str] = None
    payment_mode: str
    paid_amount: float
    credit_added: float
    wallet_balance: float
    expiry_date: str
    is_active: bool = True
    cancelled: Optional[bool] = False
    cancelled_at: Optional[str] = None
    cancel_reason: Optional[str] = None
    notified_1m_expiry: Optional[bool] = False
    notified_1w_expiry: Optional[bool] = False
    payment_confirmed: bool = True  # False for customer purchases until salon confirms
    purchased_at: str

class WalletTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    customer_phone: str
    salon_id: str
    transaction_type: str  # "credit", "debit", "expiry"
    amount: float
    balance_after: float
    description: str
    created_at: str

class LoyaltyTier(BaseModel):
    name: str  # e.g., "Bronze", "Silver", "Gold"
    spend_amount: float  # Threshold
    period_months: int  # Individual period for this tier
    topup_percentage: float  # Reward percentage

class LoyaltyProgramSettings(BaseModel):
    salon_id: str
    enabled: bool = False
    tiers: List[Dict[str, Any]] = []  # Multiple tiers with individual periods

class LoyaltyProgram(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    enabled: bool
    tiers: List[Dict[str, Any]]
    updated_at: str

@api_router.post("/salons/{salon_id}/customer-packages", response_model=CustomerPackage)
async def create_customer_package(salon_id: str, package: CustomerPackageCreate, current_user=Depends(get_current_salon_user)):
    """Create a custom package for a specific customer"""
    package_dict = package.model_dump()
    package_dict["id"] = str(uuid.uuid4())
    package_dict["is_active"] = True
    package_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.customer_packages.insert_one(package_dict)

    # Notify customer (in-app) about new custom package
    customer_phone = package_dict.get("customer_phone", "")
    if customer_phone:
        await create_in_app_notification(
            user_type="customer",
            user_id=customer_phone,
            title="Special Package Created for You",
            message=f"A custom package '{package_dict.get('name','Package')}' has been created for you. Check it out in your packages.",
            notification_type="custom_package",
            setting_key="custom_package",
            salon_id=salon_id,
            related_id=package_dict["id"],
        )

    return CustomerPackage(**package_dict)

@api_router.get("/salons/{salon_id}/customer-packages/{phone}")
async def get_customer_packages(salon_id: str, phone: str, current_user=Depends(get_current_salon_user)):
    """Get all custom packages for a customer"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    packages = await db.customer_packages.find(
        {"salon_id": salon_id, "customer_phone": phone, "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    return {"packages": packages}


# ============ MEMBERSHIP & WALLET ROUTES ============

@api_router.post("/salons/{salon_id}/membership-plans", response_model=MembershipPlan)
async def create_membership_plan(salon_id: str, plan: MembershipPlanCreate, current_user=Depends(get_current_salon_admin)):
    """Create a new membership plan"""
    plan_dict = plan.model_dump()
    plan_dict["id"] = str(uuid.uuid4())
    plan_dict["is_active"] = True
    plan_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.membership_plans.insert_one(plan_dict)
    return MembershipPlan(**plan_dict)


@api_router.put("/salons/{salon_id}/customer-packages/{package_id}")
async def update_customer_package(
    salon_id: str,
    package_id: str,
    package_data: dict,
    current_user=Depends(get_current_salon_user)
):
    """Update a custom customer package"""
    await db.customer_packages.update_one(
        {"id": package_id, "salon_id": salon_id},
        {"$set": package_data}
    )
    return {"message": "Package updated successfully"}

@api_router.delete("/salons/{salon_id}/customer-packages/{package_id}")
async def delete_customer_package(
    salon_id: str,
    package_id: str,
    current_user=Depends(get_current_salon_user)
):
    """Delete a custom customer package"""
    result = await db.customer_packages.delete_one({"id": package_id, "salon_id": salon_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Package not found")
    return {"message": "Package deleted successfully"}


@api_router.get("/salons/{salon_id}/membership-plans")
async def get_membership_plans(salon_id: str):
    """Get all active membership plans for a salon"""
    plans = await db.membership_plans.find(
        {"salon_id": salon_id, "is_active": True},
        {"_id": 0}
    ).to_list(100)
    return {"plans": plans}


@api_router.put("/salons/{salon_id}/membership-plans/{plan_id}", response_model=MembershipPlan)
async def update_membership_plan(
    salon_id: str,
    plan_id: str,
    plan: MembershipPlanCreate,
    current_user=Depends(get_current_salon_user)
):
    """Update membership plan (price locked, other fields editable)"""
    existing = await db.membership_plans.find_one({"id": plan_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Membership plan not found")

    # Update all fields but use existing price (locked)
    await db.membership_plans.update_one(
        {"id": plan_id},
        {"$set": {
            "name": plan.name,
            # "amount": existing["amount"],  # Price is LOCKED
            "credit": plan.credit,
            "validity_months": plan.validity_months,
            "terms_conditions": plan.terms_conditions,
            "tier": plan.tier or existing.get("tier", "Custom"),
            "color": plan.color or existing.get("color"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    updated = await db.membership_plans.find_one({"id": plan_id}, {"_id": 0})
    return MembershipPlan(**updated)

@api_router.get("/salons/{salon_id}/sold-memberships")
async def get_sold_memberships(salon_id: str, current_user=Depends(get_current_salon_user)):
    """Get all sold memberships for a salon"""
    memberships = await db.customer_memberships.find({
        "salon_id": salon_id
    }, {"_id": 0}).sort("purchased_at", -1).to_list(1000)
    
    return {"memberships": memberships}

@api_router.put("/salons/{salon_id}/customer-memberships/{membership_id}")
async def update_customer_membership(
    salon_id: str,
    membership_id: str,
    updates: dict,
    current_user=Depends(get_current_salon_user)
):
    """Update a sold membership (wallet balance, expiry, status)"""
    await db.customer_memberships.update_one(
        {"id": membership_id, "salon_id": salon_id},
        {"$set": updates}
    )
    
    return {"message": "Membership updated successfully"}

@api_router.post("/salons/{salon_id}/customer-memberships/{membership_id}/cancel")
async def cancel_customer_membership(
    salon_id: str,
    membership_id: str,
    body: Optional[dict] = None,
    current_user=Depends(get_current_salon_user),
):
    """Soft-cancel a sold membership.

    Sets is_active=False, cancelled=True, cancelled_at=now, and stores an optional
    cancel_reason. The record is preserved for audit. Customer is notified.
    """
    body = body or {}
    reason = (body.get("reason") or "").strip() or "Cancelled by salon"

    m = await db.customer_memberships.find_one({"id": membership_id, "salon_id": salon_id}, {"_id": 0})
    if not m:
        raise HTTPException(status_code=404, detail="Membership not found")
    if m.get("cancelled") or not m.get("is_active", True):
        return {"message": "Membership already cancelled"}

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.customer_memberships.update_one(
        {"id": membership_id},
        {"$set": {
            "is_active": False,
            "cancelled": True,
            "cancelled_at": now_iso,
            "cancel_reason": reason,
        }}
    )

    # Notify customer
    try:
        if m.get("customer_phone"):
            await create_in_app_notification(
                user_type="customer",
                user_id=m["customer_phone"],
                title="Membership Cancelled",
                message=f"Your '{m.get('membership_name','')}' membership has been cancelled by the salon. Reason: {reason}",
                notification_type="membership_cancelled",
                setting_key="membership_added",
                salon_id=salon_id,
                related_id=membership_id,
            )
    except Exception as e:
        logger.error(f"Failed membership cancellation notification: {e}")

    return {"message": "Membership cancelled", "membership_id": membership_id}

@api_router.get("/salons/{salon_id}/customers/{phone}/membership")
async def get_customer_membership_info(salon_id: str, phone: str):
    """Get customer's active membership info (no auth required for customer view)"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    membership = await db.customer_memberships.find_one({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True
    }, {"_id": 0})
    
    if not membership:
        raise HTTPException(status_code=404, detail="No active membership found")
    
    return membership

@api_router.get("/salons/{salon_id}/customers/{phone}/bookings-public")
async def get_customer_bookings_public(salon_id: str, phone: str):
    """Get customer's booking history (no auth required)"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    bookings = await db.tokens.find({
        "salon_id": salon_id,
        "phone": phone
    }, {"_id": 0}).sort("created_at", -1).limit(20).to_list(20)
    
    return {"bookings": bookings}

@api_router.delete("/salons/{salon_id}/membership-plans/{plan_id}")
async def delete_membership_plan(
    salon_id: str,
    plan_id: str,
    current_user=Depends(get_current_salon_user)
):
    """Delete/deactivate membership plan (doesn't affect existing customers)"""
    result = await db.membership_plans.delete_one({"id": plan_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Membership plan not found")
    
    return {"message": "Membership plan deleted successfully"}

@api_router.post("/salons/{salon_id}/sell-membership")
async def sell_membership(salon_id: str, membership: CustomerMembershipCreate, current_user=Depends(get_current_salon_user)):
    """Sell membership to a customer"""
    # Get membership plan
    plan = await db.membership_plans.find_one({"id": membership.membership_plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Membership plan not found")
    
    # Format phone
    phone = membership.customer_phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Calculate expiry date
    expiry_date = datetime.now(timezone.utc) + timedelta(days=plan["validity_months"] * 30)
    
    # Check if customer already has active membership
    existing = await db.customer_memberships.find_one({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True
    }, {"_id": 0})
    
    if existing:
        # Add credit to existing wallet
        new_balance = existing["wallet_balance"] + plan["credit"]
        await db.customer_memberships.update_one(
            {"id": existing["id"]},
            {"$set": {
                "wallet_balance": new_balance,
                "expiry_date": expiry_date.isoformat(),
                "tier": plan.get("tier", existing.get("tier", "Custom")),
                "color": plan.get("color") or existing.get("color"),
                # Reset expiry-warning flags on renewal
                "notified_1m_expiry": False,
                "notified_1w_expiry": False,
            }}
        )
        
        # Record transaction
        await db.wallet_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "customer_phone": phone,
            "salon_id": salon_id,
            "transaction_type": "credit",
            "amount": plan["credit"],
            "balance_after": new_balance,
            "description": f"Membership renewal: {plan['name']}",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {"message": "Membership renewed", "new_balance": new_balance}
    else:
        # Create new membership
        membership_data = {
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            "customer_phone": phone,
            "customer_name": membership.customer_name,
            "membership_plan_id": plan["id"],
            "membership_name": plan["name"],
            "tier": plan.get("tier", "Custom"),
            "color": plan.get("color"),
            "payment_mode": membership.payment_mode,
            "paid_amount": membership.paid_amount,
            "credit_added": plan["credit"],
            "wallet_balance": plan["credit"],
            "expiry_date": expiry_date.isoformat(),
            "is_active": True,
            "cancelled": False,
            "notified_1m_expiry": False,
            "notified_1w_expiry": False,
            "payment_confirmed": True,  # Salon-side sells are auto-confirmed
            "purchased_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.customer_memberships.insert_one(membership_data)
        # FastAPI can't serialize ObjectId; strip it after insert.
        membership_data.pop("_id", None)
        
        # Record transaction
        await db.wallet_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "customer_phone": phone,
            "salon_id": salon_id,
            "transaction_type": "credit",
            "amount": plan["credit"],
            "balance_after": plan["credit"],
            "description": f"Membership purchased: {plan['name']}",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        # Record financial transaction for salon-side sell (auto-confirmed)
        if membership.payment_mode != "wallet":
            await db.financial_transactions.insert_one({
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "type": "inflow",
                "category": "membership_payment",
                "amount": float(membership.paid_amount),
                "payment_mode": membership.payment_mode,
                "narration": f"Membership sold: {plan['name']} - {membership.customer_name}",
                "reference_id": membership_data["id"],
                "reference_type": "membership",
                "created_by": "system",
                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "created_at": datetime.now(timezone.utc).isoformat()
            })

        # Notify customer (in-app + WhatsApp)
        await create_in_app_notification(
            user_type="customer",
            user_id=phone,
            title="Membership Added",
            message=f"You have been enrolled in {plan['name']} membership. Wallet credit: ₹{plan['credit']}.",
            notification_type="membership_added",
            setting_key="membership_added",
            salon_id=salon_id,
            related_id=membership_data["id"],
        )
        if await should_send_customer_whatsapp(phone, "whatsapp_membership_added"):
            wa_msg = (
                f"🎉 *Membership Added!*\n\nHello {membership.customer_name},\n\n"
                f"You have been enrolled in *{plan['name']}* membership.\n"
                f"💳 Wallet Credit: ₹{plan['credit']}\n"
                f"📅 Valid for: {plan['validity_months']} months\n\n_SalonHub_"
            )
            await send_whatsapp_notification(phone, wa_msg, "membership_added")

        return {"message": "Membership created", "membership": membership_data}

@api_router.post("/salons/{salon_id}/customers/{phone}/buy-membership")
async def customer_buy_membership(
    salon_id: str, 
    phone: str,
    membership: CustomerMembershipCreate
):
    """Customer-facing endpoint to buy membership (no auth required).
    Credit is NOT added until salon confirms payment."""
    # Get membership plan
    plan = await db.membership_plans.find_one({"id": membership.membership_plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Membership plan not found")
    
    # Format phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Calculate expiry date
    expiry_date = datetime.now(timezone.utc) + timedelta(days=plan["validity_months"] * 30)
    
    # Create new membership with payment_confirmed=False
    # Credit will be added after salon confirms payment
    membership_data = {
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "customer_phone": phone,
        "customer_name": membership.customer_name,
        "membership_plan_id": plan["id"],
        "membership_name": plan["name"],
        "tier": plan.get("tier", "Custom"),
        "color": plan.get("color"),
        "payment_mode": membership.payment_mode,
        "paid_amount": membership.paid_amount,
        "credit_added": plan["credit"],
        "wallet_balance": 0,  # No credit until confirmed
        "expiry_date": expiry_date.isoformat(),
        "is_active": True,
        "cancelled": False,
        "notified_1m_expiry": False,
        "notified_1w_expiry": False,
        "payment_confirmed": False,  # Requires salon confirmation
        "purchased_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.customer_memberships.insert_one(membership_data)
    membership_data.pop("_id", None)

    # Create notification for salon about pending membership confirmation
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "user_type": "salon",
        "user_id": salon_id,
        "salon_id": salon_id,
        "title": "New Membership Purchase - Confirmation Required",
        "message": f"{membership.customer_name} ({phone}) purchased {plan['name']} membership for ₹{membership.paid_amount}. Please confirm payment.",
        "type": "membership_pending",
        "related_id": membership_data["id"],
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Create notification for customer about pending confirmation
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "user_type": "customer",
        "user_id": phone,
        "salon_id": salon_id,
        "title": "Membership Purchase - Pending Confirmation",
        "message": f"Your {plan['name']} membership purchase is pending confirmation by the salon. Wallet credit of ₹{plan['credit']} will be added after confirmation.",
        "type": "membership_pending",
        "related_id": membership_data["id"],
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {
        "message": "Membership purchase submitted. Pending salon confirmation.",
        "membership": membership_data,
        "pending_confirmation": True
    }


@api_router.get("/salons/{salon_id}/customers/{phone}/packages")
async def get_customer_available_packages(salon_id: str, phone: str):
    """Get all packages available to a customer (public + customer-specific)"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Get public salon packages
    public_packages = await db.packages.find({
        "salon_id": salon_id,
        "is_active": True
    }, {"_id": 0}).to_list(100)
    
    # Get customer-specific packages
    customer_packages = await db.customer_packages.find({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True
    }, {"_id": 0}).to_list(100)
    
    return {
        "public_packages": public_packages,
        "customer_packages": customer_packages
    }


@api_router.get("/salons/{salon_id}/customer-membership/{phone}")
async def get_customer_membership(salon_id: str, phone: str):
    """Get customer's membership and wallet details"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Get confirmed active membership
    membership = await db.customer_memberships.find_one({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True,
        "payment_confirmed": True
    }, {"_id": 0})
    
    # Also check for pending (unconfirmed) memberships
    pending_memberships = await db.customer_memberships.find({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True,
        "payment_confirmed": False
    }, {"_id": 0}).to_list(10)
    
    if not membership and not pending_memberships:
        return {"has_membership": False, "wallet_balance": 0, "pending_memberships": []}
    
    if membership:
        # Check if expired
        expiry_date = datetime.fromisoformat(membership["expiry_date"])
        if expiry_date < datetime.now(timezone.utc):
            await db.customer_memberships.update_one(
                {"id": membership["id"]},
                {"$set": {"is_active": False, "wallet_balance": 0}}
            )
            return {"has_membership": False, "wallet_balance": 0, "expired": True, "pending_memberships": pending_memberships}
        
        return {"has_membership": True, **membership, "pending_memberships": pending_memberships}
    
    return {"has_membership": False, "wallet_balance": 0, "pending_memberships": pending_memberships}

@api_router.get("/salons/{salon_id}/wallet-transactions/{phone}")
async def get_wallet_transactions(salon_id: str, phone: str):
    """Get customer's wallet transaction history"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    transactions = await db.wallet_transactions.find({
        "salon_id": salon_id,
        "customer_phone": phone
    }, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    return {"transactions": transactions}

@api_router.post("/salons/{salon_id}/use-wallet")
async def use_wallet_balance(salon_id: str, phone: str, amount: float):
    """Deduct amount from customer's wallet"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    membership = await db.customer_memberships.find_one({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True
    }, {"_id": 0})
    
    if not membership or membership["wallet_balance"] < amount:
        raise HTTPException(status_code=400, detail="Insufficient wallet balance")
    
    new_balance = membership["wallet_balance"] - amount
    
    await db.customer_memberships.update_one(
        {"id": membership["id"]},
        {"$set": {"wallet_balance": new_balance}}
    )
    
    # Record transaction
    await db.wallet_transactions.insert_one({
        "id": str(uuid.uuid4()),
        "customer_phone": phone,
        "salon_id": salon_id,
        "transaction_type": "debit",
        "amount": amount,
        "balance_after": new_balance,
        "description": "Used for booking",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"new_balance": new_balance}

@api_router.put("/salons/{salon_id}/customers/{phone}/wallet-balance")
async def update_wallet_balance(
    salon_id: str,
    phone: str,
    body: dict,
    current_user=Depends(get_current_salon_user)
):
    """Admin endpoint to manually adjust customer wallet balance"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    new_balance = body.get("wallet_balance")
    reason = body.get("reason", "Manual adjustment by admin")
    
    if new_balance is None or new_balance < 0:
        raise HTTPException(status_code=400, detail="Invalid wallet balance")
    
    membership = await db.customer_memberships.find_one({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True
    }, {"_id": 0})
    
    if not membership:
        raise HTTPException(status_code=404, detail="No active membership found")
    
    old_balance = membership["wallet_balance"]
    diff = new_balance - old_balance
    
    # Update balance
    await db.customer_memberships.update_one(
        {"id": membership["id"]},
        {"$set": {"wallet_balance": new_balance}}
    )
    
    # Record transaction for audit
    if diff != 0:
        await db.wallet_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "customer_phone": phone,
            "salon_id": salon_id,
            "transaction_type": "credit" if diff > 0 else "debit",
            "amount": abs(diff),
            "balance_after": new_balance,
            "description": reason,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    return {"message": "Wallet balance updated", "old_balance": old_balance, "new_balance": new_balance}

@api_router.get("/salons/{salon_id}/customers/{phone}/combined-history")
async def get_customer_combined_history(salon_id: str, phone: str):
    """Get combined booking + wallet transaction history for a customer"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Fetch bookings
    bookings = await db.tokens.find({
        "salon_id": salon_id,
        "phone": phone
    }, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    
    # Fetch wallet transactions
    transactions = await db.wallet_transactions.find({
        "salon_id": salon_id,
        "customer_phone": phone
    }, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)
    
    # Combine and tag
    combined = []
    for b in bookings:
        combined.append({
            **b,
            "history_type": "booking",
            "sort_date": b.get("created_at", "")
        })
    for t in transactions:
        combined.append({
            **t,
            "history_type": "transaction",
            "sort_date": t.get("created_at", "")
        })
    
    # Sort by date descending
    combined.sort(key=lambda x: x.get("sort_date", ""), reverse=True)
    
    return {"history": combined}

@api_router.get("/salons/{salon_id}/customers/{phone}/recent-services")
async def get_customer_recent_services(salon_id: str, phone: str):
    """Get recently used services by a customer at this salon"""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Get recent bookings with services (include all non-cancelled statuses)
    bookings = await db.tokens.find({
        "salon_id": salon_id,
        "phone": phone,
        "status": {"$nin": ["cancelled"]}
    }, {"_id": 0, "selected_services": 1, "created_at": 1}).sort("created_at", -1).limit(20).to_list(20)
    
    # Collect unique service IDs in order of recency
    seen = set()
    recent_service_ids = []
    for booking in bookings:
        for sid in (booking.get("selected_services") or []):
            if sid not in seen:
                seen.add(sid)
                recent_service_ids.append(sid)
    
    # Fetch service details (include even disabled ones for recent history)
    recent_services = []
    for sid in recent_service_ids[:15]:
        service = await db.services.find_one({"id": sid}, {"_id": 0})
        if service:
            recent_services.append(service)
    
    return {"recent_services": recent_services}

@api_router.post("/salons/{salon_id}/loyalty-program", response_model=LoyaltyProgram)
async def update_loyalty_program(salon_id: str, settings: LoyaltyProgramSettings, current_user=Depends(get_current_salon_admin)):
    """Update loyalty program settings"""
    existing = await db.loyalty_programs.find_one({"salon_id": salon_id}, {"_id": 0})
    
    settings_dict = settings.model_dump()
    settings_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if existing:
        settings_dict["id"] = existing["id"]
        await db.loyalty_programs.update_one(
            {"salon_id": salon_id},
            {"$set": settings_dict}
        )
    else:
        settings_dict["id"] = str(uuid.uuid4())
        await db.loyalty_programs.insert_one(settings_dict)
    
    return LoyaltyProgram(**settings_dict)

@api_router.get("/salons/{salon_id}/loyalty-program")
async def get_loyalty_program(salon_id: str, current_user=Depends(get_current_salon_user)):
    """Get loyalty program settings"""
    program = await db.loyalty_programs.find_one({"salon_id": salon_id}, {"_id": 0})
    if not program:
        return {
            "enabled": False,
            "spend_amount": 0,
            "period_months": 0,
            "topup_percentage": 0
        }
    return program

# ============ BOOKING/TOKEN ROUTES ============

@api_router.get("/shifts")
async def get_available_shifts():
    """Get available shifts for booking"""
    return {"shifts": get_shifts()}


def _fmt_hour_label(h: int) -> str:
    """Convert 24h integer to '7 AM' / '4 PM' style label."""
    try:
        h = int(h)
    except Exception:
        return ""
    if h == 0:
        return "12 AM"
    if h == 12:
        return "12 PM"
    if h < 12:
        return f"{h} AM"
    return f"{h - 12} PM"


@api_router.get("/salons/{salon_id}/shift-windows")
async def get_salon_shift_windows_api(salon_id: str, date: Optional[str] = None):
    """Return shift windows for a salon on a given date computed from its operational_hours."""
    if not date:
        date = datetime.now(timezone.utc).date().isoformat()
    windows = await get_salon_shift_windows(salon_id, date)
    shifts = []
    for shift_id in ["Morning", "Noon", "Evening"]:
        w = windows.get(shift_id, {}) or {}
        start_h = w.get("start")
        end_h = w.get("end")
        duration_hours = w.get("duration_hours", 0) or 0
        start_label = _fmt_hour_label(start_h) if start_h is not None else ""
        end_label = _fmt_hour_label(end_h) if end_h is not None else ""
        time_label = f"{start_label} - {end_label}" if start_label and end_label else ""
        shifts.append({
            "id": shift_id,
            "name": shift_id,
            "start": start_h,
            "end": end_h,
            "time": time_label,
            "duration_hours": duration_hours,
            "duration_minutes": w.get("duration_minutes", 0),
            "is_available": bool(duration_hours and duration_hours > 0),
        })
    return {"date": date, "shifts": shifts}

@api_router.get("/slots")
async def get_available_slots(date: Optional[str] = None):
    """DEPRECATED: Get 2-hour time slots (kept for backward compatibility)"""
    if date:
        parsed_date = datetime.fromisoformat(date).date()
        day_name = parsed_date.strftime("%A")
        
        # Check if Tuesday (closed)
        if day_name == "Tuesday":
            return {"slots": []}
    
    return {"slots": generate_2hour_slots()}

@api_router.post("/bookings", response_model=TokenModel)
async def create_booking(booking: BookingCreate):
    """Create new booking/token with shift-based system"""
    # Validate phone
    phone = booking.phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Handle guest booking - create or update customer account
    customer = await db.customers.find_one({"phone": phone})
    if not customer:
        # Create new customer account (unverified for guests)
        customer = {
            "id": str(uuid.uuid4()),
            "phone": phone,
            "name": booking.customer_name or "Guest",
            "gender": booking.customer_gender or "other",
            "is_verified": False if booking.is_guest else True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "last_booking_salon_id": booking.salon_id
        }
        await db.customers.insert_one(customer)
        logger.info(f"Created new customer account: {phone} (guest={booking.is_guest})")
    else:
        # Update last visited salon for smart routing
        await db.customers.update_one(
            {"phone": phone},
            {"$set": {"last_booking_salon_id": booking.salon_id}}
        )
    
    # Check booking limit: Max 2 active bookings per customer per day (1 for self + 1 for others)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    active_bookings = await db.tokens.find({
        "phone": phone,
        "date": today,
        "status": {"$in": ["waiting", "in_service"]},
        "$or": [{"cancelled": False}, {"cancelled": {"$exists": False}}]
    }, {"_id": 0}).to_list(10)
    
    # Count bookings for self and others
    bookings_for_self = sum(1 for b in active_bookings if b.get("booking_for_self", True))
    bookings_for_others = len(active_bookings) - bookings_for_self
    
    # Enforce limits: max 1 for self + max 1 for others
    if booking.booking_for_self and bookings_for_self >= 1:
        raise HTTPException(
            status_code=400,
            detail="You already have an active booking for yourself today. Please complete or cancel it before booking again."
        )
    
    if not booking.booking_for_self and bookings_for_others >= 1:
        raise HTTPException(
            status_code=400,
            detail="You already have an active booking for someone else today. Maximum 2 bookings allowed per day (1 for self + 1 for others)."
        )
    
    if len(active_bookings) >= 2:
        raise HTTPException(
            status_code=400,
            detail="You have reached the maximum limit of 2 active bookings per day. Please complete or cancel an existing booking."
        )
    
    # Check if salon is on holiday for the booking date
    booking_date = booking.date or today
    salon_holiday = await db.salon_holidays.find_one({
        "salon_id": booking.salon_id,
        "date": booking_date
    })
    if salon_holiday:
        raise HTTPException(
            status_code=400,
            detail=f"Salon is closed on {booking_date}. Please select another date."
        )
    
    # === Manual close (online/full) enforcement for online customer bookings ===
    salon_for_close = await db.salons.find_one({"id": booking.salon_id}, {"_id": 0, "manual_toggle": 1})
    mt = (salon_for_close or {}).get("manual_toggle") or {}
    if mt.get("is_overridden") and not mt.get("is_open", True):
        closed_mode = mt.get("closed_mode") or "full"
        # Allow QR-scan walk-in flow even when online_only closed
        src_lower = (booking.source or "").lower()
        is_qr = src_lower in ("qr_scan", "qr", "qr_walkin")
        if closed_mode == "full":
            raise HTTPException(
                status_code=400,
                detail="Salon is currently closed. Bookings are not being accepted right now."
            )
        if closed_mode == "online_only" and not is_qr:
            raise HTTPException(
                status_code=400,
                detail="Closed Online — Visit Salon. Online bookings are paused; please book at the salon."
            )
    
    # Check if day is marked as holiday in operational hours
    salon_for_hours = await db.salons.find_one({"id": booking.salon_id}, {"_id": 0, "operational_hours": 1})
    if salon_for_hours and salon_for_hours.get("operational_hours"):
        booking_weekday = datetime.fromisoformat(booking_date).strftime("%A").lower()
        day_hours = salon_for_hours["operational_hours"].get(booking_weekday, {})
        if day_hours.get("is_holiday"):
            raise HTTPException(
                status_code=400,
                detail=f"Salon is closed on {booking_weekday.capitalize()}s. Please select another date."
            )
    
    # Get barber details (with Fastest-Barber auto-assignment support)
    # Compute required blocked minutes for the 75% rule
    service_total_minutes = await calc_service_total_minutes(booking.selected_services)
    required_blocked = calc_blocked_minutes_from_total(service_total_minutes)

    if booking.barber_id == "any":
        # Fastest Barber Auto Assignment
        chosen = await pick_fastest_barber(
            salon_id=booking.salon_id,
            date=booking.date,
            shift=booking.shift,
            required_blocked_minutes=required_blocked,
            customer_gender=booking.customer_gender,
        )
        if not chosen:
            raise HTTPException(
                status_code=400,
                detail=f"All barbers are fully booked for {booking.shift} shift. Please choose another shift or date."
            )
        booking.barber_id = chosen["id"]
        barber_name = chosen["name"]
        barber = chosen
    else:
        barber = await db.barbers.find_one({"id": booking.barber_id}, {"_id": 0})
        if not barber:
            raise HTTPException(status_code=404, detail="Barber not found")
        barber_name = barber["name"]

        # 75% rule capacity check
        fit = await can_fit_in_barber_shift(
            salon_id=booking.salon_id,
            barber_id=booking.barber_id,
            date=booking.date,
            shift=booking.shift,
            new_blocked_minutes=required_blocked,
        )
        if not fit["allowed"]:
            raise HTTPException(
                status_code=400,
                detail=f"{barber_name}'s {booking.shift} shift is full. Please pick another barber or shift."
            )
    
    # Calculate total amount (barber is now always a real barber after fastest-barber assignment)
    total_amount = 0
    if booking.barber_id and booking.barber_id != "any":
        total_amount = await calculate_booking_total(booking.selected_services, booking.barber_id)
    # Fallback to base prices if nothing was found
    if total_amount == 0:
        for service_id in booking.selected_services:
            service = await db.services.find_one({"id": service_id}, {"_id": 0})
            if service:
                total_amount += service.get("base_price", 0)
    
    # Handle wallet payment
    payment_status = "pending"
    payment_mode = booking.payment_mode
    
    if payment_mode == "wallet":
        if not phone.startswith("+91"):
            wallet_phone = f"+91{phone}"
        else:
            wallet_phone = phone
        
        membership = await db.customer_memberships.find_one({
            "salon_id": booking.salon_id,
            "customer_phone": wallet_phone,
            "is_active": True
        }, {"_id": 0})
        
        if not membership:
            raise HTTPException(status_code=400, detail="No active wallet/membership found")
        
        if membership["wallet_balance"] < total_amount:
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient wallet balance. Available: ₹{membership['wallet_balance']}, Required: ₹{total_amount}"
            )
        
        # Deduct wallet balance
        new_balance = membership["wallet_balance"] - total_amount
        await db.customer_memberships.update_one(
            {"id": membership["id"]},
            {"$set": {"wallet_balance": new_balance}}
        )
        
        # Record wallet transaction
        await db.wallet_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "customer_phone": wallet_phone,
            "salon_id": booking.salon_id,
            "transaction_type": "debit",
            "amount": total_amount,
            "balance_after": new_balance,
            "description": f"Booking payment - {booking.shift} shift",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        payment_status = "paid"
    
    # Get GLOBAL (salon-wide) token number — salon+date+shift sequence
    token_number = await get_next_token_number(booking.salon_id, booking.date, booking.shift)

    # Resolve branch: explicit > barber's branch > salon's main branch
    booking_branch_id = await resolve_branch_id(booking.salon_id, booking.branch_id)
    if not booking_branch_id and barber and barber.get("branch_id"):
        booking_branch_id = barber.get("branch_id")

    # Snapshot the booking-time OTP-verification state of the booker so the
    # history UI can mark "express / non-OTP-verified" bookings.
    is_otp_verified_at_booking = False
    booker_user = await db.users.find_one({"phone": phone}, {"_id": 0, "is_otp_verified": 1})
    if booker_user and booker_user.get("is_otp_verified"):
        is_otp_verified_at_booking = True

    # Create token
    token_dict = {
        "id": str(uuid.uuid4()),
        "salon_id": booking.salon_id,
        "branch_id": booking_branch_id,
        "token_number": token_number,
        "customer_name": booking.customer_name,
        "phone": phone,
        "user_id": booking.user_id,
        "date": booking.date,
        "shift": booking.shift,
        "time_slot": booking.time_slot,  # Optional, for backward compatibility
        "barber_id": booking.barber_id,
        "barber_name": barber_name,
        "selected_services": booking.selected_services,
        "total_amount": total_amount,
        # 75%-rule bookkeeping (stored for consistent capacity math)
        "total_service_minutes": service_total_minutes,
        "blocked_minutes": required_blocked,
        "status": "waiting" if booking.booking_type == "instant" else "future",
        "payment_status": payment_status,
        "payment_mode": payment_mode,
        "payment_confirmed": payment_mode == "wallet",  # Auto-confirm wallet payments
        "upi_transaction_id": None,
        "source": booking.source,
        "booking_type": booking.booking_type,
        "booking_for_self": booking.booking_for_self,
        "is_otp_verified_at_booking": is_otp_verified_at_booking,
        "allocated_at": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.tokens.insert_one(token_dict)
    
    token = TokenModel(**token_dict)
    await broadcast_update("token_created", token.model_dump())
    
    # Send booking confirmation via WhatsApp
    await send_booking_notification(token_dict, 'booking_confirmation')
    
    # Notify salon (in-app) about new booking
    if token_dict.get("salon_id"):
        await create_in_app_notification(
            user_type="salon",
            user_id=token_dict["salon_id"],
            title="New Booking Received",
            message=f"{token_dict.get('customer_name','Customer')} booked Token #{token_dict.get('token_number','')} on {token_dict.get('date','')} ({token_dict.get('time_slot','')}).",
            notification_type="new_booking",
            setting_key="new_booking",
            salon_id=token_dict["salon_id"],
            related_id=token_dict.get("id", ""),
        )
    
    return token

# Slot availability endpoint
@api_router.get("/salons/{salon_id}/slot-availability")
async def get_slot_availability(
    salon_id: str, date: str, shift: str, service_ids: Optional[str] = None
):
    """Check per-barber capacity for a given date and shift using the 75%-rule.

    Query params:
      - service_ids (optional, comma-separated) — if provided, compute whether each barber
        can accommodate those specific services (more accurate display).
    """
    # Compute shift window for salon
    windows = await get_salon_shift_windows(salon_id, date)
    shift_info = windows.get(shift, {})
    capacity = int(shift_info.get("duration_minutes", 0))

    # Optional services → required blocked minutes
    required_blocked = 0
    if service_ids:
        sid_list = [s for s in service_ids.split(",") if s]
        total_mins = await calc_service_total_minutes(sid_list)
        required_blocked = calc_blocked_minutes_from_total(total_mins)

    # Find active barbers AVAILABLE on this date (joining/last-working/leave-aware)
    raw_barbers = await db.barbers.find(
        {
            "salon_id": salon_id,
            "is_active": True,
        },
        {"_id": 0},
    ).to_list(500)
    barbers = [b for b in raw_barbers if is_barber_available_on(b, date)]

    result = []
    for barber in barbers:
        used = await get_barber_blocked_minutes_used(salon_id, barber["id"], date, shift)
        remaining = max(0, capacity - used)
        # If required_blocked specified, barber is full when it can't take this booking
        already_full = used >= capacity if capacity > 0 else True
        cant_fit_new = required_blocked > 0 and remaining <= 0
        is_full = already_full or cant_fit_new
        # count of live bookings (for display)
        count = await db.tokens.count_documents({
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "date": date,
            "shift": shift,
            "status": {"$nin": ["cancelled", "skipped"]},
        })
        result.append({
            "barber_id": barber["id"],
            "barber_name": barber["name"],
            "booked": count,
            "capacity_minutes": capacity,
            "used_minutes": used,
            "remaining_minutes": remaining,
            # "available" kept for backward compatibility: approximate slots left (remaining / avg 30 min service with 75% rule ≈ 22.5m)
            "available": max(0, int(remaining / 22)) if remaining > 0 else 0,
            "is_full": is_full,
        })

    all_full = all(b["is_full"] for b in result) if result else False

    return {
        "date": date,
        "shift": shift,
        "shift_window": shift_info,
        "capacity_minutes": capacity,
        "barbers": result,
        "all_slots_full": all_full,
        # Kept for backward compat but no longer the source of truth
        "max_per_slot": 10,
    }

@api_router.get("/salons/{salon_id}/barbers/{barber_id}/queue", response_model=List[TokenModel])
async def get_barber_queue(
    salon_id: str,
    barber_id: str,
    date: Optional[str] = None,
    status: Optional[str] = None,
    branch_id: Optional[str] = None,
):
    """Get queue for specific barber"""
    if not date:
        date = datetime.now(timezone.utc).date().isoformat()
    
    query = {"salon_id": salon_id, "barber_id": barber_id, "date": date}
    if status:
        query["status"] = status
    if branch_id:
        query["branch_id"] = branch_id
    
    tokens = await db.tokens.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return tokens

@api_router.get("/salons/{salon_id}/queue", response_model=List[TokenModel])
async def get_salon_queue(
    salon_id: str,
    date: Optional[str] = None,
    status: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: Optional[dict] = Depends(get_current_salon_user_optional),
):
    """Get entire salon queue (all barbers)"""
    if not date:
        date = datetime.now(timezone.utc).date().isoformat()
    
    # Branch-manager scoping
    if current_user and is_branch_manager(current_user):
        branch_id = enforce_branch_for_manager(current_user, branch_id)
    
    query = {"salon_id": salon_id, "date": date}
    if status:
        query["status"] = status
    if branch_id:
        query["branch_id"] = branch_id
    
    tokens = await db.tokens.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return tokens

@api_router.post("/salons/{salon_id}/barbers/{barber_id}/call-next")
async def call_next_token(salon_id: str, barber_id: str, current_salon=Depends(get_current_salon)):
    """Call next token for specific barber - Does NOT auto-complete previous token"""
    date = datetime.now(timezone.utc).date().isoformat()
    
    # DON'T auto-complete previous tokens - barber must manually complete
    
    # Get next waiting token (FIFO by creation time — lexicographic sort on M1,M2,M10 is unsafe)
    next_token = await db.tokens.find_one(
        {"salon_id": salon_id, "barber_id": barber_id, "date": date, "status": "waiting"},
        {"_id": 0},
        sort=[("created_at", 1)]
    )
    
    if next_token:
        await db.tokens.update_one(
            {"id": next_token["id"]},
            {
                "$set": {
                    "status": "called",
                    "called_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        updated = TokenModel(**{**next_token, "status": "called", "called_at": datetime.now(timezone.utc).isoformat()})
        await broadcast_update("token_called", updated.model_dump())
        
        # Send WhatsApp notification that token is called
        await send_booking_notification(next_token, 'token_called')
        
        # In-app notification for customer (your turn now)
        if next_token.get("phone"):
            await create_in_app_notification(
                user_type="customer",
                user_id=next_token["phone"],
                title="It's Your Turn!",
                message=f"Your token #{next_token.get('token_number','')} is being called. Please proceed to {next_token.get('barber_name','your barber')}.",
                notification_type="turn_now",
                setting_key="booking_status_change",
                salon_id=salon_id,
                related_id=next_token.get("id", ""),
            )
        
        # Check and notify tokens that are near (3, 2, 1 away)
        await check_and_notify_nearby_tokens(salon_id, barber_id, date, next_token["token_number"])
        
        return updated
    
    return {"message": "No more tokens in queue"}


@api_router.post("/tokens/{token_id}/complete")
async def complete_token(token_id: str, current_salon=Depends(get_current_salon)):
    """Manually mark token as completed and generate invoice"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    # Check if payment is confirmed
    if not token.get("payment_confirmed", False):
        raise HTTPException(
            status_code=400, 
            detail="Payment must be confirmed before marking as complete. Please confirm payment first."
        )
    
    # Mark as completed
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {"status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()}}
    )

    # --- Auto-attendance (service_completion mode) ---------------------------
    # When the salon's attendance mode is "service_completion", marking a token
    # complete should also mark the barber present for that IST day (unless an
    # admin override already exists or the month is locked). This is the
    # behaviour the UI advertises in the "By service completion" radio option.
    try:
        barber_id = token.get("barber_id")
        salon_id_for_attn = token.get("salon_id")
        if barber_id and salon_id_for_attn:
            salon_doc = await db.salons.find_one(
                {"id": salon_id_for_attn},
                {"_id": 0, "attendance_mode": 1},
            ) or {}
            mode_now = salon_doc.get("attendance_mode") or "service_completion"
            if mode_now == "service_completion":
                # IST date of the booking (use the booking's own date when set,
                # else "today" in IST). This matches what calculate_daily_attendance
                # would compute.
                ist_today = attendance_mode_mod.current_ist_date()
                date_str = token.get("date") or ist_today
                locked = await attendance_mode_mod.is_attendance_locked(
                    db, salon_id_for_attn, barber_id, date_str
                )
                if not locked:
                    record_id = f"{salon_id_for_attn}_{barber_id}_{date_str}"
                    existing_override = await db.attendance.find_one({
                        "id": record_id,
                        "auto_calculated": False,
                    }, {"_id": 0})
                    if not existing_override:
                        calc = await calculate_barber_attendance_for_date(
                            salon_id_for_attn, barber_id, date_str
                        )
                        now_iso = datetime.now(timezone.utc).isoformat()
                        await db.attendance.update_one(
                            {"id": record_id},
                            {"$set": {
                                "id": record_id,
                                "salon_id": salon_id_for_attn,
                                "barber_id": barber_id,
                                "date": date_str,
                                "status": calc["status"],
                                "auto_calculated": True,
                                "morning_shift_completed": calc.get("morning_shift_completed", False),
                                "noon_evening_shift_completed": calc.get("noon_evening_shift_completed", False),
                                "bookings_count": calc.get("bookings_count", 0),
                                "computed_under_mode": calc.get("computed_under_mode"),
                                "updated_at": now_iso,
                            },
                             "$setOnInsert": {"created_at": now_iso}},
                            upsert=True,
                        )
    except Exception as auto_attn_err:
        # Auto-attendance must never block completing the booking. Log + continue.
        logger.warning(f"Auto-attendance upsert failed for token {token_id}: {auto_attn_err}")

    # After completion the queue advances — ping guests that are now 1 or 2
    # spots away so they can start heading over.
    try:
        salon_id_n = token.get("salon_id")
        barber_id_n = token.get("barber_id")
        date_n = token.get("date")
        tok_num_n = token.get("token_number")
        if salon_id_n and barber_id_n and date_n and tok_num_n:
            await check_and_notify_nearby_tokens(salon_id_n, barber_id_n, date_n, str(tok_num_n))
    except Exception as _e:
        logger.warning(f"nearby-alert skipped on /complete: {_e}")

    # Notify customer of status change (in-app)
    if token.get("phone"):
        await create_in_app_notification(
            user_type="customer",
            user_id=token["phone"],
            title="Service Completed",
            message=f"Your service (Token #{token.get('token_number','')}) has been completed. Thank you!",
            notification_type="booking_completed",
            setting_key="booking_status_change",
            salon_id=token.get("salon_id", ""),
            related_id=token_id,
        )
        # Send approved WhatsApp 'booking_completed' template (Item 8).
        await send_booking_notification(
            {**token, "status": "completed", "completed_at": datetime.now(timezone.utc).isoformat()},
            'booking_completed',
        )
    
    # Check and apply loyalty reward (only for non-wallet payments)
    loyalty_reward = None
    payment_mode = token.get("payment_mode", "cash")
    try:
        loyalty_reward = await check_and_apply_loyalty_reward(
            token["salon_id"], 
            token["phone"], 
            token.get("total_amount", 0),
            payment_mode
        )
    except Exception as e:
        logger.error(f"Error checking loyalty reward: {e}")
    
    # Generate and send invoice
    try:
        invoice_sent = await generate_and_send_invoice(token_id)
        await broadcast_update("token_completed", {"token_id": token_id})
        
        response = {
            "message": "Token marked as completed",
            "invoice_sent": invoice_sent
        }
        
        if loyalty_reward:
            response["loyalty_reward"] = loyalty_reward
            response["message"] += f" - Loyalty reward of ₹{loyalty_reward['topup_amount']:.2f} added to wallet!"

        # Recompute the barber's incentive payout for the booking month so the
        # Reward Plan dashboard always reflects the latest sales.
        try:
            barber_id = token.get("barber_id")
            # Tokens use `date` field (YYYY-MM-DD). Fall back to legacy `booking_date`
            # then to today's UTC date.
            tok_date = (
                token.get("date")
                or token.get("booking_date")
                or datetime.now(timezone.utc).strftime("%Y-%m-%d")
            )
            year_month = tok_date[:7]
            salon_id_for_token = token.get("salon_id")
            if barber_id and salon_id_for_token and year_month:
                await _recompute_incentive_payout(salon_id_for_token, barber_id, year_month)
        except Exception as e:
            logger.error(f"Incentive recompute failed for token {token_id}: {e}")

        return response
    except Exception as e:
        logger.error(f"Error generating invoice: {e}")
        return {
            "message": "Token marked as completed but invoice generation failed",
            "error": str(e),
            "loyalty_reward": loyalty_reward
        }

@api_router.post("/tokens/{token_id}/recall")
async def recall_token(token_id: str, current_salon=Depends(get_current_salon_user)):
    """Re-call a token (if customer not available) or recall a skipped token"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    # If token is skipped, change status to called
    if token.get("status") == "skipped":
        await db.tokens.update_one(
            {"id": token_id},
            {
                "$set": {
                    "status": "called",
                    "called_at": datetime.now(timezone.utc).isoformat(),
                    "recall_count": token.get("recall_count", 0) + 1
                }
            }
        )
        await broadcast_update("token_recalled", {
            "token_id": token_id,
            "status": "called",
            "phone": token.get("phone", ""),
            "token_number": token.get("token_number"),
            "salon_id": token.get("salon_id"),
            "barber_name": token.get("barber_name"),
        })
        return {"message": "Skipped token recalled and moved to called status"}
    
    # Otherwise, increment recall count and update called_at timestamp
    recall_count = token.get("recall_count", 0) + 1
    await db.tokens.update_one(
        {"id": token_id},
        {
            "$set": {
                "recall_count": recall_count,
                "called_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Send notification again
    await send_booking_notification(token, 'token_called')
    await broadcast_update("token_recalled", {
        "token_id": token_id,
        "recall_count": recall_count,
        "phone": token.get("phone", ""),
        "token_number": token.get("token_number"),
        "salon_id": token.get("salon_id"),
        "barber_name": token.get("barber_name"),
    })
    
    return {"message": "Token recalled and notification sent", "recall_count": recall_count}

@api_router.post("/tokens/{token_id}/skip")
async def skip_token(token_id: str, reason: Optional[str] = None, current_salon=Depends(get_current_salon)):
    """Skip/Cancel a token (if customer doesn't show up)"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    # Mark as skipped (not cancelled, different meaning)
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {
            "status": "skipped",
            "skipped_reason": reason or "Customer no-show",
            "skipped_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Send notification
    token['skipped_reason'] = reason or "Customer no-show"
    await send_booking_notification(token, 'token_skipped')
    await broadcast_update("token_skipped", {"token_id": token_id})
    
    return {"message": "Token skipped"}

@api_router.post("/tokens/{token_id}/resend-invoice")
async def resend_invoice(token_id: str, current_salon=Depends(get_current_salon)):
    """Resend invoice to customer"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    if token.get('status') != 'completed':
        raise HTTPException(status_code=400, detail="Can only resend invoice for completed bookings")
    
    try:
        invoice_sent = await generate_and_send_invoice(token_id)
        return {
            "message": "Invoice resent successfully",
            "sent": invoice_sent
        }
    except Exception as e:
        logger.error(f"Error resending invoice: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to resend invoice: {str(e)}")

@api_router.post("/tokens/{token_id}/call")
async def call_token(token_id: str, current_salon=Depends(get_current_salon_user)):
    """Call specific token"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    # Update status to "called" with timestamp
    await db.tokens.update_one(
        {"id": token_id}, 
        {"$set": {
            "status": "called",
            "called_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    await broadcast_update("token_called", {
        "token_id": token_id,
        "phone": token.get("phone", ""),
        "token_number": token.get("token_number"),
        "salon_id": token.get("salon_id"),
        "barber_name": token.get("barber_name"),
    })
    
    # Send notification
    await send_booking_notification(token, 'token_called')

    # Also ping guests waiting 2 and 1 spots away — they should start heading over.
    try:
        salon_id = token.get("salon_id")
        barber_id = token.get("barber_id")
        date = token.get("date")
        tok_num = token.get("token_number")
        if salon_id and barber_id and date and tok_num:
            await check_and_notify_nearby_tokens(salon_id, barber_id, date, str(tok_num))
    except Exception as _e:
        logger.warning(f"nearby-alert skipped on /call: {_e}")

    return {"message": "Token called"}

@api_router.post("/tokens/{token_id}/cancel")
async def cancel_token(
    token_id: str,
    current_salon: Optional[Dict[str, Any]] = Depends(get_current_salon_optional),
    current_user: Optional[Dict[str, Any]] = Depends(get_current_salon_user_optional),
):
    """Cancel token and refund wallet if paid via wallet (works for both legacy salon JWT and multi-user salon_admin/salon_staff)."""
    if not (current_salon or current_user):
        raise HTTPException(status_code=401, detail="Unauthorized")

    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")

    # Authorization
    salon_id = token.get("salon_id")
    authorized = False
    if current_salon:
        cs_salon_id = current_salon.get("sub") or current_salon.get("id")
        if cs_salon_id == salon_id:
            authorized = True
    if current_user:
        if current_user.get("role") in ("salon_admin", "admin", "salon"):
            if current_user.get("salon_id") in (None, salon_id):
                authorized = True
        elif current_user.get("permissions", {}).get("can_manage_tokens"):
            if current_user.get("salon_id") == salon_id:
                authorized = True
    if not authorized:
        raise HTTPException(status_code=403, detail="Unauthorized for this salon")
    
    # Refund wallet if payment was via wallet
    if token.get("payment_mode") == "wallet" and token.get("payment_status") == "paid":
        wallet_phone = token["phone"]
        if not wallet_phone.startswith("+91"):
            wallet_phone = f"+91{wallet_phone}"
        
        membership = await db.customer_memberships.find_one({
            "salon_id": token["salon_id"],
            "customer_phone": wallet_phone,
            "is_active": True
        }, {"_id": 0})
        
        if membership:
            refund_amount = token.get("total_amount", 0)
            new_balance = membership["wallet_balance"] + refund_amount
            
            await db.customer_memberships.update_one(
                {"id": membership["id"]},
                {"$set": {"wallet_balance": new_balance}}
            )
            
            # Record refund transaction
            await db.wallet_transactions.insert_one({
                "id": str(uuid.uuid4()),
                "customer_phone": wallet_phone,
                "salon_id": token["salon_id"],
                "transaction_type": "credit",
                "amount": refund_amount,
                "balance_after": new_balance,
                "description": f"Refund - Booking cancelled (Token {token['token_number']})",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    await db.tokens.update_one({"id": token_id}, {"$set": {"status": "cancelled"}})
    await broadcast_update("token_cancelled", {"token_id": token_id})
    
    # Send cancellation notification
    await send_booking_notification(token, 'token_cancelled')

    # Customer in-app notification
    if token.get("phone"):
        await create_in_app_notification(
            user_type="customer",
            user_id=token["phone"],
            title="Booking Cancelled",
            message=f"Your booking (Token #{token.get('token_number','')}) was cancelled by the salon.",
            notification_type="booking_cancelled",
            setting_key="booking_status_change",
            salon_id=token.get("salon_id", ""),
            related_id=token_id,
        )
    
    return {"message": "Token cancelled"}

@api_router.post("/tokens/{token_id}/customer-cancel")
async def customer_cancel_token(token_id: str):
    """Customer-facing cancel token and refund wallet if paid via wallet"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    if token.get("status") in ["completed", "cancelled"]:
        raise HTTPException(status_code=400, detail="Cannot cancel this booking")
    
    # Refund wallet if payment was via wallet
    if token.get("payment_mode") == "wallet" and token.get("payment_status") == "paid":
        wallet_phone = token["phone"]
        if not wallet_phone.startswith("+91"):
            wallet_phone = f"+91{wallet_phone}"
        
        membership = await db.customer_memberships.find_one({
            "salon_id": token["salon_id"],
            "customer_phone": wallet_phone,
            "is_active": True
        }, {"_id": 0})
        
        if membership:
            refund_amount = token.get("total_amount", 0)
            new_balance = membership["wallet_balance"] + refund_amount
            
            await db.customer_memberships.update_one(
                {"id": membership["id"]},
                {"$set": {"wallet_balance": new_balance}}
            )
            
            await db.wallet_transactions.insert_one({
                "id": str(uuid.uuid4()),
                "customer_phone": wallet_phone,
                "salon_id": token["salon_id"],
                "transaction_type": "credit",
                "amount": refund_amount,
                "balance_after": new_balance,
                "description": f"Refund - Booking cancelled by customer (Token {token['token_number']})",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    await db.tokens.update_one({"id": token_id}, {"$set": {"status": "cancelled"}})
    await broadcast_update("token_cancelled", {"token_id": token_id})
    
    await send_booking_notification(token, 'token_cancelled')

    # Customer in-app notification
    if token.get("phone"):
        await create_in_app_notification(
            user_type="customer",
            user_id=token["phone"],
            title="Booking Cancelled",
            message=f"Your booking (Token #{token.get('token_number','')}) has been cancelled.",
            notification_type="booking_cancelled",
            setting_key="booking_status_change",
            salon_id=token.get("salon_id", ""),
            related_id=token_id,
        )
    # Salon in-app notification
    if token.get("salon_id"):
        await create_in_app_notification(
            user_type="salon",
            user_id=token["salon_id"],
            title="Booking Cancelled by Customer",
            message=f"{token.get('customer_name','Customer')} cancelled booking (Token #{token.get('token_number','')}).",
            notification_type="booking_cancelled",
            setting_key="booking_change",
            salon_id=token["salon_id"],
            related_id=token_id,
        )
    
    return {"message": "Booking cancelled successfully"}


@api_router.get("/tokens/{token_id}/public-details")
async def get_token_public_details(token_id: str):
    """Public (no-auth) lookup of a booking, used by the customer-side Reschedule
    flow that's opened via WhatsApp link. Returns only the fields the booking
    page needs to hydrate its form."""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Booking not found")
    if token.get("status") in ("completed", "cancelled", "in_service", "called", "in_progress"):
        raise HTTPException(status_code=400, detail=f"Booking cannot be rescheduled (status: {token.get('status')})")
    return {
        "id": token.get("id"),
        "salon_id": token.get("salon_id"),
        "user_id": token.get("user_id"),
        "customer_name": token.get("customer_name"),
        "phone": token.get("phone"),
        "date": token.get("date"),
        "shift": token.get("shift"),
        "barber_id": token.get("barber_id"),
        "selected_services": token.get("selected_services") or [],
        "total_amount": token.get("total_amount"),
        "booking_for_self": token.get("booking_for_self", True),
        "customer_gender": token.get("customer_gender"),
        "source": token.get("source"),
        "status": token.get("status"),
        "token_number": token.get("token_number"),
    }


@api_router.put("/tokens/{token_id}/customer-reschedule")
async def customer_reschedule_token(token_id: str, body: dict):
    """Customer-driven reschedule / modification of an existing booking.
    Updates services, barber, date and shift on the SAME token (no new token
    is created). Recalculates total and re-broadcasts token_updated."""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Booking not found")
    if token.get("status") in ("completed", "cancelled", "in_service", "called", "in_progress"):
        raise HTTPException(status_code=400, detail=f"This booking cannot be modified (status: {token.get('status')})")

    updates: dict = {}
    # Services
    new_services = body.get("selected_services")
    if isinstance(new_services, list) and new_services:
        updates["selected_services"] = new_services
    # Barber
    new_barber = body.get("barber_id")
    if new_barber:
        updates["barber_id"] = new_barber
    # Date / shift
    new_date = body.get("date")
    if new_date:
        updates["date"] = new_date
    new_shift = body.get("shift")
    if new_shift:
        updates["shift"] = new_shift
    # Payment (optional)
    new_payment = body.get("payment_mode")
    if new_payment:
        updates["payment_mode"] = new_payment

    # Recalculate total ONLY when the inputs that drive pricing changed.
    # (Pure date/shift changes keep the existing price.)
    services_changed = "selected_services" in updates
    barber_changed = "barber_id" in updates
    if services_changed or barber_changed:
        effective_barber = updates.get("barber_id", token.get("barber_id"))
        effective_services = updates.get("selected_services", token.get("selected_services") or [])
        try:
            if effective_services and effective_barber and effective_barber != "any":
                updates["total_amount"] = await calculate_booking_total(effective_services, effective_barber)
            elif effective_services:
                # Fallback: sum base prices if barber == 'any' or unknown
                svc_docs = await db.services.find(
                    {"id": {"$in": effective_services}}, {"_id": 0, "base_price": 1}
                ).to_list(len(effective_services))
                updates["total_amount"] = float(sum(s.get("base_price", 0) or 0 for s in svc_docs))
        except Exception as e:
            logger.warning(f"reschedule total recompute failed: {e}")

    if not updates:
        raise HTTPException(status_code=400, detail="No changes provided")

    await db.tokens.update_one({"id": token_id}, {"$set": updates})
    updated = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    await broadcast_update("token_updated", updated)

    # Notify salon that the customer modified their booking
    if updated and updated.get("salon_id"):
        await create_in_app_notification(
            user_type="salon",
            user_id=updated["salon_id"],
            title="Booking Modified by Customer",
            message=f"{updated.get('customer_name','Customer')} rescheduled booking (Token #{updated.get('token_number','')}).",
            notification_type="booking_modified",
            setting_key="booking_change",
            salon_id=updated["salon_id"],
            related_id=token_id,
        )

    return {"message": "Booking updated successfully", "token": TokenModel(**updated)}


@api_router.post("/tokens/{token_id}/defer")
async def defer_token(token_id: str, new_slot: str, current_salon=Depends(get_current_salon)):
    """Defer token to next slot"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    await db.tokens.update_one({"id": token_id}, {"$set": {"time_slot": new_slot, "status": "waiting"}})
    await broadcast_update("token_deferred", {"token_id": token_id, "new_slot": new_slot})
    return {"message": "Token deferred"}

@api_router.post("/tokens/{token_id}/notify")
async def send_token_notification(
    token_id: str,
    current_salon: Optional[Dict[str, Any]] = Depends(get_current_salon_optional),
    current_user: Optional[Dict[str, Any]] = Depends(get_current_salon_user_optional),
):
    """
    Triggered by the salon's "Send Notification to Customer" button.
    Sends the spec'd "Salon X is calling you. Proceed to Barber Y's chair." message
    via WhatsApp + appends Reschedule / Cancel deep-link actions.
    Falls back gracefully if WhatsApp delivery fails.
    """
    if not (current_salon or current_user):
        raise HTTPException(status_code=401, detail="Unauthorized")

    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")

    # Authorization: the caller must belong to this token's salon.
    salon_id = token.get("salon_id")
    authorized = False
    if current_salon:
        cs_salon_id = current_salon.get("sub") or current_salon.get("id")
        if cs_salon_id == salon_id:
            authorized = True
    if current_user:
        if current_user.get("role") in ("salon_admin", "admin", "salon"):
            if current_user.get("salon_id") in (None, salon_id):
                authorized = True
        elif current_user.get("permissions", {}).get("can_manage_tokens"):
            if current_user.get("salon_id") == salon_id:
                authorized = True
    if not authorized:
        raise HTTPException(status_code=403, detail="Unauthorized for this salon")

    # Send the salon-calling notification (handles WhatsApp + adds action links).
    await send_booking_notification(token, 'salon_calling')

    # Persist a record so the salon can see history if needed.
    notification = {
        "id": str(uuid.uuid4()),
        "token_id": token_id,
        "notification_type": "salon_calling",
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }
    try:
        await db.notifications.insert_one(notification)
    except Exception as _e:  # noqa: BLE001
        logger.warning(f"Failed to persist notification record: {_e}")

    # Also broadcast in-app so the customer's open tab shows it instantly.
    await broadcast_update("salon_calling", {
        "token_id": token_id,
        "phone": token.get("phone", ""),
        "token_number": token.get("token_number"),
        "salon_id": salon_id,
        "barber_name": token.get("barber_name"),
    })

    return {"message": "Notification sent to customer"}

@api_router.put("/tokens/{token_id}/add-services")
async def add_services_to_token(token_id: str, request: AddServicesRequest, current_salon=Depends(get_current_salon)):
    """Add services to existing booking (before completion)"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    # Check if token is already completed
    if token.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Cannot add services to completed booking")
    
    # Get current services
    current_services = token.get("selected_services", [])
    
    # Add new services (avoid duplicates)
    for service_id in request.service_ids:
        if service_id not in current_services:
            current_services.append(service_id)
    
    # Recalculate total amount
    total_amount = 0
    if token["barber_id"] != "any":
        total_amount = await calculate_booking_total(current_services, token["barber_id"])
    
    # Update token
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {
            "selected_services": current_services,
            "total_amount": total_amount
        }}
    )
    
    # Get updated token
    updated_token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    
    # Broadcast update
    await broadcast_update("token_updated", updated_token)
    
    return {
        "message": "Services added successfully",
        "token": TokenModel(**updated_token)
    }

@api_router.put("/tokens/{token_id}/update-services")
async def update_token_services(token_id: str, request: AddServicesRequest, current_salon=Depends(get_current_salon)):
    """Update services for existing booking - can add or remove (before completion)"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    # Check if token is already completed
    if token.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Cannot modify services for completed booking")
    
    # Replace services entirely with new selection
    new_services = request.service_ids
    
    # Recalculate total amount
    total_amount = 0
    if token["barber_id"] != "any":
        total_amount = await calculate_booking_total(new_services, token["barber_id"])
    
    # Use final_amount override if provided by salon
    if request.final_amount is not None:
        total_amount = request.final_amount
    
    # Update token
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {
            "selected_services": new_services,
            "total_amount": total_amount
        }}
    )
    
    # Get updated token
    updated_token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    
    # Broadcast update
    await broadcast_update("token_updated", updated_token)
    
    return {
        "message": "Services updated successfully",
        "token": TokenModel(**updated_token)
    }

@api_router.put("/tokens/{token_id}/update-amount")
async def update_token_amount(token_id: str, body: dict, current_salon=Depends(get_current_salon)):
    """Update final amount for a token (salon override)"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    if token.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Cannot modify completed booking")
    
    final_amount = body.get("final_amount")
    if final_amount is None:
        raise HTTPException(status_code=400, detail="final_amount is required")
    
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {"total_amount": float(final_amount)}}
    )
    
    updated_token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    await broadcast_update("token_updated", updated_token)
    
    return {"message": f"Amount updated to ₹{final_amount}", "token": TokenModel(**updated_token)}


# ============ Module 7 — Unified modify booking with per-service split ============

class ServiceAssignmentIn(BaseModel):
    service_id: str
    barber_id: str
    # service_price is optional in input — if not provided, we look it up from
    # barber_services. Salon UI can override (e.g., for custom pricing).
    service_price: Optional[float] = None


class TokenModifyIn(BaseModel):
    """Payload for PUT /api/tokens/{token_id}/modify."""
    main_barber_id: Optional[str] = None
    service_assignments: Optional[List[ServiceAssignmentIn]] = None
    order_discount_percent: Optional[float] = None  # 0-100
    final_amount: Optional[float] = None  # If supplied, treated as authoritative
    payment_mode: Optional[str] = None
    payment_confirmed: Optional[bool] = None


async def _resolve_assignment_price(barber_id: str, service_id: str) -> float:
    """Look up barber-specific price for a service; fall back to base price."""
    bs = await db.barber_services.find_one(
        {"barber_id": barber_id, "service_id": service_id}, {"_id": 0, "price": 1}
    )
    if bs and bs.get("price") is not None:
        return float(bs.get("price") or 0)
    svc = await db.services.find_one({"id": service_id}, {"_id": 0, "base_price": 1})
    return float((svc or {}).get("base_price") or 0)


@api_router.put("/tokens/{token_id}/modify")
async def modify_token_v2(
    token_id: str,
    payload: TokenModifyIn,
    current_salon=Depends(get_current_salon_user),
):
    """Unified modify-booking endpoint (Module 7).

    Replaces the prior chain of `update-services` + `change-barber` +
    `update-amount` with a single transactional update that supports:
      • Main barber change (customer-facing)
      • Per-service barber assignment (revenue split for reports + incentives)
      • Order-level discount (% AND/OR final ₹ override — last-edited wins on the client)
      • Payment mode + payment confirmation toggle
    """
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    if token.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Cannot modify a completed booking")

    salon_id = token["salon_id"]

    # --- 1. Main barber (customer-facing). "any" is NOT accepted on this path.
    new_main_barber_id = payload.main_barber_id or token.get("barber_id")
    if new_main_barber_id == "any":
        raise HTTPException(status_code=400, detail="A specific main barber must be selected")
    main_barber = await db.barbers.find_one(
        {"id": new_main_barber_id, "salon_id": salon_id, "is_active": True}, {"_id": 0}
    )
    if not main_barber:
        raise HTTPException(status_code=400, detail="Main barber not found or inactive")

    # --- 2. Resolve service_assignments
    assignments_in = payload.service_assignments or []
    if not assignments_in:
        # No split provided — credit everything to the main barber (legacy shape)
        services = token.get("selected_services") or []
        assignments_in = [
            ServiceAssignmentIn(service_id=sid, barber_id=new_main_barber_id)
            for sid in services
        ]

    # New flat services list = the assignments' service_ids (de-duped, preserving order)
    new_services_list: List[str] = []
    seen = set()
    for a in assignments_in:
        if a.service_id not in seen:
            seen.add(a.service_id)
            new_services_list.append(a.service_id)

    # --- 3. Validate all line barbers belong to the salon and are active
    line_barber_ids = list({a.barber_id for a in assignments_in})
    line_barbers = await db.barbers.find(
        {"id": {"$in": line_barber_ids}, "salon_id": salon_id}, {"_id": 0}
    ).to_list(100)
    barber_map = {b["id"]: b for b in line_barbers}
    for bid in line_barber_ids:
        b = barber_map.get(bid)
        if not b:
            raise HTTPException(status_code=400, detail=f"Barber {bid} not found in this salon")
        if not b.get("is_active", True):
            raise HTTPException(status_code=400, detail=f"Barber {b.get('name')} is inactive")

    # --- 4. Compute line prices, subtotal, discount, total
    resolved: List[Dict[str, Any]] = []
    subtotal = 0.0
    for a in assignments_in:
        price = a.service_price
        if price is None:
            price = await _resolve_assignment_price(a.barber_id, a.service_id)
        price = float(price or 0)
        subtotal += price
        resolved.append({
            "service_id": a.service_id,
            "barber_id": a.barber_id,
            "barber_name_snapshot": (barber_map.get(a.barber_id) or {}).get("name", ""),
            "service_price": round(price, 2),
        })

    subtotal = round(subtotal, 2)

    # Discount resolution priority:
    #  • If final_amount provided AND no order_discount_percent → derive % from final.
    #  • Else use order_discount_percent (0–100), final = subtotal - discount.
    #  • Final wins if BOTH provided (matches "last-edited wins" UX).
    if payload.final_amount is not None:
        final_amount = max(0.0, float(payload.final_amount))
        diff = max(0.0, subtotal - final_amount)
        discount_pct = (diff / subtotal * 100.0) if subtotal > 0 else 0.0
        discount_amt = diff
    else:
        discount_pct = max(0.0, min(100.0, float(payload.order_discount_percent or 0)))
        discount_amt = round(subtotal * discount_pct / 100.0, 2)
        final_amount = round(subtotal - discount_amt, 2)

    # Pro-rata per-line discount allocation → compute line_total for each row
    for row in resolved:
        sp = row["service_price"]
        share = (discount_amt * (sp / subtotal)) if subtotal > 0 else 0.0
        row["discount_amount"] = round(share, 2)
        row["line_total"] = round(sp - share, 2)

    # --- 5. Persist
    update_doc: Dict[str, Any] = {
        "barber_id": new_main_barber_id,
        "barber_name": main_barber.get("name", ""),
        "selected_services": new_services_list,
        "service_assignments": resolved,
        "subtotal": subtotal,
        "order_discount_percent": round(discount_pct, 2),
        "order_discount_amount": round(discount_amt, 2),
        "total_amount": final_amount,
    }
    if payload.payment_mode is not None:
        update_doc["payment_mode"] = payload.payment_mode
    if payload.payment_confirmed is not None:
        update_doc["payment_confirmed"] = bool(payload.payment_confirmed)
        if payload.payment_confirmed:
            update_doc["payment_status"] = "paid"

    await db.tokens.update_one({"id": token_id}, {"$set": update_doc})

    # --- 6. Recompute incentive payouts for ALL barbers touched (new + old)
    touched_barber_ids = set(line_barber_ids)
    old_assignments = token.get("service_assignments") or []
    for a in old_assignments:
        if a.get("barber_id"):
            touched_barber_ids.add(a["barber_id"])
    if token.get("barber_id") and token.get("barber_id") != "any":
        touched_barber_ids.add(token["barber_id"])

    ym = (token.get("date") or "")[:7] or datetime.now(timezone.utc).strftime("%Y-%m")
    for bid in touched_barber_ids:
        try:
            await _recompute_incentive_payout(salon_id, bid, ym)
        except Exception as ex:
            logging.warning("Incentive recompute failed for barber=%s month=%s: %s", bid, ym, ex)

    updated = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    await broadcast_update("token_updated", updated)

    return {
        "message": "Booking updated",
        "token": TokenModel(**updated),
    }


@api_router.get("/salons/{salon_id}/customers/{phone}/wallet")
async def get_customer_wallet(salon_id: str, phone: str):
    """Return customer's active wallet balance with this salon.

    Wallet balance comes from two sources:
    1. Membership wallet (if customer has active membership)
    2. Loyalty wallet (independent of membership, from loyalty rewards)
    
    Returns combined balance for customer's usage.
    """
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Check for membership wallet
    membership = await db.customer_memberships.find_one(
        {"salon_id": salon_id, "customer_phone": phone, "is_active": True, "payment_confirmed": True},
        {"_id": 0, "id": 1, "wallet_balance": 1, "membership_name": 1, "tier": 1, "color": 1, "expiry_date": 1},
    )
    
    # Check for loyalty wallet (independent of membership)
    loyalty_wallet = await db.customer_wallets.find_one(
        {"salon_id": salon_id, "customer_phone": phone},
        {"_id": 0, "id": 1, "wallet_balance": 1}
    )
    
    membership_balance = float(membership.get("wallet_balance", 0)) if membership else 0.0
    loyalty_balance = float(loyalty_wallet.get("wallet_balance", 0)) if loyalty_wallet else 0.0
    total_balance = membership_balance + loyalty_balance
    
    if not membership and not loyalty_wallet:
        return {
            "has_membership": False,
            "has_loyalty_wallet": False,
            "wallet_balance": 0.0,
            "membership_balance": 0.0,
            "loyalty_balance": 0.0,
            "membership_id": None,
            "loyalty_wallet_id": None,
            "membership_name": None,
            "tier": None,
            "color": None,
            "expiry_date": None,
        }
    
    return {
        "has_membership": bool(membership),
        "has_loyalty_wallet": bool(loyalty_wallet),
        "wallet_balance": total_balance,
        "membership_balance": membership_balance,
        "loyalty_balance": loyalty_balance,
        "membership_id": membership.get("id") if membership else None,
        "loyalty_wallet_id": loyalty_wallet.get("id") if loyalty_wallet else None,
        "membership_name": membership.get("membership_name") if membership else None,
        "tier": membership.get("tier") if membership else None,
        "color": membership.get("color") if membership else None,
        "expiry_date": membership.get("expiry_date") if membership else None,
    }


@api_router.post("/tokens/{token_id}/confirm-payment")
async def confirm_token_payment(token_id: str, body: dict, current_salon=Depends(get_current_salon)):
    """Salon confirms payment for a token. Supports full wallet, full cash/upi, or split.

    Body shape:
      - payment_mode: "cash" | "upi" | "wallet" | "split" | "card"
      - wallet_amount (float, optional): portion paid from wallet (for "wallet" or "split")
      - cash_amount (float, optional): portion paid in cash (for "split")
      - upi_amount (float, optional): portion paid via UPI (for "split")

    Rules:
      - For "wallet", the customer's wallet_balance must be >= total_amount, else 400.
      - For "split", wallet_amount must be <= wallet_balance AND
        wallet_amount + cash_amount + upi_amount must equal total_amount (±1 tolerance).
      - Wallet deduction is recorded on `customer_memberships.wallet_balance` and logged
        in `wallet_transactions`. Cash/UPI portions are logged to `financial_transactions`.
    """
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    if token.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Token is already completed")

    already_confirmed = token.get("payment_confirmed", False)
    total_amount = float(token.get("total_amount", 0))
    payment_mode = (body.get("payment_mode") or token.get("payment_mode") or "cash").lower()

    wallet_amount = float(body.get("wallet_amount") or 0)
    cash_amount = float(body.get("cash_amount") or 0)
    upi_amount = float(body.get("upi_amount") or 0)

    # Backfill amounts from legacy body shapes when only payment_mode is provided
    if payment_mode == "wallet" and wallet_amount <= 0:
        wallet_amount = total_amount
    elif payment_mode == "cash" and cash_amount <= 0 and wallet_amount <= 0 and upi_amount <= 0:
        cash_amount = total_amount
    elif payment_mode == "upi" and upi_amount <= 0 and wallet_amount <= 0 and cash_amount <= 0:
        upi_amount = total_amount
    elif payment_mode == "card" and cash_amount <= 0 and upi_amount <= 0 and wallet_amount <= 0:
        # record card portion as cash bucket for now
        cash_amount = total_amount

    # Fetch wallet (active & confirmed membership only)
    salon_id = token.get("salon_id", "")
    phone = token.get("phone") or ""
    wallet_doc = None
    if phone:
        wallet_doc = await db.customer_memberships.find_one(
            {"salon_id": salon_id, "customer_phone": phone, "is_active": True, "payment_confirmed": True},
            {"_id": 0, "id": 1, "wallet_balance": 1, "membership_name": 1},
        )
    wallet_balance = float(wallet_doc["wallet_balance"]) if wallet_doc else 0.0

    # Validation
    if wallet_amount < 0 or cash_amount < 0 or upi_amount < 0:
        raise HTTPException(status_code=400, detail="Payment amounts cannot be negative")

    if wallet_amount > 0 and not wallet_doc:
        raise HTTPException(
            status_code=400,
            detail="Customer has no active wallet. Select cash or UPI instead.",
        )
    if wallet_amount > wallet_balance + 0.001:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Insufficient wallet balance. Available: ₹{wallet_balance:.2f}, "
                f"attempted to use: ₹{wallet_amount:.2f}. Please pay the remainder via cash/UPI."
            ),
        )

    total_paid = wallet_amount + cash_amount + upi_amount
    if abs(total_paid - total_amount) > 1.0:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Payment total ₹{total_paid:.2f} does not match bill ₹{total_amount:.2f}. "
                f"Please adjust amounts."
            ),
        )

    # Determine effective payment_mode label
    modes_used = []
    if wallet_amount > 0:
        modes_used.append("wallet")
    if cash_amount > 0:
        modes_used.append("cash")
    if upi_amount > 0:
        modes_used.append("upi")
    if len(modes_used) == 1:
        effective_mode = modes_used[0]
    elif len(modes_used) > 1:
        effective_mode = "split"
    else:
        effective_mode = payment_mode or "cash"

    # Apply wallet deduction (first-time only)
    new_wallet_balance = wallet_balance
    if wallet_amount > 0 and not already_confirmed:
        new_wallet_balance = wallet_balance - wallet_amount
        await db.customer_memberships.update_one(
            {"id": wallet_doc["id"]},
            {"$set": {"wallet_balance": new_wallet_balance}},
        )
        await db.wallet_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "customer_phone": phone,
            "salon_id": salon_id,
            "transaction_type": "debit",
            "amount": wallet_amount,
            "balance_after": new_wallet_balance,
            "description": f"Service payment: Token {token.get('token_number','')} - {token.get('customer_name','')}",
            "reference_type": "token",
            "reference_id": token_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    # Update token
    update_data = {
        "payment_confirmed": True,
        "payment_status": "paid",
        "payment_mode": effective_mode,
        "payment_breakdown": {
            "wallet_amount": wallet_amount,
            "cash_amount": cash_amount,
            "upi_amount": upi_amount,
        },
    }
    await db.tokens.update_one({"id": token_id}, {"$set": update_data})
    updated_token = await db.tokens.find_one({"id": token_id}, {"_id": 0})

    # Customer notification
    if phone:
        breakdown_bits = []
        if wallet_amount > 0:
            breakdown_bits.append(f"₹{wallet_amount:.0f} wallet")
        if cash_amount > 0:
            breakdown_bits.append(f"₹{cash_amount:.0f} cash")
        if upi_amount > 0:
            breakdown_bits.append(f"₹{upi_amount:.0f} UPI")
        breakdown_str = " + ".join(breakdown_bits) if breakdown_bits else effective_mode
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "user_type": "customer",
            "user_id": phone,
            "salon_id": salon_id,
            "title": "Payment Confirmed",
            "message": f"Payment of ₹{total_amount:.0f} for token {token.get('token_number','')} confirmed ({breakdown_str}).",
            "type": "payment_confirmed",
            "related_id": token_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })

    await broadcast_update("token_updated", updated_token)

    # Financial transactions — log each non-wallet portion
    if not already_confirmed:
        for mode_name, amt in (("cash", cash_amount), ("upi", upi_amount)):
            if amt > 0:
                await db.financial_transactions.insert_one({
                    "id": str(uuid.uuid4()),
                    "salon_id": salon_id,
                    "branch_id": token.get("branch_id"),
                    "type": "inflow",
                    "category": "booking_payment",
                    "amount": float(amt),
                    "payment_mode": mode_name,
                    "narration": f"Booking {token.get('token_number','')} - {token.get('customer_name','')}"
                                + (" (split)" if effective_mode == "split" else ""),
                    "reference_id": token_id,
                    "reference_type": "token",
                    "created_by": "system",
                    "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
    else:
        # Just update the payment_mode on existing financial transactions for this token
        await db.financial_transactions.update_many(
            {"reference_id": token_id, "reference_type": "token"},
            {"$set": {"payment_mode": effective_mode}},
        )
    
    return {
        "message": "Payment confirmed successfully",
        "token": TokenModel(**updated_token),
        "payment": {
            "mode": effective_mode,
            "wallet_amount": wallet_amount,
            "cash_amount": cash_amount,
            "upi_amount": upi_amount,
            "wallet_balance_after": new_wallet_balance,
        },
    }

@api_router.put("/tokens/{token_id}/change-barber")
async def change_token_barber(token_id: str, body: dict, current_salon=Depends(get_current_salon)):
    """Change the barber assigned to a token and recalculate total + capacity.

    Validations:
      - Target barber must have capacity for this token in the same shift (75% rule).
    Side effects:
      - Recalculates total_amount using new barber's pricing.
      - Sends customer notification: "Your service will now be provided by Barber X".
    """
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    if token.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Cannot change barber for completed booking")
    
    new_barber_id = body.get("barber_id")
    if not new_barber_id:
        raise HTTPException(status_code=400, detail="barber_id is required")

    old_barber_id = token.get("barber_id")

    # Resolve new barber (supports auto-assign via "any")
    if new_barber_id == "any":
        # Use fastest-barber logic, passing this token's blocked minutes
        required_blocked = int(token.get("blocked_minutes") or 0)
        if required_blocked <= 0:
            required_blocked = calc_blocked_minutes_from_total(
                await calc_service_total_minutes(token.get("selected_services", []))
            )
        chosen = await pick_fastest_barber(
            salon_id=token["salon_id"],
            date=token["date"],
            shift=token.get("shift", "Morning"),
            required_blocked_minutes=required_blocked,
        )
        if not chosen:
            raise HTTPException(status_code=400, detail="No barber has capacity for this shift.")
        new_barber_id = chosen["id"]
        new_barber_name = chosen["name"]
    else:
        barber = await db.barbers.find_one({"id": new_barber_id}, {"_id": 0})
        if not barber:
            raise HTTPException(status_code=404, detail="Barber not found")
        new_barber_name = barber.get("name", "Unknown")

        # Capacity check (skip if same barber as before)
        if new_barber_id != old_barber_id:
            required_blocked = int(token.get("blocked_minutes") or 0)
            if required_blocked <= 0:
                required_blocked = calc_blocked_minutes_from_total(
                    await calc_service_total_minutes(token.get("selected_services", []))
                )
            fit = await can_fit_in_barber_shift(
                salon_id=token["salon_id"],
                barber_id=new_barber_id,
                date=token["date"],
                shift=token.get("shift", "Morning"),
                new_blocked_minutes=required_blocked,
            )
            if not fit["allowed"]:
                raise HTTPException(
                    status_code=400,
                    detail=f"{new_barber_name}'s shift is full. Please choose another barber."
                )

    # Recalculate total amount based on new barber's pricing
    selected_services = token.get("selected_services", [])
    total_amount = await calculate_booking_total(selected_services, new_barber_id)
    if total_amount == 0:
        for service_id in selected_services:
            service = await db.services.find_one({"id": service_id}, {"_id": 0})
            if service:
                total_amount += service.get("base_price", 0)

    # Update token — also RESET near-notification flags so the new barber's queue re-triggers them
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {
            "barber_id": new_barber_id,
            "barber_name": new_barber_name,
            "total_amount": total_amount,
            "notified_3_away": False,
            "notified_2_away": False,
            "notified_1_away": False,
        }}
    )

    # Get updated token
    updated_token = await db.tokens.find_one({"id": token_id}, {"_id": 0})

    await broadcast_update("token_updated", updated_token)

    # Notify customer of barber change
    try:
        if updated_token.get("phone") and new_barber_id != old_barber_id:
            await create_in_app_notification(
                user_type="customer",
                user_id=updated_token["phone"],
                title="Barber Changed",
                message=f"Your service will now be provided by Barber {new_barber_name}.",
                notification_type="barber_changed",
                setting_key="booking_status_change",
                salon_id=updated_token.get("salon_id", ""),
                related_id=token_id,
            )
    except Exception as e:
        logger.error(f"Failed to send barber-change notification: {e}")

    return {
        "message": f"Barber changed to {new_barber_name}. Total recalculated: ₹{total_amount}",
        "token": TokenModel(**updated_token)
    }

# ============ MEMBERSHIP PAYMENT CONFIRMATION ============

@api_router.post("/salons/{salon_id}/memberships/{membership_id}/confirm-payment")
async def confirm_membership_payment(salon_id: str, membership_id: str, body: dict = None, current_user=Depends(get_current_salon_user)):
    if body is None:
        body = {}
    """Salon confirms payment for a customer-purchased membership. Credits wallet after confirmation."""
    membership = await db.customer_memberships.find_one({
        "id": membership_id,
        "salon_id": salon_id
    }, {"_id": 0})
    
    if not membership:
        raise HTTPException(status_code=404, detail="Membership not found")
    
    if membership.get("payment_confirmed", True):
        raise HTTPException(status_code=400, detail="Payment is already confirmed")
    
    # Get the plan to know credit amount
    plan = await db.membership_plans.find_one({"id": membership["membership_plan_id"]}, {"_id": 0})
    credit_amount = plan["credit"] if plan else membership.get("credit_added", 0)
    
    # Optionally change payment mode
    payment_mode = body.get("payment_mode", membership.get("payment_mode"))
    
    # Check for existing active membership to add credit to
    phone = membership["customer_phone"]
    existing = await db.customer_memberships.find_one({
        "salon_id": salon_id,
        "customer_phone": phone,
        "is_active": True,
        "payment_confirmed": True,
        "id": {"$ne": membership_id}
    }, {"_id": 0})
    
    if existing:
        # Add credit to existing confirmed membership wallet
        new_balance = existing["wallet_balance"] + credit_amount
        await db.customer_memberships.update_one(
            {"id": existing["id"]},
            {"$set": {"wallet_balance": new_balance, "expiry_date": membership.get("expiry_date")}}
        )
        # Mark this one as merged/confirmed
        await db.customer_memberships.update_one(
            {"id": membership_id},
            {"$set": {
                "payment_confirmed": True,
                "payment_mode": payment_mode,
                "wallet_balance": 0,
                "is_active": False  # merged into existing
            }}
        )
        wallet_balance_after = new_balance
    else:
        # Confirm and activate this membership with credit
        await db.customer_memberships.update_one(
            {"id": membership_id},
            {"$set": {
                "payment_confirmed": True,
                "payment_mode": payment_mode,
                "wallet_balance": credit_amount
            }}
        )
        wallet_balance_after = credit_amount
    
    # Record wallet transaction
    await db.wallet_transactions.insert_one({
        "id": str(uuid.uuid4()),
        "customer_phone": phone,
        "salon_id": salon_id,
        "transaction_type": "credit",
        "amount": credit_amount,
        "balance_after": wallet_balance_after,
        "description": f"Membership confirmed: {membership.get('membership_name', 'N/A')}",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Create notification for customer
    await db.notifications.insert_one({
        "id": str(uuid.uuid4()),
        "user_type": "customer",
        "user_id": phone,
        "salon_id": salon_id,
        "title": "Membership Payment Confirmed",
        "message": f"Your {membership.get('membership_name', '')} membership payment has been confirmed. ₹{credit_amount} has been added to your wallet.",
        "type": "membership_confirmed",
        "related_id": membership_id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Record financial transaction for membership (cash/upi/card impact cash flow)
    if payment_mode != "wallet":
        await db.financial_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "salon_id": salon_id,
            "type": "inflow",
            "category": "membership_payment",
            "amount": float(membership.get("paid_amount", 0)),
            "payment_mode": payment_mode,
            "narration": f"Membership: {membership.get('membership_name', '')} - {membership.get('customer_name', '')}",
            "reference_id": membership_id,
            "reference_type": "membership",
            "created_by": "system",
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    return {
        "message": "Membership payment confirmed and wallet credited",
        "credit_added": credit_amount,
        "wallet_balance": wallet_balance_after
    }


# ============ FINANCIALS SYSTEM ============

class FinancialTransactionCreate(BaseModel):
    type: str  # inflow, outflow, withdrawal, deposit, adjustment
    category: str  # salary, staff_refreshment, consumables, utilities, rent, maintenance, products, custom, withdrawal, deposit, adjustment
    amount: float
    payment_mode: str = "cash"  # cash, upi, card, wallet
    narration: Optional[str] = ""
    date: Optional[str] = None  # YYYY-MM-DD, defaults to today
    branch_id: Optional[str] = None  # Defaults to main branch

@api_router.get("/salons/{salon_id}/financials/settings")
async def get_financial_settings(salon_id: str, current_user=Depends(get_current_salon_user)):
    """Get financial settings (opening balance, etc.)"""
    settings = await db.financial_settings.find_one({"salon_id": salon_id}, {"_id": 0})
    if not settings:
        return {"salon_id": salon_id, "opening_balance": 0, "opening_balance_date": datetime.now(timezone.utc).strftime("%Y-%m-%d")}
    return settings

@api_router.put("/salons/{salon_id}/financials/settings")
async def update_financial_settings(salon_id: str, body: dict, current_user=Depends(get_current_salon_user)):
    """Update financial settings (opening balance)"""
    if not has_module_permission(current_user, "financials", "edit_transaction"):
        raise HTTPException(status_code=403, detail="Permission denied: financials.edit_transaction")
    opening_balance = body.get("opening_balance", 0)
    opening_balance_date = body.get("opening_balance_date", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    
    await db.financial_settings.update_one(
        {"salon_id": salon_id},
        {"$set": {
            "salon_id": salon_id,
            "opening_balance": float(opening_balance),
            "opening_balance_date": opening_balance_date,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"message": "Financial settings updated", "opening_balance": opening_balance}

@api_router.get("/salons/{salon_id}/financials/transactions")
async def get_financial_transactions(
    salon_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    type: Optional[str] = None,
    category: Optional[str] = None,
    payment_mode: Optional[str] = None,
    branch_id: Optional[str] = None,
    limit: int = 200,
    current_user=Depends(get_current_salon_user)
):
    """Get financial transactions with filters"""
    if not has_module_permission(current_user, "financials", "view_transactions"):
        raise HTTPException(status_code=403, detail="Permission denied: financials.view_transactions")
    if is_branch_manager(current_user):
        branch_id = enforce_branch_for_manager(current_user, branch_id)
    query = {"salon_id": salon_id}
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    if type:
        query["type"] = type
    if category:
        query["category"] = category
    if payment_mode:
        query["payment_mode"] = payment_mode
    if branch_id:
        query["branch_id"] = branch_id
    
    transactions = await db.financial_transactions.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    
    return {"transactions": transactions}

@api_router.post("/salons/{salon_id}/financials/transactions")
async def create_financial_transaction(salon_id: str, txn: FinancialTransactionCreate, current_user=Depends(get_current_salon_user)):
    """Create a manual financial transaction (expense, withdrawal, deposit, adjustment)"""
    if not has_module_permission(current_user, "financials", "create_transaction"):
        raise HTTPException(status_code=403, detail="Permission denied: financials.create_transaction")
    txn_date = txn.date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    txn_data = {
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "branch_id": await resolve_branch_id(salon_id, txn.branch_id),
        "type": txn.type,
        "category": txn.category,
        "amount": txn.amount if txn.type == "adjustment" else abs(txn.amount),
        "payment_mode": txn.payment_mode,
        "narration": txn.narration or "",
        "reference_id": "",
        "reference_type": "manual",
        "created_by": current_user.get("id", "admin"),
        "date": txn_date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.financial_transactions.insert_one(txn_data)
    
    # Remove MongoDB _id before returning
    txn_data.pop("_id", None)
    
    return {"message": "Transaction recorded", "transaction": txn_data}

@api_router.delete("/salons/{salon_id}/financials/transactions/{txn_id}")
async def delete_financial_transaction(salon_id: str, txn_id: str, current_user=Depends(get_current_salon_user)):
    """Delete a manual financial transaction (admin only)"""
    if not has_module_permission(current_user, "financials", "delete_transaction"):
        raise HTTPException(status_code=403, detail="Permission denied: financials.delete_transaction")
    txn = await db.financial_transactions.find_one({"id": txn_id, "salon_id": salon_id}, {"_id": 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Transaction not found")
    if txn.get("reference_type") != "manual":
        raise HTTPException(status_code=400, detail="Cannot delete system-generated transactions")
    
    await db.financial_transactions.delete_one({"id": txn_id})
    return {"message": "Transaction deleted"}

@api_router.get("/salons/{salon_id}/financials/dashboard")
async def get_financial_dashboard(
    salon_id: str,
    period: str = "daily",  # daily or monthly
    date: Optional[str] = None,  # YYYY-MM-DD for daily, YYYY-MM for monthly
    branch_id: Optional[str] = None,
    current_user=Depends(get_current_salon_user)
):
    """Get financial dashboard data with cash in/out summary"""
    if not has_module_permission(current_user, "financials", "view_dashboard"):
        raise HTTPException(status_code=403, detail="Permission denied: financials.view_dashboard")
    if is_branch_manager(current_user):
        branch_id = enforce_branch_for_manager(current_user, branch_id)
    today = datetime.now(timezone.utc)
    
    if period == "daily":
        target_date = date or today.strftime("%Y-%m-%d")
        query = {"salon_id": salon_id, "date": target_date}
    else:  # monthly
        target_month = date or today.strftime("%Y-%m")
        query = {"salon_id": salon_id, "date": {"$regex": f"^{target_month}"}}
    if branch_id:
        query["branch_id"] = branch_id
    
    transactions = await db.financial_transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Calculate summaries
    total_inflow = 0
    total_outflow = 0
    inflow_by_mode = {"cash": 0, "upi": 0, "card": 0, "wallet": 0}
    outflow_by_category = {}
    inflow_by_category = {}
    
    for txn in transactions:
        amount = txn.get("amount", 0)
        mode = txn.get("payment_mode", "cash")
        cat = txn.get("category", "other")
        txn_type = txn.get("type", "")
        
        if txn_type in ("inflow", "deposit"):
            total_inflow += amount
            inflow_by_mode[mode] = inflow_by_mode.get(mode, 0) + amount
            inflow_by_category[cat] = inflow_by_category.get(cat, 0) + amount
        elif txn_type in ("outflow", "withdrawal"):
            total_outflow += amount
            outflow_by_category[cat] = outflow_by_category.get(cat, 0) + amount
        elif txn_type == "adjustment":
            if amount >= 0:
                total_inflow += amount
            else:
                total_outflow += abs(amount)
    
    # Get opening balance
    settings = await db.financial_settings.find_one({"salon_id": salon_id}, {"_id": 0})
    opening_balance = settings.get("opening_balance", 0) if settings else 0
    
    # For daily view, calculate running balance from opening + all previous transactions
    if period == "daily":
        target = target_date
        prev_query = {"salon_id": salon_id, "date": {"$lt": target}}
        if branch_id:
            prev_query["branch_id"] = branch_id
        prev_txns = await db.financial_transactions.find(prev_query, {"_id": 0, "type": 1, "amount": 1}).to_list(10000)
        
        running_balance = opening_balance
        for pt in prev_txns:
            if pt["type"] in ("inflow", "deposit"):
                running_balance += pt["amount"]
            elif pt["type"] in ("outflow", "withdrawal"):
                running_balance -= pt["amount"]
        
        day_opening = running_balance
        day_closing = day_opening + total_inflow - total_outflow
    else:
        day_opening = opening_balance
        day_closing = opening_balance + total_inflow - total_outflow
    
    # Daily breakdown for monthly view
    daily_breakdown = []
    if period == "monthly":
        daily_data = {}
        for txn in transactions:
            d = txn.get("date", "")
            if d not in daily_data:
                daily_data[d] = {"date": d, "inflow": 0, "outflow": 0}
            if txn["type"] in ("inflow", "deposit"):
                daily_data[d]["inflow"] += txn["amount"]
            elif txn["type"] in ("outflow", "withdrawal"):
                daily_data[d]["outflow"] += txn["amount"]
        daily_breakdown = sorted(daily_data.values(), key=lambda x: x["date"])
    
    return {
        "period": period,
        "date": date or (today.strftime("%Y-%m-%d") if period == "daily" else today.strftime("%Y-%m")),
        "opening_balance": day_opening,
        "closing_balance": day_closing,
        "total_inflow": total_inflow,
        "total_outflow": total_outflow,
        "net": total_inflow - total_outflow,
        "inflow_by_mode": inflow_by_mode,
        "inflow_by_category": inflow_by_category,
        "outflow_by_category": outflow_by_category,
        "daily_breakdown": daily_breakdown,
        "transactions": transactions
    }

@api_router.get("/salons/{salon_id}/financials/report/csv")
async def download_financial_report_csv(
    salon_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Download financial report as CSV"""
    from fastapi.responses import StreamingResponse
    import io
    import csv
    
    query = {"salon_id": salon_id}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    transactions = await db.financial_transactions.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Time", "Type", "Category", "Amount", "Payment Mode", "Narration", "Reference"])
    
    for txn in transactions:
        # Extract time from created_at
        created_at = txn.get("created_at", "")
        time_str = ""
        if created_at:
            try:
                dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                time_str = dt.strftime("%H:%M:%S")
            except (ValueError, AttributeError):
                time_str = ""
        
        writer.writerow([
            txn.get("date", ""),
            time_str,
            txn.get("type", ""),
            txn.get("category", ""),
            txn.get("amount", 0),
            txn.get("payment_mode", ""),
            txn.get("narration", ""),
            txn.get("reference_type", "")
        ])
    
    output.seek(0)
    filename = f"financials_{salon_id}_{start_date or 'all'}_{end_date or 'all'}.csv"
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/salons/{salon_id}/today-sales")
async def get_today_sales(
    salon_id: str,
    branch_id: Optional[str] = None,
    current_user: Optional[dict] = Depends(get_current_salon_user_optional),
):
    """Get today's sales from completed tokens (analytics logic)"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Branch-manager scoping
    if current_user and is_branch_manager(current_user):
        branch_id = enforce_branch_for_manager(current_user, branch_id)
    
    # Find all completed tokens for today (matching analytics logic)
    query = {
        "salon_id": salon_id,
        "date": today,
        "status": "completed"
    }
    if branch_id:
        query["branch_id"] = branch_id
    completed_tokens = await db.tokens.find(query, {"_id": 0, "total_amount": 1}).to_list(1000)
    
    # Sum up total_amount from all completed tokens (same as analytics)
    total_sales = sum(token.get("total_amount", 0) for token in completed_tokens)
    
    return {"today_sales": total_sales}


# ============ NOTIFICATIONS SYSTEM ============

@api_router.get("/notifications/{user_type}/{user_id}")
async def get_notifications(user_type: str, user_id: str, limit: int = 50):
    """Get notifications for a user (customer or salon)"""
    # Normalize phone for customer
    if user_type == "customer" and not user_id.startswith("+91"):
        user_id = f"+91{user_id}"
    
    notifications = await db.notifications.find({
        "user_type": user_type,
        "user_id": user_id
    }, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    
    return {"notifications": notifications}

@api_router.get("/notifications/{user_type}/{user_id}/unread-count")
async def get_unread_notification_count(user_type: str, user_id: str):
    """Get unread notification count"""
    if user_type == "customer" and not user_id.startswith("+91"):
        user_id = f"+91{user_id}"
    
    count = await db.notifications.count_documents({
        "user_type": user_type,
        "user_id": user_id,
        "is_read": False
    })
    
    return {"unread_count": count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str):
    """Mark a notification as read"""
    result = await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"is_read": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/{user_type}/{user_id}/read-all")
async def mark_all_notifications_read(user_type: str, user_id: str):
    """Mark all notifications as read for a user"""
    if user_type == "customer" and not user_id.startswith("+91"):
        user_id = f"+91{user_id}"
    
    await db.notifications.update_many(
        {"user_type": user_type, "user_id": user_id, "is_read": False},
        {"$set": {"is_read": True}}
    )
    
    return {"message": "All notifications marked as read"}


# ============ NOTIFICATION SETTINGS ============

@api_router.get("/salons/{salon_id}/notification-settings")
async def get_salon_notif_settings(salon_id: str):
    """Get salon notification preferences (defaults all True if not set)."""
    settings = await get_salon_notification_settings(salon_id)
    return settings


@api_router.put("/salons/{salon_id}/notification-settings")
async def update_salon_notif_settings(
    salon_id: str,
    body: dict,
    current_user=Depends(get_current_salon_user),
):
    """Update salon notification preferences."""
    if not has_module_permission(current_user, "salon_settings", "edit_notifications"):
        raise HTTPException(status_code=403, detail="Permission denied: salon_settings.edit_notifications")
    update_doc = {k: bool(v) for k, v in body.items() if k in DEFAULT_SALON_NOTIFICATION_SETTINGS}
    update_doc["salon_id"] = salon_id
    update_doc["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.salon_notification_settings.update_one(
        {"salon_id": salon_id},
        {"$set": update_doc},
        upsert=True,
    )
    settings = await get_salon_notification_settings(salon_id)
    return settings


@api_router.get("/customers/{phone}/notification-settings")
async def get_customer_notif_settings(phone: str):
    """Get customer notification preferences (defaults all True)."""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    settings = await get_customer_notification_settings(phone)
    return settings


@api_router.put("/customers/{phone}/notification-settings")
async def update_customer_notif_settings(phone: str, body: dict):
    """Update customer notification preferences (in-app + WhatsApp toggles)."""
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    update_doc = {k: bool(v) for k, v in body.items() if k in DEFAULT_CUSTOMER_NOTIFICATION_SETTINGS}
    update_doc["phone"] = phone
    update_doc["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.customer_notification_settings.update_one(
        {"phone": phone},
        {"$set": update_doc},
        upsert=True,
    )
    settings = await get_customer_notification_settings(phone)
    return settings


# ============ CANCEL VIA WHATSAPP LINK ============

@api_router.get("/tokens/{token_id}/cancel-link", response_class=HTMLResponse)
async def cancel_token_via_link(token_id: str):
    """
    GET handler: renders a "Are you sure?" confirmation page.
    Actual cancellation happens on POST to /cancel-link (see below) which the
    form on this page submits.
    """
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        return HTMLResponse(content=_render_cancel_page(success=False, message="Booking not found."), status_code=404)

    current_status = token.get("status")
    if current_status == "cancelled":
        return HTMLResponse(content=_render_cancel_page(
            success=True,
            message=f"Your booking (Token #{token.get('token_number','')}) is already cancelled."
        ))
    if current_status in ("completed", "in_service", "called", "in_progress"):
        return HTMLResponse(content=_render_cancel_page(
            success=False,
            message=f"This booking cannot be cancelled (status: {current_status})."
        ), status_code=400)

    # Fetch salon name for a nicer confirmation UX
    salon = await db.salons.find_one({"id": token.get("salon_id")}, {"_id": 0, "salon_name": 1}) or {}
    salon_name = salon.get("salon_name", "SalonHub")

    return HTMLResponse(content=_render_cancel_confirm_page(
        token=token,
        salon_name=salon_name,
    ))


@api_router.post("/tokens/{token_id}/cancel-link", response_class=HTMLResponse)
async def cancel_token_via_link_confirm(token_id: str):
    """
    POST handler: actually cancels the booking after the customer confirms on
    the HTML interstitial. Returns result HTML page.
    """
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        return HTMLResponse(content=_render_cancel_page(success=False, message="Booking not found."), status_code=404)

    current_status = token.get("status")
    if current_status == "cancelled":
        return HTMLResponse(content=_render_cancel_page(
            success=True,
            message=f"Your booking (Token #{token.get('token_number','')}) is already cancelled."
        ))
    if current_status in ("completed", "in_service", "called", "in_progress"):
        return HTMLResponse(content=_render_cancel_page(
            success=False,
            message=f"This booking cannot be cancelled (status: {current_status})."
        ), status_code=400)

    # Refund wallet if paid via wallet
    if token.get("payment_mode") == "wallet" and token.get("payment_status") == "paid":
        wallet_phone = token.get("phone", "")
        if wallet_phone and not wallet_phone.startswith("+91"):
            wallet_phone = f"+91{wallet_phone}"
        membership = await db.customer_memberships.find_one({
            "salon_id": token.get("salon_id"),
            "customer_phone": wallet_phone,
            "is_active": True,
        }, {"_id": 0})
        if membership:
            refund_amount = token.get("total_amount", 0) or 0
            new_balance = (membership.get("wallet_balance", 0) or 0) + refund_amount
            await db.customer_memberships.update_one(
                {"id": membership["id"]},
                {"$set": {"wallet_balance": new_balance}},
            )
            await db.wallet_transactions.insert_one({
                "id": str(uuid.uuid4()),
                "customer_phone": wallet_phone,
                "salon_id": token.get("salon_id"),
                "transaction_type": "credit",
                "amount": refund_amount,
                "balance_after": new_balance,
                "description": f"Refund - Cancelled via WhatsApp (Token {token.get('token_number','')})",
                "created_at": datetime.now(timezone.utc).isoformat(),
            })

    # Cancel the booking
    await db.tokens.update_one({"id": token_id}, {"$set": {"status": "cancelled"}})
    await broadcast_update("token_cancelled", {"token_id": token_id})

    # Notify customer (in-app + whatsapp)
    if token.get("phone"):
        await create_in_app_notification(
            user_type="customer",
            user_id=token["phone"],
            title="Booking Cancelled",
            message=f"Your booking (Token #{token.get('token_number','')}) has been cancelled via WhatsApp.",
            notification_type="booking_cancelled",
            setting_key="booking_status_change",
            salon_id=token.get("salon_id", ""),
            related_id=token_id,
        )
    await send_booking_notification(token, "token_cancelled")

    # Notify salon (in-app)
    if token.get("salon_id"):
        await create_in_app_notification(
            user_type="salon",
            user_id=token["salon_id"],
            title="Booking Cancelled by Customer",
            message=f"{token.get('customer_name','Customer')} cancelled booking (Token #{token.get('token_number','')}) via WhatsApp.",
            notification_type="booking_cancelled",
            setting_key="booking_change",
            salon_id=token["salon_id"],
            related_id=token_id,
        )

    return HTMLResponse(content=_render_cancel_page(
        success=True,
        message=f"Your booking (Token #{token.get('token_number','')}) has been cancelled successfully."
    ))


def _render_cancel_page(success: bool, message: str) -> str:
    icon = "✅" if success else "❌"
    color = "#16a34a" if success else "#dc2626"
    title = "Cancellation Confirmed" if success else "Cancellation Failed"
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
  body {{ margin:0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
         background: linear-gradient(135deg,#1f2937 0%, #111827 100%); min-height: 100vh;
         display:flex; align-items:center; justify-content:center; color:#fff; padding:20px; }}
  .card {{ background:#1f2937; border:1px solid rgba(184,134,11,0.3); border-radius:16px;
         padding:36px 28px; max-width:420px; width:100%; box-shadow:0 20px 50px rgba(0,0,0,0.4);
         text-align:center; }}
  .icon {{ font-size:64px; line-height:1; margin-bottom:12px; }}
  h1 {{ color:{color}; margin:0 0 12px 0; font-size:22px; }}
  p {{ color:#cbd5e1; margin:8px 0 24px; line-height:1.5; }}
  .btn {{ display:inline-block; padding:12px 24px; border-radius:8px;
         background:#b8860b; color:#000; text-decoration:none; font-weight:700; }}
  .small {{ color:#94a3b8; font-size:12px; margin-top:18px; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">{icon}</div>
  <h1>{title}</h1>
  <p>{message}</p>
  <a class="btn" href="/salons">Book Another Appointment</a>
  <div class="small">SalonHub · You can close this tab.</div>
</div>
</body>
</html>"""


def _render_cancel_confirm_page(token: dict, salon_name: str) -> str:
    """Interstitial "Are you sure?" page shown when the customer clicks the
    WhatsApp cancel link. Submits a POST to the same URL to actually cancel."""
    import html as _html
    token_id = token.get("id", "")
    token_number = token.get("token_number", "")
    customer_name = _html.escape(token.get("customer_name", "Customer"))
    date = _html.escape(token.get("date", ""))
    shift = _html.escape(str(token.get("shift", "") or "").title())
    services_count = len(token.get("selected_services", []) or [])
    total_amount = token.get("total_amount", 0) or 0
    salon_name_html = _html.escape(salon_name)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Cancel Booking?</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin:0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
         background: linear-gradient(135deg,#1f2937 0%, #111827 100%); min-height: 100vh;
         display:flex; align-items:center; justify-content:center; color:#fff; padding:20px; }}
  .card {{ background:#1f2937; border:1px solid rgba(184,134,11,0.3); border-radius:16px;
         padding:32px 24px; max-width:440px; width:100%; box-shadow:0 20px 50px rgba(0,0,0,0.4);
         text-align:center; }}
  .icon {{ font-size:56px; line-height:1; margin-bottom:8px; }}
  h1 {{ color:#f59e0b; margin:0 0 8px; font-size:22px; }}
  .sub {{ color:#94a3b8; margin:0 0 20px; font-size:13px; }}
  .details {{ background:#0f172a; border:1px solid rgba(184,134,11,0.2); border-radius:12px;
         padding:16px; margin: 16px 0 20px; text-align:left; }}
  .row {{ display:flex; justify-content:space-between; padding:6px 0; font-size:14px; }}
  .row .k {{ color:#94a3b8; }}
  .row .v {{ color:#f1f5f9; font-weight:600; }}
  .actions {{ display:flex; gap:10px; margin-top:16px; }}
  .btn {{ flex:1; padding:14px 16px; border-radius:10px; border:none; font-weight:700;
         font-size:15px; cursor:pointer; text-decoration:none; display:inline-flex;
         align-items:center; justify-content:center; }}
  .btn-no {{ background:#334155; color:#e2e8f0; }}
  .btn-no:hover {{ background:#475569; }}
  .btn-yes {{ background:#dc2626; color:#fff; }}
  .btn-yes:hover {{ background:#b91c1c; }}
  .small {{ color:#64748b; font-size:12px; margin-top:18px; }}
</style>
</head>
<body>
<div class="card">
  <div class="icon">⚠️</div>
  <h1>Cancel this booking?</h1>
  <p class="sub">Please confirm — this action cannot be undone.</p>
  <div class="details">
    <div class="row"><span class="k">Salon</span><span class="v">{salon_name_html}</span></div>
    <div class="row"><span class="k">Token</span><span class="v">#{token_number}</span></div>
    <div class="row"><span class="k">Customer</span><span class="v">{customer_name}</span></div>
    <div class="row"><span class="k">Date</span><span class="v">{date}</span></div>
    <div class="row"><span class="k">Shift</span><span class="v">{shift or '—'}</span></div>
    <div class="row"><span class="k">Services</span><span class="v">{services_count}</span></div>
    <div class="row"><span class="k">Amount</span><span class="v">₹{total_amount}</span></div>
  </div>
  <form method="POST" action="/api/tokens/{token_id}/cancel-link" style="margin:0;">
    <div class="actions">
      <a href="/salons" class="btn btn-no">No, Keep Booking</a>
      <button type="submit" class="btn btn-yes">Yes, Cancel</button>
    </div>
  </form>
  <div class="small">SalonHub · Safe to close if you change your mind.</div>
</div>
</body>
</html>"""



@api_router.get("/users/last-salon")
async def get_last_salon_by_phone(phone: str):
    """Get last visited salon by user phone number"""
    # Normalize phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    # Find most recent booking for this phone
    recent_booking = await db.tokens.find_one(
        {"phone": phone},
        {"_id": 0, "salon_id": 1},
        sort=[("created_at", -1)]
    )
    
    if not recent_booking:
        return {"salon": None}
    
    # Get salon details
    salon = await db.salons.find_one(
        {"id": recent_booking["salon_id"]},
        {"_id": 0}
    )
    
    if salon:
        return {"salon": Salon(**salon)}
    
    return {"salon": None}

@api_router.get("/users/{user_id}/recent-services")
async def get_user_recent_services(user_id: str):
    """Get user's recent services from last 5 bookings"""
    # Find last 5 bookings for this user
    recent_bookings = await db.tokens.find(
        {"user_id": user_id},
        {"_id": 0, "selected_services": 1}
    ).sort("created_at", -1).limit(5).to_list(5)
    
    # Collect all unique service IDs
    service_ids = set()
    for booking in recent_bookings:
        service_ids.update(booking.get("selected_services", []))
    
    # Get service details
    if service_ids:
        services = await db.services.find(
            {"id": {"$in": list(service_ids)}, "is_active": True},
            {"_id": 0}
        ).to_list(100)
        return {"services": [Service(**s) for s in services]}
    
    return {"services": []}

@api_router.get("/salons/{salon_id}/token-status")
async def get_salon_token_status(salon_id: str, shift: Optional[str] = None, date: Optional[str] = None, branch_id: Optional[str] = None):
    """Get current token status for salon (overall and per barber)"""
    today = date or datetime.now().date().isoformat()
    
    # Get all barbers for this salon
    barber_query = {"salon_id": salon_id, "is_active": True}
    if branch_id:
        barber_query["branch_id"] = branch_id
    barbers = await db.barbers.find(
        barber_query,
        {"_id": 0}
    ).to_list(100)
    
    result = {
        "date": today,
        "overall": {},
        "barbers": []
    }
    
    # Get overall salon status
    query = {"salon_id": salon_id, "date": today, "status": {"$in": ["waiting", "called"]}}
    if shift:
        query["shift"] = shift
    if branch_id:
        query["branch_id"] = branch_id
    
    waiting_tokens = await db.tokens.find(query, {"_id": 0}).to_list(1000)
    
    # Get currently called token
    called_query = {"salon_id": salon_id, "date": today, "status": "called"}
    if shift:
        called_query["shift"] = shift
    if branch_id:
        called_query["branch_id"] = branch_id
    
    called_token = await db.tokens.find_one(called_query, {"_id": 0}, sort=[("called_at", -1)])
    
    result["overall"] = {
        "current_token": called_token.get("token_number") if called_token else None,
        "waiting_count": len(waiting_tokens),
        "shift": shift
    }
    
    # Get per-barber status
    for barber in barbers:
        barber_query = {
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "date": today
        }
        if shift:
            barber_query["shift"] = shift
        
        barber_waiting = await db.tokens.find(
            {**barber_query, "status": {"$in": ["waiting", "called"]}},
            {"_id": 0}
        ).to_list(1000)
        
        barber_called = await db.tokens.find_one(
            {**barber_query, "status": {"$in": ["called", "in_progress", "in_service"]}},
            {"_id": 0},
            sort=[("called_at", -1)]
        )
        
        # Count total tokens for today (all statuses)
        total_tokens_query = {
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "date": today
        }
        total_tokens_today = await db.tokens.count_documents(total_tokens_query)

        # Compute approx finish time for the currently-serving token using 75% rule residual
        approx_finish_minutes = None
        currently_serving_payload = None
        if barber_called:
            try:
                remaining = None
                if barber_called.get("blocked_minutes") is not None:
                    remaining = int(barber_called["blocked_minutes"])
                else:
                    mins = await calc_service_total_minutes(barber_called.get("selected_services", []))
                    remaining = calc_blocked_minutes_from_total(mins)
                # Subtract elapsed since called_at
                if barber_called.get("called_at"):
                    try:
                        called_dt = datetime.fromisoformat(str(barber_called["called_at"]).replace("Z", "+00:00"))
                        elapsed = (datetime.now(timezone.utc) - called_dt).total_seconds() / 60
                        remaining = max(0, int(remaining - elapsed))
                    except Exception:
                        pass
                approx_finish_minutes = remaining
                # Service names (not sensitive)
                svc_names = []
                for sid in barber_called.get("selected_services", []) or []:
                    svc = await db.services.find_one({"id": sid}, {"_id": 0, "service_name": 1})
                    if svc:
                        svc_names.append(svc.get("service_name"))
                currently_serving_payload = {
                    "token_number": barber_called.get("token_number"),
                    "services": svc_names,
                    "started_at": barber_called.get("called_at"),
                    "approx_finish_minutes": approx_finish_minutes,
                }
            except Exception as e:
                logger.error(f"live-status finish calc failed: {e}")

        result["barbers"].append({
            "barber_id": barber["id"],
            "barber_name": barber["name"],
            "category": barber.get("category"),
            "specialization": barber.get("specialization"),
            "current_token": barber_called.get("token_number") if barber_called else None,
            "waiting_count": len(barber_waiting),
            "total_tokens_today": total_tokens_today,
            "queue_status": barber.get("queue_status", "available"),
            "currently_serving": currently_serving_payload,
            "approx_finish_minutes": approx_finish_minutes,
        })
    
    return result

@api_router.get("/salons/{salon_id}/live-status")
async def get_salon_live_status(salon_id: str, shift: Optional[str] = None, date: Optional[str] = None, branch_id: Optional[str] = None):
    """Get current live status for salon (alias for token-status)"""
    return await get_salon_token_status(salon_id, shift, date, branch_id)


# ============ SALON HOME KPIs + DIRECT INVOICE ============

@api_router.get("/salons/{salon_id}/home-kpis")
async def get_salon_home_kpis(
    salon_id: str,
    date_mode: str = "today",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_salon=Depends(get_current_salon_user)
):
    """One-shot KPIs for the redesigned salon Home page.

    date_mode: today | yesterday | tomorrow | range | week
      When date_mode='range', supply date_from / date_to (YYYY-MM-DD).
    """
    # -- date basis --
    today = datetime.now(timezone.utc).date()
    if date_mode == "yesterday":
        start_date = end_date = today - timedelta(days=1)
    elif date_mode == "tomorrow":
        start_date = end_date = today + timedelta(days=1)
    elif date_mode == "week":
        start_date = today - timedelta(days=6)
        end_date = today
    elif date_mode == "range":
        try:
            start_date = date.fromisoformat(date_from) if date_from else today
            end_date = date.fromisoformat(date_to) if date_to else today
        except Exception:
            start_date = end_date = today
        if end_date < start_date:
            start_date, end_date = end_date, start_date
    else:
        start_date = end_date = today
    basis_date = end_date.isoformat()  # for legacy fields
    date_start_iso = start_date.isoformat()
    date_end_iso = end_date.isoformat()

    # -- fetch tokens in the basis range --
    tokens_basis = await db.tokens.find(
        {"salon_id": salon_id, "date": {"$gte": date_start_iso, "$lte": date_end_iso}},
        {"_id": 0}
    ).to_list(20000)

    completed_basis = [t for t in tokens_basis if t.get("status") == "completed"]
    total_amounts = [float(t.get("total_amount") or 0) for t in completed_basis]

    today_sales = round(sum(total_amounts), 2)
    avg_ticket = round((today_sales / len(completed_basis)), 2) if completed_basis else 0.0

    # no-show rate — skipped/cancelled over total tokens (today)
    total_ct = len(tokens_basis)
    ns_ct = sum(1 for t in tokens_basis if t.get("status") in ("skipped", "cancelled", "no_show"))
    no_show_rate = round((ns_ct * 100.0 / total_ct), 1) if total_ct else 0.0

    # -- rebooking / retention rate: fraction of completed today's customers
    #    who had at least one earlier completed token (last 180 days) --
    prior_start = (today - timedelta(days=180)).isoformat()
    rebook_hit = 0
    unique_phones = set()
    for t in completed_basis:
        ph = (t.get("phone") or "").strip()
        if not ph:
            continue
        unique_phones.add(ph)
    if unique_phones:
        prior_agg = await db.tokens.find(
            {
                "salon_id": salon_id,
                "phone": {"$in": list(unique_phones)},
                "status": "completed",
                "date": {"$gte": prior_start, "$lt": date_start_iso},
            },
            {"_id": 0, "phone": 1}
        ).to_list(20000)
        prior_phones = {p["phone"] for p in prior_agg}
        rebook_hit = len(prior_phones & unique_phones)
    rebooking_rate = round((rebook_hit * 100.0 / len(unique_phones)), 1) if unique_phones else 0.0
    retention_rate = rebooking_rate  # alias (short-hand)

    # -- new clients today: unique phones today with NO prior tokens ever --
    new_clients_count = 0
    if unique_phones:
        seen_ever = await db.tokens.find(
            {
                "salon_id": salon_id,
                "phone": {"$in": list(unique_phones)},
                "date": {"$lt": date_start_iso},
            },
            {"_id": 0, "phone": 1}
        ).to_list(20000)
        prior_ever = {p["phone"] for p in seen_ever}
        new_clients_count = sum(1 for p in unique_phones if p not in prior_ever)

    # -- chair utilization: (completed service-minutes today) / (barbers * 8h * 60 min) --
    barbers = await db.barbers.find({"salon_id": salon_id, "is_active": True}, {"_id": 0}).to_list(1000)
    barber_count = max(1, len(barbers))
    # Try to derive minutes from selected_services duration; fallback to 30m/service
    service_dur_cache: Dict[str, int] = {}
    used_minutes = 0
    for t in completed_basis:
        for svc_obj in (t.get("selected_services") or []):
            # Handle both string ID and dict formats for backward compatibility
            if isinstance(svc_obj, dict):
                sid = svc_obj.get("service_id") or svc_obj.get("id")
                dur = svc_obj.get("default_duration") or svc_obj.get("duration")
            else:
                sid = svc_obj
                dur = None
            if not sid:
                continue
            if sid not in service_dur_cache:
                if dur is not None:
                    service_dur_cache[sid] = int(dur or 30)
                else:
                    svc = await db.services.find_one({"id": sid}, {"_id": 0, "default_duration": 1})
                    service_dur_cache[sid] = int((svc or {}).get("default_duration") or 30)
            used_minutes += service_dur_cache[sid]
    available_minutes = barber_count * 8 * 60
    chair_utilization = round(min(100.0, used_minutes * 100.0 / available_minutes), 1) if available_minutes else 0.0

    # -- retail sales (from salon_orders or salon_store_sales if exists) --
    retail_sales = 0.0
    try:
        rs_cursor = db.salon_store_orders.find(
            {"salon_id": salon_id, "created_at": {"$gte": date_start_iso}},
            {"_id": 0, "total_amount": 1}
        )
        async for r in rs_cursor:
            retail_sales += float(r.get("total_amount") or 0)
    except Exception:
        retail_sales = 0.0

    # -- reminder / confirmation rate (marketing messages sent today: delivered/sent%) --
    reminder_confirmation_rate = None
    try:
        sent = await db.marketing_messages.count_documents(
            {"salon_id": salon_id, "sent_at": {"$regex": f"^{basis_date}"}}
        )
        delivered = await db.marketing_messages.count_documents(
            {"salon_id": salon_id, "sent_at": {"$regex": f"^{basis_date}"},
             "status": {"$in": ["delivered", "read"]}}
        )
        reminder_confirmation_rate = round((delivered * 100.0 / sent), 1) if sent else 0.0
    except Exception:
        reminder_confirmation_rate = 0.0

    # -- waitlist (feature not present in DB yet) --
    waitlist_count = 0
    try:
        waitlist_count = await db.salon_waitlist.count_documents(
            {"salon_id": salon_id, "status": "waiting"}
        )
    except Exception:
        waitlist_count = 0

    # -- payment mix (today) --
    payment_mix: Dict[str, float] = {}
    for t in completed_basis:
        pm = (t.get("payment_mode") or "unknown").lower()
        payment_mix[pm] = payment_mix.get(pm, 0.0) + float(t.get("total_amount") or 0)
    payment_mix = {k: round(v, 2) for k, v in payment_mix.items()}

    # -- top services today (by count + revenue) --
    svc_ct: Dict[str, Dict[str, Any]] = {}
    for t in completed_basis:
        for svc_obj in (t.get("selected_services") or []):
            if isinstance(svc_obj, dict):
                sid = svc_obj.get("service_id") or svc_obj.get("id")
            else:
                sid = svc_obj
            if not sid:
                continue
            if sid not in svc_ct:
                svc_ct[sid] = {"service_id": sid, "count": 0}
            svc_ct[sid]["count"] += 1
    if svc_ct:
        svc_docs = await db.services.find(
            {"id": {"$in": list(svc_ct.keys())}},
            {"_id": 0, "id": 1, "service_name": 1, "base_price": 1}
        ).to_list(200)
        by_id = {s["id"]: s for s in svc_docs}
        for sid, row in svc_ct.items():
            info = by_id.get(sid) or {}
            row["service_name"] = info.get("service_name") or "Service"
            row["revenue"] = round(row["count"] * float(info.get("base_price") or 0), 2)
    top_services = sorted(svc_ct.values(), key=lambda r: r.get("revenue", 0), reverse=True)[:5]

    # -- staff leaderboard (today) --
    staff_leaderboard: List[Dict[str, Any]] = []
    barber_stats: Dict[str, Dict[str, Any]] = {}
    for t in completed_basis:
        bid = t.get("barber_id")
        if not bid:
            continue
        row = barber_stats.setdefault(bid, {
            "barber_id": bid,
            "barber_name": t.get("barber_name") or "Barber",
            "sales": 0.0,
            "tips": 0.0,
            "bookings": 0,
            "rebook_count": 0,
        })
        row["sales"] += float(t.get("total_amount") or 0)
        row["tips"] += float(t.get("tip_amount") or 0)
        row["bookings"] += 1
        # rebook: customer had past completed tokens with same barber (last 180d)
        # keep this cheap — just flag if phone existed anywhere in prior set
        if t.get("phone") and t["phone"] in unique_phones and rebook_hit:
            pass  # aggregate rebook % below
    # Compute rebook % per barber (simple)
    for bid, row in barber_stats.items():
        row["sales"] = round(row["sales"], 2)
        row["tips"] = round(row["tips"], 2)
        # % of returning customers among today's completed for this barber
        row["rebook_pct"] = rebooking_rate
    staff_leaderboard = sorted(barber_stats.values(), key=lambda r: r["sales"], reverse=True)

    # -- reviews summary --
    reviews_summary = {"avg_rating": 0.0, "total_reviews": 0, "distribution": {"5":0,"4":0,"3":0,"2":0,"1":0}}
    try:
        all_ratings = await db.ratings.find({"salon_id": salon_id}, {"_id": 0, "rating": 1}).to_list(5000)
        if all_ratings:
            total = len(all_ratings)
            avg = sum(float(r.get("rating") or 0) for r in all_ratings) / total
            dist = {"5":0,"4":0,"3":0,"2":0,"1":0}
            for r in all_ratings:
                k = str(int(round(float(r.get("rating") or 0))))
                if k in dist:
                    dist[k] += 1
            reviews_summary = {"avg_rating": round(avg, 1), "total_reviews": total, "distribution": dist}
    except Exception:
        pass

    # -- targets vs actual (from salon.settings.targets or fallback) --
    salon_doc = await db.salons.find_one({"id": salon_id}, {"_id": 0}) or {}
    targets = salon_doc.get("targets") or {}
    daily_target = float(targets.get("daily_revenue") or 15000)
    monthly_target = float(targets.get("monthly_revenue") or 300000)
    membership_target = float(targets.get("monthly_memberships") or 20)

    # month-to-date revenue
    first_of_month = today.replace(day=1).isoformat()
    mtd_docs = await db.tokens.find(
        {"salon_id": salon_id, "status": "completed", "date": {"$gte": first_of_month, "$lte": today.isoformat()}},
        {"_id": 0, "total_amount": 1}
    ).to_list(20000)
    mtd_revenue = round(sum(float(d.get("total_amount") or 0) for d in mtd_docs), 2)

    # memberships sold this month
    try:
        mtd_memberships = await db.customer_memberships.count_documents({
            "salon_id": salon_id,
            "purchased_at": {"$gte": first_of_month}
        })
    except Exception:
        mtd_memberships = 0

    # -- revenue 7d sparkline --
    revenue_7d: List[Dict[str, Any]] = []
    for i in range(6, -1, -1):
        d = (today - timedelta(days=i)).isoformat()
        docs = await db.tokens.find(
            {"salon_id": salon_id, "status": "completed", "date": d},
            {"_id": 0, "total_amount": 1}
        ).to_list(5000)
        revenue_7d.append({"date": d, "total": round(sum(float(x.get("total_amount") or 0) for x in docs), 2)})

    # -- busy hours histogram (last 7d) --
    hist_start = (today - timedelta(days=6)).isoformat()
    hist_docs = await db.tokens.find(
        {"salon_id": salon_id, "date": {"$gte": hist_start, "$lte": today.isoformat()}},
        {"_id": 0, "created_at": 1}
    ).to_list(20000)
    busy_hours = {str(h): 0 for h in range(24)}
    for d in hist_docs:
        try:
            hh = datetime.fromisoformat((d.get("created_at") or "").replace("Z", "+00:00")).hour
            busy_hours[str(hh)] += 1
        except Exception:
            continue

    # ---- Customer count by source (new home page KPI) ------------------------
    # customers.source enum: online | qr | owner | direct
    #   online = booked from customer website / app
    #   qr     = walked in via QR scan
    #   owner  = booked by salon owner from admin panel
    #   direct = direct invoice (walk-out, no queue)
    # Falls back to token.source when customers doc is missing.
    source_map = {"online": 0, "qr": 0, "owner": 0, "direct": 0}
    _phone_source_cache: Dict[str, str] = {}
    for t in tokens_basis:
        # Prefer explicit source stored on the token, otherwise look up the
        # customer's stored source, otherwise infer from token metadata.
        src = (t.get("source") or "").lower().strip()
        if src not in source_map:
            ph = (t.get("phone") or "").strip()
            if ph:
                if ph not in _phone_source_cache:
                    cust = await db.customers.find_one(
                        {"phone": ph}, {"_id": 0, "source": 1}
                    )
                    _phone_source_cache[ph] = ((cust or {}).get("source") or "").lower().strip()
                src = _phone_source_cache[ph]
        if src not in source_map:
            # Heuristic fallback: booking_type or direct_invoice flag
            if t.get("is_direct_invoice") or t.get("booking_type") == "direct":
                src = "direct"
            elif t.get("booking_type") == "walk_in" or t.get("via_qr"):
                src = "qr"
            elif t.get("created_via_admin") or t.get("booking_type") == "admin":
                src = "owner"
            else:
                src = "online"
        source_map[src] = source_map.get(src, 0) + 1
    customer_count_by_source = source_map
    customer_count_total = sum(source_map.values())

    # ---- Staff attendance (In / Late / Out for today) -------------------------
    # attendance_mode.py stores docs in db.staff_attendance
    #   {barber_id, salon_id, date, check_in_at, check_out_at, status}
    attendance_today: List[Dict[str, Any]] = []
    try:
        att_rows = await db.staff_attendance.find(
            {"salon_id": salon_id, "date": today.isoformat()},
            {"_id": 0}
        ).to_list(500)
        att_by_barber = {a.get("barber_id"): a for a in att_rows}
        # Build per-barber row from all active barbers so admin can toggle
        for b in barbers:
            a = att_by_barber.get(b.get("id")) or {}
            ci = a.get("check_in_at")
            co = a.get("check_out_at")
            if co:
                status = "out"
            elif ci:
                status = "in"
            else:
                status = "late"
            attendance_today.append({
                "barber_id": b.get("id"),
                "name": b.get("name") or "Staff",
                "status": status,
                "check_in_at": ci,
                "check_out_at": co,
            })
    except Exception:
        attendance_today = []

    # ---- Marketing performance for the period --------------------------------
    marketing_perf: Dict[str, Any] = {
        "sent": 0, "delivered": 0, "clicked": 0, "redeemed": 0, "revenue": 0.0,
        "delivered_pct": 0.0, "click_pct": 0.0,
        "campaigns": [], "channels": {},
    }
    try:
        mm_query = {
            "salon_id": salon_id,
            "sent_at": {"$gte": date_start_iso, "$lte": date_end_iso + "T23:59:59Z"},
        }
        mm_docs = await db.marketing_messages.find(mm_query, {"_id": 0}).to_list(50000)
        sent = len(mm_docs)
        delivered = sum(1 for m in mm_docs if m.get("status") in ("delivered", "read", "clicked"))
        clicked = sum(1 for m in mm_docs if m.get("clicked_at") or m.get("status") == "clicked")
        redeemed = sum(1 for m in mm_docs if m.get("redeemed_at"))
        revenue = sum(float(m.get("attributed_revenue") or 0) for m in mm_docs)
        marketing_perf["sent"] = sent
        marketing_perf["delivered"] = delivered
        marketing_perf["clicked"] = clicked
        marketing_perf["redeemed"] = redeemed
        marketing_perf["revenue"] = round(revenue, 2)
        marketing_perf["delivered_pct"] = round((delivered * 100.0 / sent), 1) if sent else 0.0
        marketing_perf["click_pct"] = round((clicked * 100.0 / sent), 1) if sent else 0.0
        # Channel mix (WhatsApp / SMS / Email)
        chans: Dict[str, int] = {}
        for m in mm_docs:
            ch = (m.get("channel") or m.get("provider") or "WhatsApp").capitalize()
            chans[ch] = chans.get(ch, 0) + 1
        marketing_perf["channels"] = chans
        # Active campaigns list (rolled up)
        camp_rollup: Dict[str, Dict[str, Any]] = {}
        for m in mm_docs:
            cid = m.get("campaign_id") or "adhoc"
            row = camp_rollup.setdefault(cid, {
                "id": cid, "name": None, "channel": (m.get("channel") or "WhatsApp").capitalize(),
                "sent": 0, "delivered": 0, "redeemed": 0, "revenue": 0.0
            })
            row["sent"] += 1
            if m.get("status") in ("delivered", "read", "clicked"):
                row["delivered"] += 1
            if m.get("redeemed_at"):
                row["redeemed"] += 1
            row["revenue"] += float(m.get("attributed_revenue") or 0)
        # Fetch campaign names
        camp_ids = [c for c in camp_rollup.keys() if c != "adhoc"]
        if camp_ids:
            camp_docs = await db.marketing_campaigns.find(
                {"id": {"$in": camp_ids}}, {"_id": 0, "id": 1, "name": 1, "channel": 1}
            ).to_list(200)
            for cd in camp_docs:
                if cd["id"] in camp_rollup:
                    camp_rollup[cd["id"]]["name"] = cd.get("name")
                    camp_rollup[cd["id"]]["channel"] = (cd.get("channel") or camp_rollup[cd["id"]]["channel"]).capitalize()
        for cid, row in camp_rollup.items():
            if not row["name"]:
                row["name"] = "Ad-hoc messages" if cid == "adhoc" else "Campaign"
            row["revenue"] = round(row["revenue"], 2)
        marketing_perf["campaigns"] = sorted(
            camp_rollup.values(), key=lambda r: r["sent"], reverse=True
        )[:6]
    except Exception:
        pass

    # ---- Booking link URLs (used by the compact Send booking link chip) ------
    # These are the 3 links the salon can share with a customer via WhatsApp.
    frontend_origin = os.environ.get("PUBLIC_APP_URL") or (
        # Fall back to REACT_APP_BACKEND_URL sans /api since ingress rewrites /api → :8001.
        # The customer routes live at /(salons/:id), /(salons/:id/menu), /(salons/:id/book).
        ""
    )
    salon_slug = salon_id  # keep simple; salons/{id} routes work everywhere
    booking_links = {
        "book_url":  f"{frontend_origin}/salons/{salon_slug}/book"  if frontend_origin else f"/salons/{salon_slug}/book",
        "home_url":  f"{frontend_origin}/salons/{salon_slug}"       if frontend_origin else f"/salons/{salon_slug}",
        "menu_url":  f"{frontend_origin}/salons/{salon_slug}/menu"  if frontend_origin else f"/salons/{salon_slug}/menu",
    }

    return {
        "date_basis": basis_date,
        "date_range": {"from": date_start_iso, "to": date_end_iso},
        "primary": {
            "today_sales": today_sales,
            "avg_ticket": avg_ticket,
            "rebooking_rate": rebooking_rate,
            "no_show_rate": no_show_rate,
            "chair_utilization": chair_utilization,
        },
        "customer_count": {
            "total": customer_count_total,
            "by_source": customer_count_by_source,
        },
        "staff_attendance": attendance_today,
        "marketing_perf": marketing_perf,
        "booking_links": booking_links,
        "secondary": {
            "appointments_count": total_ct,
            "new_clients_count": new_clients_count,
            "retention_rate": retention_rate,
            "retail_sales": round(retail_sales, 2),
            "reminder_confirmation_rate": reminder_confirmation_rate or 0.0,
            "waitlist_count": waitlist_count,
        },
        "staff_leaderboard": staff_leaderboard,
        "reviews": reviews_summary,
        "targets": {
            "daily_target": daily_target,
            "daily_actual": today_sales,
            "monthly_target": monthly_target,
            "monthly_actual": mtd_revenue,
            "membership_target": membership_target,
            "membership_actual": mtd_memberships,
        },
        "revenue_7d": revenue_7d,
        "payment_mix": payment_mix,
        "top_services": top_services,
        "busy_hours": busy_hours,
    }


# ---------------------------------------------------------------------------
# Send booking link — used by the compact "Send booking link" chip on Home.
#   Sends one of 3 links (booking page / salon home / menu) to a guest via
#   WhatsApp. Accepts either a saved customer's phone or a raw 10-digit mobile.
# ---------------------------------------------------------------------------
class SendBookingLinkIn(BaseModel):
    phone: str                # 10-digit or +91-prefixed
    link_type: str = "book"   # book | home | menu
    name: Optional[str] = None
    save_as_lead: Optional[bool] = False


@api_router.post("/salons/{salon_id}/send-booking-link")
async def send_booking_link(
    salon_id: str,
    body: SendBookingLinkIn,
    current_salon=Depends(get_current_salon_user),
):
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    # Normalize phone → +91 prefix (India only for now)
    raw = (body.phone or "").strip().replace(" ", "").replace("-", "")
    if not raw:
        raise HTTPException(status_code=400, detail="Phone is required")
    if not raw.startswith("+"):
        digits = "".join(c for c in raw if c.isdigit())
        if len(digits) < 10:
            raise HTTPException(status_code=400, detail="Enter a valid 10-digit phone")
        raw = f"+91{digits[-10:]}"

    salon_name = salon.get("salon_name") or salon.get("name") or "our salon"
    frontend_origin = os.environ.get("PUBLIC_APP_URL") or ""
    kind = (body.link_type or "book").lower()
    if kind not in ("book", "home", "menu"):
        raise HTTPException(status_code=400, detail="link_type must be book|home|menu")
    path_by_kind = {
        "book": f"/salons/{salon_id}/book",
        "home": f"/salons/{salon_id}",
        "menu": f"/salons/{salon_id}/menu",
    }
    link_url = f"{frontend_origin}{path_by_kind[kind]}" if frontend_origin else path_by_kind[kind]
    template = {
        "book": f"Hi{(' ' + body.name) if body.name else ''}, book your next visit at {salon_name} here: {link_url}",
        "home": f"Hi{(' ' + body.name) if body.name else ''}, check out {salon_name}: {link_url}",
        "menu": f"Hi{(' ' + body.name) if body.name else ''}, here's our service menu at {salon_name}: {link_url}",
    }
    message = template[kind]

    # Optionally record as a lead for later marketing
    if body.save_as_lead:
        try:
            await db.customers.update_one(
                {"phone": raw},
                {"$setOnInsert": {
                    "id": str(uuid.uuid4()),
                    "phone": raw,
                    "name": body.name or "",
                    "source": "owner",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }},
                upsert=True,
            )
        except Exception:
            pass

    # Send via WhatsApp
    try:
        from whatsapp_service import send_whatsapp_message
        result = await send_whatsapp_message(raw, text=message)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"WhatsApp send failed: {exc}")
    if isinstance(result, dict) and result.get("error"):
        # Not fatal for UX — surface but 200 with delivery_status so the UI can show a hint
        return {"ok": False, "sent_to": raw, "link_url": link_url, "delivery_status": "failed", "note": result.get("error")}
    return {"ok": True, "sent_to": raw, "link_url": link_url, "delivery_status": "sent"}


# ---------------------------------------------------------------------------
# Home-page staff attendance toggle — one-tap Check-in / Check-out per staff.
#   POST /api/salons/{salon_id}/home/staff-attendance/toggle
#   body = {barber_id: str, action: "in" | "out"}
#
# RBAC:
#   - Admins / branch_managers  → can toggle any staff (bypasses geo-fence).
#   - Staff role with `staff.attendance` (module) + `staff.view_all` OR the
#     legacy `can_view_all_staff` flag → can toggle any staff.
#   - Staff role without `view_all` → can ONLY toggle their OWN linked staff_id
#     (self check-in). Trying to toggle a peer returns 403.
#   - Staff role without any staff-module access at all → 403 for everyone.
#
# Multi-session fix: after a full check-in → check-out, the same staff CAN
# check in again the same day. We now maintain a `sessions[]` array (matching
# attendance_mode.py) so a fresh "in" appends a new open session instead of
# short-circuiting on `check_in_at`.
# ---------------------------------------------------------------------------
@api_router.post("/salons/{salon_id}/home/staff-attendance/toggle")
async def home_toggle_attendance(
    salon_id: str,
    body: dict,
    current_salon=Depends(get_current_salon_user),
):
    barber_id = (body or {}).get("barber_id")
    action = ((body or {}).get("action") or "in").lower()
    if not barber_id:
        raise HTTPException(status_code=400, detail="barber_id required")
    if action not in ("in", "out"):
        raise HTTPException(status_code=400, detail="action must be in|out")

    # ---- RBAC ----
    role = current_salon.get("role")
    is_admin_like = role in ("salon_admin", "admin", "salon", "salon_branch_manager")
    if not is_admin_like:
        # Must at least have staff.attendance access.
        if not has_module_permission(current_salon, "staff", "attendance"):
            raise HTTPException(status_code=403, detail="Permission denied: staff.attendance")
        # If they don't have view_all, restrict to their OWN staff record.
        can_toggle_others = has_module_permission(current_salon, "staff", "view_all")
        if not can_toggle_others:
            own_staff_id = current_salon.get("staff_id")
            if not own_staff_id or own_staff_id != barber_id:
                raise HTTPException(status_code=403, detail="You can only check in/out your own attendance.")

    today_iso = datetime.now(timezone.utc).date().isoformat()
    now_iso = datetime.now(timezone.utc).isoformat()

    doc = await db.staff_attendance.find_one(
        {"salon_id": salon_id, "barber_id": barber_id, "date": today_iso},
        {"_id": 0}
    ) or {}

    # Migrate legacy check_in_at / check_out_at (no sessions[]) into sessions[]
    # so the multi-session logic works uniformly.
    sessions = list(doc.get("sessions") or [])
    if not sessions and (doc.get("check_in_at") or doc.get("check_out_at")):
        sessions = [{
            "ci": doc.get("check_in_at"),
            "co": doc.get("check_out_at"),
            "ci_method": doc.get("check_in_method") or "home_toggle",
            "co_method": doc.get("check_out_method") or "home_toggle",
        }]

    last = sessions[-1] if sessions else None
    has_open = bool(last and last.get("ci") and not last.get("co"))

    if action == "in":
        if has_open:
            # Already checked in and not yet out — idempotent no-op.
            return {"ok": True, "already_in": True, "check_in_at": last.get("ci"), "sessions": sessions}
        # Append a new open session (allows re-check-in after a check-out today).
        sessions.append({"ci": now_iso, "co": None, "ci_method": "home_toggle"})
        await db.staff_attendance.update_one(
            {"salon_id": salon_id, "barber_id": barber_id, "date": today_iso},
            {"$set": {
                "check_in_at": now_iso,   # keep legacy field pointing at latest CI
                "check_out_at": None,     # clear legacy CO because we're back "in"
                "status": "in",
                "sessions": sessions,
                "salon_id": salon_id, "barber_id": barber_id, "date": today_iso,
            }},
            upsert=True,
        )
        return {"ok": True, "check_in_at": now_iso, "status": "in", "sessions": sessions}

    # action == "out"
    if not has_open:
        # No open session — record an "out" without altering sessions[] (defensive).
        await db.staff_attendance.update_one(
            {"salon_id": salon_id, "barber_id": barber_id, "date": today_iso},
            {"$set": {"check_out_at": now_iso, "status": "out",
                      "salon_id": salon_id, "barber_id": barber_id, "date": today_iso}},
            upsert=True,
        )
        return {"ok": True, "check_out_at": now_iso, "status": "out", "sessions": sessions}

    # Close the currently-open session.
    sessions[-1]["co"] = now_iso
    sessions[-1]["co_method"] = "home_toggle"
    await db.staff_attendance.update_one(
        {"salon_id": salon_id, "barber_id": barber_id, "date": today_iso},
        {"$set": {"check_out_at": now_iso, "status": "out", "sessions": sessions}},
    )
    return {"ok": True, "check_out_at": now_iso, "status": "out", "sessions": sessions}


# ---------------------------------------------------------------------------
# Guest profile aggregate — used by the New-Appointment drawer's right-hand
# customer details panel and by the "View full details" popup.
#   Returns: base master + last_visit / last_barber_name / last_invoice /
#            membership_active / membership_name / wallet_balance /
#            total_visits / total_spend / history (last 20 tokens)
# ---------------------------------------------------------------------------
@api_router.get("/salons/{salon_id}/customers/profile")
async def get_customer_profile(
    salon_id: str,
    phone: str,
    current_salon=Depends(get_current_salon_user),
):
    ph = (phone or "").strip()
    if not ph:
        raise HTTPException(status_code=400, detail="phone query param required")
    alt = ph
    if not ph.startswith("+"):
        d = "".join(c for c in ph if c.isdigit())
        if len(d) >= 10:
            alt = f"+91{d[-10:]}"
    or_phones = list({ph, alt})

    master = await db.salon_customers.find_one(
        {"salon_id": salon_id, "phone": {"$in": or_phones}}, {"_id": 0}
    ) or {}

    tks = await db.tokens.find(
        {"salon_id": salon_id, "phone": {"$in": or_phones}},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)

    total_visits = len(tks)
    total_spend = 0.0
    last_visit_iso = None
    last_barber_id = None
    last_barber_name = None
    last_invoice_amount = None
    last_invoice_date = None
    for t in tks:
        try:
            total_spend += float(t.get("total_amount") or t.get("total") or 0)
        except Exception:
            pass
        dt = t.get("date") or t.get("created_at") or ""
        if isinstance(dt, str) and dt and (not last_visit_iso or dt > last_visit_iso):
            last_visit_iso = dt
            last_barber_id = t.get("barber_id")
            last_barber_name = t.get("barber_name")
            last_invoice_amount = t.get("total_amount") or t.get("total")
            last_invoice_date = dt

    wallet_doc = await db.salon_wallets.find_one(
        {"salon_id": salon_id, "phone": {"$in": or_phones}}, {"_id": 0, "balance": 1}
    ) or {}
    wallet_balance = float(wallet_doc.get("balance") or 0)

    mem = await db.customer_memberships.find_one(
        {"salon_id": salon_id, "phone": {"$in": or_phones}, "status": {"$in": ["active", "Active"]}},
        {"_id": 0}
    )
    membership_active = bool(mem)
    membership_name = None
    membership_expires = None
    if mem:
        plan_id = mem.get("plan_id")
        if plan_id:
            plan = await db.membership_plans.find_one({"id": plan_id}, {"_id": 0, "name": 1})
            if plan:
                membership_name = plan.get("name")
        membership_expires = mem.get("expires_at") or mem.get("valid_till")

    history_tokens = [{
        "id": t.get("id"),
        "date": t.get("date") or t.get("created_at"),
        "barber_name": t.get("barber_name"),
        "services_count": len(t.get("selected_services") or []),
        "total": float(t.get("total_amount") or t.get("total") or 0),
        "status": t.get("status"),
    } for t in tks[:20]]

    return {
        "phone": master.get("phone") or (or_phones[0] if or_phones else ph),
        "name": master.get("name") or (tks[0].get("customer_name") if tks else None),
        "gender": master.get("gender"),
        "email": master.get("email"),
        "dob": master.get("dob") or master.get("date_of_birth"),
        "anniversary": master.get("anniversary"),
        "tags": master.get("tags") or [],
        "notes": master.get("notes"),
        "photo_url": master.get("photo_url"),
        "instagram_id": master.get("instagram_id"),
        "facebook_id": master.get("facebook_id"),
        "preferred_barber_id": master.get("preferred_barber_id"),
        "source": master.get("source"),
        "last_visit": last_visit_iso,
        "last_barber_id": last_barber_id,
        "last_barber_name": last_barber_name,
        "last_invoice_amount": last_invoice_amount,
        "last_invoice_date": last_invoice_date,
        "wallet_balance": wallet_balance,
        "membership_active": membership_active,
        "membership_name": membership_name,
        "membership_expires": membership_expires,
        "total_visits": total_visits,
        "total_spend": round(total_spend, 2),
        "history_tokens": history_tokens,
    }






@api_router.post("/salons/{salon_id}/direct-invoice")
async def create_direct_invoice(
    salon_id: str,
    body: dict,
    current_user=Depends(get_current_salon_user)
):
    """Create a direct invoice bypassing the queue.

    Payload:
      customer_name (str, required)
      phone (str, optional)
      gender (str, optional)
      barber_id (str, optional)
      selected_services (List[str], required)
      payment_mode (str, required) — cash|upi|card|wallet
      coupon_code (str, optional)
      membership_plan_id (str, optional) — sold + discount applied to THIS order
      tip_amount (float, optional)
      notes (str, optional)

    Effect:
      - Computes bill with services + optional membership benefit + optional coupon
      - Creates a `completed` token record (status=completed, payment_confirmed=True)
        so it shows in sales/analytics/history.
      - Generates the invoice PDF via generate_and_send_invoice(token_id).
      - If membership_plan_id provided, sells the membership (credits wallet) and
        applies the plan's discount % to this order (top-up wallet or discount).
      - Records coupon redemption.
    Returns invoice_id + token_id.
    """
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")

    customer_name = (body.get("customer_name") or "Walk-in").strip() or "Walk-in"
    phone = (body.get("phone") or "").strip()
    if phone:
        phone = phone.replace(" ", "").replace("-", "")
        if not phone.startswith("+91"):
            phone = f"+91{phone}"
    gender = body.get("gender") or "Men"
    barber_id = body.get("barber_id") or "any"
    selected_services = body.get("selected_services") or []
    selected_products = body.get("selected_products") or []  # [{product_id, name, qty, unit_price}]
    payment_mode = (body.get("payment_mode") or "cash").lower()
    coupon_code = (body.get("coupon_code") or "").strip().upper() or None
    membership_plan_id = body.get("membership_plan_id") or None
    tip_amount = float(body.get("tip_amount") or 0)
    notes = body.get("notes") or ""

    if not selected_services and not selected_products:
        raise HTTPException(status_code=400, detail="Add at least one service or product")

    # -- Compute base subtotal (services) --
    subtotal = 0.0
    service_details = []
    for sid in selected_services:
        # Accept either a raw string ID or a dict {service_id, price?}
        if isinstance(sid, dict):
            svc_id = sid.get("service_id") or sid.get("id")
            fallback_price = sid.get("price") or sid.get("base_price")
        else:
            svc_id = sid
            fallback_price = None
        if not svc_id:
            continue
        # Barber-specific pricing
        price = None
        if barber_id and barber_id != "any":
            barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
            if barber:
                bs = next((s for s in (barber.get("services") or []) if s.get("service_id") == svc_id), None)
                if bs:
                    price = float(bs.get("price") or 0)
        if price is None or price == 0:
            svc = await db.services.find_one({"id": svc_id}, {"_id": 0})
            if svc:
                # Robust price lookup — try multiple field names
                price = float(
                    svc.get("base_price")
                    or svc.get("price")
                    or svc.get("selling_price")
                    or svc.get("default_price")
                    or 0
                )
        if (price is None or price == 0) and fallback_price:
            price = float(fallback_price)
        price = float(price or 0)
        subtotal += price
        service_details.append({"service_id": svc_id, "price": price})

    # -- Product subtotal (also decrements inventory) --
    product_subtotal = 0.0
    product_details = []
    for p in selected_products:
        pid = p.get("product_id") or p.get("id")
        qty = int(p.get("qty") or 0)
        unit_price = float(p.get("unit_price") or 0)
        if not pid or qty <= 0:
            continue
        item = None
        if unit_price <= 0 or True:  # always look up for stock check
            item = await db.salon_inventory.find_one({"id": pid, "salon_id": salon_id}, {"_id": 0})
            if item and unit_price <= 0:
                unit_price = float(item.get("retail_price") or item.get("selling_price") or 0)
        # Stock validation — soft guard (does not block sale for a non-existent item)
        if item is not None:
            available = int(item.get("stock_quantity") or 0)
            if available < qty:
                # Allow but log; UI should ideally block this
                logger.warning(f"Insufficient stock for {pid}: requested {qty}, have {available}")
        line_total = qty * unit_price
        product_subtotal += line_total
        product_details.append({
            "product_id": pid,
            "name": p.get("name") or (item or {}).get("name"),
            "qty": qty,
            "unit_price": unit_price,
            "line_total": line_total,
        })
        # Decrement inventory (best effort)
        try:
            await db.salon_inventory.update_one(
                {"id": pid, "salon_id": salon_id},
                {"$inc": {"stock_quantity": -qty}}
            )
        except Exception:
            pass
    subtotal += product_subtotal

    membership_discount = 0.0
    membership_sale_amount = 0.0
    membership_info = None

    # -- Optional membership upsell: sell membership + apply discount to THIS order --
    if membership_plan_id:
        plan = await db.membership_plans.find_one({"id": membership_plan_id}, {"_id": 0})
        if not plan:
            raise HTTPException(status_code=404, detail="Membership plan not found")
        if not phone:
            raise HTTPException(status_code=400, detail="Customer phone required to sell membership")

        # Apply plan's service discount to THIS order (services only, not products)
        disc_pct = float(plan.get("discount_percentage") or plan.get("service_discount_pct") or 0)
        service_only_subtotal = subtotal - product_subtotal
        membership_discount = round(service_only_subtotal * disc_pct / 100.0, 2)
        # The membership itself is a paid line (paid_amount from plan price)
        membership_sale_amount = float(plan.get("price") or plan.get("plan_price") or 0)

        expiry_date = datetime.now(timezone.utc) + timedelta(days=int(plan.get("validity_months") or 6) * 30)
        existing = await db.customer_memberships.find_one({
            "salon_id": salon_id, "customer_phone": phone, "is_active": True
        }, {"_id": 0})
        if existing:
            new_balance = float(existing.get("wallet_balance") or 0) + float(plan.get("credit") or 0)
            await db.customer_memberships.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "wallet_balance": new_balance,
                    "expiry_date": expiry_date.isoformat(),
                    "tier": plan.get("tier", existing.get("tier", "Custom")),
                    "color": plan.get("color") or existing.get("color"),
                }}
            )
        else:
            await db.customer_memberships.insert_one({
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "customer_phone": phone,
                "customer_name": customer_name,
                "membership_plan_id": plan["id"],
                "membership_name": plan.get("name"),
                "tier": plan.get("tier", "Custom"),
                "color": plan.get("color"),
                "payment_mode": payment_mode,
                "paid_amount": membership_sale_amount,
                "credit_added": float(plan.get("credit") or 0),
                "wallet_balance": float(plan.get("credit") or 0),
                "expiry_date": expiry_date.isoformat(),
                "is_active": True,
                "cancelled": False,
                "payment_confirmed": True,
                "purchased_at": datetime.now(timezone.utc).isoformat(),
            })
        membership_info = {
            "plan_id": plan["id"],
            "name": plan.get("name"),
            "discount_pct": disc_pct,
            "discount_amount": membership_discount,
            "sold_at": membership_sale_amount,
        }

    # -- Optional coupon --
    coupon_discount = 0.0
    coupon_doc = None
    if coupon_code:
        try:
            from marketing import _coupon_active as _coupon_active_fn  # type: ignore
            from marketing import _compute_coupon_discount as _compute_coupon_discount_fn  # type: ignore
        except Exception:
            _coupon_active_fn = None
            _compute_coupon_discount_fn = None
        c = await db.salon_coupons.find_one({"salon_id": salon_id, "code": coupon_code})
        if not c:
            raise HTTPException(status_code=404, detail="Invalid coupon code")
        if _coupon_active_fn and not _coupon_active_fn(c):
            raise HTTPException(status_code=400, detail="Coupon expired or inactive")
        bill_before_coupon = max(0.0, subtotal - membership_discount)
        if bill_before_coupon < float(c.get("min_bill_amount") or 0):
            raise HTTPException(status_code=400, detail=f"Minimum bill amount for coupon is ₹{c.get('min_bill_amount')}")
        if _compute_coupon_discount_fn:
            coupon_discount = await _compute_coupon_discount_fn(c, bill_before_coupon)
        else:
            # Fallback: percentage
            if c.get("type") == "percentage":
                coupon_discount = round(bill_before_coupon * float(c.get("value") or 0) / 100.0, 2)
            else:
                coupon_discount = float(c.get("value") or 0)
        coupon_doc = c

    services_total = round(max(0.0, subtotal - membership_discount - coupon_discount), 2)
    grand_total = round(services_total + membership_sale_amount + tip_amount, 2)

    # -- Auto-assign a barber if 'any' so history displays a name --
    if not barber_id or barber_id == "any":
        anyb = await db.barbers.find_one({"salon_id": salon_id, "is_active": True}, {"_id": 0})
        barber_id = anyb["id"] if anyb else None

    barber_name = "Direct Invoice"
    if barber_id:
        b = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
        if b:
            barber_name = b.get("name") or barber_name

    # -- Create synthetic completed token --
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    now_iso = datetime.now(timezone.utc).isoformat()

    from datetime import datetime as _dt
    _h = _dt.now().hour
    shift = "Morning" if _h < 12 else ("Noon" if _h < 16 else "Evening")

    token_id = str(uuid.uuid4())
    token_number = await get_next_token_number(salon_id, today_str, shift)

    token_doc = {
        "id": token_id,
        "salon_id": salon_id,
        "branch_id": body.get("branch_id") or salon.get("main_branch_id"),
        "token_number": token_number,
        "customer_name": customer_name,
        "phone": phone,
        "gender": gender,
        "barber_id": barber_id,
        "barber_name": barber_name,
        "selected_services": selected_services,
        "selected_products": product_details,
        "total_amount": grand_total,
        "subtotal": subtotal,
        "membership_discount": membership_discount,
        "coupon_discount": coupon_discount,
        "coupon_code": coupon_code,
        "tip_amount": tip_amount,
        "membership_sale_amount": membership_sale_amount,
        "date": today_str,
        "shift": shift,
        "status": "completed",
        "payment_mode": payment_mode,
        "payment_confirmed": True,
        "payment_status": "paid",
        "is_direct_invoice": True,
        "notes": notes,
        "created_at": now_iso,
        "completed_at": now_iso,
    }
    await db.tokens.insert_one(token_doc)

    # Coupon redemption record
    if coupon_doc:
        try:
            from marketing import record_coupon_redemption as _rec
            await _rec(
                salon_id=salon_id,
                coupon_id=coupon_doc.get("id"),
                customer_phone=phone,
                booking_id=token_id,
                amount=coupon_discount,
            )
        except Exception:
            pass

    # Upsert customer master
    if phone:
        try:
            await db.customer_master.update_one(
                {"salon_id": salon_id, "phone": phone},
                {
                    "$set": {"name": customer_name, "gender": gender, "last_visit_at": now_iso},
                    "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now_iso},
                    "$inc": {"total_visits": 1, "total_spent": grand_total},
                },
                upsert=True,
            )
        except Exception:
            pass

    # Generate invoice PDF (best-effort; DB stores invoice even if WhatsApp fails)
    invoice_id = None
    try:
        await generate_and_send_invoice(token_id)
        refreshed = await db.tokens.find_one({"id": token_id}, {"_id": 0, "invoice_id": 1})
        invoice_id = (refreshed or {}).get("invoice_id")
    except Exception as e:
        logger.warning(f"Direct-invoice PDF gen/send failed for token {token_id}: {e}")

    return {
        "success": True,
        "token_id": token_id,
        "token_number": token_number,
        "invoice_id": invoice_id,
        "totals": {
            "subtotal": round(subtotal, 2),
            "membership_discount": membership_discount,
            "coupon_discount": coupon_discount,
            "services_total": services_total,
            "membership_sale": membership_sale_amount,
            "tip_amount": tip_amount,
            "grand_total": grand_total,
        },
        "membership": membership_info,
        "coupon": {"code": coupon_code, "discount": coupon_discount} if coupon_code else None,
    }


# ============ RATING/REVIEW ROUTES ============

@api_router.post("/ratings", response_model=RatingResponse)
async def create_rating(rating_data: RatingCreate):
    """Create a rating/review for a completed booking"""
    # Verify the token exists and is completed
    token = await db.tokens.find_one({"id": rating_data.token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if token.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only rate completed bookings")
    
    # Check if rating already exists for this token
    existing_rating = await db.ratings.find_one({"token_id": rating_data.token_id})
    if existing_rating:
        raise HTTPException(status_code=400, detail="You have already rated this booking")
    
    # Get barber name
    barber = await db.barbers.find_one({"id": rating_data.barber_id}, {"_id": 0})
    barber_name = barber.get("name", "Unknown") if barber else "Unknown"
    
    # Get user info from token
    user_id = token.get("user_id", "")
    user_name = token.get("customer_name", "Customer")
    
    # Create rating
    rating_dict = {
        "id": str(uuid.uuid4()),
        "token_id": rating_data.token_id,
        "user_id": user_id,
        "user_name": user_name,
        "barber_id": rating_data.barber_id,
        "barber_name": barber_name,
        "salon_id": rating_data.salon_id,
        "rating": rating_data.rating,
        "review": rating_data.review,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.ratings.insert_one(rating_dict)
    
    # Update barber's average rating
    await update_barber_average_rating(rating_data.barber_id)

    # Notify salon (in-app) about new review
    if rating_data.salon_id:
        stars = "⭐" * int(rating_data.rating)
        await create_in_app_notification(
            user_type="salon",
            user_id=rating_data.salon_id,
            title=f"New Review ({rating_data.rating}/5)",
            message=f"{user_name} rated {barber_name} {stars}: {(rating_data.review or '')[:100]}",
            notification_type="review_added",
            setting_key="review_added",
            salon_id=rating_data.salon_id,
            related_id=rating_dict["id"],
        )
    
    return RatingResponse(**rating_dict)

async def update_barber_average_rating(barber_id: str):
    """Update barber's average rating after a new review, and also update the salon-level aggregate."""
    pipeline = [
        {"$match": {"barber_id": barber_id}},
        {"$group": {
            "_id": "$barber_id",
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1}
        }}
    ]
    result = await db.ratings.aggregate(pipeline).to_list(1)

    salon_id = None
    if result:
        avg_rating = round(result[0]["average_rating"], 1)
        total_reviews = result[0]["total_reviews"]
        await db.barbers.update_one(
            {"id": barber_id},
            {"$set": {"rating": avg_rating, "total_reviews": total_reviews}}
        )
        barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0, "salon_id": 1})
        if barber:
            salon_id = barber.get("salon_id")

    if not salon_id:
        # Fallback: look up salon via ratings
        sample = await db.ratings.find_one({"barber_id": barber_id}, {"_id": 0, "salon_id": 1})
        if sample:
            salon_id = sample.get("salon_id")

    if salon_id:
        await update_salon_average_rating(salon_id)


async def update_salon_average_rating(salon_id: str):
    """Recompute the salon's aggregate rating across all its ratings."""
    pipeline = [
        {"$match": {"salon_id": salon_id}},
        {"$group": {
            "_id": "$salon_id",
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1}
        }}
    ]
    result = await db.ratings.aggregate(pipeline).to_list(1)
    if result:
        avg = round(result[0]["average_rating"], 1)
        total = result[0]["total_reviews"]
    else:
        avg = 0
        total = 0
    await db.salons.update_one(
        {"id": salon_id},
        {"$set": {"rating": avg, "total_reviews": total}}
    )

@api_router.get("/barbers/{barber_id}/ratings", response_model=BarberRatingSummary)
async def get_barber_ratings(barber_id: str, limit: int = 20, skip: int = 0):
    """Get all ratings and reviews for a barber"""
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    # Get ratings
    reviews = await db.ratings.find(
        {"barber_id": barber_id},
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Calculate average
    pipeline = [
        {"$match": {"barber_id": barber_id}},
        {"$group": {
            "_id": "$barber_id",
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1}
        }}
    ]
    stats = await db.ratings.aggregate(pipeline).to_list(1)
    
    avg_rating = round(stats[0]["average_rating"], 1) if stats else 0
    total_reviews = stats[0]["total_reviews"] if stats else 0
    
    return BarberRatingSummary(
        barber_id=barber_id,
        barber_name=barber.get("name", "Unknown"),
        average_rating=avg_rating,
        total_reviews=total_reviews,
        reviews=[RatingResponse(**r) for r in reviews]
    )

@api_router.get("/salons/{salon_id}/barbers/{barber_id}/profile")
async def get_barber_profile(salon_id: str, barber_id: str):
    """Get detailed barber profile with ratings"""
    barber = await db.barbers.find_one(
        {"id": barber_id, "salon_id": salon_id},
        {"_id": 0}
    )
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    # Get barber's services
    barber_services = await db.barber_services.find(
        {"barber_id": barber_id, "is_available": True},
        {"_id": 0}
    ).to_list(100)
    
    service_ids = [bs["service_id"] for bs in barber_services]
    services = await db.services.find(
        {"id": {"$in": service_ids}},
        {"_id": 0}
    ).to_list(100)
    
    # Add prices to services
    price_map = {bs["service_id"]: bs["price"] for bs in barber_services}
    for service in services:
        service["barber_price"] = price_map.get(service["id"], service.get("base_price", 0))
    
    # Get ratings summary
    pipeline = [
        {"$match": {"barber_id": barber_id}},
        {"$group": {
            "_id": "$barber_id",
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1}
        }}
    ]
    stats = await db.ratings.aggregate(pipeline).to_list(1)
    
    # Get recent reviews
    reviews = await db.ratings.find(
        {"barber_id": barber_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        **barber,
        "services": services,
        "average_rating": round(stats[0]["average_rating"], 1) if stats else barber.get("rating", 0),
        "total_reviews": stats[0]["total_reviews"] if stats else 0,
        "recent_reviews": reviews
    }

@api_router.get("/tokens/{token_id}")
async def get_token_details(token_id: str):
    """Get details of a specific token"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    return token

@api_router.get("/tokens/{token_id}/can-rate")
async def check_can_rate_token(token_id: str):
    """Check if a token can be rated"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    # Check if already rated
    existing_rating = await db.ratings.find_one({"token_id": token_id})
    
    return {
        "can_rate": token.get("status") == "completed" and not existing_rating,
        "is_completed": token.get("status") == "completed",
        "already_rated": existing_rating is not None
    }

@api_router.get("/users/{user_id}/pending-ratings")
async def get_user_pending_ratings(user_id: str):
    """Get completed bookings that haven't been rated yet"""
    # Get completed tokens for this user
    completed_tokens = await db.tokens.find(
        {"user_id": user_id, "status": "completed"},
        {"_id": 0}
    ).sort("completed_at", -1).limit(20).to_list(20)
    
    # Filter out already rated ones
    pending_ratings = []
    for token in completed_tokens:
        existing_rating = await db.ratings.find_one({"token_id": token["id"]})
        if not existing_rating:
            pending_ratings.append(token)
    
    return pending_ratings

# Direct barber review endpoint (for customers viewing barber profile)
class DirectBarberReview(BaseModel):
    user_id: Optional[str] = None
    user_name: str
    rating: int
    review: str

@api_router.post("/barbers/{barber_id}/reviews")
async def create_barber_review(barber_id: str, review_data: DirectBarberReview):
    """Create a direct review for a barber (without requiring a completed booking)"""
    # Verify barber exists
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    # Validate rating
    if review_data.rating < 1 or review_data.rating > 5:
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    
    # Create review
    review = {
        "id": str(uuid.uuid4()),
        "barber_id": barber_id,
        "salon_id": barber.get("salon_id"),
        "user_id": review_data.user_id,
        "user_name": review_data.user_name,
        "rating": review_data.rating,
        "review": review_data.review,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "type": "direct"  # To distinguish from booking-based reviews
    }
    
    await db.ratings.insert_one(review)
    
    # Update barber's average rating
    pipeline = [
        {"$match": {"barber_id": barber_id}},
        {"$group": {
            "_id": "$barber_id",
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1}
        }}
    ]
    stats = await db.ratings.aggregate(pipeline).to_list(1)
    
    if stats:
        await db.barbers.update_one(
            {"id": barber_id},
            {"$set": {
                "rating": round(stats[0]["average_rating"], 1),
                "total_reviews": stats[0]["total_reviews"]
            }}
        )
    
    return {"message": "Review submitted successfully", "review_id": review["id"]}

# ============ ANALYTICS ROUTES ============

# ============ EMPLOYEE REWARD PLAN (INCENTIVES) ============

class IncentiveSlab(BaseModel):
    from_pct: float = Field(ge=0)
    to_pct: float = Field(ge=0)        # use 9999 for "+" / open-ended
    type: str                           # "total_pct" | "additional_pct" | "fixed_amount"
    value: float = Field(ge=0)

class IncentivePlanConfig(BaseModel):
    target_type: str = "salary_multiplier"  # "salary_multiplier" | "manual"
    multiplier: Optional[float] = None
    manual_target: Optional[float] = None
    slabs: List[IncentiveSlab] = []

class RewardPlanCreate(BaseModel):
    mode: str  # "all" | "individual" | "partial"
    global_plan: Optional[IncentivePlanConfig] = None
    assigned_barber_ids: List[str] = []
    individual_plans: Dict[str, IncentivePlanConfig] = {}

class IncentivePayoutStatusUpdate(BaseModel):
    status: str  # "Pending" | "Approved" | "Paid" | "Hold"
    payment_method: Optional[str] = None  # "cash" | "upi" | "bank"
    notes: Optional[str] = None
    # Admin-overridden incentive amount (used when approving / paying out).
    # If null, the auto-calculated incentive_earned is used.
    manual_amount: Optional[float] = None


# ============ STAFF ATTENDANCE MODELS ============

class AttendanceStatus(str, Enum):
    PRESENT = "present"
    HALF_DAY = "half_day"
    ABSENT = "absent"
    HOLIDAY = "holiday"

class AttendanceRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    barber_id: str
    date: str  # YYYY-MM-DD
    status: str  # present/half_day/absent/holiday
    auto_calculated: bool = True  # True if system calculated, False if admin override
    morning_shift_completed: bool = False
    noon_evening_shift_completed: bool = False
    bookings_count: int = 0
    # Module 4 — Mode B (geo check-in) raw + computed fields (all optional, coexist with Mode A above)
    check_in_at: Optional[str] = None
    check_in_lat: Optional[float] = None
    check_in_lng: Optional[float] = None
    check_in_distance_meters: Optional[int] = None
    check_in_method: Optional[str] = None  # "self" | "admin_on_behalf" | "self_override" | "admin_edit"
    check_out_at: Optional[str] = None
    check_out_lat: Optional[float] = None
    check_out_lng: Optional[float] = None
    check_out_method: Optional[str] = None
    total_minutes: Optional[int] = None
    computed_under_mode: Optional[str] = None  # "service_completion" | "geo_checkin"
    half_day_reason: Optional[str] = None  # "late_checkin" | "short_hours" | "single_shift" | None
    created_at: str
    updated_at: str
    override_by: Optional[str] = None  # Admin user ID who overrode
    override_note: Optional[str] = None

class AttendanceOverride(BaseModel):
    status: str  # present/half_day/absent/holiday
    note: Optional[str] = None

class SalaryPaymentRequest(BaseModel):
    payment_method: str  # cash/upi/bank
    note: Optional[str] = None

class SalaryRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    salon_id: str
    barber_id: str
    month: str  # YYYY-MM
    base_salary: float
    working_days: int
    present_days: int
    half_days: int
    absent_days: int
    holidays: int
    calculated_salary: float
    incentive_amount: float = 0
    total_payable: float
    is_paid: bool = False
    paid_at: Optional[str] = None
    payment_method: Optional[str] = None
    financial_transaction_id: Optional[str] = None
    # Module 4 — payroll integration (additive; total_payable stays = final_payable)
    base_compensation: Optional[float] = None
    working_days_in_month: Optional[int] = None
    paid_leave_days: Optional[float] = None
    unpaid_leave_days: Optional[float] = None
    leave_breakdown: Optional[Dict[str, float]] = None
    lop_deduction: Optional[float] = None
    final_payable: Optional[float] = None
    attendance_mode_snapshot: Optional[str] = None
    created_at: str
    updated_at: str


def _compute_incentive_amount(plan_config: dict, target: float, actual_sales: float) -> dict:
    """Calculate incentive earned based on plan slabs.

    Slab types:
    - additional_pct: % applied only on the incremental sale within that slab range (cumulative)
    - total_pct:      % applied on the FULL actual_sales (only the matching range applies)
    - fixed_amount:   flat ₹ bonus when achievement reaches from_pct (additive)
    """
    target = float(target or 0)
    actual_sales = float(actual_sales or 0)
    if target <= 0:
        return {"earned": 0.0, "achievement_pct": 0.0, "breakdown": []}

    achievement_pct = (actual_sales / target) * 100.0
    slabs = sorted(plan_config.get("slabs", []) or [], key=lambda s: float(s.get("from_pct", 0)))

    earned = 0.0
    breakdown = []

    # 1) Additional % slabs (cumulative)
    for slab in slabs:
        if slab.get("type") != "additional_pct":
            continue
        from_pct = float(slab.get("from_pct", 0))
        to_pct = float(slab.get("to_pct", 9999))
        rate = float(slab.get("value", 0))
        if achievement_pct < from_pct:
            continue
        slab_lower_amt = (from_pct / 100.0) * target
        slab_upper_amt = (to_pct / 100.0) * target
        applicable = min(actual_sales, slab_upper_amt) - slab_lower_amt
        if applicable > 0:
            amt = applicable * (rate / 100.0)
            earned += amt
            breakdown.append({
                "slab": f"{from_pct:g}%-{to_pct:g}%",
                "type": "% of Additional Sale",
                "applied_on": round(applicable, 2),
                "rate": f"{rate:g}%",
                "amount": round(amt, 2),
            })

    # 2) Total % slab
    # Pick the slab with the HIGHEST from_pct that the barber has crossed.
    # If achievement_pct lands inside an explicit [from, to) range, that slab wins.
    # If achievement_pct is past the highest defined to_pct, the highest crossed
    # slab still applies (the barber doesn't lose the bonus for over-performing).
    total_pct_match = None
    for slab in slabs:
        if slab.get("type") != "total_pct":
            continue
        from_pct = float(slab.get("from_pct", 0))
        to_pct = float(slab.get("to_pct", 9999))
        if achievement_pct < from_pct:
            continue
        # This slab is "crossed" — keep it as the best match so far.
        # Slabs are sorted ascending by from_pct, so later iterations may overwrite.
        total_pct_match = slab
        if achievement_pct < to_pct or to_pct >= 9999:
            # Exact range match — no higher slab can be better.
            break

    if total_pct_match is not None:
        from_pct = float(total_pct_match.get("from_pct", 0))
        to_pct = float(total_pct_match.get("to_pct", 9999))
        rate = float(total_pct_match.get("value", 0))
        amt = actual_sales * (rate / 100.0)
        earned += amt
        breakdown.append({
            "slab": f"{from_pct:g}%-{to_pct:g}%",
            "type": "% of Total Sale",
            "applied_on": round(actual_sales, 2),
            "rate": f"{rate:g}%",
            "amount": round(amt, 2),
        })

    # 3) Fixed bonuses (additive when achievement crosses from_pct)
    for slab in slabs:
        if slab.get("type") != "fixed_amount":
            continue
        from_pct = float(slab.get("from_pct", 0))
        amount = float(slab.get("value", 0))
        if achievement_pct >= from_pct:
            earned += amount
            breakdown.append({
                "slab": f"{from_pct:g}%+",
                "type": "Fixed Bonus",
                "applied_on": "—",
                "rate": "—",
                "amount": round(amount, 2),
            })

    return {
        "earned": round(earned, 2),
        "achievement_pct": round(achievement_pct, 2),
        "breakdown": breakdown,
    }


async def _get_effective_plan_for_barber(salon_id: str, barber_id: str) -> Optional[dict]:
    """Return the incentive plan config that applies to this barber.

    Override priority: individual_plans[barber_id] > global_plan (if barber is
    assigned/all). Returns None if no plan applies.
    """
    config = await db.salon_reward_plans.find_one({"salon_id": salon_id}, {"_id": 0})
    if not config:
        return None
    individual_plans = config.get("individual_plans") or {}
    if barber_id in individual_plans and individual_plans[barber_id]:
        return individual_plans[barber_id]
    mode = config.get("mode", "all")
    if mode == "individual":
        return None
    assigned = config.get("assigned_barber_ids") or []
    # If empty assigned_barber_ids in "all" mode, treat as "applies to everyone"
    if mode == "all" and not assigned:
        return config.get("global_plan")
    if barber_id in assigned:
        return config.get("global_plan")
    return None


async def _get_barber_target(plan_config: dict, barber: dict) -> float:
    """Compute monthly target from plan + barber salary."""
    target_type = (plan_config or {}).get("target_type", "salary_multiplier")
    if target_type == "manual":
        return float(plan_config.get("manual_target") or 0)
    salary = float(barber.get("compensation") or 0)
    multiplier = float(plan_config.get("multiplier") or 0)
    return salary * multiplier


async def _get_barber_actual_sales(salon_id: str, barber_id: str, year_month: str) -> float:
    """Sum of revenue credited to this barber on completed tokens in the month.

    Module 7: tokens may have `service_assignments` that split revenue across
    barbers. We sum `line_attributed_revenue` per assignment for this barber
    (pro-rata discount applied). Legacy tokens (no `service_assignments`) fall
    back to crediting the full `total_amount` to the token's main `barber_id`.
    """
    start = f"{year_month}-01"
    try:
        y, m = year_month.split("-")
        y_i, m_i = int(y), int(m)
        if m_i == 12:
            next_first = f"{y_i + 1}-01-01"
        else:
            next_first = f"{y_i}-{m_i + 1:02d}-01"
    except Exception:
        return 0.0

    cursor = db.tokens.find({
        "salon_id": salon_id,
        "status": "completed",
        # Match either as the main barber OR as a line-assignment barber.
        "$and": [
            {"$or": [
                {"barber_id": barber_id},
                {"service_assignments.barber_id": barber_id},
            ]},
            {"$or": [
                {"date": {"$gte": start, "$lt": next_first}},
                {"booking_date": {"$gte": start, "$lt": next_first}},
                {
                    "date": {"$in": [None, ""]},
                    "booking_date": {"$in": [None, ""]},
                    "created_at": {"$gte": start, "$lt": next_first + "T99"},
                },
            ]},
        ],
    }, {"_id": 0})

    total = 0.0
    async for tok in cursor:
        per_barber = attribute_token_revenue_to_barbers(tok)
        total += per_barber.get(barber_id, 0.0)
    return total


async def _recompute_incentive_payout(salon_id: str, barber_id: str, year_month: str) -> Optional[dict]:
    """Recompute and upsert the incentive_payouts row for a barber+month. Preserves status."""
    plan = await _get_effective_plan_for_barber(salon_id, barber_id)
    if not plan:
        return None

    barber = await db.barbers.find_one({"id": barber_id, "salon_id": salon_id}, {"_id": 0})
    if not barber:
        return None

    target = await _get_barber_target(plan, barber)
    actual_sales = await _get_barber_actual_sales(salon_id, barber_id, year_month)
    calc = _compute_incentive_amount(plan, target, actual_sales)

    existing = await db.incentive_payouts.find_one(
        {"salon_id": salon_id, "barber_id": barber_id, "month": year_month}, {"_id": 0}
    )

    payout = {
        "id": existing.get("id") if existing else str(uuid.uuid4()),
        "salon_id": salon_id,
        "barber_id": barber_id,
        "barber_name": barber.get("name", ""),
        "month": year_month,
        "salary": float(barber.get("compensation") or 0),
        "target": round(float(target), 2),
        "actual_sales": round(actual_sales, 2),
        "achievement_pct": calc["achievement_pct"],
        "incentive_earned": calc["earned"],
        "breakdown": calc["breakdown"],
        "status": (existing or {}).get("status") or ("Pending" if calc["earned"] > 0 else "Pending"),
        "payment_method": (existing or {}).get("payment_method"),
        "paid_at": (existing or {}).get("paid_at"),
        "linked_expense_id": (existing or {}).get("linked_expense_id"),
        "manual_amount": (existing or {}).get("manual_amount"),
        "notes": (existing or {}).get("notes", ""),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "created_at": (existing or {}).get("created_at") or datetime.now(timezone.utc).isoformat(),
    }
    await db.incentive_payouts.update_one(
        {"salon_id": salon_id, "barber_id": barber_id, "month": year_month},
        {"$set": payout},
        upsert=True,
    )
    return payout


@api_router.get("/salons/{salon_id}/reward-plan")
async def get_reward_plan(salon_id: str, current_user=Depends(get_current_salon_user)):
    config = await db.salon_reward_plans.find_one({"salon_id": salon_id}, {"_id": 0})
    if not config:
        return {
            "salon_id": salon_id,
            "mode": "all",
            "global_plan": None,
            "assigned_barber_ids": [],
            "individual_plans": {},
        }
    return config


@api_router.post("/salons/{salon_id}/reward-plan")
async def save_reward_plan(salon_id: str, body: RewardPlanCreate, current_user=Depends(get_current_salon_user)):
    """Save the salon's reward plan configuration (admin only)."""
    if not has_module_permission(current_user, "staff", "access_control"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.access_control")

    if body.mode not in ("all", "individual", "partial"):
        raise HTTPException(status_code=400, detail="mode must be one of: all, individual, partial")

    doc = {
        "salon_id": salon_id,
        "mode": body.mode,
        "global_plan": body.global_plan.dict() if body.global_plan else None,
        "assigned_barber_ids": body.assigned_barber_ids or [],
        "individual_plans": {bid: cfg.dict() for bid, cfg in (body.individual_plans or {}).items()},
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.salon_reward_plans.update_one(
        {"salon_id": salon_id}, {"$set": doc}, upsert=True
    )
    return {"message": "Reward plan saved", "config": doc}


@api_router.get("/salons/{salon_id}/reward-plan/eligible-barbers")
async def get_eligible_barbers(salon_id: str, current_user=Depends(get_current_salon_user)):
    """List barbers that can be assigned to a reward plan (those marked Visible to Customers)."""
    cursor = db.barbers.find(
        {"salon_id": salon_id, "is_barber": True},
        {"_id": 0, "id": 1, "name": 1, "compensation": 1, "is_barber": 1, "designation": 1}
    )
    items = []
    async for b in cursor:
        items.append(b)
    return {"barbers": items}


@api_router.get("/salons/{salon_id}/reward-plan/incentives")
async def list_incentives(
    salon_id: str,
    month: Optional[str] = None,    # YYYY-MM
    barber_id: Optional[str] = None,
    current_user=Depends(get_current_salon_user),
):
    """Recompute all eligible barbers for the month and return the latest payout rows.

    This computes on demand so the dashboard always reflects the latest sales.
    """
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")

    # Get all eligible (is_barber=True) members
    barber_filter = {"salon_id": salon_id, "is_barber": True}
    if barber_id:
        barber_filter["id"] = barber_id

    barbers_cursor = db.barbers.find(barber_filter, {"_id": 0})
    rows = []
    async for b in barbers_cursor:
        payout = await _recompute_incentive_payout(salon_id, b["id"], month)
        if payout:
            rows.append(payout)

    # Also include any barbers who have payouts saved historically but might not be in barber_filter
    if not barber_id:
        existing_cursor = db.incentive_payouts.find(
            {"salon_id": salon_id, "month": month}, {"_id": 0}
        )
        seen = {r["barber_id"] for r in rows}
        async for p in existing_cursor:
            if p["barber_id"] not in seen:
                rows.append(p)

    return {"month": month, "incentives": rows}


@api_router.put("/salons/{salon_id}/reward-plan/incentives/{barber_id}/{month}/status")
async def update_incentive_status(
    salon_id: str,
    barber_id: str,
    month: str,
    body: IncentivePayoutStatusUpdate,
    current_user=Depends(get_current_salon_user),
):
    """Update payout status. On 'Paid', creates a financial expense entry and links it."""
    if body.status not in ("Pending", "Approved", "Paid", "Hold"):
        raise HTTPException(status_code=400, detail="Invalid status")

    # RBAC: staff.salary_pay
    if not has_module_permission(current_user, "staff", "salary_pay"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.salary_pay")

    payout = await db.incentive_payouts.find_one(
        {"salon_id": salon_id, "barber_id": barber_id, "month": month}, {"_id": 0}
    )
    if not payout:
        # Recompute if missing
        payout = await _recompute_incentive_payout(salon_id, barber_id, month)
        if not payout:
            raise HTTPException(status_code=404, detail="No incentive computed for this barber+month")

    update = {
        "status": body.status,
        "notes": body.notes if body.notes is not None else payout.get("notes", ""),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    # If the admin passed a manual_amount, persist it (None = clear back to auto).
    # Allowed only for Approved / Paid / Pending — not Hold.
    if body.manual_amount is not None:
        if float(body.manual_amount) < 0:
            raise HTTPException(status_code=400, detail="manual_amount cannot be negative")
        update["manual_amount"] = float(body.manual_amount)
    elif body.status in ("Approved", "Paid"):
        # Keep whatever was previously stored
        pass

    # Effective payout amount = manual override (if set) else auto-calculated
    effective_manual = (
        update.get("manual_amount")
        if "manual_amount" in update
        else payout.get("manual_amount")
    )
    effective_amount = (
        float(effective_manual)
        if effective_manual is not None
        else float(payout.get("incentive_earned") or 0)
    )

    # Handle "Paid" — must have payment_method, create expense entry
    if body.status == "Paid":
        if not body.payment_method:
            raise HTTPException(status_code=400, detail="payment_method is required when marking Paid")
        if body.payment_method not in ("cash", "upi", "bank"):
            raise HTTPException(status_code=400, detail="payment_method must be cash, upi or bank")

        # Strict idempotency: never create more than ONE financial txn per payout.id.
        # Even if linked_expense_id was somehow lost, look up by reference_id+type.
        existing_txn = None
        if payout.get("linked_expense_id"):
            existing_txn = await db.financial_transactions.find_one(
                {"id": payout["linked_expense_id"], "salon_id": salon_id}, {"_id": 0}
            )
        if not existing_txn:
            existing_txn = await db.financial_transactions.find_one(
                {
                    "salon_id": salon_id,
                    "reference_type": "incentive_payout",
                    "reference_id": payout.get("id"),
                },
                {"_id": 0},
            )

        if existing_txn:
            # Already booked — keep the existing entry; ensure the payout points at it.
            update["linked_expense_id"] = existing_txn["id"]
        elif effective_amount > 0:
            txn = {
                "id": str(uuid.uuid4()),
                "salon_id": salon_id,
                "type": "outflow",
                "category": "staff_incentive",
                "amount": effective_amount,
                "payment_mode": body.payment_method,
                "narration": f"Incentive payout to {payout.get('barber_name', '')} for {month}",
                "reference_id": payout.get("id"),
                "reference_type": "incentive_payout",
                "created_by": current_user.get("id") if isinstance(current_user, dict) else "admin",
                "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.financial_transactions.insert_one(txn)
            update["linked_expense_id"] = txn["id"]

        update["payment_method"] = body.payment_method
        update["paid_at"] = datetime.now(timezone.utc).isoformat()
    else:
        # Reset paid fields if moved away from Paid
        if payout.get("status") == "Paid":
            # Keep linked expense (admin can manually delete it from Financials if needed)
            pass

    await db.incentive_payouts.update_one(
        {"salon_id": salon_id, "barber_id": barber_id, "month": month}, {"$set": update}
    )
    fresh = await db.incentive_payouts.find_one(
        {"salon_id": salon_id, "barber_id": barber_id, "month": month}, {"_id": 0}
    )
    return {"message": "Status updated", "payout": fresh}


# ============ ANALYTICS ROUTES (continued) ============

@api_router.get("/analytics/day-wise-sales")
async def get_day_wise_sales(
    salon_id: str,
    start_date: str,
    end_date: Optional[str] = None,
    current_salon=Depends(get_current_salon_user)
):
    """Get day-wise sales for the salon"""
    if not end_date:
        # Default to 10 days from start_date
        start = datetime.fromisoformat(start_date)
        end = start + timedelta(days=10)
        end_date = end.date().isoformat()
    
    # Get all completed tokens in date range
    tokens = await db.tokens.find({
        "salon_id": salon_id,
        "date": {"$gte": start_date, "$lte": end_date},
        "status": "completed"
    }, {"_id": 0}).to_list(10000)
    
    # Group by date
    day_wise = {}
    for token in tokens:
        date = token["date"]
        if date not in day_wise:
            day_wise[date] = {
                "date": date,
                "total_sales": 0,
                "total_bookings": 0
            }
        day_wise[date]["total_sales"] += token.get("total_amount", 0)
        day_wise[date]["total_bookings"] += 1
    
    # Convert to list and sort
    result = sorted(day_wise.values(), key=lambda x: x["date"])
    return {"data": result}

@api_router.get("/analytics/barber-wise-sales")
async def get_barber_wise_sales(
    salon_id: str,
    start_date: str,
    end_date: str,
    current_salon=Depends(get_current_salon_user)
):
    """Get barber-wise sales for the salon.

    Module 7: respects per-service `service_assignments` so split-barber tokens
    correctly credit each barber's share (pro-rata discount applied).
    """
    tokens = await db.tokens.find({
        "salon_id": salon_id,
        "date": {"$gte": start_date, "$lte": end_date},
        "status": "completed"
    }, {"_id": 0}).to_list(10000)

    # Pre-load barber names so split-token credits show the right label even if
    # `barber_name_snapshot` is missing.
    barber_id_set: set = set()
    for tok in tokens:
        if tok.get("barber_id"):
            barber_id_set.add(tok["barber_id"])
        for a in (tok.get("service_assignments") or []):
            if a.get("barber_id"):
                barber_id_set.add(a["barber_id"])
    barbers = await db.barbers.find(
        {"id": {"$in": list(barber_id_set)}}, {"_id": 0, "id": 1, "name": 1}
    ).to_list(1000)
    name_by_id = {b["id"]: b.get("name", "Unknown") for b in barbers}

    barber_wise: Dict[str, Dict[str, Any]] = {}
    for token in tokens:
        per_barber = attribute_token_revenue_to_barbers(token)
        # Track which barbers participated in this token for booking-count math.
        # A split token counts as ONE booking for every participating barber.
        contributing = set(per_barber.keys())
        if not contributing and token.get("barber_id"):
            contributing.add(token["barber_id"])
        for bid in contributing:
            if bid not in barber_wise:
                barber_wise[bid] = {
                    "barber_id": bid,
                    "barber_name": name_by_id.get(bid) or token.get("barber_name") or "Unknown",
                    "total_sales": 0.0,
                    "total_bookings": 0,
                }
            barber_wise[bid]["total_sales"] += per_barber.get(bid, 0.0)
            barber_wise[bid]["total_bookings"] += 1

    for row in barber_wise.values():
        row["total_sales"] = round(row["total_sales"], 2)

    result = sorted(barber_wise.values(), key=lambda x: x["total_sales"], reverse=True)
    return {"data": result}

@api_router.get("/analytics/service-wise-sales")
async def get_service_wise_sales(
    salon_id: str,
    start_date: str,
    end_date: str,
    current_salon=Depends(get_current_salon_user)
):
    """Get top 10 services by count + revenue.

    Module 7: revenue uses real per-service line_total (pro-rata discount)
    when `service_assignments` is available; legacy tokens fall back to an
    even distribution of total_amount across selected_services.
    """
    tokens = await db.tokens.find({
        "salon_id": salon_id,
        "date": {"$gte": start_date, "$lte": end_date},
        "status": "completed"
    }, {"_id": 0}).to_list(10000)

    all_services = await db.services.find({}, {"_id": 0}).to_list(2000)
    name_by_id = {s["id"]: s.get("service_name", "Unknown Service") for s in all_services}

    bucket: Dict[str, Dict[str, Any]] = {}
    for token in tokens:
        rows = attribute_token_revenue_to_services(token)
        for r in rows:
            sid = r.get("service_id")
            if not sid:
                continue
            if sid not in bucket:
                bucket[sid] = {
                    "service_id": sid,
                    "service_name": name_by_id.get(sid, "Unknown Service"),
                    "count": 0,
                    "revenue": 0.0,
                }
            bucket[sid]["count"] += 1
            bucket[sid]["revenue"] += float(r.get("line_total") or 0)

    for row in bucket.values():
        row["revenue"] = round(row["revenue"], 2)

    result = sorted(bucket.values(), key=lambda x: x["count"], reverse=True)[:10]
    return {"data": result}

@api_router.get("/analytics/gender-distribution")
async def get_gender_distribution(
    salon_id: str,
    start_date: str,
    end_date: str,
    current_salon=Depends(get_current_salon_user)
):
    """Get gender distribution of customers"""
    # Get all completed tokens in date range
    tokens = await db.tokens.find({
        "salon_id": salon_id,
        "date": {"$gte": start_date, "$lte": end_date},
        "status": "completed"
    }, {"_id": 0}).to_list(10000)
    
    # Get unique user IDs
    user_ids = list(set([t.get("user_id") for t in tokens if t.get("user_id")]))
    
    # Get user gender data
    users = await db.users.find(
        {"id": {"$in": user_ids}},
        {"_id": 0, "id": 1, "gender": 1}
    ).to_list(10000)
    
    # Count by gender
    gender_count = {"Men": 0, "Women": 0, "Not Specified": 0}
    
    for user in users:
        gender = user.get("gender", "Not Specified")
        if gender in gender_count:
            gender_count[gender] += 1
        else:
            gender_count["Not Specified"] += 1
    
    result = [
        {"name": "Men", "value": gender_count["Men"]},
        {"name": "Women", "value": gender_count["Women"]},
        {"name": "Not Specified", "value": gender_count["Not Specified"]}
    ]
    
    return {"data": result}

@api_router.get("/analytics/detailed-report")
async def get_detailed_report(
    salon_id: str,
    start_date: str,
    end_date: str,
    current_salon=Depends(get_current_salon_user)
):
    """Get detailed booking report for export"""
    # Get all tokens in date range
    tokens = await db.tokens.find({
        "salon_id": salon_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Get all services for name mapping
    services_dict = {}
    all_services = await db.services.find({}, {"_id": 0}).to_list(1000)
    for service in all_services:
        services_dict[service["id"]] = service["service_name"]
    
    # Format data for export
    detailed_data = []
    for token in tokens:
        # Get service names
        service_names = [services_dict.get(sid, "Unknown") for sid in token.get("selected_services", [])]
        
        # Calculate time taken
        time_taken = ""
        if token.get("called_at") and token.get("completed_at"):
            called = datetime.fromisoformat(token["called_at"])
            completed = datetime.fromisoformat(token["completed_at"])
            duration = completed - called
            minutes = int(duration.total_seconds() / 60)
            time_taken = f"{minutes} min"
        
        detailed_data.append({
            "date": token["date"],
            "token_number": token.get("token_number", ""),
            "customer_name": token["customer_name"],
            "phone": token["phone"],
            "barber_name": token.get("barber_name", ""),
            "services": ", ".join(service_names),
            "amount": token.get("total_amount", 0),
            "status": token.get("status", ""),
            "shift": token.get("shift", token.get("time_slot", "")),
            "call_time": token.get("called_at", ""),
            "complete_time": token.get("completed_at", ""),
            "time_taken": time_taken,
            "payment_status": token.get("payment_status", "")
        })
    
    return {"data": detailed_data}

# ============ PAYMENT ROUTES ============

@api_router.post("/payments/generate-upi-qr")
async def generate_upi_qr(token_id: str):
    """Generate UPI QR code for payment"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    salon = await db.salons.find_one({"id": token["salon_id"]}, {"_id": 0})
    if not salon or not salon.get("upi_id"):
        raise HTTPException(status_code=400, detail="Salon UPI not configured")
    
    # Generate UPI URL
    upi_url = f"upi://pay?pa={salon['upi_id']}&pn={salon['salon_name']}&am={token['total_amount']}&cu=INR&tn=Token_{token['token_number']}"
    
    # Generate QR
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(upi_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    img_str = base64.b64encode(buffer.getvalue()).decode()
    
    return {
        "qr_code": f"data:image/png;base64,{img_str}",
        "upi_url": upi_url,
        "amount": token['total_amount']
    }

@api_router.post("/payments/verify")
async def verify_payment(token_id: str, transaction_id: str, current_salon=Depends(get_current_salon)):
    """Verify payment (manual for UPI) - salon admin"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {
            "payment_status": "paid",
            "payment_mode": "upi",
            "upi_transaction_id": transaction_id
        }}
    )
    
    return {"message": "Payment verified"}

# ============ CUSTOMER PAYMENT CONFIRM ============

@api_router.post("/payments/customer-confirm-upi")
async def customer_confirm_upi(body: dict):
    """Customer confirms UPI payment was done"""
    token_id = body.get("token_id")
    upi_ref = body.get("upi_reference", "")
    
    if not token_id:
        raise HTTPException(status_code=400, detail="Token ID required")
    
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    await db.tokens.update_one(
        {"id": token_id},
        {"$set": {
            "payment_status": "paid",
            "payment_mode": "upi",
            "upi_transaction_id": upi_ref or "Customer confirmed"
        }}
    )
    
    return {"message": "Payment confirmed", "token_id": token_id}

# ============ USER HISTORY ============

@api_router.get("/user/{user_id}/history", response_model=List[TokenModel])
async def get_user_history(user_id: str):
    tokens = await db.tokens.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return tokens

@api_router.get("/user/{user_id}/active", response_model=List[TokenModel])
async def get_user_active_bookings(user_id: str):
    today = datetime.now().date().isoformat()
    tokens = await db.tokens.find(
        {"user_id": user_id, "date": {"$gte": today}, "status": {"$in": ["waiting", "in_progress"]}},
        {"_id": 0}
    ).sort("date", 1).to_list(10)
    return tokens

@api_router.get("/customers/{phone}/active-bookings")
async def get_customer_active_bookings(phone: str):
    """Get active bookings for a customer by phone, enriched with live queue context."""
    # Normalize phone
    if not phone.startswith("+91"):
        phone = f"+91{phone}"
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get active bookings (waiting or in_service)
    tokens = await db.tokens.find(
        {
            "phone": phone,
            "date": today,
            "status": {"$in": ["waiting", "called", "in_progress", "in_service"]},
            "$or": [{"cancelled": False}, {"cancelled": {"$exists": False}}]
        },
        {"_id": 0}
    ).sort("created_at", 1).to_list(10)
    
    # Enrich with salon + live queue details
    for token in tokens:
        salon = await db.salons.find_one({"id": token["salon_id"]}, {"_id": 0})
        if salon:
            token["salon_details"] = {
                "salon_name": salon.get("salon_name"),
                "address": salon.get("address"),
                "city": salon.get("city"),
                "logo_url": salon.get("logo_url")
            }
        # Branch enrichment so the customer card can show "Salon · Whitefield" + branch address
        if token.get("branch_id"):
            branch = await db.salon_branches.find_one(
                {"id": token["branch_id"]}, {"_id": 0, "branch_name": 1, "address": 1, "is_main_branch": 1}
            )
            if branch and not branch.get("is_main_branch"):
                token["branch_name"] = branch.get("branch_name")
                token["branch_address"] = branch.get("address")
        # Live queue context (customer-facing only)
        try:
            q = await compute_queue_status_for_token(token)
            # Inject fields the frontend already reads, plus new ones
            token["queue_position"] = q["people_before"]       # "N ahead of you"
            token["people_before"] = q["people_before"]
            token["barber_position"] = q["position"]            # 1-indexed, #3 style
            token["estimated_wait_minutes"] = q["estimated_wait_minutes"]
            token["queue_status_message"] = q["status_message"]
            token["currently_serving"] = q["currently_serving"]
            token["approx_finish_minutes"] = q["approx_finish_minutes"]
        except Exception as e:
            logger.error(f"Queue enrichment failed for token {token.get('id')}: {e}")

    return {"active_bookings": tokens, "count": len(tokens)}


@api_router.get("/tokens/{token_id}/queue-status")
async def get_token_queue_status(token_id: str):
    """Customer-facing queue status for a single token.

    Returns total token, barber name, barber-queue position, people before,
    estimated wait time (75% rule), live serving info, and a status message.
    """
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    try:
        q = await compute_queue_status_for_token(token)
    except Exception as e:
        logger.error(f"Error computing queue status: {e}")
        raise HTTPException(status_code=500, detail="Unable to compute queue status")
    return q


# ============ INVOICE ROUTES ============

@api_router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str):
    """Get invoice data"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Remove base64 from response (too large)
    invoice_data = {k: v for k, v in invoice.items() if k != 'pdf_base64'}
    return invoice_data

@api_router.get("/invoices/{invoice_id}/view")
async def view_invoice(invoice_id: str):
    """View invoice PDF in browser"""
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if not invoice.get('pdf_base64'):
        raise HTTPException(status_code=404, detail="Invoice PDF not found")
    
    # Decode base64 PDF
    import base64
    pdf_bytes = base64.b64decode(invoice['pdf_base64'])
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"inline; filename=invoice_{invoice['invoice_no']}.pdf"
        }
    )

@api_router.get("/invoices/{invoice_id}/download")
async def download_invoice(invoice_id: str):
    """Download invoice PDF"""
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if not invoice.get('pdf_base64'):
        raise HTTPException(status_code=404, detail="Invoice PDF not found")
    
    # Decode base64 PDF
    import base64
    pdf_bytes = base64.b64decode(invoice['pdf_base64'])
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=invoice_{invoice['invoice_no']}.pdf"
        }
    )

@api_router.get("/tokens/{token_id}/invoice")
async def get_token_invoice(token_id: str):
    """Get invoice for a specific token"""
    token = await db.tokens.find_one({"id": token_id}, {"_id": 0})
    if not token:
        raise HTTPException(status_code=404, detail="Token not found")
    
    if not token.get('invoice_id'):
        raise HTTPException(status_code=404, detail="Invoice not generated yet")
    
    invoice = await db.invoices.find_one({"id": token['invoice_id']}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Return invoice data without base64
    invoice_data = {k: v for k, v in invoice.items() if k != 'pdf_base64'}
    invoice_data['view_link'] = f"/api/invoices/{invoice['id']}/view"
    invoice_data['download_link'] = f"/api/invoices/{invoice['id']}/download"
    
    return invoice_data

# Moved app.include_router to end of all routes (before scheduler)

fastapi_app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# MEMBERSHIP EXPIRY NOTIFICATIONS — runs daily to notify customers whose
# membership expires in 30 or 7 days. Honors customer notification settings.
# ============================================================================
async def notify_expiring_memberships():
    """Send 30-days-before and 7-days-before expiry reminders."""
    try:
        now = datetime.now(timezone.utc)
        # Fetch all active, non-cancelled memberships with an expiry date
        cursor = db.customer_memberships.find(
            {"is_active": True, "$or": [{"cancelled": False}, {"cancelled": {"$exists": False}}]},
            {"_id": 0},
        )
        async for m in cursor:
            try:
                expiry = datetime.fromisoformat(str(m["expiry_date"]).replace("Z", "+00:00"))
            except Exception:
                continue
            days_left = (expiry - now).days
            phone = m.get("customer_phone")
            if not phone:
                continue

            # 30-day reminder
            if 28 <= days_left <= 30 and not m.get("notified_1m_expiry"):
                await _send_expiry_reminder(m, days_left, "1m")
                await db.customer_memberships.update_one(
                    {"id": m["id"]}, {"$set": {"notified_1m_expiry": True}}
                )
            # 7-day reminder
            if 6 <= days_left <= 7 and not m.get("notified_1w_expiry"):
                await _send_expiry_reminder(m, days_left, "1w")
                await db.customer_memberships.update_one(
                    {"id": m["id"]}, {"$set": {"notified_1w_expiry": True}}
                )
    except Exception as e:
        logger.error(f"notify_expiring_memberships failed: {e}")


async def _send_expiry_reminder(m: Dict[str, Any], days_left: int, key: str):
    phone = m["customer_phone"]
    salon_id = m.get("salon_id", "")
    name = m.get("membership_name", "Membership")
    label = "30 days" if key == "1m" else "7 days"
    title = f"Your {name} membership expires in ~{label}"
    body = (
        f"Your {name} membership will expire in about {label}. "
        f"Renew with the salon to keep your wallet benefits."
    )
    try:
        await create_in_app_notification(
            user_type="customer",
            user_id=phone,
            title=title,
            message=body,
            notification_type="membership_expiry",
            setting_key="membership_expiry",
            salon_id=salon_id,
            related_id=m.get("id", ""),
        )
    except Exception as e:
        logger.error(f"in-app expiry notification failed: {e}")

    # WhatsApp (respects customer's whatsapp_membership_expiry setting)
    try:
        if await should_send_customer_whatsapp(phone, "whatsapp_membership_expiry"):
            msg = (
                f"Hi {m.get('customer_name','')}, your {name} membership expires in ~{label} "
                f"(on {str(m.get('expiry_date',''))[:10]}). Please visit the salon to renew."
            )
            await send_whatsapp_notification(phone, msg, f"membership_expiry_{key}")
    except Exception as e:
        logger.error(f"whatsapp expiry notification failed: {e}")


# ============ STAFF ATTENDANCE ENDPOINTS ============

async def calculate_barber_attendance_for_date(salon_id: str, barber_id: str, date_str: str) -> dict:
    """Calculate attendance for a barber on a specific date.

    Module 4 — Mode-aware dispatch:
      • Looks up which attendance_mode was active on `date_str` (via the
        salon's attendance_mode_history; defaults to "service_completion").
      • Mode A ("service_completion"): the existing booking-completion rule.
      • Mode B ("geo_checkin"): read raw check-in/out fields and apply the
        late-checkin / short-hours rules from compute_mode_b_status.

    Always returns the same dict shape (status, bookings_count, shift flags)
    so existing callers stay happy; under Mode B we additionally surface
    total_minutes / half_day_reason / computed_under_mode.
    """
    # Look up barber to apply employment-window / leave gating.
    barber = await db.barbers.find_one(
        {"id": barber_id, "salon_id": salon_id},
        {"_id": 0, "doj": 1, "last_working_date": 1, "leave_dates": 1, "branch_id": 1}
    ) or {}

    leave_dates = barber.get("leave_dates") or []
    doj = (barber.get("doj") or "").strip()
    lwd = (barber.get("last_working_date") or "").strip()

    # Outside employment window → absent regardless of mode.
    if doj and date_str < doj:
        return {"status": "absent", "bookings_count": 0,
                "morning_shift_completed": False, "noon_evening_shift_completed": False,
                "computed_under_mode": None}
    if lwd and date_str > lwd:
        return {"status": "absent", "bookings_count": 0,
                "morning_shift_completed": False, "noon_evening_shift_completed": False,
                "computed_under_mode": None}

    # Module 2 — leave_records take priority across both modes (active record on this date).
    on_leave_record = await db.leave_records.find_one({
        "salon_id": salon_id, "barber_id": barber_id, "date": date_str,
        "status": {"$ne": "cancelled"},
    }, {"_id": 0, "id": 1})

    # Resolve the mode that governs this date.
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0}) or {}
    mode = attendance_mode_mod.resolve_mode_for_date(salon, date_str)

    if mode == "geo_checkin":
        # Mode B — read raw check-in/out from the existing attendance doc (if any).
        record_id = f"{salon_id}_{barber_id}_{date_str}"
        existing = await db.attendance.find_one({"id": record_id}, {"_id": 0}) or {}
        # Legacy back-compat: leave_dates list also marks absent under Mode B.
        on_leave = bool(on_leave_record) or (isinstance(leave_dates, list) and date_str in leave_dates)
        day_passed = date_str < attendance_mode_mod.current_ist_date()
        computed = attendance_mode_mod.compute_mode_b_status(salon, existing, day_has_passed=day_passed, on_leave=on_leave)
        return {
            "status": computed["status"],
            "bookings_count": int(existing.get("bookings_count") or 0),
            "morning_shift_completed": bool(existing.get("morning_shift_completed", False)),
            "noon_evening_shift_completed": bool(existing.get("noon_evening_shift_completed", False)),
            "total_minutes": computed["total_minutes"],
            "half_day_reason": computed["half_day_reason"],
            "computed_under_mode": "geo_checkin",
        }

    # ---- Mode A — service_completion (existing logic, unchanged) -------------
    # Match by the booking's IST `date` field directly — comparing the UTC
    # `completed_at` ISO-string against an IST-boundary range fails on
    # early-morning IST completions where UTC is still on the previous day.
    completed_bookings = await db.tokens.find({
        "salon_id": salon_id,
        "barber_id": barber_id,
        "status": "completed",
        "date": date_str,
    }, {"_id": 0, "shift": 1, "token_number": 1}).to_list(100)

    bookings_count = len(completed_bookings)
    shifts_with_bookings = set()
    for booking in completed_bookings:
        shift = booking.get("shift", "").lower()
        if shift:
            shifts_with_bookings.add(shift)
    morning_shift = "morning" in shifts_with_bookings
    noon_evening_shift = "noon" in shifts_with_bookings or "evening" in shifts_with_bookings

    on_leave = bool(on_leave_record) or (isinstance(leave_dates, list) and date_str in leave_dates)
    if on_leave:
        return {"status": "absent", "bookings_count": bookings_count,
                "morning_shift_completed": morning_shift,
                "noon_evening_shift_completed": noon_evening_shift,
                "computed_under_mode": "service_completion"}

    status_val = "present" if bookings_count >= 1 else "absent"
    return {
        "status": status_val,
        "bookings_count": bookings_count,
        "morning_shift_completed": morning_shift,
        "noon_evening_shift_completed": noon_evening_shift,
        "computed_under_mode": "service_completion",
    }


@api_router.get("/salons/{salon_id}/staff-attendance/month/{month}")
async def get_monthly_attendance(
    salon_id: str, 
    month: str,  # YYYY-MM format
    barber_id: Optional[str] = None
):
    """Get attendance records for a month. If barber_id specified, returns only that barber's records."""
    # Validate month format
    try:
        year, mon = month.split("-")
        year, mon = int(year), int(mon)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    
    # Get all barbers for this salon if no specific barber requested
    query = {"salon_id": salon_id}
    if barber_id:
        query["barber_id"] = barber_id
    
    # Get existing attendance records
    records = await db.attendance.find({
        **query,
        "date": {"$regex": f"^{month}"}
    }, {"_id": 0}).to_list(1000)
    
    # Get barbers
    barber_query = {"salon_id": salon_id, "is_barber": True, "is_active": True}
    if barber_id:
        barber_query["id"] = barber_id
    barbers = await db.barbers.find(barber_query, {"_id": 0, "id": 1, "name": 1, "compensation": 1, "branch_id": 1}).to_list(100)

    # Module 4 — surface "has_login" so the UI can warn admins that a
    # staff under Mode B will rely on admin-manual marking.
    barber_ids = [b["id"] for b in barbers]
    if barber_ids:
        users_with_login = await db.salon_users.find(
            {"salon_id": salon_id, "staff_id": {"$in": barber_ids}, "status": "active"},
            {"_id": 0, "staff_id": 1},
        ).to_list(length=500)
    else:
        users_with_login = []
    staff_with_login = {u.get("staff_id") for u in users_with_login if u.get("staff_id")}

    # Resolve current mode for the UI to render the right hint.
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0, "attendance_mode": 1}) or {}
    current_mode = salon.get("attendance_mode") or "service_completion"

    # Build response
    response = {
        "month": month,
        "attendance_mode": current_mode,
        "barbers": []
    }

    for barber in barbers:
        barber_records = [r for r in records if r.get("barber_id") == barber["id"]]
        has_login = barber["id"] in staff_with_login
        # Under Mode B, a staff without login can't self check-in — admin
        # must mark manually.  Surface this so the UI/admin sees it.
        no_checkin_capability = (current_mode == "geo_checkin") and not has_login
        response["barbers"].append({
            "barber_id": barber["id"],
            "barber_name": barber["name"],
            "branch_id": barber.get("branch_id"),
            "compensation": barber.get("compensation", 0),
            "has_login": has_login,
            "no_checkin_capability": no_checkin_capability,
            "attendance": barber_records
        })

    return response


@api_router.post("/salons/{salon_id}/staff-attendance/calculate/{date}")
async def calculate_daily_attendance(
    salon_id: str,
    date: str,  # YYYY-MM-DD format
    current_user=Depends(get_current_salon_user)
):
    """Calculate attendance for all barbers for a specific date based on completed bookings."""
    # RBAC: staff.attendance
    if not has_module_permission(current_user, "staff", "attendance"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.attendance")
    
    # Get all active barbers
    barbers = await db.barbers.find({
        "salon_id": salon_id,
        "is_barber": True,
        "is_active": True
    }, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    
    results = []
    for barber in barbers:
        # Module 4 — lock-on-paid: skip barbers whose month's salary is already paid.
        locked = await attendance_mode_mod.is_attendance_locked(db, salon_id, barber["id"], date)
        if locked:
            existing_locked = await db.attendance.find_one(
                {"id": f"{salon_id}_{barber['id']}_{date}"}, {"_id": 0}
            )
            if existing_locked:
                results.append(existing_locked)
            continue
        # Check if already has an override
        existing = await db.attendance.find_one({
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "date": date,
            "auto_calculated": False  # Manual override exists
        }, {"_id": 0})
        
        if existing:
            results.append(existing)
            continue
        
        # Calculate attendance
        calc = await calculate_barber_attendance_for_date(salon_id, barber["id"], date)

        # Create or update record
        record_id = f"{salon_id}_{barber['id']}_{date}"
        record = {
            "id": record_id,
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "date": date,
            "status": calc["status"],
            "auto_calculated": True,
            "morning_shift_completed": calc.get("morning_shift_completed", False),
            "noon_evening_shift_completed": calc.get("noon_evening_shift_completed", False),
            "bookings_count": calc.get("bookings_count", 0),
            "computed_under_mode": calc.get("computed_under_mode"),
            "half_day_reason": calc.get("half_day_reason"),
            "total_minutes": calc.get("total_minutes"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.attendance.update_one(
            {"id": record_id},
            {"$set": record},
            upsert=True
        )
        results.append(record)
    
    return {"date": date, "attendance": results}


@api_router.put("/salons/{salon_id}/staff-attendance/override/{barber_id}/{date}")
async def override_attendance(
    salon_id: str,
    barber_id: str,
    date: str,  # YYYY-MM-DD format
    body: AttendanceOverride,
    current_user=Depends(get_current_salon_user)
):
    """Admin override for attendance. Click on calendar date to change status."""
    # RBAC: staff.attendance
    if not has_module_permission(current_user, "staff", "attendance"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.attendance")
    
    # Validate status
    if body.status not in ["present", "half_day", "absent", "holiday", "on_leave"]:
        raise HTTPException(status_code=400, detail="Invalid status. Use: present, half_day, absent, holiday, on_leave")
    
    # Fetch barber to validate doj / last_working_date when status implies attending work
    barber = await db.barbers.find_one({"id": barber_id, "salon_id": salon_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    if body.status in ("present", "half_day"):
        doj = (barber.get("doj") or "").strip()
        if doj and date < doj:
            raise HTTPException(status_code=400, detail=f"Cannot mark present before joining date ({doj})")
        lwd = (barber.get("last_working_date") or "").strip()
        if lwd and date > lwd:
            raise HTTPException(status_code=400, detail=f"Cannot mark present after last working day ({lwd})")
    
    # Module 4 — lock-on-paid: refuse to modify a month whose salary is paid.
    locked = await attendance_mode_mod.is_attendance_locked(db, salon_id, barber_id, date)
    if locked:
        raise HTTPException(status_code=423, detail=f"Salary for {locked} is already paid; attendance is locked")

    record_id = f"{salon_id}_{barber_id}_{date}"
    
    # Check if record exists
    existing = await db.attendance.find_one({"id": record_id}, {"_id": 0})
    
    now = datetime.now(timezone.utc).isoformat()
    
    if existing:
        # Update existing record
        await db.attendance.update_one(
            {"id": record_id},
            {"$set": {
                "status": body.status,
                "auto_calculated": False,
                "override_by": current_user.get("id"),
                "marked_by_role": current_user.get("role"),
                "marked_by_name": current_user.get("name") or current_user.get("identifier"),
                "override_note": body.note,
                "updated_at": now
            }}
        )
    else:
        # Create new record
        record = {
            "id": record_id,
            "salon_id": salon_id,
            "barber_id": barber_id,
            "date": date,
            "status": body.status,
            "auto_calculated": False,
            "morning_shift_completed": False,
            "noon_evening_shift_completed": False,
            "bookings_count": 0,
            "override_by": current_user.get("id"),
            "marked_by_role": current_user.get("role"),
            "marked_by_name": current_user.get("name") or current_user.get("identifier"),
            "override_note": body.note,
            "created_at": now,
            "updated_at": now
        }
        await db.attendance.insert_one(record)
    
    updated = await db.attendance.find_one({"id": record_id}, {"_id": 0})
    return updated


@api_router.delete("/salons/{salon_id}/staff-attendance/override/{barber_id}/{date}")
async def clear_attendance_override(
    salon_id: str,
    barber_id: str,
    date: str,  # YYYY-MM-DD
    current_user=Depends(get_current_salon_user),
):
    """
    Clear an attendance entry for the given barber+date.
    Used by the calendar's status cycle:
        present → half_day → absent → holiday → blank (this endpoint).
    """
    if not has_module_permission(current_user, "staff", "attendance"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.attendance")

    # Module 4 — lock-on-paid.
    locked = await attendance_mode_mod.is_attendance_locked(db, salon_id, barber_id, date)
    if locked:
        raise HTTPException(status_code=423, detail=f"Salary for {locked} is already paid; attendance is locked")

    record_id = f"{salon_id}_{barber_id}_{date}"
    res = await db.attendance.delete_one({"id": record_id})
    return {
        "deleted": res.deleted_count > 0,
        "id": record_id,
    }


@api_router.post("/salons/{salon_id}/staff-attendance/mark-all-present/{date}")
async def mark_all_present(
    salon_id: str,
    date: str,  # YYYY-MM-DD
    current_user=Depends(get_current_salon_user)
):
    """Bulk mark every eligible barber 'present' on this date.
    Eligibility: barber.is_active=True AND (no joining date OR doj <= date) AND (no last_working_date OR date <= last_working_date) AND date NOT in barber.leave_dates.
    Returns counts of marked / skipped.
    """
    if current_user.get("role") not in ["admin", "salon_admin", "salon"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Basic date format check
    try:
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    barbers = await db.barbers.find(
        {"salon_id": salon_id, "is_active": True}, {"_id": 0}
    ).to_list(500)

    now = datetime.now(timezone.utc).isoformat()
    marked, skipped = 0, []

    for barber in barbers:
        b_id = barber.get("id")
        # Module 4 — lock-on-paid skip.
        locked = await attendance_mode_mod.is_attendance_locked(db, salon_id, b_id, date)
        if locked:
            skipped.append({"barber_id": b_id, "name": barber.get("name"), "reason": f"locked_{locked}"})
            continue
        # Eligibility checks
        doj = (barber.get("doj") or "").strip()
        if doj and date < doj:
            skipped.append({"barber_id": b_id, "name": barber.get("name"), "reason": "before_joining"})
            continue
        lwd = (barber.get("last_working_date") or "").strip()
        if lwd and date > lwd:
            skipped.append({"barber_id": b_id, "name": barber.get("name"), "reason": "after_last_working_day"})
            continue
        leave_dates = barber.get("leave_dates") or []
        if isinstance(leave_dates, list) and date in leave_dates:
            skipped.append({"barber_id": b_id, "name": barber.get("name"), "reason": "on_leave"})
            continue
        
        record_id = f"{salon_id}_{b_id}_{date}"
        existing = await db.attendance.find_one({"id": record_id}, {"_id": 0})
        if existing:
            await db.attendance.update_one(
                {"id": record_id},
                {"$set": {
                    "status": "present",
                    "auto_calculated": False,
                    "override_by": current_user.get("id"),
                    "override_note": "Bulk mark all present",
                    "updated_at": now
                }}
            )
        else:
            await db.attendance.insert_one({
                "id": record_id,
                "salon_id": salon_id,
                "barber_id": b_id,
                "date": date,
                "status": "present",
                "auto_calculated": False,
                "morning_shift_completed": False,
                "noon_evening_shift_completed": False,
                "bookings_count": 0,
                "override_by": current_user.get("id"),
                "override_note": "Bulk mark all present",
                "created_at": now,
                "updated_at": now,
            })
        marked += 1
    
    return {"date": date, "marked": marked, "skipped": skipped, "total": len(barbers)}


class LeaveToggleBody(BaseModel):
    date: str  # YYYY-MM-DD
    is_on_leave: bool


@api_router.put("/salons/{salon_id}/barbers/{barber_id}/leave-date")
async def toggle_barber_leave_date(
    salon_id: str,
    barber_id: str,
    body: LeaveToggleBody,
    current_user=Depends(get_current_salon_user)
):
    """Add or remove a date from the barber's leave_dates list.
    When marked on leave: also creates / overrides today/that-date's attendance record to 'absent'.
    """
    if current_user.get("role") not in ["admin", "salon_admin", "salon"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Basic date validation
    try:
        datetime.strptime(body.date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    barber = await db.barbers.find_one({"id": barber_id, "salon_id": salon_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    
    # Validate against doj / last working date — can't take leave outside employment
    doj = (barber.get("doj") or "").strip()
    if doj and body.date < doj:
        raise HTTPException(status_code=400, detail=f"Cannot set leave before joining date ({doj})")
    lwd = (barber.get("last_working_date") or "").strip()
    if lwd and body.date > lwd:
        raise HTTPException(status_code=400, detail=f"Cannot set leave after last working day ({lwd})")
    
    leave_dates = list(barber.get("leave_dates") or [])
    
    if body.is_on_leave:
        if body.date not in leave_dates:
            leave_dates.append(body.date)
        # Also write attendance record as 'absent' for that date (admin-overridden)
        record_id = f"{salon_id}_{barber_id}_{body.date}"
        now = datetime.now(timezone.utc).isoformat()
        existing = await db.attendance.find_one({"id": record_id}, {"_id": 0})
        if existing:
            await db.attendance.update_one(
                {"id": record_id},
                {"$set": {
                    "status": "absent",
                    "auto_calculated": False,
                    "override_by": current_user.get("id"),
                    "override_note": "Marked on leave",
                    "updated_at": now,
                }}
            )
        else:
            await db.attendance.insert_one({
                "id": record_id,
                "salon_id": salon_id,
                "barber_id": barber_id,
                "date": body.date,
                "status": "absent",
                "auto_calculated": False,
                "morning_shift_completed": False,
                "noon_evening_shift_completed": False,
                "bookings_count": 0,
                "override_by": current_user.get("id"),
                "override_note": "Marked on leave",
                "created_at": now,
                "updated_at": now,
            })
    else:
        if body.date in leave_dates:
            leave_dates = [d for d in leave_dates if d != body.date]
    
    # Decide legacy on_leave flag based on whether today is in leave_dates
    ist = timezone(timedelta(hours=5, minutes=30))
    today_ist = datetime.now(ist).strftime("%Y-%m-%d")
    legacy_on_leave = today_ist in leave_dates
    
    await db.barbers.update_one(
        {"id": barber_id, "salon_id": salon_id},
        {"$set": {"leave_dates": leave_dates, "on_leave": legacy_on_leave}}
    )
    
    return {
        "barber_id": barber_id,
        "date": body.date,
        "is_on_leave": body.is_on_leave,
        "leave_dates": leave_dates,
    }


@api_router.get("/salons/{salon_id}/barbers/{barber_id}/leave-dates")
async def get_barber_leave_dates(salon_id: str, barber_id: str, current_user=Depends(get_current_salon_user)):
    """Return barber's leave_dates list and key employment dates."""
    if current_user.get("role") not in ["admin", "salon_admin", "salon"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    barber = await db.barbers.find_one({"id": barber_id, "salon_id": salon_id}, {"_id": 0})
    if not barber:
        raise HTTPException(status_code=404, detail="Barber not found")
    return {
        "barber_id": barber_id,
        "leave_dates": list(barber.get("leave_dates") or []),
        "doj": barber.get("doj"),
        "last_working_date": barber.get("last_working_date"),
        "on_leave": bool(barber.get("on_leave")),
    }


@api_router.get("/salons/{salon_id}/staff-attendance/report")
async def staff_attendance_report(
    salon_id: str,
    start_date: str,  # YYYY-MM-DD
    end_date: str,    # YYYY-MM-DD
    branch_id: Optional[str] = None,
    barber_ids: Optional[str] = None,  # comma-separated
    format: str = "json",  # "json" | "csv"
    current_user=Depends(get_current_salon_user),
):
    """Module 4 Phase 7 — Consolidated staff attendance report.

    Returns one row per (barber × date) within the range, with the status
    that was *computed under the mode active on that date* (so months
    that span a mode switch read correctly).  Columns:
      Branch, Date, Staff Code, Staff Name, Status (P/H/A/L/HOL),
      Leave Type (if on leave), Check-in, Check-out, Worked Minutes,
      Override By, Override Note, Mode.
    """
    if current_user.get("role") not in ("admin", "salon_admin", "salon", "salon_branch_manager"):
        raise HTTPException(status_code=403, detail="Admin / branch manager access required")
    # Validate date format
    try:
        d_start = datetime.strptime(start_date, "%Y-%m-%d").date()
        d_end = datetime.strptime(end_date, "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid date format; expected YYYY-MM-DD")
    if d_end < d_start:
        raise HTTPException(status_code=400, detail="end_date must be on or after start_date")

    # Resolve barbers (optionally filter by branch / explicit list).
    barber_q: dict = {"salon_id": salon_id, "is_active": True}
    if branch_id:
        barber_q["branch_id"] = branch_id
    if barber_ids:
        ids_list = [x.strip() for x in barber_ids.split(",") if x.strip()]
        barber_q["id"] = {"$in": ids_list}
    barbers = await db.barbers.find(
        barber_q, {"_id": 0, "id": 1, "name": 1, "branch_id": 1}
    ).to_list(length=1000)
    barber_by_id = {b["id"]: b for b in barbers}

    # Branch names cache.
    branch_ids_used = {b.get("branch_id") for b in barbers if b.get("branch_id")}
    branches_map: dict = {}
    if branch_ids_used:
        for br in await db.branches.find(
            {"id": {"$in": list(branch_ids_used)}}, {"_id": 0, "id": 1, "branch_name": 1}
        ).to_list(length=200):
            branches_map[br["id"]] = br.get("branch_name") or br["id"]

    # All attendance docs in the window.
    att_docs = await db.attendance.find({
        "salon_id": salon_id,
        "barber_id": {"$in": list(barber_by_id.keys())},
        "date": {"$gte": start_date, "$lte": end_date},
    }, {"_id": 0}).to_list(length=20000)
    by_key: dict = {(d["barber_id"], d["date"]): d for d in att_docs}

    # Leaves in the window (for status='leave' rows and leave-type labelling).
    leave_docs = await db.leave_records.find({
        "salon_id": salon_id,
        "barber_id": {"$in": list(barber_by_id.keys())},
        "date": {"$gte": start_date, "$lte": end_date},
        "status": {"$ne": "cancelled"},
    }, {"_id": 0}).to_list(length=20000)
    leave_by_key: dict = {(d["barber_id"], d["date"]): d for d in leave_docs}

    # Public holidays in window.
    pub_hols = await db.salon_holidays.find({
        "salon_id": salon_id,
        "date": {"$gte": start_date, "$lte": end_date},
    }, {"_id": 0, "date": 1}).to_list(length=400)
    public_holiday_dates = {h["date"] for h in pub_hols}

    rows: list[dict] = []
    cur = d_start
    one_day = timedelta(days=1)
    while cur <= d_end:
        ds = cur.strftime("%Y-%m-%d")
        for b in barbers:
            att = by_key.get((b["id"], ds)) or {}
            lv = leave_by_key.get((b["id"], ds))
            if lv:
                status_code = "L"
                leave_label = lv.get("leave_type_code")
            elif ds in public_holiday_dates or att.get("status") == "holiday":
                status_code = "HOL"
                leave_label = None
            elif att.get("status") == "present":
                status_code = "P"
                leave_label = None
            elif att.get("status") == "half_day":
                status_code = "H"
                leave_label = None
            elif att.get("status") == "absent":
                status_code = "A"
                leave_label = None
            else:
                status_code = ""
                leave_label = None
            # Item 2 — Marked By label: "Auto" (auto_calculated true), "Admin"
            # (admin/salon role override), "Staff" (any other override role), "—" if blank.
            mb_role = (att.get("marked_by_role") or "").lower()
            mb_name = att.get("marked_by_name")
            if not att:
                marked_by_label = "—"
            elif att.get("auto_calculated") is True:
                marked_by_label = "Auto"
            elif mb_role in ("admin", "salon_admin", "salon"):
                marked_by_label = "Admin"
            elif mb_role:
                marked_by_label = "Staff"
            elif att.get("override_by"):
                marked_by_label = "Staff"
            else:
                marked_by_label = "—"

            rows.append({
                "branch": branches_map.get(b.get("branch_id"), "—"),
                "branch_id": b.get("branch_id"),
                "date": ds,
                "staff_id": b["id"],
                "staff_name": b["name"],
                "status": status_code,
                "leave_type": leave_label,
                "check_in": att.get("check_in_at"),
                "check_out": att.get("check_out_at"),
                "worked_minutes": att.get("total_minutes"),
                "half_day_reason": att.get("half_day_reason"),
                "override_by": att.get("override_by"),
                "override_note": att.get("override_note"),
                "mode": att.get("computed_under_mode"),
                "marked_by_label": marked_by_label,
                "marked_by_name": mb_name,
            })
        cur += one_day

    if format.lower() == "csv":
        import io
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Branch", "Date", "Staff ID", "Staff Name", "Status",
            "Leave Type", "Check-in", "Check-out", "Worked (min)",
            "Half-day Reason", "Marked By", "Marked By Name", "Override Note", "Mode",
        ])
        for r in rows:
            writer.writerow([
                r["branch"], r["date"], r["staff_id"], r["staff_name"], r["status"],
                r["leave_type"] or "", r["check_in"] or "", r["check_out"] or "",
                r["worked_minutes"] if r["worked_minutes"] is not None else "",
                r["half_day_reason"] or "", r.get("marked_by_label") or "—",
                r.get("marked_by_name") or "", r["override_note"] or "",
                r["mode"] or "",
            ])
        output.seek(0)
        filename = f"attendance_{salon_id}_{start_date}_to_{end_date}.csv"
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {
        "salon_id": salon_id,
        "start_date": start_date,
        "end_date": end_date,
        "rows": rows,
    }


@api_router.get("/salons/{salon_id}/staff-salary/month/{month}")
async def get_monthly_salary(
    salon_id: str,
    month: str,  # YYYY-MM format
    barber_id: Optional[str] = None
):
    """Calculate and return salary for all barbers for a month.

    Module 4 — Payroll Integration:
      • working_days_in_month = total_days_in_month − weekly_offs − public_holidays
        (weekly_offs read from branch.operational_hours where available, else
        salon.operational_hours; public_holidays from `salon_holidays`.)
      • Reads `leave_records` and buckets by `leave_types_config.is_paid` →
          paid_leave_days, unpaid_leave_days, leave_breakdown {code: days}.
      • lop_deduction = (base_compensation / working_days_in_month) * unpaid_leave_days.
      • final_payable = base_compensation − lop_deduction + incentive_amount.
      • Skips recalc when is_paid=true (preserves history exactly).
      • Stamps `attendance_mode_snapshot` so audit knows which mode governed
        the bulk of the month (current attendance_mode at compute time).
    """
    # Validate month format
    try:
        year, mon = month.split("-")
        year, mon = int(year), int(mon)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")

    num_days = calendar.monthrange(year, mon)[1]

    # Pull salon + holidays once.
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0}) or {}
    salon_op_hours = salon.get("operational_hours") or {}
    public_holiday_docs = await db.salon_holidays.find({
        "salon_id": salon_id,
        "date": {"$regex": f"^{month}"},
    }, {"_id": 0, "date": 1}).to_list(length=400)
    public_holiday_dates = {h["date"] for h in public_holiday_docs}

    # Leave-type config (for is_paid lookup, per salon).
    leave_type_cfgs = await db.leave_types_config.find(
        {"salon_id": salon_id}, {"_id": 0, "code": 1, "is_paid": 1},
    ).to_list(length=200)
    is_paid_by_code = {c["code"]: bool(c.get("is_paid")) for c in leave_type_cfgs}

    # Get barbers.
    barber_query = {"salon_id": salon_id, "is_barber": True, "is_active": True}
    if barber_id:
        barber_query["id"] = barber_id
    barbers = await db.barbers.find(
        barber_query, {"_id": 0, "id": 1, "name": 1, "compensation": 1, "branch_id": 1}
    ).to_list(100)

    # Cache branch operational hours so we don't re-fetch per barber.
    branch_op_cache: dict[str, dict] = {}

    async def _weekly_offs_for_barber(barber: dict) -> set[int]:
        """Return the set of weekday integers (Mon=0..Sun=6) that are
        weekly offs for this barber's branch (fallback to salon-level)."""
        op = None
        bid = barber.get("branch_id")
        if bid:
            if bid not in branch_op_cache:
                br = await db.branches.find_one({"id": bid}, {"_id": 0, "operational_hours": 1})
                branch_op_cache[bid] = (br or {}).get("operational_hours") or {}
            op = branch_op_cache[bid]
        if not op:
            op = salon_op_hours
        weekday_names = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        offs: set[int] = set()
        if isinstance(op, dict):
            for i, name in enumerate(weekday_names):
                day_cfg = op.get(name) or {}
                if isinstance(day_cfg, dict) and day_cfg.get("is_holiday"):
                    offs.add(i)
        return offs

    results = []
    for barber in barbers:
        salary_id = f"{salon_id}_{barber['id']}_{month}"
        existing = await db.salary_records.find_one({"id": salary_id}, {"_id": 0})

        # Lock-on-paid — do not recompute already-paid months.
        if existing and existing.get("is_paid"):
            existing["barber_name"] = barber["name"]
            results.append(existing)
            continue

        # ---- Working-day calculation ----
        weekly_offs = await _weekly_offs_for_barber(barber)
        weekly_off_count = 0
        for day in range(1, num_days + 1):
            d = datetime(year, mon, day).weekday()
            if d in weekly_offs:
                weekly_off_count += 1
        public_holiday_count_in_month = sum(
            1 for d in range(1, num_days + 1)
            if f"{year:04d}-{mon:02d}-{d:02d}" in public_holiday_dates
        )
        working_days_in_month = max(1, num_days - weekly_off_count - public_holiday_count_in_month)

        # ---- Attendance counts (legacy fields kept) ----
        attendance_records = await db.attendance.find({
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "date": {"$regex": f"^{month}"}
        }, {"_id": 0}).to_list(200)
        present_days = sum(1 for r in attendance_records if r.get("status") == "present")
        half_days = sum(1 for r in attendance_records if r.get("status") == "half_day")
        absent_days = sum(1 for r in attendance_records if r.get("status") == "absent")
        holidays = sum(1 for r in attendance_records if r.get("status") == "holiday")

        # ---- Leave breakdown (Module 2 source of truth) ----
        leave_rows = await db.leave_records.find({
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "date": {"$regex": f"^{month}"},
            "status": {"$ne": "cancelled"},
        }, {"_id": 0, "leave_type_code": 1, "half_day": 1, "leave_type_snapshot": 1}).to_list(length=400)
        leave_breakdown: dict[str, float] = {}
        paid_leave_days = 0.0
        unpaid_leave_days = 0.0
        for lr in leave_rows:
            code = (lr.get("leave_type_code") or "").upper()
            qty = 0.5 if lr.get("half_day") else 1.0
            leave_breakdown[code] = leave_breakdown.get(code, 0.0) + qty
            # Module 4 — prefer the snapshot stored on the record (immune to
            # admins editing the leave-type config mid-month).  Fall back to
            # the live config for legacy records that pre-date the snapshot.
            snap = lr.get("leave_type_snapshot") or {}
            if "is_paid" in snap:
                is_paid_for_record = bool(snap["is_paid"])
            else:
                is_paid_for_record = is_paid_by_code.get(code, True)
            if is_paid_for_record:
                paid_leave_days += qty
            else:
                unpaid_leave_days += qty

        # ---- Salary math (prorated to actual attendance) ----
        base_compensation = float(barber.get("compensation") or 0)
        per_day_rate = base_compensation / working_days_in_month if working_days_in_month > 0 else 0
        lop_deduction = round(per_day_rate * unpaid_leave_days, 2)  # informational

        # Days the staff "earned" pay for. Policy:
        #   present_days + half_days * 0.5 + paid_leave_days
        # (weekly offs and holidays are excluded from working_days_in_month, so
        #  per_day_rate already accounts for them being paid implicitly.)
        earned_days = max(0.0, present_days + (half_days * 0.5) + paid_leave_days)
        earned_salary = round(per_day_rate * earned_days, 2)

        # Legacy "calculated_salary" — keep computing the way it used to so
        # existing readers don't regress.  Effective working days = present + 0.5*half.
        # We use num_days as the daily-rate divisor (old behaviour).
        old_daily_rate = base_compensation / num_days if num_days > 0 else 0
        effective_days = present_days + (half_days * 0.5)
        calculated_salary = round(old_daily_rate * effective_days, 2)

        # Incentive (existing).
        incentive = await db.incentive_payouts.find_one({
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "month": month,
            "status": {"$in": ["Approved", "Paid"]}
        }, {"_id": 0, "incentive_earned": 1})
        incentive_amount = float(incentive.get("incentive_earned", 0)) if incentive else 0.0

        # ---- Canonical total: prorated earnings + incentive ----
        # Previously this was `base − lop_deduction` which only deducted unpaid
        # leave; absent days were silently paid. Now it scales linearly with
        # earned_days so 10 present days in a 26-working-day month pays
        # base × 10/26 + incentive.
        final_payable = round(earned_salary + incentive_amount, 2)
        total_payable = final_payable  # keep total_payable in lock-step with final_payable

        attendance_mode_snapshot = salon.get("attendance_mode") or "service_completion"

        now = datetime.now(timezone.utc).isoformat()
        record = {
            "id": salary_id,
            "salon_id": salon_id,
            "barber_id": barber["id"],
            "barber_name": barber["name"],
            "month": month,
            "base_salary": base_compensation,
            "base_compensation": base_compensation,
            "working_days": working_days_in_month,
            "working_days_in_month": working_days_in_month,
            "present_days": present_days,
            "half_days": half_days,
            "absent_days": absent_days,
            "holidays": holidays,
            "paid_leave_days": paid_leave_days,
            "unpaid_leave_days": unpaid_leave_days,
            "leave_breakdown": leave_breakdown,
            "lop_deduction": lop_deduction,
            "calculated_salary": calculated_salary,
            "incentive_amount": incentive_amount,
            "final_payable": final_payable,
            "total_payable": total_payable,
            "attendance_mode_snapshot": attendance_mode_snapshot,
            "is_paid": False,
            "updated_at": now,
        }
        if not existing:
            record["created_at"] = now
            await db.salary_records.insert_one(record.copy())
        else:
            await db.salary_records.update_one({"id": salary_id}, {"$set": record})

        record.pop("_id", None)
        results.append(record)

    return {"month": month, "salary_records": results}


@api_router.post("/salons/{salon_id}/staff-salary/pay/{barber_id}/{month}")
async def mark_salary_paid(
    salon_id: str,
    barber_id: str,
    month: str,
    body: SalaryPaymentRequest,
    current_user=Depends(get_current_salon_user)
):
    """Mark salary as paid and create financial transaction."""
    # RBAC: staff.salary_pay
    if not has_module_permission(current_user, "staff", "salary_pay"):
        raise HTTPException(status_code=403, detail="Permission denied: staff.salary_pay")
    
    if body.payment_method not in ["cash", "upi", "bank"]:
        raise HTTPException(status_code=400, detail="Invalid payment method. Use: cash, upi, bank")
    
    salary_id = f"{salon_id}_{barber_id}_{month}"
    
    # Get salary record
    salary_record = await db.salary_records.find_one({"id": salary_id}, {"_id": 0})
    if not salary_record:
        raise HTTPException(status_code=404, detail="Salary record not found. Calculate salary first.")
    
    if salary_record.get("is_paid"):
        raise HTTPException(status_code=400, detail="Salary already paid")
    
    # Get barber name for narration
    barber = await db.barbers.find_one({"id": barber_id}, {"_id": 0, "name": 1})
    barber_name = barber.get("name", "Unknown") if barber else "Unknown"
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Create financial transaction
    transaction_id = str(uuid.uuid4())
    transaction = {
        "id": transaction_id,
        "salon_id": salon_id,
        "type": "expense",
        "category": "staff_salary",
        "amount": salary_record["total_payable"],
        "payment_method": body.payment_method,
        "description": f"Salary payment for {barber_name} - {month}",
        "narration": f"Monthly salary paid to {barber_name} for {month}. Base: ₹{salary_record['calculated_salary']}, Incentive: ₹{salary_record['incentive_amount']}",
        "linked_salary_id": salary_id,
        "created_at": now,
        "updated_at": now
    }
    await db.financial_transactions.insert_one(transaction)
    transaction.pop("_id", None)  # avoid leaking BSON ObjectId in JSON response
    
    # Update salary record
    await db.salary_records.update_one(
        {"id": salary_id},
        {"$set": {
            "is_paid": True,
            "paid_at": now,
            "payment_method": body.payment_method,
            "financial_transaction_id": transaction_id,
            "updated_at": now
        }}
    )
    
    updated = await db.salary_records.find_one({"id": salary_id}, {"_id": 0})
    updated["barber_name"] = barber_name
    updated["transaction"] = transaction
    
    return updated


@api_router.get("/salons/{salon_id}/staff-holidays")
async def get_salon_holidays(salon_id: str, year: Optional[int] = None):
    """Get marked holidays for a salon."""
    if not year:
        year = datetime.now().year
    
    holidays = await db.salon_holidays.find({
        "salon_id": salon_id,
        "date": {"$regex": f"^{year}"}
    }, {"_id": 0}).to_list(365)
    
    return {"year": year, "holidays": holidays}


@api_router.get("/salons/{salon_id}/check-holiday/{date}")
async def check_salon_holiday(salon_id: str, date: str):
    """Check if a salon is closed/holiday on a specific date - for customer booking UI"""
    # Check explicit holiday
    holiday = await db.salon_holidays.find_one({
        "salon_id": salon_id,
        "date": date
    }, {"_id": 0})
    
    if holiday:
        return {
            "is_closed": True,
            "reason": holiday.get("description", "Holiday"),
            "type": "holiday"
        }
    
    # Check operational hours for that day
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0, "operational_hours": 1})
    if salon and salon.get("operational_hours"):
        try:
            weekday = datetime.fromisoformat(date).strftime("%A").lower()
            day_hours = salon["operational_hours"].get(weekday, {})
            if day_hours.get("is_holiday"):
                return {
                    "is_closed": True,
                    "reason": f"Closed on {weekday.capitalize()}s",
                    "type": "weekly_off"
                }
        except (ValueError, AttributeError, KeyError):
            pass
    
    return {
        "is_closed": False,
        "reason": None,
        "type": None
    }


@api_router.post("/salons/{salon_id}/staff-holidays")
async def add_salon_holiday(
    salon_id: str,
    date: str,  # YYYY-MM-DD
    description: Optional[str] = None,
    current_user=Depends(get_current_salon_user)
):
    """Add a holiday for the salon."""
    if current_user.get("role") not in ["admin", "salon_admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    holiday_id = f"{salon_id}_{date}"
    
    holiday = {
        "id": holiday_id,
        "salon_id": salon_id,
        "date": date,
        "description": description or "Holiday",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.salon_holidays.update_one(
        {"id": holiday_id},
        {"$set": holiday},
        upsert=True
    )
    
    # Also mark all barbers as holiday for this date
    barbers = await db.barbers.find({
        "salon_id": salon_id,
        "is_barber": True,
        "is_active": True
    }, {"_id": 0, "id": 1}).to_list(100)
    
    for barber in barbers:
        # Skip locked months — don't overwrite paid attendance.
        locked = await attendance_mode_mod.is_attendance_locked(db, salon_id, barber["id"], date)
        if locked:
            continue
        record_id = f"{salon_id}_{barber['id']}_{date}"
        await db.attendance.update_one(
            {"id": record_id},
            {"$set": {
                "id": record_id,
                "salon_id": salon_id,
                "barber_id": barber["id"],
                "date": date,
                "status": "holiday",
                "auto_calculated": False,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
    
    return holiday


@api_router.delete("/salons/{salon_id}/staff-holidays/{date}")
async def remove_salon_holiday(
    salon_id: str,
    date: str,
    current_user=Depends(get_current_salon_user)
):
    """Remove a holiday."""
    if current_user.get("role") not in ["admin", "salon_admin"]:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    await db.salon_holidays.delete_one({"id": f"{salon_id}_{date}"})
    
    # Remove holiday status from attendance records
    await db.attendance.update_many(
        {"salon_id": salon_id, "date": date, "status": "holiday"},
        {"$set": {"status": "absent", "auto_calculated": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Holiday removed"}


async def cancel_active_bookings_end_of_day():
    """End-of-day cleanup at 00:00 IST.
    Cancels every booking still in an "active" (non-final) state for the day that just ended,
    so the queue resets fresh tomorrow and no stale bookings carry forward.
    """
    try:
        # Use IST for "today's date" — at 00:00 IST we cancel the day that just ended
        ist = timezone(timedelta(hours=5, minutes=30))
        now_ist = datetime.now(ist)
        # The day that just ended (yesterday relative to IST midnight)
        target_date = (now_ist - timedelta(minutes=1)).strftime("%Y-%m-%d")
        
        # Final/terminal statuses that we should NOT touch
        final_statuses = ["completed", "cancelled", "skipped"]
        
        active_tokens = await db.tokens.find({
            "date": target_date,
            "status": {"$nin": final_statuses}
        }, {"_id": 0}).to_list(5000)
        
        if not active_tokens:
            logger.info(f"[EOD] No active bookings to cancel for {target_date}")
            return
        
        cancelled_count = 0
        for token in active_tokens:
            try:
                # Refund wallet balance if the booking was paid via wallet
                if token.get("payment_mode") == "wallet" and token.get("payment_status") == "paid":
                    wallet_phone = token.get("phone") or ""
                    if wallet_phone and not wallet_phone.startswith("+91"):
                        wallet_phone = f"+91{wallet_phone}"
                    if wallet_phone:
                        membership = await db.customer_memberships.find_one({
                            "salon_id": token.get("salon_id"),
                            "customer_phone": wallet_phone,
                            "is_active": True
                        }, {"_id": 0})
                        if membership:
                            refund_amount = token.get("total_amount", 0) or 0
                            new_balance = (membership.get("wallet_balance") or 0) + refund_amount
                            await db.customer_memberships.update_one(
                                {"id": membership["id"]},
                                {"$set": {"wallet_balance": new_balance}}
                            )
                            await db.wallet_transactions.insert_one({
                                "id": str(uuid.uuid4()),
                                "customer_phone": wallet_phone,
                                "salon_id": token.get("salon_id"),
                                "transaction_type": "credit",
                                "amount": refund_amount,
                                "balance_after": new_balance,
                                "description": f"Refund - End-of-day auto cancel (Token {token.get('token_number','')})",
                                "created_at": datetime.now(timezone.utc).isoformat()
                            })
                
                await db.tokens.update_one(
                    {"id": token["id"]},
                    {"$set": {
                        "status": "cancelled",
                        "cancelled_reason": "Auto-cancelled at end of day (00:00 IST)",
                        "cancelled_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                cancelled_count += 1
                
                # Notify customer (best-effort)
                try:
                    await send_booking_notification(token, 'token_cancelled')
                except Exception as nerr:
                    logger.warning(f"[EOD] Notification failed for token {token.get('id')}: {nerr}")
                
                if token.get("phone"):
                    try:
                        await create_in_app_notification(
                            user_type="customer",
                            user_id=token["phone"],
                            title="Booking Cancelled",
                            message=f"Your booking (Token #{token.get('token_number','')}) was auto-cancelled at the end of the day.",
                            notification_type="booking_cancelled",
                            setting_key="booking_status_change",
                            salon_id=token.get("salon_id", ""),
                            related_id=token["id"],
                        )
                    except Exception as ierr:
                        logger.warning(f"[EOD] In-app notif failed: {ierr}")
                
                await broadcast_update("token_cancelled", {"token_id": token["id"]})
            except Exception as terr:
                logger.error(f"[EOD] Error cancelling token {token.get('id')}: {terr}")
        
        logger.info(f"[EOD] Auto-cancelled {cancelled_count} bookings for {target_date}")
    except Exception as e:
        logger.error(f"[EOD] cancel_active_bookings_end_of_day failed: {e}")


# ============ SUBSCRIPTION ROUTES (SalonHub Pro) ============

@api_router.get("/subscription-plans")
async def list_subscription_plans():
    """Public list of active subscription plans."""
    plans = await db.subscription_plans.find(
        {"status": "active"}, {"_id": 0}
    ).to_list(50)
    return plans

@api_router.get("/subscription-plans/active")
async def get_active_subscription_plan():
    """Return the default active plan (single source of truth for pricing)."""
    plan = await get_active_plan()
    if not plan:
        raise HTTPException(status_code=404, detail="No active subscription plan configured")
    return plan

@api_router.get("/salons/{salon_id}/subscription/status")
async def api_subscription_status(salon_id: str):
    """Public endpoint - get current subscription state for a salon."""
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0, "id": 1})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    return await get_subscription_status(salon_id)


@api_router.post("/salons/{salon_id}/subscription/start-trial")
async def api_subscription_start_trial(
    salon_id: str,
    current_user=Depends(get_current_salon_user),
):
    """Grant a one-time 30-day free trial subscription to this salon.

    Idempotent guards:
    - Only the salon's own admin can start the trial.
    - Salon must NOT have used a trial before (`salons.trial_used != true`).
    - Salon must NOT already have an active paid subscription.
    Returns the newly-created trial subscription document.
    """
    _check_salon_admin_for_salon(current_user, salon_id)

    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")
    if salon.get("trial_used"):
        raise HTTPException(status_code=400, detail="Free trial already used for this salon")

    existing = await db.salon_subscriptions.find_one(
        {"salon_id": salon_id, "payment_status": "paid"},
        {"_id": 0, "id": 1, "expiry_date": 1},
        sort=[("expiry_date", -1)],
    )
    if existing and existing.get("expiry_date"):
        try:
            exp_dt = datetime.fromisoformat(existing["expiry_date"].replace("Z", "+00:00"))
            if exp_dt.tzinfo is None:
                exp_dt = exp_dt.replace(tzinfo=timezone.utc)
            if exp_dt > datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="You already have an active subscription")
        except HTTPException:
            raise
        except Exception:
            pass

    plan = await get_active_plan()
    if not plan:
        raise HTTPException(status_code=404, detail="No active plan configured")

    now_dt = datetime.now(timezone.utc)
    expiry_dt = now_dt + timedelta(days=30)
    price_per_branch = float(plan.get("price_per_branch") or plan.get("price") or 0)
    branch_ids = await get_active_branch_ids(salon_id)
    billable = max(len(branch_ids), 1)

    trial_doc = {
        "id": str(uuid.uuid4()),
        "salon_id": salon_id,
        "plan_id": plan.get("id"),
        "plan_name": plan.get("plan_name"),
        "price": price_per_branch,
        "subscription_status": "active",
        "start_date": now_dt.isoformat(),
        "expiry_date": expiry_dt.isoformat(),
        "payment_status": "paid",  # treat as paid so paywall doesn't block
        "cashfree_order_id": None,
        "cashfree_payment_id": "FREE_TRIAL_30D",
        "auto_renew": False,
        "is_trial": True,
        "billable_branch_count": billable,
        "price_per_branch_snapshot": price_per_branch,
        "branch_ids_snapshot": branch_ids,
        "base_amount": 0.0,
        "discount_code_applied": None,
        "discount_amount": 0.0,
        "total_amount": 0.0,
        "created_at": now_dt.isoformat(),
        "updated_at": now_dt.isoformat(),
    }
    await db.salon_subscriptions.insert_one(trial_doc.copy())
    await db.salons.update_one(
        {"id": salon_id},
        {"$set": {"trial_used": True, "trial_started_at": now_dt.isoformat()}},
    )
    logger.info(f"[Trial] Started 30-day free trial for salon {salon_id} until {expiry_dt.isoformat()}")
    return {
        "success": True,
        "subscription": trial_doc,
        "expires_at": expiry_dt.isoformat(),
        "trial_days": 30,
    }


@api_router.get("/salons/{salon_id}/subscription/quote")
async def api_subscription_quote(
    salon_id: str,
    plan_id: Optional[str] = None,
    discount_code: Optional[str] = None,
):
    """Phase 2 (Part C) — pricing preview for a salon.

    Returns the per-branch breakdown the UI shows on the subscription page
    and inside the paywall modal. Discount handling is wired up here as a
    pass-through stub now; the real validation lands in Phase 4 (Part D).
    """
    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0, "id": 1})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")

    if plan_id:
        plan = await db.subscription_plans.find_one(
            {"id": plan_id, "status": "active"}, {"_id": 0}
        )
    else:
        plan = await get_active_plan()
    if not plan:
        raise HTTPException(status_code=404, detail="No active plan available")

    price_per_branch = float(
        plan.get("price_per_branch") or plan.get("price") or 0
    )
    billable_branch_count = await count_billable_branches(salon_id)
    branch_ids_snapshot = await get_active_branch_ids(salon_id)
    base_amount = round(price_per_branch * billable_branch_count, 2)

    discount_amount = 0.0
    discount_details = None
    discount_code_applied: Optional[str] = None
    if discount_code:
        # Phase 4 (Part D) — real validation + math
        result = await discount_codes_mod.validate_discount_code(
            code=discount_code,
            salon_id=salon_id,
            billable_branch_count=billable_branch_count,
        )
        if result["valid"]:
            comp = discount_codes_mod.compute_discount(
                code_doc=result["code_doc"],
                base_amount=base_amount,
                price_per_branch=price_per_branch,
                billable_branch_count=billable_branch_count,
            )
            discount_amount = comp["discount_amount"]
            discount_details = comp["discount_details"]
            discount_code_applied = result["code_doc"]["code"]
        else:
            discount_details = {
                "code": discount_code.strip().upper(),
                "valid": False,
                "reason": result["reason"] or "Code could not be applied",
            }

    total_amount = round(max(base_amount - discount_amount, 0.0), 2)

    return {
        "salon_id": salon_id,
        "plan_id": plan.get("id"),
        "plan_name": plan.get("plan_name"),
        "billing_cycle": plan.get("billing_cycle", "monthly"),
        "billable_branch_count": billable_branch_count,
        "branch_ids_snapshot": branch_ids_snapshot,
        "price_per_branch": price_per_branch,
        "base_amount": base_amount,
        "discount_code_applied": discount_code_applied,
        "discount_amount": discount_amount,
        "total_amount": total_amount,
        "discount_details": discount_details,
    }


@api_router.get("/salons/{salon_id}/subscription/transactions")
async def api_subscription_transactions(
    salon_id: str, current_user=Depends(get_current_salon_user)
):
    """List payment transactions for a salon (admin only)."""
    _check_salon_admin_for_salon(current_user, salon_id)
    txs = await db.payment_transactions.find(
        {"salon_id": salon_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(200)
    return txs

@api_router.post("/salons/{salon_id}/subscription/create-order")
async def create_subscription_order(
    salon_id: str,
    payload: CreateOrderRequest,
    current_user=Depends(get_current_salon_user),
):
    """
    Create a Cashfree order for subscribing the salon to a plan.
    Returns the payment_session_id which the frontend uses with Cashfree JS SDK.
    For free_months discount codes, skips Cashfree entirely (Phase 7).
    """
    _check_salon_admin_for_salon(current_user, salon_id)

    salon = await db.salons.find_one({"id": salon_id}, {"_id": 0})
    if not salon:
        raise HTTPException(status_code=404, detail="Salon not found")

    # Resolve plan
    if payload.plan_id:
        plan = await db.subscription_plans.find_one(
            {"id": payload.plan_id, "status": "active"}, {"_id": 0}
        )
        if not plan:
            raise HTTPException(status_code=404, detail="Subscription plan not found")
    else:
        plan = await get_active_plan()
        if not plan:
            raise HTTPException(status_code=404, detail="No active plan available")

    amount = float(plan.get("price", 0))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Invalid plan price")

    # Phase 2 (Part C) — per-branch repricing
    price_per_branch = float(plan.get("price_per_branch") or amount)
    billable_branch_count = await count_billable_branches(salon_id)
    branch_ids_snapshot = await get_active_branch_ids(salon_id)
    base_amount = round(price_per_branch * billable_branch_count, 2)

    # Phase 7 (Part D) — apply discount code if provided
    discount_amount = 0.0
    discount_code_applied: Optional[str] = None
    discount_code_doc: Optional[dict] = None
    is_free_months = False
    free_months_granted = 0
    discount_details: Optional[dict] = None

    if payload.discount_code:
        try:
            validation = await discount_codes_mod.validate_discount_code(
                code=payload.discount_code,
                salon_id=salon_id,
                billable_branch_count=billable_branch_count,
            )
        except Exception as e:
            logger.error(f"[Subscription] discount validation error: {e}")
            validation = {"valid": False, "code_doc": None, "reason": "Could not validate code"}

        if not validation["valid"]:
            raise HTTPException(
                status_code=400,
                detail=validation.get("reason") or "Invalid discount code",
            )

        discount_code_doc = validation["code_doc"]
        comp = discount_codes_mod.compute_discount(
            code_doc=discount_code_doc,
            base_amount=base_amount,
            price_per_branch=price_per_branch,
            billable_branch_count=billable_branch_count,
        )
        discount_amount = float(comp["discount_amount"])
        is_free_months = bool(comp["is_free_months"])
        free_months_granted = int(comp["free_months"])
        discount_details = comp["discount_details"]
        discount_code_applied = discount_code_doc["code"]

    total_amount = round(max(base_amount - discount_amount, 0.0), 2)

    # Phase 7 — Free months short-circuit. Skip Cashfree entirely.
    if is_free_months and free_months_granted > 0:
        now = datetime.now(timezone.utc)
        start_date_iso = now.isoformat()
        expiry_date_iso = (now + timedelta(days=30 * free_months_granted)).isoformat()
        sub_id = str(uuid.uuid4())
        sub_doc = {
            "id": sub_id,
            "salon_id": salon_id,
            "plan_id": plan["id"],
            "plan_name": plan.get("plan_name"),
            "price": 0.0,
            "subscription_status": "active",
            "start_date": start_date_iso,
            "expiry_date": expiry_date_iso,
            "payment_status": "discounted_free",
            "cashfree_order_id": None,
            "cashfree_payment_id": None,
            "auto_renew": False,
            "billing_cycle": plan.get("billing_cycle", "monthly"),
            "billable_branch_count": billable_branch_count,
            "price_per_branch_snapshot": price_per_branch,
            "branch_ids_snapshot": branch_ids_snapshot,
            "base_amount": base_amount,
            "discount_code_applied": discount_code_applied,
            "discount_amount": discount_amount,
            "total_amount": 0.0,
            "free_months_granted": free_months_granted,
            "created_at": start_date_iso,
            "updated_at": start_date_iso,
        }
        await db.salon_subscriptions.insert_one(sub_doc.copy())

        # Record usage atomically
        try:
            await discount_codes_mod.record_discount_usage(
                code_doc=discount_code_doc,
                salon_id=salon_id,
                subscription_id=sub_id,
                base_amount=base_amount,
                discount_amount=discount_amount,
                final_amount=0.0,
                branch_count_at_use=billable_branch_count,
            )
        except Exception as e:
            logger.error(f"[Subscription] record_discount_usage failed: {e}")

        # Best-effort notification to the salon admin
        try:
            await create_in_app_notification(
                user_type="salon",
                user_id=salon_id,
                title="🎉 Subscription Activated (Free)",
                message=f"Your {free_months_granted}-month free subscription is now active using code {discount_code_applied}.",
                notification_type="subscription_activated",
                setting_key="subscription_activated",
                salon_id=salon_id,
                related_id=sub_id,
            )
        except Exception:
            pass

        return {
            "order_id": None,
            "payment_session_id": None,
            "amount": 0.0,
            "currency": "INR",
            "plan": plan,
            "subscription_id": sub_id,
            "cashfree_env": cashfree_service.CASHFREE_ENV,
            "billable_branch_count": billable_branch_count,
            "price_per_branch": price_per_branch,
            "base_amount": base_amount,
            "discount_amount": discount_amount,
            "discount_code_applied": discount_code_applied,
            "total_amount": 0.0,
            "is_free_months": True,
            "free_months_granted": free_months_granted,
            "discount_details": discount_details,
            "payment_status": "discounted_free",
            "expiry_date": expiry_date_iso,
        }

    amount = total_amount  # what actually gets charged

    if amount <= 0:
        # This is the non-free-months edge case (e.g., 100% percent code).
        # Treat the same way: skip Cashfree, mark discounted_free with same expiry as a paid month.
        raise HTTPException(status_code=400, detail="Total amount must be greater than 0")

    # Phase 7 — Cashfree config check happens here so free_months path above works
    # even when the payment gateway isn't configured (no payment needed).
    if not cashfree_service.is_configured():
        raise HTTPException(
            status_code=503,
            detail="Payment gateway is not configured. Please contact support.",
        )

    # Build order
    order_id = f"SH-{salon_id[:8]}-{int(datetime.now(timezone.utc).timestamp())}"
    customer_id = f"salon_{salon_id[:12]}"
    customer_name = (salon.get("owner_name") or salon.get("salon_name") or "Salon Owner")[:50]
    customer_email = salon.get("email") or "noreply@salonhub.in"
    customer_phone = (salon.get("phone") or "9999999999").replace("+91", "").replace("+", "")[-10:]
    if len(customer_phone) < 10:
        customer_phone = "9999999999"

    backend_base = os.environ.get("BACKEND_PUBLIC_URL", "")
    frontend_base = os.environ.get("FRONTEND_PUBLIC_URL", "")
    return_url = f"{frontend_base}/subscription/callback?order_id={order_id}" if frontend_base else None
    notify_url = f"{backend_base}/api/subscriptions/webhook" if backend_base else None

    try:
        cf_response = await cashfree_service.create_order(
            order_id=order_id,
            order_amount=amount,
            customer_id=customer_id,
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            return_url=return_url or "https://salonhub.in/subscription/callback",
            notify_url=notify_url,
            order_note=f"{plan['plan_name']} - {plan.get('billing_cycle','monthly')}",
            order_currency="INR",
        )
    except Exception as e:
        logger.error(f"[Subscription] Cashfree create_order error: {e}")
        raise HTTPException(status_code=502, detail=f"Payment gateway error: {str(e)}")

    payment_session_id = cf_response.get("payment_session_id")
    cf_order_id = cf_response.get("order_id") or order_id

    # Persist a pending subscription record
    sub_id = str(uuid.uuid4())
    sub_doc = {
        "id": sub_id,
        "salon_id": salon_id,
        "plan_id": plan["id"],
        "plan_name": plan.get("plan_name"),
        "price": amount,
        "subscription_status": "pending",
        "start_date": None,
        "expiry_date": None,
        "payment_status": "pending",
        "cashfree_order_id": cf_order_id,
        "cashfree_payment_id": None,
        "auto_renew": False,
        "billing_cycle": plan.get("billing_cycle", "monthly"),
        # Phase 2 (Part C) snapshot fields
        "billable_branch_count": billable_branch_count,
        "price_per_branch_snapshot": price_per_branch,
        "branch_ids_snapshot": branch_ids_snapshot,
        "base_amount": base_amount,
        "discount_code_applied": discount_code_applied,
        "discount_amount": discount_amount,
        "total_amount": total_amount,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None,
    }
    await db.salon_subscriptions.insert_one(sub_doc.copy())

    # Persist a pending transaction record
    tx_id = str(uuid.uuid4())
    tx_doc = {
        "id": tx_id,
        "salon_id": salon_id,
        "subscription_id": sub_id,
        "plan_id": plan["id"],
        "amount": amount,
        "currency": "INR",
        "payment_gateway": "cashfree",
        "gateway_order_id": cf_order_id,
        "gateway_payment_id": None,
        "payment_status": "pending",
        "payment_response": cf_response,
        "payment_method": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": None,
    }
    await db.payment_transactions.insert_one(tx_doc.copy())

    return {
        "order_id": cf_order_id,
        "payment_session_id": payment_session_id,
        "amount": amount,
        "currency": "INR",
        "plan": plan,
        "subscription_id": sub_id,
        "cashfree_env": cashfree_service.CASHFREE_ENV,
        # Phase 2 (Part C) breakdown so the frontend can confirm what's being charged
        "billable_branch_count": billable_branch_count,
        "price_per_branch": price_per_branch,
        "base_amount": base_amount,
        "discount_amount": discount_amount,
        "discount_code_applied": discount_code_applied,
        "total_amount": total_amount,
    }


async def _activate_subscription_from_payment(
    salon_id: str, sub_id: str, plan: dict, cf_payment: Optional[dict] = None
) -> dict:
    """Internal helper: mark subscription paid + extend expiry."""
    now = datetime.now(timezone.utc)
    # Stack from latest existing active expiry (renewal)
    existing = await db.salon_subscriptions.find_one(
        {
            "salon_id": salon_id,
            "payment_status": "paid",
            "subscription_status": "active",
        },
        {"_id": 0},
        sort=[("expiry_date", -1)],
    )
    base_dt = now
    if existing and existing.get("expiry_date"):
        try:
            existing_exp = datetime.fromisoformat(existing["expiry_date"].replace("Z", "+00:00"))
            if existing_exp.tzinfo is None:
                existing_exp = existing_exp.replace(tzinfo=timezone.utc)
            if existing_exp > now:
                base_dt = existing_exp
        except Exception:
            pass

    billing_cycle = plan.get("billing_cycle", "monthly")
    expiry_dt = _add_billing_cycle(base_dt, billing_cycle)

    update_doc = {
        "subscription_status": "active",
        "payment_status": "paid",
        "start_date": now.isoformat(),
        "expiry_date": expiry_dt.isoformat(),
        "updated_at": now.isoformat(),
    }
    if cf_payment:
        update_doc["cashfree_payment_id"] = (
            cf_payment.get("cf_payment_id") or cf_payment.get("payment_id")
        )

    await db.salon_subscriptions.update_one({"id": sub_id}, {"$set": update_doc})

    # Phase 7 — record discount usage on payment success (idempotent guard)
    try:
        sub_doc_now = await db.salon_subscriptions.find_one({"id": sub_id}, {"_id": 0})
        code_applied = (sub_doc_now or {}).get("discount_code_applied")
        if code_applied:
            already = await db.discount_code_usages.find_one(
                {"subscription_id": sub_id}, {"_id": 0, "id": 1}
            )
            if not already:
                code_doc = await db.discount_codes.find_one({"code": code_applied}, {"_id": 0})
                if code_doc:
                    await discount_codes_mod.record_discount_usage(
                        code_doc=code_doc,
                        salon_id=salon_id,
                        subscription_id=sub_id,
                        base_amount=float(sub_doc_now.get("base_amount") or 0),
                        discount_amount=float(sub_doc_now.get("discount_amount") or 0),
                        final_amount=float(sub_doc_now.get("total_amount") or 0),
                        branch_count_at_use=int(sub_doc_now.get("billable_branch_count") or 1),
                    )
    except Exception as e:
        logger.error(f"[Subscription] record_discount_usage on activation failed: {e}")

    return update_doc


@api_router.post("/salons/{salon_id}/subscription/verify-payment")
async def verify_subscription_payment(
    salon_id: str,
    body: dict = Body(...),
    current_user=Depends(get_current_salon_user),
):
    """Verify payment with Cashfree by fetching order/payment status."""
    _check_salon_admin_for_salon(current_user, salon_id)

    order_id = (body or {}).get("order_id")
    if not order_id:
        raise HTTPException(status_code=400, detail="order_id is required")

    sub = await db.salon_subscriptions.find_one(
        {"salon_id": salon_id, "cashfree_order_id": order_id}, {"_id": 0}
    )
    if not sub:
        raise HTTPException(status_code=404, detail="Subscription order not found")

    try:
        cf_order = await cashfree_service.fetch_order(order_id)
    except Exception as e:
        logger.error(f"[Subscription] verify fetch_order failed: {e}")
        raise HTTPException(status_code=502, detail="Failed to fetch order status")

    order_status = cf_order.get("order_status")  # PAID, ACTIVE, EXPIRED, TERMINATED
    payments = []
    try:
        payments = await cashfree_service.fetch_payments_for_order(order_id)
    except Exception as e:
        logger.warning(f"[Subscription] verify fetch_payments warning: {e}")

    successful_payment = None
    for p in payments:
        if (p.get("payment_status") or "").upper() == "SUCCESS":
            successful_payment = p
            break

    plan = await db.subscription_plans.find_one({"id": sub["plan_id"]}, {"_id": 0})
    if not plan:
        plan = await get_active_plan()

    now_iso = datetime.now(timezone.utc).isoformat()

    if order_status == "PAID" and successful_payment:
        update = await _activate_subscription_from_payment(
            salon_id, sub["id"], plan, successful_payment
        )
        await db.payment_transactions.update_many(
            {"gateway_order_id": order_id},
            {
                "$set": {
                    "payment_status": "success",
                    "gateway_payment_id": successful_payment.get("cf_payment_id"),
                    "payment_method": (
                        successful_payment.get("payment_group")
                        or (successful_payment.get("payment_method") or {}).get("type")
                        if isinstance(successful_payment.get("payment_method"), dict)
                        else successful_payment.get("payment_method")
                    ),
                    "payment_response": successful_payment,
                    "updated_at": now_iso,
                }
            },
        )
        return {
            "success": True,
            "status": "active",
            "expiry_date": update["expiry_date"],
            "order_status": order_status,
        }

    # Handle failed/cancelled
    if order_status in ("EXPIRED", "TERMINATED") or (
        payments and not successful_payment
    ):
        await db.salon_subscriptions.update_one(
            {"id": sub["id"]},
            {
                "$set": {
                    "subscription_status": "payment_failed",
                    "payment_status": "failed",
                    "updated_at": now_iso,
                }
            },
        )
        await db.payment_transactions.update_many(
            {"gateway_order_id": order_id, "payment_status": "pending"},
            {
                "$set": {
                    "payment_status": "failed",
                    "payment_response": cf_order,
                    "updated_at": now_iso,
                }
            },
        )
        return {"success": False, "status": "payment_failed", "order_status": order_status}

    # Otherwise still pending
    return {"success": False, "status": "pending", "order_status": order_status}


@fastapi_app.post("/api/subscriptions/webhook")
async def cashfree_webhook(request: Request):
    """
    Cashfree webhook endpoint. Mounted directly on fastapi_app so we can
    read the raw body for signature verification.
    """
    raw_body = await request.body()
    timestamp = request.headers.get("x-webhook-timestamp", "")
    signature = request.headers.get("x-webhook-signature", "")

    if not cashfree_service.verify_webhook_signature(raw_body, timestamp, signature):
        logger.warning("[Cashfree Webhook] Invalid signature")
        # Always return 200 so Cashfree doesn't retry storms in dev when secret mismatches
        return {"received": True, "verified": False}

    try:
        payload = json.loads(raw_body.decode("utf-8"))
    except Exception:
        return {"received": True, "verified": True, "parsed": False}

    data = payload.get("data") or {}
    order = data.get("order") or {}
    payment = data.get("payment") or {}
    order_id = order.get("order_id")
    if not order_id:
        return {"received": True, "verified": True}

    # Customer → salon service payments (Cashfree Easy Split) share this
    # webhook URL. Route them first; if handled, we're done. Any handler error
    # is logged but does NOT block subscription handling below.
    try:
        svc_result = await service_payments_mod.handle_service_payment_webhook(payload)
        if svc_result:
            return {"received": True, "verified": True, **svc_result}
    except Exception as e:
        logger.error(f"[Cashfree Webhook] service payment handler error: {e}")

    sub = await db.salon_subscriptions.find_one(
        {"cashfree_order_id": order_id}, {"_id": 0}
    )
    if not sub:
        logger.warning(f"[Cashfree Webhook] Unknown order_id {order_id}")
        return {"received": True, "verified": True, "matched": False}

    plan = await db.subscription_plans.find_one({"id": sub["plan_id"]}, {"_id": 0})
    if not plan:
        plan = await get_active_plan()

    payment_status = (payment.get("payment_status") or "").upper()
    now_iso = datetime.now(timezone.utc).isoformat()

    if payment_status == "SUCCESS":
        await _activate_subscription_from_payment(sub["salon_id"], sub["id"], plan, payment)
        await db.payment_transactions.update_many(
            {"gateway_order_id": order_id},
            {
                "$set": {
                    "payment_status": "success",
                    "gateway_payment_id": payment.get("cf_payment_id"),
                    "payment_response": payload,
                    "updated_at": now_iso,
                }
            },
        )
    elif payment_status in ("FAILED", "USER_DROPPED", "CANCELLED"):
        await db.salon_subscriptions.update_one(
            {"id": sub["id"]},
            {
                "$set": {
                    "subscription_status": "payment_failed",
                    "payment_status": "failed",
                    "updated_at": now_iso,
                }
            },
        )
        await db.payment_transactions.update_many(
            {"gateway_order_id": order_id, "payment_status": "pending"},
            {
                "$set": {
                    "payment_status": "failed",
                    "payment_response": payload,
                    "updated_at": now_iso,
                }
            },
        )

    return {"received": True, "verified": True, "matched": True}


# ----- Admin endpoints to manage subscription plans (configurable pricing) -----
@api_router.put("/admin/subscription-plans/{plan_id}")
async def admin_update_subscription_plan(
    plan_id: str,
    payload: SubscriptionPlanUpdate,
    current_user=Depends(get_current_salon_user),
):
    """Update a subscription plan (price, features, etc.). Admin-only."""
    if current_user.get("role") not in ("admin", "salon_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    update_doc = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_doc:
        raise HTTPException(status_code=400, detail="No fields provided")
    update_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = await db.subscription_plans.update_one({"id": plan_id}, {"$set": update_doc})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    plan = await db.subscription_plans.find_one({"id": plan_id}, {"_id": 0})
    return plan




# Include router - MUST be after ALL @api_router routes are defined
fastapi_app.include_router(api_router)

# Phase 1 (Part A) — mount platform admin router. Uses its own /api/platform prefix.
platform_admin_mod.init_platform_admin_router(
    db=db,
    send_whatsapp_otp=send_whatsapp_otp,
    otp_is_valid=_otp_is_valid,
    secret_key=SECRET_KEY,
    algorithm=ALGORITHM,
)
fastapi_app.include_router(platform_admin_mod.platform_router)

# Phase 5 (Part A) — mount platform admin management router.
platform_admin_mgmt_mod.init_platform_management_router(
    db=db,
    get_subscription_status=get_subscription_status,
    count_billable_branches=count_billable_branches,
    get_active_branch_ids=get_active_branch_ids,
    get_active_plan=get_active_plan,
    secret_key=SECRET_KEY,
    algorithm=ALGORITHM,
    create_in_app_notification=create_in_app_notification,
    send_whatsapp_notification=send_whatsapp_notification,
)
fastapi_app.include_router(platform_admin_mgmt_mod.management_router)

# Phase 4 (Part D) — discount codes router
discount_codes_mod.init_discount_codes_router(db=db)
fastapi_app.include_router(discount_codes_mod.discount_codes_router)

# Phase 8 (Part B) — Supplier auth + signup router
supplier_auth_mod.init_supplier_auth_router(
    db=db,
    send_whatsapp_otp=send_whatsapp_otp,
    secret_key=SECRET_KEY,
    algorithm=ALGORITHM,
)
fastapi_app.include_router(supplier_auth_mod.supplier_auth_router)

# Phase 9 (Part B) — Supplier products / dashboard / samples router
supplier_products_mod.init_supplier_products_router(db=db)
supplier_products_mod.set_notification_hook(in_app_notifier=create_in_app_notification)
fastapi_app.include_router(supplier_products_mod.supplier_products_router)

# Phase 10 — Salon-facing Marketplace router (browse supplier catalogs + inquiries)
import marketplace as marketplace_mod  # noqa: E402
marketplace_mod.init_marketplace_router(db=db, get_current_salon_user=get_current_salon_user)
fastapi_app.include_router(marketplace_mod.marketplace_router)

# Jul 2026 (continuation_request) — Global salon-side search (ribbon → magnifier)
import salon_search as salon_search_mod  # noqa: E402
salon_search_mod.init_search_router(db=db, get_current_salon_user=get_current_salon_user)
fastapi_app.include_router(salon_search_mod.search_router)

# Phase 10/11/12 — Salon Store (browse + cart + checkout + order lifecycle)
import salon_store as salon_store_mod  # noqa: E402
from supplier_auth import require_supplier as _require_supplier_dep  # noqa: E402
salon_store_mod.init_salon_store_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
    create_in_app_notification=create_in_app_notification,
    require_supplier=_require_supplier_dep,
)
fastapi_app.include_router(salon_store_mod.salon_store_router)

# Customer → salon service payments (Cashfree Easy Split). Shares the Cashfree
# webhook mounted above; routes registered under /api/... via its own router.
import service_payments as service_payments_mod  # noqa: E402
service_payments_mod.init_service_payments_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
    create_in_app_notification=create_in_app_notification,
    check_salon_admin_for_salon=_check_salon_admin_for_salon,
    broadcast_update=broadcast_update,
)
fastapi_app.include_router(service_payments_mod.service_payments_router)

# Phase 14 — Salon Inventory (browse + lifecycle + movement history)
import salon_inventory as salon_inventory_mod  # noqa: E402
salon_inventory_mod.init_salon_inventory_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
    resolve_branch_id=resolve_branch_id,
)
fastapi_app.include_router(salon_inventory_mod.salon_inventory_router)

# Reports Router (merged Financials + Analytics)
import reports_router as reports_router_mod  # noqa: E402
_reports_router = reports_router_mod.init_reports_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
    has_module_permission=has_module_permission,
    is_branch_manager=is_branch_manager,
    enforce_branch_for_manager=enforce_branch_for_manager,
    attribute_token_revenue_to_barbers=attribute_token_revenue_to_barbers,
    attribute_token_revenue_to_services=attribute_token_revenue_to_services,
)
fastapi_app.include_router(_reports_router)

# Phase 13 — wire auto-post hook into salon_store.supplier_deliver_order
salon_store_mod.set_auto_post_hook(
    lambda order: salon_inventory_mod.auto_post_on_delivery(order, db=db)
)

# Phase 15/16/17 — Customer in-salon shop (products + memberships unified)
import customer_shop as customer_shop_mod  # noqa: E402
customer_shop_mod.init_customer_shop_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
    create_in_app_notification=create_in_app_notification,
    resolve_branch_id=resolve_branch_id,
    send_low_stock_alert=None,
)
fastapi_app.include_router(customer_shop_mod.customer_shop_router)

# Phase 16 — wire cross-module notification hooks now that both are loaded.
salon_inventory_mod.set_notification_hooks(
    fire_customer_restock=customer_shop_mod.fire_restock_notifications_for_product,
    in_app_notifier=create_in_app_notification,
)

# Module 2 — Leave Tracker (leave types config + balance ledger + records)
import leave_tracker as leave_tracker_mod  # noqa: E402
leave_tracker_mod.init_leave_tracker_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
)
fastapi_app.include_router(leave_tracker_mod.leave_tracker_router)

# Module 4 — Attendance Mode (service_completion / geo_checkin toggle + check-in/out)
import attendance_mode as attendance_mode_mod  # noqa: E402
attendance_mode_mod.init_attendance_mode(
    db=db,
    get_current_salon_user=get_current_salon_user,
    get_current_salon_admin=get_current_salon_admin,
)
fastapi_app.include_router(attendance_mode_mod.attendance_mode_router)

# Marketing Module (SalonHub 2.0) — segments, coupons, channels, webhook, overview
import marketing as marketing_mod  # noqa: E402
marketing_mod.init_marketing_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
    get_current_salon_admin=get_current_salon_admin,
)
fastapi_app.include_router(marketing_mod.marketing_router)

# Marketing Settings Module (Twilio sub-account, Cashfree wallet, DLT, email, sending windows)
import salon_marketing_settings as mkt_settings_mod  # noqa: E402
mkt_settings_mod.init_marketing_settings_router(
    db=db,
    get_current_salon_user=get_current_salon_user,
    get_current_salon_admin=get_current_salon_admin,
)
fastapi_app.include_router(mkt_settings_mod.settings_router)

# Register marketing scheduler jobs (M6 automations daily + M5 scheduled campaigns every 5m)
try:
    marketing_mod.register_marketing_jobs(scheduler)
except Exception as _e:
    logger.error(f"[STARTUP] marketing scheduler registration failed: {_e}")

# Health check endpoint for Kubernetes liveness/readiness probes
@fastapi_app.get("/health")
async def health_check():
    """
    Health check endpoint for Kubernetes probes.
    Returns 200 OK if the application is healthy.
    """
    return {"status": "healthy", "service": "salon-backend"}

# Scheduler for token allocation
scheduler = AsyncIOScheduler()
scheduler.add_job(allocate_future_tokens, 'cron', hour=5, minute=30)  # Run at 5:30 AM daily
# Membership expiry reminders (once daily at 9:00 AM UTC)
scheduler.add_job(notify_expiring_memberships, 'cron', hour=9, minute=0)
# End-of-day cleanup at 00:00 IST = 18:30 UTC. Cancels all active (non-final) bookings
# for the day so they don't carry forward to the next day.
scheduler.add_job(cancel_active_bookings_end_of_day, 'cron', hour=18, minute=30)

# Module 2 — Leave Tracker scheduled jobs.
# Monthly accrual on the 1st of every month at 00:30 IST = 19:00 UTC on day-end-of-prev-month.
# We schedule on the 1st 00:30 IST → UTC = 1st 19:00 UTC of previous day; APScheduler doesn't
# directly support local-tz cron without pytz, so we keep IST math in the job (`run_monthly_accrual`
# is idempotent — uses `last_accrual_date`). Schedule at 19:00 UTC daily; it will only do work on
# the first IST day.
async def _leave_accrual_job_wrapper():
    today_ist = (datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)).date()
    if today_ist.day != 1:
        return
    try:
        await leave_tracker_mod.run_monthly_accrual(db=db)
    except Exception as e:
        logger.error(f"[scheduler] monthly accrual failed: {e}")


async def _leave_year_end_wrapper():
    today_ist = (datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)).date()
    if not (today_ist.month == 4 and today_ist.day == 1):
        return
    try:
        await leave_tracker_mod.run_year_end_close(db=db)
    except Exception as e:
        logger.error(f"[scheduler] year-end close failed: {e}")


# 00:30 IST == 19:00 UTC previous day.  Run daily — job no-ops outside the trigger date.
scheduler.add_job(_leave_accrual_job_wrapper, 'cron', hour=19, minute=0)
# 01:00 IST == 19:30 UTC previous day.
scheduler.add_job(_leave_year_end_wrapper, 'cron', hour=19, minute=30)

# Module 4 — Mode B auto-close open check-ins.  Runs at 18:25 UTC == 23:55 IST,
# right before the typical `auto_close_at = 23:59`.  Idempotent per the
# implementation in attendance_mode.py.
async def _attendance_auto_close_wrapper():
    try:
        await attendance_mode_mod.auto_close_open_checkins_job(db=db)
    except Exception as e:
        logger.error(f"[scheduler] attendance auto-close failed: {e}")

scheduler.add_job(_attendance_auto_close_wrapper, 'cron', hour=18, minute=25)

async def cleanup_legacy_predefined_services():
    """One-time cleanup for the July 2026 "reduce predefined services" change.
    
    Removes all GLOBAL (i.e. `salon_id` missing/null) predefined services whose
    category is not "General" from the `db.services` collection, and clears any
    salon_service mappings that pointed at them so the salon services list is
    not left with dangling references.
    
    Idempotent: the flag doc `db.system_migrations.cleanup_predefined_v1` stops
    re-runs.
    """
    marker = await db.system_migrations.find_one({"_id": "cleanup_predefined_v1"})
    if marker:
        return
    # 1) Delete global (unowned) services outside General
    stale = await db.services.find(
        {"$or": [{"salon_id": {"$exists": False}}, {"salon_id": None}], "category": {"$ne": "General"}},
        {"_id": 0, "id": 1},
    ).to_list(length=None)
    stale_ids = [s["id"] for s in stale if s.get("id")]
    if stale_ids:
        res_svc = await db.services.delete_many({"id": {"$in": stale_ids}})
        res_ss = await db.salon_services.delete_many({"service_id": {"$in": stale_ids}})
        res_bs = await db.barber_services.delete_many({"service_id": {"$in": stale_ids}})
        logger.info(
            f"[cleanup_predefined_v1] Removed {res_svc.deleted_count} legacy services, "
            f"{res_ss.deleted_count} salon_service links, {res_bs.deleted_count} barber_service links."
        )
    else:
        logger.info("[cleanup_predefined_v1] No legacy predefined services found.")
    await db.system_migrations.insert_one({"_id": "cleanup_predefined_v1", "at": datetime.now(timezone.utc).isoformat()})


@fastapi_app.on_event("startup")
async def startup_event():
    await initialize_data()
    await migrate_branches()
    # Jul 2026 — cleanup legacy predefined services (all non-General global services)
    try:
        await cleanup_legacy_predefined_services()
    except Exception as e:
        logger.error(f"[STARTUP] cleanup_legacy_predefined_services failed: {e}")
    # Phase 2 (Part C) — per-branch pricing migration (idempotent)
    try:
        await migrate_subscription_pricing_v2()
    except Exception as e:
        logger.error(f"[STARTUP] subscription_v2 migration failed: {e}")
    # Phase 1 (Part A) — bootstrap platform owner from PLATFORM_OWNER_MOBILE
    try:
        await platform_admin_mod.bootstrap_platform_owner()
    except Exception as e:
        logger.error(f"[STARTUP] platform owner bootstrap failed: {e}")
    # Phase 9 — seed the 30 permanent supplier product samples (idempotent)
    try:
        from data.product_samples_seed import seed_supplier_product_samples
        n = await seed_supplier_product_samples(db)
        logger.info(f"[STARTUP] Seeded {n} supplier product samples")
    except Exception as e:
        logger.error(f"[STARTUP] product samples seed failed: {e}")
    # Also seed a minimal set of live supplier_products so the salon-store
    # Shop tab has real product cards to browse (idempotent).
    try:
        from seed_store_fixtures import SUPPLIER_FIXTURES, _now_iso
        from passlib.context import CryptContext as _CC
        _pwd = _CC(schemes=["bcrypt"], deprecated="auto")
        seeded_p = 0
        for sup in SUPPLIER_FIXTURES:
            await db.suppliers.update_one(
                {"id": sup["id"]},
                {"$set": {
                    "id": sup["id"],
                    "business_name": sup["business_name"],
                    "owner_name": sup["owner_name"],
                    "phone": sup["phone"],
                    "mobile": sup["phone"],
                    "email": sup["email"],
                    "city": sup["city"],
                    "state": sup["state"],
                    "country": "India",
                    "rating_avg": sup["rating_avg"],
                    "rating_count": sup["rating_count"],
                    "status": "active",
                    "approved_at": _now_iso(),
                    "password_hash": _pwd.hash("supplier123"),
                    "login_id": sup["phone"],
                    "created_at": _now_iso(),
                    "updated_at": _now_iso(),
                }},
                upsert=True,
            )
            for p in sup["products"]:
                pid = f"{sup['id']}::{p['name']}".replace(" ", "_")[:80]
                await db.supplier_products.update_one(
                    {"id": pid},
                    {"$set": {
                        "id": pid,
                        "supplier_id": sup["id"],
                        "name": p["name"],
                        "brand": p["brand"],
                        "category": p["category"],
                        "description": f"Sample product from {sup['business_name']} — {p['name']}.",
                        "images": [],
                        "selling_price": p["selling_price"],
                        "mrp": p["mrp"],
                        "gst_percent": p["gst_percent"],
                        "inventory_available": p["inventory_available"],
                        "inventory_reserved": 0,
                        "min_order_qty": 1,
                        "pack_size": None,
                        "unit": p["unit"],
                        "is_active": True,
                        "is_deleted": False,
                        "created_at": _now_iso(),
                        "updated_at": _now_iso(),
                    }},
                    upsert=True,
                )
                seeded_p += 1
        logger.info(f"[STARTUP] Seeded {seeded_p} live supplier products for Shop")
    except Exception as e:
        logger.error(f"[STARTUP] live supplier products seed failed: {e}")
    scheduler.start()
    # Marketing M0 — one-time migration to add can_access_marketing=true to
    # admin/branch-manager records. We overwrite existing values (default:
    # admins/BM = True) because the field was added AFTER their records were
    # created — the previous default was implicitly "true" via role-based
    # checks. Staff records keep their explicit value if already set.
    try:
        r_admin = await db.salon_users.update_many(
            {"role": {"$in": ["salon_admin", "admin", "platform_admin", "branch_manager"]}},
            {"$set": {"permissions.can_access_marketing": True}},
        )
        r_staff = await db.salon_users.update_many(
            {"role": "staff", "permissions.can_access_marketing": {"$exists": False}},
            {"$set": {"permissions.can_access_marketing": False}},
        )
        logger.info(
            f"[STARTUP] can_access_marketing migration: admins/managers={r_admin.modified_count} staff={r_staff.modified_count}"
        )
    except Exception as e:
        logger.error(f"[STARTUP] can_access_marketing migration failed: {e}")
    # Phase 10/11 — start the salon-store reservation sweeper
    try:
        salon_store_mod.start_sweeper()
    except Exception as e:
        logger.error(f"[STARTUP] salon_store sweeper failed: {e}")
    # Jul 2026 (continuation) — auto-seed a few services, active barbers and
    # 3 sample bookings for the admin salon so testers can immediately
    # explore the UI. Safe & idempotent (see seed_test_data.py).
    try:
        from seed_test_data import main as seed_test_data_main
        await seed_test_data_main()
    except Exception as e:
        logger.error(f"[STARTUP] seed_test_data failed: {e}")
    logger.info("Application started with multi-salon support")

@fastapi_app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()

# Note: Socket.IO is already integrated via socketio.ASGIApp wrapping
# at line 63: socket_app = socketio.ASGIApp(sio, fastapi_app)
# The main app export is socket_app which wraps fastapi_app
